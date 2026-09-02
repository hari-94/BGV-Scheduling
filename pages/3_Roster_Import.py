"""
Roster Import — the weekly staff Schedule.xlsx, stored, diffed and editable.

Upload the workbook weekly; the page stores each week, shows what changed since
last time, renders any week as an Excel-style grid, lets you correct cells in
the app, and exports a workbook with those corrections written back in.
"""
import streamlit as st
import sys, os, io, re, json, datetime
import html as _html
import pandas as pd
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
import auth, db, clock, roster_import as ri
import staffing
import ui

# After a deploy the server can still hold the previous roster_import in
# sys.modules while serving the new page file. The first call to a helper added
# in this release then raises AttributeError from inside a widget callback, and
# Streamlit redacts the message. Reload from disk instead of failing.
if getattr(ri, "__version__", 0) < 14:
    import importlib
    ri = importlib.reload(ri)

st.set_page_config(page_title="Roster Import · Cleaning Schedule",
                   page_icon="GC8", layout="wide")

st.markdown("""<style>
[data-testid="stSidebarNav"]{display:none !important;}
</style>""", unsafe_allow_html=True)

# One guard for every page: it restores a signed-in browser from its cookie
# and, failing that, goes to the sign-in screen. The hand-rolled version that
# used to sit here seeded logged_in=False into the session before init_auth
# ran, which made the restore skip itself -- so a refresh on this page could
# never have brought anybody back, cookie or no cookie.
auth.require_login()
# Admin and RQS both run schedules day to day, so both get this page.
# can_generate is exactly that distinction: true for admin and rqs, false for
# housekeepers (who get My Home instead).
if not auth.can("can_generate"):
    st.error("This page is for admins and RQS. Your own schedule is on **My Home**.")
    st.stop()
IS_ADMIN = auth.can("can_manage_users")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Mono:wght@300;400;500&family=DM+Sans:wght@300;400;500;600&display=swap');
:root{--bg:#f4f5f7;--bg2:#fff;--border:#e2e5ea;--border-hi:#c3c9d4;
  --indigo:#2563a8;--txt:#1f2733;--txt2:#5b6675;--txt3:#8a93a1;--radius:14px;--radius-sm:8px;}
*{box-sizing:border-box;}
html,body,[class*="css"]{font-family:'DM Sans',sans-serif!important;color:var(--txt)!important;}
.stApp{background:#f4f5f7!important;}
.block-container{padding-top:1.2rem!important;max-width:1400px;}
.pg-title{font-family:'Syne',sans-serif!important;font-size:1.6rem;font-weight:700;
  letter-spacing:-.03em;color:#16202e;margin:0 0 2px}
.pg-sub{font-size:.82rem;color:var(--txt2);margin:0 0 .8rem}
.sec{font-family:'DM Mono',monospace!important;font-size:.6rem;font-weight:500;
  text-transform:uppercase;letter-spacing:.16em;color:var(--txt2);
  padding-bottom:6px;border-bottom:1px solid var(--border);margin:1.2rem 0 .6rem}
.stamp{display:flex;flex-wrap:wrap;align-items:center;gap:10px 22px;background:#fff;
  border:1px solid var(--border);border-left:3px solid var(--indigo);border-radius:10px;
  padding:11px 16px;margin-bottom:14px}
.stamp .k{font-family:'DM Mono',monospace;font-size:.58rem;text-transform:uppercase;
  letter-spacing:.12em;color:var(--txt3);display:block;margin-bottom:2px}
.stamp .v{font-size:.85rem;font-weight:600;color:#16202e}
.pill{display:inline-block;border-radius:6px;padding:1px 8px;font-size:.68rem;font-weight:700}
.mono{font-family:'DM Mono',monospace;font-size:.72rem;color:var(--txt3)}
.gridwrap{overflow-x:auto;border:1px solid var(--border);border-radius:10px;background:#fff}
table.wk{border-collapse:collapse;width:100%;font-size:.76rem}
table.wk th{position:sticky;top:0;background:#f8fafc;padding:7px 9px;text-align:left;
  font-family:'DM Mono',monospace;font-size:.6rem;text-transform:uppercase;letter-spacing:.09em;
  color:var(--txt2);border-bottom:1px solid var(--border);white-space:nowrap}
table.wk td{padding:5px 9px;border-bottom:1px solid #eef1f5;vertical-align:top}
table.wk td.nm{font-weight:600;color:#16202e;white-space:nowrap;position:sticky;left:0;background:#fff}
table.wk tr.grp td{background:#eef2f7;font-family:'DM Mono',monospace;font-size:.6rem;
  text-transform:uppercase;letter-spacing:.11em;color:var(--txt2);font-weight:600}
.cell{display:block;border-radius:5px;padding:2px 6px;font-size:.72rem;line-height:1.3}
.chg{outline:2px solid #f59e0b;outline-offset:1px}
.ovr{box-shadow:inset 3px 0 0 #7c3aed}
.cov{box-shadow:inset 0 -2px 0 #7c3aed}
.nocall{box-shadow:inset 3px 0 0 #b91c1c;font-weight:700}
table.wk tbody tr:hover td{background:#f7f9fc}
table.wk tbody tr:hover td.nm{background:#eef2f7}
.cell{transition:transform .1s ease}
.cell:hover{transform:translateY(-1px)}
/* Tab styling lives in ui.py — one source for every page. */
.stamp{animation:fadeUp .35s cubic-bezier(.16,1,.3,1) both}
@keyframes fadeUp{from{opacity:0;transform:translateY(5px)}to{opacity:1;transform:none}}
[data-testid="stMetricValue"]{font-family:'Syne',sans-serif!important;font-weight:700!important}
footer{visibility:hidden!important;}#MainMenu{visibility:hidden!important;}
header[data-testid="stHeader"]{background:transparent!important;height:0!important;}
section[data-testid="stSidebar"]{background:#fff!important;border-right:1px solid var(--border)!important;}
.stButton>button{border-radius:var(--radius-sm)!important;font-weight:600!important;background:#fff!important;
  color:var(--txt)!important;border:1px solid var(--border-hi)!important;}
.stButton>button:hover{background:#2563a8!important;color:#fff!important;border-color:#2563a8!important;}
.stButton>button[kind="primary"]{background:#2563a8!important;border:none!important;color:#fff!important;}
@media (max-width:768px){
  .block-container{padding-left:.4rem!important;padding-right:.4rem!important;}
  .pg-title{font-size:1.3rem!important;}
  [data-testid="stHorizontalBlock"]{flex-direction:column!important;}
  [data-testid="stHorizontalBlock"] > [data-testid="column"]{width:100%!important;min-width:100%!important;}
}
</style>""", unsafe_allow_html=True)

ui.topnav("Roster Import")
def e(s): return _html.escape(str(s) if s is not None else "")

def person_label(info):
    """Name for a picker: the person, their team, and their home property if
    they are visiting from one."""
    txt = f'{info["label"]}  ·  {info["sections"][0]}'
    return txt + f'  ·  {info["home"]}' if info.get("home") else txt


# Brighter, more saturated so the states carry at a glance in a dense grid.
# Foregrounds are the darkened form of each hue, which keeps the text legible
# on top of it. "Off" stays muted on purpose — it is the absence of a state.
KIND_STYLE = {
    ri.KIND_WORKING: ("#a7f0bf", "#08532c"),   # green
    ri.KIND_DAILY:   ("#93ebe5", "#04504b"),   # teal
    ri.KIND_OTHER:   ("#ffdc73", "#6f3d00"),   # amber
    ri.KIND_VTO:     ("#b9c2ff", "#1d24a3"),   # indigo
    ri.KIND_NOCALL:  ("#ff8f8f", "#6d0b0b"),   # red
    ri.KIND_OFF:     ("#e9eef4", "#7d8896"),   # grey
    ri.KIND_UNKNOWN: ("#ddb3ff", "#4c1580"),   # violet
}
KIND_LABEL = {ri.KIND_DAILY: "Daily Service", ri.KIND_WORKING: "Working",
              ri.KIND_OFF: "Off", ri.KIND_OTHER: "Other duty (counts as worked)",
              ri.KIND_NOCALL: "No call / no show", ri.KIND_VTO: "VTO — paid",
              ri.KIND_UNKNOWN: "Unrecognised"}
#: Must match the Schedule page's RQS selectbox "nobody" option exactly.
RQS_NONE = "— none —"

XLSX_MIME = ("application/vnd.openxmlformats-officedocument"
             ".spreadsheetml.sheet")

def week_download(week, wk_key, key_prefix):
    """Offer this week as an .xlsx in the workbook's own layout."""
    raw = st.session_state.get("ri_raw_bytes")
    if not raw:
        raw, _info = db.load_staff_file()
    if not raw:
        st.caption("Upload the workbook once on **Upload & sync** and the "
                   "Excel download appears here.")
        return
    title = ri.week_range_text(wk_key, with_year=False)
    if st.button("Build Excel file", key=f"{key_prefix}_build"):
        try:
            data, sheet = ri.export_week_xlsx(raw, week, title=title)
            st.session_state[f"{key_prefix}_xlsx"] = (data, sheet)
        except Exception as ex:
            st.error(f"Could not build the file: {ex}")
    got = st.session_state.get(f"{key_prefix}_xlsx")
    if got:
        st.download_button(
            f"Download {got[1]}.xlsx", data=got[0],
            file_name=f"Schedule_{wk_key}.xlsx", mime=XLSX_MIME,
            key=f"{key_prefix}_dl")
        st.caption("Same layout as the source workbook — section headers, "
                   "formulas and the red no-call marks all intact.")

def human_ago(iso):
    try:
        then = datetime.datetime.fromisoformat(str(iso))
    except Exception:
        return str(iso or "—")
    now = datetime.datetime.now(then.tzinfo) if then.tzinfo else datetime.datetime.now()
    secs = (now - then).total_seconds()
    if secs < 90:            return "just now"
    if secs < 5400:          return f"{int(secs//60)} min ago"
    if secs < 172800:        return f"{int(secs//3600)} hours ago"
    return f"{int(secs//86400)} days ago"

# ══════════════════════════════════════════════════════════════════════════════
#  HEADER — last-loaded stamp
# ══════════════════════════════════════════════════════════════════════════════
if not db.settings_table_ready():
    st.markdown('<p class="pg-title">Roster Import</p>', unsafe_allow_html=True)
    st.error(
        "**One-time database setup needed.**\n\n"
        "This page stores the staff schedule under text keys, but the "
        "`app_settings` table does not exist yet, so nothing can be saved.\n\n"
        "Open **Supabase → SQL Editor**, paste the contents of "
        "`migration_app_settings.sql` from the repo, and run it. Then reload "
        "this page.")
    with st.expander("Show the SQL"):
        try:
            _sql = open(os.path.join(os.path.dirname(os.path.dirname(__file__)),
                                     "migration_app_settings.sql"), encoding="utf-8").read()
        except Exception:
            _sql = ("CREATE TABLE IF NOT EXISTS app_settings (\n"
                    "  key        TEXT        PRIMARY KEY,\n"
                    "  payload    JSONB       NOT NULL,\n"
                    "  updated_at TIMESTAMPTZ DEFAULT now()\n);")
        st.code(_sql, language="sql")
    st.caption("Note: the standing roster (add/remove staff, building moves) has "
               "never persisted either — it relies on the same missing table. "
               "Running this SQL fixes both.")
    st.stop()

meta = db.load_staff_meta() or {}
overrides = db.load_staff_overrides()
custom_opts = db.load_custom_options()
week_keys = db.staff_week_keys()

# ── Backfill the red no-call marks ────────────────────────────────────────────
# Weeks stored before cell fills were read carry none. The workbook itself is
# in the database, so re-derive them rather than asking for another upload.
# Runs once; the marker stops it repeating for everyone who opens the page.
def _backfill_fills():
    info = db.staff_file_info()
    if not info.get("size") or not week_keys:
        return None
    stamp = str(info.get("saved_at", "")) + "|" + str(len(week_keys))
    if (db.load_autoapply() or {}).get("fills_backfilled") == stamp:
        return None
    sample = db.load_staff_week(week_keys[-1])
    if not ri.needs_fill_backfill(sample):
        return None
    with st.spinner("Reading the red no-call marks from the stored workbook…"):
        raw, _i = db.load_staff_file()
        if not raw:
            return None
        import openpyxl
        parsed = ri.parse_all_weeks(
            openpyxl.load_workbook(io.BytesIO(raw), data_only=True),
            openpyxl.load_workbook(io.BytesIO(raw), data_only=False))
        total = weeks_done = 0
        for wk in week_keys:
            stored = db.load_staff_week(wk)
            if not ri.needs_fill_backfill(stored) or wk not in parsed:
                continue
            updated, n = ri.merge_fill_data(stored, parsed[wk])
            if updated is not None:
                db.save_staff_week(wk, updated)
                total += n; weeks_done += 1
        marker = dict(db.load_autoapply() or {})
        marker["fills_backfilled"] = stamp
        db.save_autoapply(marker)
        st.cache_data.clear()
    return total, weeks_done

try:
    _bf = _backfill_fills()
except Exception as _ex:
    _bf = None
    st.warning(f"Could not read the no-call marks from the stored workbook: {_ex}")
if _bf:
    st.success(f"Picked up **{_bf[0]}** no-call/no-show marks across "
               f"**{_bf[1]}** stored weeks, straight from the saved workbook — "
               f"no re-upload needed.")

st.markdown('<p class="pg-title">Roster Import</p>', unsafe_allow_html=True)
st.markdown('<p class="pg-sub">The weekly staff schedule — stored, compared week to week, '
            'and editable here.</p>', unsafe_allow_html=True)

if meta.get("uploaded_at"):
    when = str(meta["uploaded_at"])[:16].replace("T", " ")
    fields = [
        ("Schedule last loaded", f'{when} <span style="color:#8a93a1;font-weight:400">'
                                 f'({human_ago(meta["uploaded_at"])})</span>'),
        ("By", e(meta.get("uploaded_by", "—"))),
        ("File", e(meta.get("file_name", "—"))),
        ("Weeks stored", f'{len(week_keys)}'),
        ("Covering", f'{e(meta.get("date_min","—"))} → {e(meta.get("date_max","—"))}'),
    ]
    if overrides:
        fields.append(("In-app edits", f'<span style="color:#7c3aed">{len(overrides)}</span>'))
    _fi = db.staff_file_info()
    fields.append(("Workbook stored",
                   f'<span style="color:#15803d">yes · {_fi["size"]/1024/1024:.1f} MB</span>'
                   if _fi.get("size") else '<span style="color:#b45309">no</span>'))
    _aa = db.load_autoapply() or {}
    fields.append(("Today auto-loaded",
                   f'<span style="color:#15803d">{e(_aa.get("date",""))}</span>'
                   if _aa.get("date") else '<span style="color:#8a93a1">not yet</span>'))
    st.markdown('<div class="stamp">' + "".join(
        f'<div><span class="k">{k}</span><span class="v">{v}</span></div>' for k, v in fields
    ) + '</div>', unsafe_allow_html=True)
else:
    st.markdown('<div class="stamp" style="border-left-color:#f59e0b">'
                '<div><span class="k">Schedule last loaded</span>'
                '<span class="v">Never — upload the workbook below</span></div></div>',
                unsafe_allow_html=True)

# Ordered by how often each is actually used: the week grid is a daily glance,
# attendance drives time-off decisions, applying is occasional now that it
# happens automatically, and uploading is weekly.
@st.cache_data(ttl=300, show_spinner="Reading attendance history…")
def _history(_token: str):
    """Flatten every stored week once. Keyed on a token so it refreshes when
    the workbook or the in-app edits change."""
    weeks = db.load_staff_weeks()
    return ri.history_rows(weeks, db.load_staff_overrides())

(tab_week, tab_month, tab_plan, tab_att, tab_apply, tab_changes,
 tab_sync) = st.tabs(
    ["Week view", "Month view", "Plan a week", "Attendance", "Apply to roster",
     "What changed", "Upload & sync"])

# ══════════════════════════════════════════════════════════════════════════════
#  UPLOAD & SYNC
# ══════════════════════════════════════════════════════════════════════════════
with tab_sync:
    st.markdown('<p class="sec">Upload the weekly workbook</p>', unsafe_allow_html=True)
    up = st.file_uploader("Schedule.xlsx", type=["xlsx", "xlsm"], key="ri_xlsx",
                          help="One sheet per week, Sunday–Saturday in columns B–H.")
    if up is not None:
        st.session_state["ri_raw_bytes"] = up.getvalue()
        st.session_state["ri_file_name"] = up.name

    raw = st.session_state.get("ri_raw_bytes")
    if not raw:
        st.info("Upload the workbook to compare it against what is stored. "
                "Nothing is written until you press **Save**.")
    else:
        @st.cache_data(show_spinner="Reading workbook…", max_entries=2)
        def _parse(raw_bytes: bytes):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
            # Second open without data_only: openpyxl gives cached values OR
            # styles, never both, and the red no-call fills are styles.
            try:
                wbs = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=False)
            except Exception:
                wbs = None
            return ri.parse_all_weeks(wb, wbs), len(wb.worksheets)

        try:
            incoming, n_sheets = _parse(raw)
        except Exception as ex:
            st.error(f"Could not read that workbook: {ex}")
            st.stop()
        if not incoming:
            st.error("No dated sheets found — is this the weekly schedule workbook?")
            st.stop()

        stored = db.load_staff_weeks()
        d = ri.diff_all(stored, incoming)
        st.success(f"Read **{n_sheets}** sheets · **{len(incoming)}** dated weeks "
                   f"from `{e(st.session_state.get('ri_file_name',''))}`")

        c = st.columns(4)
        c[0].metric("New weeks", len(d["new_weeks"]))
        c[1].metric("Weeks changed", len(d["changed_weeks"]))
        c[2].metric("Cells changed", d["n_changed_cells"])
        c[3].metric("Already stored", len(stored))

        if d["new_weeks"]:
            st.markdown(f'<div style="background:#ecfdf5;border:1px solid #a7f3d0;border-radius:8px;'
                        f'padding:9px 13px;font-size:.79rem;color:#065f46">'
                        f'<b>New week(s):</b> {e(", ".join(ri.week_label(w) for w in d["new_weeks"]))}</div>',
                        unsafe_allow_html=True)
        if d["gone_weeks"]:
            st.markdown(f'<div style="background:#fef2f2;border:1px solid #fecaca;border-radius:8px;'
                        f'padding:9px 13px;font-size:.79rem;color:#991b1b;margin-top:6px">'
                        f'<b>Stored but not in this file:</b> {e(", ".join(ri.week_range_text(w) for w in d["gone_weeks"]))} — '
                        f'these are kept, not deleted.</div>', unsafe_allow_html=True)

        if d["changed_weeks"]:
            st.markdown('<p class="sec">Changed cells</p>', unsafe_allow_html=True)
            rows = []
            for wk, dd in sorted(d["changed_weeks"].items()):
                for ch in dd["changed"]:
                    rows.append({"Week": ri.week_range_text(wk), "Person": ch["name"], "Date": ch["date"],
                                 "Was": ch["old"] or "—", "Now": ch["new"] or "—"})
                for n in dd["added_people"]:
                    rows.append({"Week": ri.week_range_text(wk), "Person": n, "Date": "—",
                                 "Was": "not on sheet", "Now": "added"})
                for n in dd["removed_people"]:
                    rows.append({"Week": ri.week_range_text(wk), "Person": n, "Date": "—",
                                 "Was": "on sheet", "Now": "removed"})
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True,
                         height=min(60 + 35 * len(rows), 420))

        clash = [k for k in overrides
                 if k.split("|", 2)[0] in d["changed_weeks"] and any(
                     ch["name"] == k.split("|", 2)[1] and ch["date"] == k.split("|", 2)[2]
                     for ch in d["changed_weeks"][k.split("|", 2)[0]]["changed"])]
        if clash:
            st.warning(f"**{len(clash)}** of your in-app edits sit on cells that also changed "
                       f"in this file. Your edits win and are kept — the Week view shows the "
                       f"Excel value underneath each one so you can revert if you'd rather.")

        if not d["new_weeks"] and not d["changed_weeks"]:
            st.info("Nothing new — the stored copy already matches this workbook.")

        if st.button("Save to app", type="primary", key="ri_save"):
            touched = set(d["new_weeks"]) | set(d["changed_weeks"])
            with st.spinner(f"Saving {len(touched) or len(incoming)} week(s)…"):
                to_write = touched if stored else set(incoming)
                ok, failed = 0, []
                for wk in sorted(to_write):
                    try:
                        db.save_staff_week(wk, incoming[wk]); ok += 1
                    except Exception as ex:
                        failed.append(f"{wk}: {ex}")
                all_dates = sorted(x for w in incoming.values() for x in w["dates"])
                try:
                    db.save_staff_meta({
                        "uploaded_at": datetime.datetime.now().isoformat(timespec="seconds"),
                        "uploaded_by": st.session_state.get("username", "unknown"),
                        "file_name":   st.session_state.get("ri_file_name", ""),
                        "n_sheets":    n_sheets,
                        "n_weeks":     len(incoming),
                        "date_min":    all_dates[0] if all_dates else "",
                        "date_max":    all_dates[-1] if all_dates else "",
                        "last_diff": {
                            "new_weeks": d["new_weeks"],
                            "n_changed_cells": d["n_changed_cells"],
                            # Bounded so the meta row stays small.
                            "changed": [
                                {"week": wk, **ch}
                                for wk, dd in sorted(d["changed_weeks"].items())
                                for ch in dd["changed"]][:500],
                        },
                    })
                except Exception as ex:
                    failed.append(f"meta: {ex}")
                # Keep the workbook itself so the Excel export works any day,
                # not only in a session where someone happened to upload it.
                try:
                    db.save_staff_file(raw, st.session_state.get("ri_file_name", ""))
                except Exception as ex:
                    failed.append(f"workbook: {ex}")
                # A fresh upload supersedes today's auto-apply, so let it re-run.
                try:
                    db.save_autoapply({})
                except Exception:
                    pass
            if failed:
                st.error("Some writes failed:\n\n" + "\n\n".join(f"- {f}" for f in failed))
            else:
                st.success(f"Saved {ok} week(s). Timestamp updated.")
                st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
#  WEEK VIEW — Excel-style grid + editing
# ══════════════════════════════════════════════════════════════════════════════
with tab_week:
    if not week_keys:
        st.info("No weeks stored yet. Upload the workbook on the **Upload & sync** tab.")
    else:
        today_iso = clock.today().isoformat()
        default_wk = max([w for w in week_keys if w <= today_iso], default=week_keys[-1])
        # Newest first: the weeks people actually open sit at the top instead of
        # at the bottom of a two-year list.
        wk_order = sorted(week_keys, reverse=True)
        cc = st.columns([2, 2, 3])
        with cc[0]:
            sel_wk = st.selectbox("Week", wk_order,
                                  index=wk_order.index(default_wk), key="ri_wk",
                                  format_func=lambda w: ri.week_label(w))
        week = db.load_staff_week(sel_wk)
        if not week:
            st.error(f"Could not load week {sel_wk}.")
            st.stop()
        eff, applied = ri.apply_overrides(week, overrides, sel_wk)
        last_diff = (meta.get("last_diff") or {})
        changed_set = {(ch["name"], ch["date"]) for ch in last_diff.get("changed", [])
                       if ch.get("week") == sel_wk}
        with cc[1]:
            groups = ["All", "Housekeepers", "RQS", "Other teams"]
            sel_grp = st.selectbox("Show", groups, key="ri_grp")
        with cc[2]:
            st.markdown(
                f'<div style="padding-top:26px;font-size:.76rem;color:#5b6675">'
                f'Sheet <b>{e(week["sheet"])}</b> &nbsp;·&nbsp; {len(eff["people"])} people'
                f'{" &nbsp;·&nbsp; <span style=\"color:#7c3aed\">" + str(len(applied)) + " edited here</span>" if applied else ""}'
                f'{" &nbsp;·&nbsp; <span style=\"color:#b45309\">" + str(len(changed_set)) + " changed at last sync</span>" if changed_set else ""}'
                f'</div>', unsafe_allow_html=True)

        def wanted(rec):
            if sel_grp == "All": return True
            if sel_grp == "Housekeepers": return rec["group"] == "hk"
            if sel_grp == "RQS": return rec["group"] == "rqs"
            return rec["group"] == "other"

        st.markdown(
            '<div style="display:flex;gap:14px;flex-wrap:wrap;font-size:.7rem;color:#5b6675;margin:.5rem 0">'
            # A hairline keeps the palest swatches from vanishing on white.
            + "".join(f'<span><span class="pill" style="background:{b};color:{f};'
                      f'border:1px solid rgba(22,32,46,.12)">&nbsp;&nbsp;</span> '
                      f'{KIND_LABEL[k]}</span>'
                      for k, (b, f) in KIND_STYLE.items())
            + '<span><span class="pill" style="background:#fff;'
              'outline:2px solid #f59e0b">&nbsp;&nbsp;</span> changed at last sync</span>'
              '<span><span class="pill" style="background:#fff;'
              'border:1px solid rgba(22,32,46,.12);box-shadow:inset 3px 0 0 #7c3aed">'
              '&nbsp;&nbsp;</span> edited in app</span>'
            + '</div>', unsafe_allow_html=True)

        with st.expander("Download this week as Excel"):
            week_download(eff, sel_wk, f"wkdl_{sel_wk}")

        with st.expander("What the numbers mean (from the workbook's own legend)"):
            lc = st.columns(2)
            with lc[0]:
                st.markdown('<div class="mono">HOUSEKEEPER — 8AM TO 10AM EXTRA TASK</div>',
                            unsafe_allow_html=True)
                st.markdown("".join(
                    f'<div style="font-size:.79rem;margin:3px 0"><b>{k}</b> &nbsp;{e(v)}</div>'
                    for k, v in ri.HK_TASK_LEGEND.items()), unsafe_allow_html=True)
                st.caption("They clean rooms the rest of the day — the number never "
                           "means unavailable.")
            with lc[1]:
                st.markdown('<div class="mono">HOUSEPERSON — ZONE</div>', unsafe_allow_html=True)
                st.markdown("".join(
                    f'<div style="font-size:.79rem;margin:3px 0"><b>{k}</b> &nbsp;{e(v)}</div>'
                    for k, v in ri.HP_ZONE_LEGEND.items()), unsafe_allow_html=True)

        # ── grid ──────────────────────────────────────────────────────────────
        dates = eff["dates"]
        head = "".join(
            f'<th>{datetime.date.fromisoformat(dt).strftime("%a")}<br>'
            f'<span style="font-weight:400;color:#8a93a1">{dt[5:]}</span></th>' for dt in dates)
        body, current = [], None
        for name, rec in eff["people"].items():
            if not wanted(rec): continue
            if rec["section"] != current:
                current = rec["section"]
                body.append(f'<tr class="grp"><td colspan="{len(dates)+1}">{e(current)}</td></tr>')
            tds = []
            for dt in dates:
                raw = rec["cells"].get(dt, "")
                kind = ri.cell_kind(rec, dt)[0]
                bg, fg = KIND_STYLE[kind]
                cls = "cell"
                if (name, dt) in changed_set: cls += " chg"
                tips = []
                if kind == ri.KIND_NOCALL:
                    cls += " nocall"
                    tips.append("No call / no show")
                means = ri.legend_for(rec, raw)
                if means: tips.append(means)
                if rec["group"] != "hk" and ri.is_room_cover(raw):
                    cls += " cov"; tips.append("On guest rooms today")
                if (name, dt) in applied:
                    cls += " ovr"
                    tips.append(f'Edited in app · Excel says: '
                                f'{applied[(name, dt)]["excel"] or "(blank)"}')
                tip = f' title="{e(" — ".join(tips))}"' if tips else ""
                # Most red cells are blank, so give them a label to show.
                shown = e(raw) or ("NC / NS" if kind == ri.KIND_NOCALL else "&nbsp;")
                tds.append(f'<td><span class="{cls}" style="background:{bg};color:{fg}"{tip}>'
                           f'{shown}</span></td>')
            body.append(f'<tr><td class="nm">{e(name)}</td>{"".join(tds)}</tr>')
        st.markdown(f'<div class="gridwrap"><table class="wk"><thead><tr><th>Name</th>{head}</tr>'
                    f'</thead><tbody>{"".join(body)}</tbody></table></div>',
                    unsafe_allow_html=True)

        # ── editing ───────────────────────────────────────────────────────────
        st.markdown('<p class="sec">Edit this week</p>', unsafe_allow_html=True)
        st.caption("Every cell is a dropdown, with the choices that fit that role. "
                   "Pick and press Save. Edits are kept as overrides — they survive "
                   "re-uploading the workbook, and the grid marks them in purple.")
        cols_lbl = [f'{datetime.date.fromisoformat(dt).strftime("%a")} {dt[5:]}' for dt in dates]

        def to_cell(v):     return v if str(v or "").strip() else ri.BLANK_LABEL
        def from_cell(v):   return "" if str(v) == ri.BLANK_LABEL else str(v or "").strip()

        # One editor per section: st.data_editor applies column_config per COLUMN,
        # so a single grid could not give housekeepers and housepersons different
        # choices. Splitting by section is what makes the options role-specific.
        sections, order = {}, []
        for name, rec in eff["people"].items():
            if not wanted(rec): continue
            sections.setdefault(rec["section"], []).append((name, rec))
            if rec["section"] not in order: order.append(rec["section"])

        editors = []
        for section in order:
            members = sections[section]
            # Union of the curated list and everything this section already uses,
            # so no existing one-off vanishes when its cell is opened.
            present = {rec["cells"].get(dt, "") for _n, rec in members for dt in dates}
            rkey = ri.role_key(members[0][1])
            opts = ri.options_for(members[0][1], sorted(v for v in present if v),
                                  extra=custom_opts.get(rkey, []))
            rows = [{"Name": n, **{cols_lbl[i]: to_cell(rec["cells"].get(dt, ""))
                                   for i, dt in enumerate(dates)}}
                    for n, rec in members]
            base = pd.DataFrame(rows, columns=["Name"] + cols_lbl)
            st.markdown(f'<div class="sec" style="margin:.9rem 0 .3rem">{e(section)} '
                        f'<span style="color:#8a93a1;text-transform:none;letter-spacing:0">'
                        f'· {len(opts)-1} choices</span></div>', unsafe_allow_html=True)
            with st.popover(f"＋ Add a choice for {section}"):
                st.caption("Type anything the dropdown does not offer. It is added to "
                           "this role's list and stays available from then on.")
                nc1, nc2 = st.columns([3, 1])
                newval = nc1.text_input("New choice", key=f"ri_new_{sel_wk}_{section}",
                                        label_visibility="collapsed",
                                        placeholder="e.g. TOUCH UP, Deep clean lobby")
                if nc2.button("Add", key=f"ri_addbtn_{sel_wk}_{section}"):
                    v = " ".join(str(newval or "").split())
                    if not v:
                        st.warning("Type a value first.")
                    elif v.casefold() in {o.casefold() for o in opts}:
                        st.info(f"'{v}' is already a choice.")
                    else:
                        merged = dict(custom_opts)
                        merged[rkey] = sorted(set(merged.get(rkey, []) + [v]))
                        try:
                            db.save_custom_options(merged)
                            st.success(f"Added '{v}'.")
                            st.rerun()
                        except Exception as ex:
                            st.error(f"Could not save: {ex}")
                if custom_opts.get(rkey):
                    st.caption("Yours: " + ", ".join(custom_opts[rkey]))
            ed = st.data_editor(
                base, hide_index=True, use_container_width=True,
                key=f"ri_ed_{sel_wk}_{section}", num_rows="fixed",
                height=min(90 + 35 * len(base), 460),
                column_config={
                    "Name": st.column_config.TextColumn("Name", disabled=True, width="medium"),
                    **{c: st.column_config.SelectboxColumn(c, options=opts, required=False,
                                                           width="medium") for c in cols_lbl},
                })
            editors.append(ed)

        b1, b2, _ = st.columns([1, 1, 3])
        with b1:
            if st.button("Save edits", type="primary", key="ri_save_ed"):
                new_ov = dict(overrides)
                added = removed = 0
                for ed in editors:
                    for _, row in ed.iterrows():
                        name = row["Name"]
                        src = week["people"].get(name, {}).get("cells", {})
                        for i, dt in enumerate(dates):
                            val = from_cell(row[cols_lbl[i]])
                            excel_val = str(src.get(dt, "") or "")
                            key = ri.override_key(sel_wk, name, dt)
                            if val == excel_val:
                                # Back to what the workbook says — drop the override
                                # instead of storing a no-op.
                                if key in new_ov: del new_ov[key]; removed += 1
                            elif new_ov.get(key, {}).get("value", None) != val:
                                new_ov[key] = {
                                    "value": val,
                                    "by": st.session_state.get("username", "unknown"),
                                    "at": datetime.datetime.now().isoformat(timespec="seconds")}
                                added += 1
                if not added and not removed:
                    st.info("No changes to save.")
                else:
                    try:
                        db.save_staff_overrides(new_ov)
                        st.success(f"Saved — {added} edit(s) recorded, {removed} reverted.")
                        st.rerun()
                    except Exception as ex:
                        st.error(f"Could not save edits: {ex}")
        with b2:
            wk_ov = [k for k in overrides if k.split("|", 2)[0] == sel_wk]
            if wk_ov and st.button(f"Revert all ({len(wk_ov)})", key="ri_revert"):
                try:
                    db.save_staff_overrides({k: v for k, v in overrides.items() if k not in wk_ov})
                    st.success(f"Reverted {len(wk_ov)} edit(s) for this week.")
                    st.rerun()
                except Exception as ex:
                    st.error(f"Could not revert: {ex}")

# ══════════════════════════════════════════════════════════════════════════════
#  PLAN A WEEK — build a week that has not been written yet

# ══════════════════════════════════════════════════════════════════════════════
with tab_month:
    # A week at a time answers "who is in on Tuesday". It cannot answer "is
    # this person carrying the month", which is the question behind planning
    # the next one -- so the same grid, four or five weeks wide, with the names
    # and the tallies pinned while the days slide past them.
    if not week_keys:
        st.info("Upload the workbook first — a month is built from the weeks in it.")
    else:
        _all_days = sorted({d for k in week_keys
                            for d in (db.load_staff_week(k) or {}).get("dates", [])}) \
            if False else None

        @st.cache_data(ttl=600, show_spinner="Building the month…")
        def _month_grid(_token, ym, grp):
            """Every stored day in one month, as a person-by-day grid."""
            weeks = {}
            for k in db.staff_week_keys():
                w = db.load_staff_week(k)
                if not w:
                    continue
                if any(d[:7] == ym for d in w.get("dates", [])):
                    weeks[k] = ri.apply_overrides(w, db.load_staff_overrides(), k)[0]
            days = sorted({d for w in weeks.values() for d in w["dates"]
                           if d[:7] == ym})
            people = {}
            for k in sorted(weeks):
                for name, rec in weeks[k]["people"].items():
                    p = people.setdefault(name, {
                        "label": rec.get("label") or name,
                        "group": rec.get("group"), "section": rec.get("section", ""),
                        "cells": {}, "nocall": set()})
                    for d, raw in (rec.get("cells") or {}).items():
                        if d[:7] == ym:
                            p["cells"][d] = raw
                    for d in (rec.get("nocall") or []):
                        if d[:7] == ym:
                            p["nocall"].add(d)
            if grp == "Housekeepers":
                people = {k: v for k, v in people.items() if v["group"] == "hk"}
            elif grp == "RQS":
                people = {k: v for k, v in people.items() if v["group"] == "rqs"}
            elif grp == "Other teams":
                people = {k: v for k, v in people.items() if v["group"] == "other"}
            return days, people

        _months = sorted({d[:7] for k in week_keys
                          for d in (db.load_staff_week(k) or {}).get("dates", [])},
                         reverse=True)
        mc1, mc2, mc3 = st.columns([1.4, 1.4, 2])
        with mc1:
            _ym = st.selectbox(
                "Month", _months, key="mv_month",
                format_func=lambda x: datetime.date(int(x[:4]), int(x[5:7]), 1)
                .strftime("%B %Y"))
        with mc2:
            _grp = st.selectbox("Show", ["Housekeepers", "RQS", "Other teams", "All"],
                                key="mv_grp")
        _days, _people = _month_grid(
            f'{meta.get("uploaded_at","")}|{len(overrides)}|{len(week_keys)}', _ym, _grp)
        with mc3:
            st.markdown(f'<div style="padding-top:26px;font-size:.76rem;color:#5b6675">'
                        f'{len(_people)} people &nbsp;·&nbsp; {len(_days)} days stored'
                        f'</div>', unsafe_allow_html=True)

        if not _days or not _people:
            st.info("No stored days in that month.")
        else:
            #: One or two characters per day, so a month fits across a screen.
            def _code(raw, kind):
                if kind == ri.KIND_DAILY:
                    return "DS"
                if kind == ri.KIND_NOCALL:
                    return "NC"
                if kind == ri.KIND_VTO:
                    return "VTO"
                if kind == ri.KIND_OFF:
                    return ""
                t = str(raw or "").strip()
                if not t:
                    return ""
                m = re.match(r"^(?:hsp|hp)\s*(\d+)", t, re.I)
                if m:
                    return "H" + m.group(1)
                if re.match(r"^\d+$", t):        # a building number
                    return t
                if re.match(r"^rqs", t, re.I):
                    return "RQS"
                return t[:3].upper()

            _order = sorted(_people.items(),
                            key=lambda kv: (kv[1]["group"] != "hk",
                                            kv[1]["section"] or "", kv[1]["label"].lower()))
            _rows, _kinds = [], []
            for _name, _p in _order:
                worked = ds = 0
                row, kinds = {}, {}
                for _d in _days:
                    raw = _p["cells"].get(_d, "")
                    kind = (ri.KIND_NOCALL if _d in _p["nocall"]
                            else ri.classify(raw))
                    if kind in ri.WORKED_KINDS:
                        worked += 1
                    if kind == ri.KIND_DAILY:
                        ds += 1
                    lbl = datetime.date.fromisoformat(_d).strftime("%-d") \
                        if os.name != "nt" else str(int(_d[8:]))
                    row[lbl] = _code(raw, kind)
                    kinds[lbl] = kind
                _rows.append({"Person": _p["label"], "In": worked, "DS": ds, **row})
                _kinds.append(kinds)

            _df = pd.DataFrame(_rows)
            _daycols = [c for c in _df.columns if c not in ("Person", "In", "DS")]

            def _paint(_):
                out = pd.DataFrame("", index=_df.index, columns=_df.columns)
                for i, kinds in enumerate(_kinds):
                    for c in _daycols:
                        bg, fg = KIND_STYLE.get(kinds.get(c), ("#ffffff", "#1f2733"))
                        if kinds.get(c) == ri.KIND_OFF:
                            out.iloc[i, out.columns.get_loc(c)] = \
                                "background-color:#f7f9fc;color:#c3cedd"
                        else:
                            out.iloc[i, out.columns.get_loc(c)] = \
                                f"background-color:{bg};color:{fg};font-weight:600"
                return out

            st.dataframe(
                _df.style.apply(_paint, axis=None),
                hide_index=True, use_container_width=True,
                height=min(90 + 27 * len(_df), 640), row_height=27,
                column_config={
                    "Person": st.column_config.TextColumn("Person", pinned=True,
                                                          width=140),
                    "In": st.column_config.NumberColumn(
                        "In", pinned=True, width=48,
                        help="Days worked this month"),
                    "DS": st.column_config.NumberColumn(
                        "DS", pinned=True, width=48,
                        help="Days on daily service this month"),
                    **{c: st.column_config.TextColumn(c, width=34) for c in _daycols}})

            st.caption("The name and the two tallies stay put while the days slide "
                       "past. **In** is days worked, **DS** days on daily service — "
                       "the pair to compare people on. Codes: a building number, "
                       "**DS** daily service, **H1** houseperson zone, **VTO** paid "
                       "time off, **NC** no call, blank for off.")

            _hk = [r for r in _rows if r["In"]]
            if _hk and _grp == "Housekeepers":
                _busy = sorted(_hk, key=lambda r: -r["DS"])[:3]
                _idle = [r for r in sorted(_hk, key=lambda r: r["DS"]) if r["In"] >= 8][:3]
                st.markdown(
                    '<div style="background:#f7f9fc;border:1px solid #e6ebf2;'
                    'border-radius:9px;padding:9px 13px;font-size:.78rem;color:#42536a">'
                    'Most daily service this month: '
                    + ", ".join(f'<b>{e(r["Person"])}</b> {r["DS"]}' for r in _busy)
                    + ' &nbsp;·&nbsp; least, of those in 8+ days: '
                    + ", ".join(f'<b>{e(r["Person"])}</b> {r["DS"]}' for r in _idle)
                    + '</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
with tab_plan:
    if not week_keys:
        st.info("Store at least one week first — a plan is built from what came before.")
    elif not IS_ADMIN:
        st.info("Planning a new week is an admin task. You can review it here once saved.")
    else:
        missing = ri.next_missing_weeks(week_keys, count=8)
        latest = max(week_keys)
        p1, p2 = st.columns([1, 1])
        with p1:
            choices = missing + [w for w in sorted(week_keys, reverse=True)[:6]]
            tgt = st.selectbox(
                "Week to plan", choices, key="pl_week",
                format_func=lambda w: ri.week_label(
                    w, note="not created yet" if w in missing
                       else "already stored, will overwrite"))
        with p2:
            basis = st.radio("Start from", ["Suggest from past patterns",
                                            "Copy the previous week", "Blank week"],
                             key="pl_basis")
        prev_key = max([w for w in week_keys if w < tgt], default=latest)
        template = db.load_staff_week(prev_key)
        if not template:
            st.error("Could not load week " + prev_key + " to build from.")
            st.stop()

        st.caption("Built on **" + template["sheet"] + "** (" +
                   ri.week_label(prev_key) + ") for its list of people, buildings "
                   "and row positions.")

        idx_all = {}
        if basis == "Copy the previous week":
            draft = ri.copy_week(template, tgt)
            sugg = {}
        elif basis == "Blank week":
            draft = ri.copy_week(template, tgt)
            for rec in draft["people"].values():
                rec["cells"] = {}
            sugg = {}
        else:
            all_rows = _history(f'{meta.get("uploaded_at","")}|{len(overrides)}|{len(week_keys)}')
            sugg = ri.suggest_week(all_rows, tgt)
            idx_all = ri.people_index(all_rows)
            draft = ri.suggestion_to_week(idx_all, sugg, tgt, template, rows=all_rows)

        dates = draft["dates"]
        filled = sum(1 for rec in draft["people"].values() for d in dates
                     if rec["cells"].get(d))
        working = sum(1 for rec in draft["people"].values() for d in dates
                      if ri.classify(rec["cells"].get(d, "")) in ri.WORKED_KINDS)
        low = []
        if sugg:
            low = [(pid, iso, g) for pid, days in sugg.items()
                   for iso, g in days.items() if iso in dates and g["confidence"] < 0.6]
        mm = st.columns(4)
        mm[0].metric("People", len(draft["people"]))
        mm[1].metric("Cells filled", filled)
        mm[2].metric("Working days", working)
        mm[3].metric("Need a look", len(low))

        if sugg:
            st.markdown(
                '<div style="background:#eef4fb;border:1px solid #cddff0;border-radius:8px;'
                'padding:10px 13px;font-size:.78rem;color:#1c4a78">'
                'Each cell is what that person has most often had on that weekday recently, '
                'weighted so the last few weeks count most. Backtested on 20 real weeks it '
                'gets <b>81%</b> of working-or-off calls right, but only <b>39%</b> of exact '
                'wordings — a first draft, not a finished rota.</div>',
                unsafe_allow_html=True)
            if low:
                with st.expander(str(len(low)) + " cells the pattern is unsure about"):
                    lab = {p: i["label"] for p, i in idx_all.items()}
                    st.dataframe(pd.DataFrame(
                        [{"Person": lab.get(pid, pid), "Date": iso,
                          "Suggested": g["value"] or "(off)",
                          "Confidence": str(int(g["confidence"] * 100)) + "%",
                          "Also seen": ", ".join((v or "(off)") + " " + str(int(w * 100)) + "%"
                                                 for v, w in g["basis"][1:]) or "—"}
                         for pid, iso, g in sorted(low, key=lambda x: x[2]["confidence"])]),
                        use_container_width=True, hide_index=True, height=300)


        # ── How many people the week needs ────────────────────────────────
        # The workbook works this out in rows it keeps hidden: HSKP Needed is
        # =<labour minutes>/360 and RQS needed is <checkouts>/12. Nothing had
        # ever read them, so the numbers behind a week's staffing were
        # invisible to everyone including the person planning it.
        st.markdown('<div class="sec" style="margin:1.1rem 0 .3rem">'
                    'How many people this week needs</div>', unsafe_allow_html=True)

        _tpl_metrics = template.get("metrics") or {}
        _tpl_dates = list(template.get("dates") or [])

        def _seed(i):
            """Last week's numbers for the same weekday, as a starting point."""
            if i < len(_tpl_dates):
                return _tpl_metrics.get(_tpl_dates[i]) or {}
            return {}

        # Who the plan currently has working, per day, straight from the draft.
        def _on_hand(iso, group):
            return sum(1 for rec in draft["people"].values()
                       if rec.get("group") == group
                       and ri.classify(rec["cells"].get(iso, "")) in ri.WORKED_KINDS)

        _mrows = []
        for _i, _d in enumerate(dates):
            _m = _seed(_i)
            _mrows.append({
                "Day": datetime.date.fromisoformat(_d).strftime("%a %d %b"),
                "Labour minutes": int(_m.get("minutes") or 0),
                "Checkouts": int(_m.get("checkouts") or 0),
                "Daily services": int(_m.get("dailies") or 0),
            })
        st.caption("Seeded from the same weekday last week — type over any of it. "
                   "Labour minutes and checkouts are the two numbers the sheet "
                   "hides; daily services is the row it leaves blank, and filling "
                   "it in is what makes the inspector count right.")
        _med = st.data_editor(
            pd.DataFrame(_mrows), hide_index=True, use_container_width=True,
            num_rows="fixed", key="pl_metrics_" + tgt,
            column_config={
                "Day": st.column_config.TextColumn("Day", disabled=True),
                "Labour minutes": st.column_config.NumberColumn(
                    "Labour minutes", min_value=0, max_value=40000, step=10,
                    help="The day's total cleaning time. The sheet keeps this "
                         "inside the HSKP Needed formula, as =4460/360."),
                "Checkouts": st.column_config.NumberColumn(
                    "Checkouts", min_value=0, max_value=500, step=1),
                "Daily services": st.column_config.NumberColumn(
                    "Daily services", min_value=0, max_value=500, step=1),
            })

        _ests, _plan_metrics, _warn = [], {}, []
        for _i, _d in enumerate(dates):
            _row = _med.iloc[_i]
            _mins = float(_row["Labour minutes"] or 0)
            _chk = int(_row["Checkouts"] or 0)
            _dly = int(_row["Daily services"] or 0)
            _e = staffing.estimate(_mins, _chk, _dly,
                                   on_hand_hskp=_on_hand(_d, "hk"),
                                   on_hand_rqs=_on_hand(_d, "rqs"))
            _e["date"] = _d
            _ests.append(_e)
            _plan_metrics[_d] = {"minutes": _mins, "checkouts": _chk,
                                 "dailies": _dly,
                                 "hskp_needed": _e["hskp"], "rqs_needed": _e["rqs"]}
            _note = staffing.consistency(_mins, _chk, _dly)
            if _note:
                _warn.append((_d, _note))

        def _gap(n):
            if n > 0:
                return f"+{n} spare"
            if n < 0:
                return f"{abs(n)} SHORT"
            return "exact"

        st.dataframe(pd.DataFrame([{
            "Day": datetime.date.fromisoformat(x["date"]).strftime("%a %d %b"),
            "Minutes": f'{int(x["minutes"]):,}' if x["minutes"] else "—",
            "Rooms": (str(x["checkouts"] + x["dailies"]) if x["checkouts"] or x["dailies"]
                      else "—"),
            "HK needed": x["hskp"] or "—",
            "If it runs badly": (f'{x["hskp_low"]}–{x["hskp_high"]}'
                                 if x["minutes"] else "—"),
            "HK planned": x.get("on_hand_hskp", 0),
            "HK gap": _gap(x.get("extra_hskp", 0)),
            "RQS needed": x["rqs"] or "—",
            "RQS planned": x.get("on_hand_rqs", 0),
            "RQS gap": _gap(x.get("extra_rqs", 0)),
            "Sheet would say": (f'{x["sheet_hskp"]:.1f} HK / {x["sheet_rqs"]:.1f} RQS'
                                if x["minutes"] or x["checkouts"] else "—"),
        } for x in _ests]), hide_index=True, use_container_width=True,
            height=38 * len(_ests) + 40)

        _tot = staffing.week_totals(_ests)
        _short = [x for x in _ests if x.get("extra_hskp", 0) < 0]
        _short_r = [x for x in _ests if x.get("extra_rqs", 0) < 0]
        st.markdown(
            f'<div style="background:{"#fff5f5" if _short or _short_r else "#f2fbf5"};'
            f'border:1px solid {"#f6c6c6" if _short or _short_r else "#c3e9d0"};'
            f'border-radius:9px;padding:9px 13px;font-size:.79rem;'
            f'color:{"#8a1c1c" if _short or _short_r else "#0a5c32"}">'
            f'Week needs <b>{_tot["hskp"]}</b> housekeeper-days and '
            f'<b>{_tot["rqs"]}</b> inspector-days; the plan plots '
            f'<b>{_tot["on_hand_hskp"]}</b> and <b>{_tot["on_hand_rqs"]}</b>. '
            + ("Short on " + ", ".join(
                datetime.date.fromisoformat(x["date"]).strftime("%a") for x in _short)
               + " for housekeepers. " if _short else "")
            + ("Short of an inspector on " + ", ".join(
                datetime.date.fromisoformat(x["date"]).strftime("%a") for x in _short_r)
               + ". " if _short_r else "")
            + ("" if _short or _short_r else "Every day is covered.")
            + '</div>', unsafe_allow_html=True)

        for _d, _note in _warn:
            st.warning(datetime.date.fromisoformat(_d).strftime("%A") + ": " + _note)

        with st.expander("Why these numbers differ from the sheet's"):
            st.markdown(
                "The sheet divides the whole day's labour by **360** and the "
                "checkouts by **12**. That is a good first cut and it is kept "
                "above, in the last column, so nothing is lost.\n\n"
                "Four things are done differently here:\n\n"
                "1. **A Full Clean and a Daily Service are not the same job.** "
                "The scheduler loads a Full Clean housekeeper to 330–380 "
                "minutes and a Daily Service round to about 460, so the two "
                "pools are divided by their own targets instead of one 360 "
                "standing for both.\n"
                "2. **People are whole.** 5.75 housekeepers is not an answer, "
                "and the fraction leaks into the sheet's Extras row — it reads "
                "*0.49 extras* for a thirty-person day that is exactly covered.\n"
                "3. **An inspector walks the daily services too**, so the RQS "
                "count comes from every room to be looked at, not the checkouts "
                "alone. On a day with 60 daily services the sheet asks for one "
                "inspector; this asks for five.\n"
                "4. **A range, because the risk is lopsided.** Being one short "
                "costs a missed checkout or overtime; being one over costs a few "
                "idle hours. *If it runs badly* is that day with everyone loaded "
                "light.\n\n"
                "With the daily-services row left blank — as it is in most "
                "sheets — this lands on the same housekeeper count as the sheet. "
                "The gain is in the inspectors, the whole numbers, and knowing "
                "when the two figures behind a day disagree.")


        # ── Daily service: whose turn ─────────────────────────────────────
        st.markdown('<div class="sec" style="margin:1.1rem 0 .3rem">'
                    'Daily service — whose turn</div>', unsafe_allow_html=True)

        # Anything already applied is folded into the draft before the section
        # editors below are built from it, so the suggestion shows up in the
        # grid like any other cell and can be typed over there.
        _ds_key = "pl_ds_" + tgt
        _ds_applied = st.session_state.get(_ds_key) or {}
        for _iso, _names in _ds_applied.items():
            for _n in _names:
                if _n in draft["people"]:
                    draft["people"][_n]["cells"][_iso] = ri.DS_LABEL

        # How many each day: what the same weekday carried last week, which is
        # the honest starting point, and editable.
        @st.cache_data(ttl=300, show_spinner=False)
        def _ds_shape(_token, _before):
            """How many the most recent filled-in week put on each weekday.

            The week just gone is the obvious template, but upcoming weeks
            often have the daily-service column still empty -- both of the
            last two here do -- and seeding every day with the same number
            would put two people on a Tuesday that has taken six.
            """
            for _k in sorted(db.staff_week_keys(), reverse=True):
                if _k >= _before:
                    continue
                _w = db.load_staff_week(_k)
                if not _w:
                    continue
                _n = {}
                for _d in _w["dates"]:
                    _n[datetime.date.fromisoformat(_d).weekday()] = sum(
                        1 for r in _w["people"].values() if r.get("group") == "hk"
                        and "daily service" in str(r["cells"].get(_d, "")).lower())
                if any(_n.values()):
                    return {"week": _k, "sheet": _w.get("sheet", _k), "by_weekday": _n}
            return {"week": None, "sheet": "", "by_weekday": {}}

        _shape = _ds_shape(f'{meta.get("uploaded_at","")}|{len(week_keys)}', tgt)

        def _tpl_ds_count(i):
            return _shape["by_weekday"].get(
                datetime.date.fromisoformat(dates[i]).weekday(), 2)

        _dsc = st.data_editor(
            pd.DataFrame([{"Day": datetime.date.fromisoformat(d).strftime("%a %d %b"),
                           "How many": _tpl_ds_count(i)}
                          for i, d in enumerate(dates)]),
            hide_index=True, use_container_width=True, num_rows="fixed",
            key="pl_dscount_" + tgt,
            column_config={
                "Day": st.column_config.TextColumn("Day", disabled=True),
                "How many": st.column_config.NumberColumn(
                    "How many", min_value=0, max_value=20, step=1,
                    help="How many housekeepers this day needs on daily service.")})
        if _shape.get("week"):
            st.caption("Seeded from **" + _shape["sheet"] + "**, the most recent "
                       "week with daily service filled in — the last two weeks "
                       "have that column empty, which is what this is for.")
        _per_day = {d: int(_dsc.iloc[i]["How many"] or 0) for i, d in enumerate(dates)}

        _hist_rows = _history(f'{meta.get("uploaded_at","")}|{len(overrides)}|{len(week_keys)}')
        _ds = ri.suggest_daily_service([r for r in _hist_rows if r["date"] < tgt],
                                       draft, dates, _per_day)
        _lab = {p: (i.get("label") or p) for p, i in (idx_all or {}).items()}

        _ds_tbl = []
        for _i, _d in enumerate(dates):
            _names = _ds["picks"].get(_d, [])
            _cells = []
            for _n in _names:
                _w = _ds["why"].get((_d, _n), {})
                _last = _w.get("last")
                _ago = ("never in the last four weeks" if not _last else
                        str((datetime.date.fromisoformat(_d)
                             - datetime.date.fromisoformat(_last)).days) + "d ago")
                _cells.append(f'{_lab.get(_n, _n)} ({_w.get("ds", 0)}/{_w.get("worked", 0)}, '
                              f'{_ago})' + (" · second turn" if _w.get("second_turn") else ""))
            _ds_tbl.append({
                "Day": datetime.date.fromisoformat(_d).strftime("%a %d %b"),
                "Needs": _per_day[_d],
                "Suggested": " · ".join(_cells) or "—",
            })
        st.dataframe(pd.DataFrame(_ds_tbl), hide_index=True, use_container_width=True,
                     height=38 * len(_ds_tbl) + 40)
        st.caption("Each name shows their turns against days worked in the last four "
                   "weeks, and how long since the last one. Only housekeepers the "
                   "plan already has working that day are eligible.")

        _repeats = [p for p, n in (_ds.get("week_counts") or {}).items() if n > 1]
        _c1, _c2 = st.columns([1, 2])
        with _c1:
            if st.button("Put these in the plan", key="pl_ds_apply", type="primary"):
                st.session_state[_ds_key] = {d: list(v) for d, v in _ds["picks"].items()}
                st.rerun()
        with _c2:
            if _ds_applied:
                if st.button("Take them back out", key="pl_ds_clear"):
                    st.session_state.pop(_ds_key, None)
                    st.rerun()
        if _ds_applied:
            st.success("In the plan — edit any of them in the grids below like "
                       "any other cell.")
        if _repeats:
            st.info("Everyone available had a turn before " +
                    ", ".join(_lab.get(p, p) for p in _repeats) +
                    " took a second, which is the rotation starting over.")

        with st.expander("How the turn is decided, and whether it is fairer"):
            st.markdown(
                "Two rules. **Nobody takes a second turn in the week until "
                "everyone who could have taken a first has had one** — and when "
                "the pool runs out it simply starts again, which is the only "
                "way a short crew can cover the week. Between people level on "
                "the week, the turn goes to whoever has done it least often "
                "against the days they worked in the **last four weeks**, then "
                "to whoever has waited longest.\n\n"
                "Two things that took some finding. Counting a whole career "
                "made it *worse* than the sheet: it spends months repaying old "
                "debts, hammering the few people it thinks are owed and never "
                "asking anybody else. And raw turns-per-day lets somebody with "
                "three days on the books outrank a veteran, so thin records are "
                "pulled toward the house average until there is enough of one.\n\n"
                "Replayed over four separate ten-week stretches of this "
                "workbook, with the same crews and the same daily-service "
                "counts the sheet actually used:\n\n"
                "| | spread across the team | people never asked |\n"
                "|---|---|---|\n"
                "| what the sheet did | 5.8 – 6.6 | 1 to 7 |\n"
                "| this rotation | 2.0 – 2.8 | none, in any stretch |\n\n"
                "Roughly half the spread, and nobody left out. For scale, today "
                "one housekeeper spends 18% of her days on daily service while "
                "another has worked 77 days and has never once been asked.")

        _zmoves = draft.get("zone_moves") or []
        if _zmoves:
            st.markdown(f'<div style="background:#f5f3ff;border:1px solid #ddd6fe;'
                        f'border-radius:8px;padding:9px 13px;font-size:.78rem;color:#5b21b6">'
                        f'Moved <b>{len(_zmoves)}</b> houseperson(s) off a zone another '
                        f'person already had, onto one nobody was covering.</div>',
                        unsafe_allow_html=True)
        cover_rep = ri.zone_coverage(draft)
        if cover_rep:
            gaps = sum(len(v["missing"]) for d in cover_rep.values() for v in d.values())
            dups = sum(len(v["doubled"]) for d in cover_rep.values() for v in d.values())
            with st.expander(f"Houseperson zone cover — {gaps} gap(s), {dups} doubled"):
                st.caption("Zones run 1–7 (1 Runner Bld 2, 2 Lobby, 3 Pool Bld 1, "
                           "4 Runner Bld 1, 5 Runner Bld 3, 6 Pool Bld 3, 7 Plaza Bld 3). "
                           "Full cover happens on only 39% of real shift-days, so a gap is "
                           "normal on a short day — a doubled zone usually is not.")
                st.dataframe(pd.DataFrame(
                    [{"Shift": sec, "Day": iso,
                      "Zones with nobody": ", ".join(v["missing"]) or "—",
                      "Zones doubled": ", ".join(v["doubled"]) or "—"}
                     for sec, days in cover_rep.items() for iso, v in sorted(days.items())]),
                    use_container_width=True, hide_index=True, height=260)

        with st.expander("What the patterns actually say"):
            pc1, pc2 = st.columns(2)
            with pc1:
                st.markdown('<div class="mono">HOUSEKEEPERS — BUILDINGS BARELY MOVE</div>',
                            unsafe_allow_html=True)
                bsug = ri.suggest_building(_history(
                    f'{meta.get("uploaded_at","")}|{len(overrides)}|{len(week_keys)}'))
                moved = [v for v in bsug.values() if v["moved"]]
                st.markdown(
                    f'<div style="font-size:.8rem;color:#5b6675;line-height:1.55">'
                    f'Across the whole workbook only <b>1.5%</b> of week-to-week '
                    f'transitions are a building change, and <b>106 of 121</b> '
                    f'housekeepers spend 90%+ of their weeks in one building. A move is '
                    f'a transfer, not a rotation — so a person\'s building is simply '
                    f'where they have settled.</div>', unsafe_allow_html=True)
                if moved:
                    st.markdown('<div class="mono" style="margin-top:10px">'
                                'RECENTLY TRANSFERRED</div>', unsafe_allow_html=True)
                    st.dataframe(pd.DataFrame(
                        [{"Person": v["person"], "Now": f'Building {v["building"]}',
                          "Was": f'Building {v["settled"]}',
                          "Confidence": f'{v["confidence"]*100:.0f}%'} for v in moved]),
                        use_container_width=True, hide_index=True,
                        height=min(60 + 33 * len(moved), 200))
                else:
                    st.caption("Nobody has changed building recently.")
            with pc2:
                st.markdown('<div class="mono">HOUSEPERSONS — ZONES PERSIST, THEN SHIFT</div>',
                            unsafe_allow_html=True)
                st.markdown(
                    '<div style="font-size:.8rem;color:#5b6675;line-height:1.55">'
                    'Zones do <b>not</b> follow the weekday and they do not cycle. A '
                    'houseperson holds a zone for a stretch — median <b>2</b> consecutive '
                    'working days, mean 3.4 — then moves to another. So the best guess is '
                    'the zone they last worked, which scores <b>54%</b> against <b>36%</b> '
                    'for a weekday-based guess. Whether they work at all is still a '
                    'weekday question, so only the zone itself is carried forward.</div>',
                    unsafe_allow_html=True)
                dups = ri.duplicate_candidates(idx_all) if idx_all else []
                if dups:
                    st.markdown('<div class="mono" style="margin-top:12px">'
                                'POSSIBLE DUPLICATE NAMES</div>', unsafe_allow_html=True)
                    st.caption("Building suffixes like \"Jose M - BLD 1\" are merged "
                               "automatically. These look similar but are NOT merged — "
                               "some are genuinely different people.")
                    st.dataframe(pd.DataFrame(
                        [{"Name": a, "Looks like": b, "Similarity": f"{r*100:.0f}%",
                          "Days A": na, "Days B": nb} for a, b, r, na, nb in dups[:12]]),
                        use_container_width=True, hide_index=True,
                        height=min(60 + 33 * len(dups[:12]), 260))
                if sugg:
                    nlast = sum(1 for days in sugg.values() for iso, g in days.items()
                                if iso in dates and g.get("model") == "last zone")
                    st.markdown(f'<div style="margin-top:10px;font-size:.8rem;color:#5b21b6">'
                                f'<b>{nlast}</b> cell(s) in this draft came from the '
                                f'carried-forward zone rather than the weekday pattern.</div>',
                                unsafe_allow_html=True)

        st.markdown('<p class="sec">Draft — change anything before saving</p>',
                    unsafe_allow_html=True)
        cols_lbl = [f'{datetime.date.fromisoformat(d).strftime("%a")} {d[5:]}' for d in dates]

        def pl_to_cell(v):
            return v if str(v or "").strip() else ri.BLANK_LABEL

        def pl_from_cell(v):
            return "" if str(v) == ri.BLANK_LABEL else str(v or "").strip()

        sections, order = {}, []
        for name, rec in draft["people"].items():
            sections.setdefault(rec["section"], []).append((name, rec))
            if rec["section"] not in order:
                order.append(rec["section"])
        show = st.multiselect("Sections to edit", order, default=order[:3], key="pl_secs")
        pl_editors = []
        for section in [x for x in order if x in show]:
            members = sections[section]
            present = {rec["cells"].get(d, "") for _n, rec in members for d in dates}
            opts = ri.options_for(members[0][1], sorted(v for v in present if v),
                                  extra=custom_opts.get(ri.role_key(members[0][1]), []))
            base = pd.DataFrame(
                [{"Name": n, **{cols_lbl[i]: pl_to_cell(rec["cells"].get(d, ""))
                                for i, d in enumerate(dates)}} for n, rec in members],
                columns=["Name"] + cols_lbl)
            st.markdown('<div class="sec" style="margin:.9rem 0 .3rem">' + e(section) +
                        '</div>', unsafe_allow_html=True)
            pl_editors.append(st.data_editor(
                base, hide_index=True, use_container_width=True, num_rows="fixed",
                key="pl_ed_" + tgt + "_" + section, height=min(90 + 35 * len(base), 420),
                column_config={
                    "Name": st.column_config.TextColumn("Name", disabled=True,
                                                        width="medium"),
                    **{c: st.column_config.SelectboxColumn(c, options=opts, required=False,
                                                           width="medium")
                       for c in cols_lbl}}))

        st.caption("Sections you do not open are saved exactly as drafted above.")
        if st.button("Save week " + ri.week_range_text(tgt), type="primary", key="pl_save"):
            edited_by_name = {}
            for ed in pl_editors:
                for _, row in ed.iterrows():
                    edited_by_name[row["Name"]] = {
                        d: pl_from_cell(row[cols_lbl[i]]) for i, d in enumerate(dates)}
            out = {"sheet": draft["sheet"], "dates": dates,
                   "cols": draft.get("cols", {}), "people": {},
                   "planned": True, "from_sheet": draft.get("from_sheet", ""),
                   "created_by": st.session_state.get("username", "unknown"),
                   "created_at": datetime.datetime.now().isoformat(timespec="seconds"),
                   "basis": basis,
                   # The staffing numbers are part of the plan, not a scratch
                   # calculation: stored, they can be read back next week and
                   # written into the sheet's hidden rows.
                   "metrics": _plan_metrics}
            for name, rec in draft["people"].items():
                cells = edited_by_name.get(name, rec["cells"])
                out["people"][name] = {**rec,
                                       "cells": {k: v for k, v in cells.items() if v}}
            try:
                db.save_staff_week(tgt, out)
                n_fill = sum(1 for r in out["people"].values()
                             for v in r["cells"].values() if v)
                st.cache_data.clear()
                st.success("Saved week " + ri.week_range_text(tgt) + " — " + str(len(out["people"])) +
                           " people, " + str(n_fill) + " cells filled. It now shows in "
                           "Week view, My Home and the roster.")
            except Exception as ex:
                st.error("Could not save: " + str(ex))

        with st.expander("Download this plan as Excel"):
            st.caption("Uses the sheet it was built from, with the dates "
                       "rewritten to this week and last week's absences cleared.")
            week_download(draft, tgt, f"pldl_{tgt}")

# ══════════════════════════════════════════════════════════════════════════════
#  ATTENDANCE — how much has each person actually worked
# ══════════════════════════════════════════════════════════════════════════════
with tab_att:
    if not week_keys:
        st.info("No weeks stored yet. Upload the workbook on the **Upload & sync** tab.")
    else:
        token = f'{meta.get("uploaded_at","")}|{len(overrides)}|{len(week_keys)}'
        rows = _history(token)
        today = clock.today()
        wk_s, mo_s, yr_s = ri.period_bounds(today)
        t_iso = today.isoformat()

        c = st.columns([2, 1, 2])
        with c[0]:
            period = st.radio("Period", ["This week", "This month", "This year", "All stored"],
                              horizontal=True, key="ri_att_period", label_visibility="collapsed")
        start = {"This week": wk_s, "This month": mo_s,
                 "This year": yr_s, "All stored": "0000-01-01"}[period]
        with c[2]:
            groups2 = ["Everyone", "Housekeepers", "RQS", "Other teams"]
            gsel = st.selectbox("Team", groups2, key="ri_att_grp", label_visibility="collapsed")

        # Whole period, not just up to today: on a Wednesday a manager wants the
        # week a person is down for, not only the days already behind them.
        end = {"This week": (datetime.date.fromisoformat(wk_s)
                             + datetime.timedelta(days=6)).isoformat(),
               "This month": ((today.replace(day=28) + datetime.timedelta(days=4))
                              .replace(day=1) - datetime.timedelta(days=1)).isoformat(),
               "This year": today.replace(month=12, day=31).isoformat(),
               "All stored": "9999-12-31"}[period]
        sub = [r for r in rows if start <= r["date"] <= end]
        if gsel == "Housekeepers":  sub = [r for r in sub if r["group"] == "hk"]
        elif gsel == "RQS":         sub = [r for r in sub if r["group"] == "rqs"]
        elif gsel == "Other teams": sub = [r for r in sub if r["group"] == "other"]

        if not sub:
            st.info(f"Nothing recorded for **{period.lower()}** yet.")
        else:
            idx_ppl = ri.people_index(sub)
            people = sorted(idx_ppl, key=lambda p: idx_ppl[p]["label"].lower())
            summ = {p: ri.summarise_person(sub, pid=p) for p in people}
            worked_days = sum(s["n_worked"] for s in summ.values())
            done_days = sum(1 for r in sub if r["worked"] and r["date"] <= t_iso)
            m = st.columns(4)
            m[0].metric("People", len(people))
            m[1].metric("Days in period", worked_days,
                        delta=f'{done_days} already worked', delta_color="off")
            m[2].metric("Avg days / person", f'{worked_days/max(len(people),1):.1f}')
            m[3].metric("No call / no show", sum(1 for r in sub if r.get("no_call")),
                        delta=f'{sum(1 for r in sub if r.get("paid_off"))} VTO days',
                        delta_color="off")

            # Fills are only read on upload, so weeks stored before this release
            # carry no no-call marks. Say so rather than implying nobody has any.
            if not any(r.get("no_call") for r in rows):
                st.info("No no-call/no-show marks found. These come from the **red "
                        "cells** in the workbook, which are only read when the file is "
                        "uploaded — re-upload on **Upload & sync** to pick them up for "
                        "the weeks already stored.")

            st.markdown('<p class="sec">Who has worked how much</p>', unsafe_allow_html=True)
            st.caption("Sorted by days worked. Use it to see who is due time off "
                       "and who is carrying the week.")
            tbl = []
            for p in people:
                s = summ[p]
                if not s["rows"]: continue
                done = sum(1 for r in s["worked"] if r["date"] <= t_iso)
                tbl.append({
                    "Person": idx_ppl[p]["label"],
                    "Team": idx_ppl[p]["sections"][0],
                    "From": idx_ppl[p]["home"] or "—",
                    "Worked": s["n_worked"],
                    "So far": done,
                    "Ahead": s["n_worked"] - done,
                    "Off": s["off"],
                    "No call": s["n_no_call"],
                    "VTO (paid)": s["n_paid_off"],
                    "Daily Service": s["shifts"].get("Daily Service", 0),
                    "Room cover": s["shifts"].get("Room cover", 0),
                    "Other duty": s["shifts"].get("Other duty", 0),
                    "Usual days": ", ".join(s["usual_days"]) or "—",
                })
            df = pd.DataFrame(tbl).sort_values("Worked", ascending=False)
            st.dataframe(
                df, use_container_width=True, hide_index=True,
                height=min(60 + 33 * len(df), 520),
                column_config={
                    "Worked": st.column_config.ProgressColumn(
                        "Worked", format="%d",
                        min_value=0, max_value=int(df["Worked"].max() or 1)),
                })

            st.markdown('<p class="sec">One person in detail</p>', unsafe_allow_html=True)
            labels = [person_label(idx_ppl[p]) for p in people]
            pick = st.selectbox("Person", labels, key="ri_att_person",
                                label_visibility="collapsed")
            pk = people[labels.index(pick)]
            s = summ[pk]
            pick = idx_ppl[pk]["label"]
            d1, d2 = st.columns([3, 2])
            with d1:
                st.markdown('<div class="mono" style="margin-bottom:4px">DAYS OF THE WEEK '
                            'USUALLY WORKED</div>', unsafe_allow_html=True)
                mx = max(s["by_dow"].values()) or 1
                bars = ""
                for d, n in s["by_dow"].items():
                    pct = int(100 * n / mx)
                    hot = n >= mx * 0.6
                    bars += (f'<div style="display:flex;align-items:center;gap:9px;margin:3px 0">'
                             f'<span style="width:34px;font-size:.74rem;color:#5b6675">{d}</span>'
                             f'<div style="flex:1;background:#eef1f5;border-radius:4px;height:14px">'
                             f'<div style="width:{pct}%;height:14px;border-radius:4px;'
                             f'background:{"#2563a8" if hot else "#b9c6d6"}"></div></div>'
                             f'<span style="width:26px;text-align:right;font-size:.74rem;'
                             f'color:#16202e;font-weight:600">{n}</span></div>')
                st.markdown(bars, unsafe_allow_html=True)
            with d2:
                st.markdown('<div class="mono" style="margin-bottom:4px">ROLES &amp; SHIFTS</div>',
                            unsafe_allow_html=True)
                chips = ""
                for k2, v2 in list(s["roles"].items())[:4]:
                    chips += (f'<span class="pill" style="background:#eef2ff;color:#3730a3;'
                              f'margin:0 4px 4px 0;display:inline-block">{e(k2)} · {v2}</span>')
                for k2, v2 in s["shifts"].items():
                    bg, fg = {"Daily Service": ("#ccfbf1", "#115e59"),
                              "Room cover": ("#ede9fe", "#5b21b6"),
                              "Other duty": ("#fef3c7", "#92400e")}.get(k2, ("#dcfce7", "#15803d"))
                    chips += (f'<span class="pill" style="background:{bg};color:{fg};'
                              f'margin:0 4px 4px 0;display:inline-block">{e(k2)} · {v2}</span>')
                st.markdown(chips, unsafe_allow_html=True)
                st.markdown(f'<div style="margin-top:8px;font-size:.8rem;color:#5b6675">'
                            f'Worked <b>{s["n_worked"]}</b> of {s["n_days"]} days '
                            f'({100*s["n_worked"]/max(s["n_days"],1):.0f}%) · '
                            f'off <b>{s["off"]}</b></div>', unsafe_allow_html=True)

            with st.expander(f"Every recorded day for {pick}"):
                dd = pd.DataFrame([{"Date": r["date"], "Day": r["dow"], "Team": r["section"],
                                    "Cell": r["raw"] or "—",
                                    "Status": KIND_LABEL[r["kind"]],
                                    "Means": ri.legend_for(r, r["raw"]) or "—"}
                                   for r in sorted(s["rows"], key=lambda x: x["date"],
                                                   reverse=True)])
                st.dataframe(dd, use_container_width=True, hide_index=True,
                             height=min(60 + 33 * len(dd), 420))

# ══════════════════════════════════════════════════════════════════════════════
#  WHAT CHANGED
# ══════════════════════════════════════════════════════════════════════════════
with tab_changes:
    last_diff = (meta.get("last_diff") or {})
    if not meta.get("uploaded_at"):
        st.info("Nothing loaded yet.")
    else:
        st.markdown(f'<p class="sec">At the last sync — {str(meta["uploaded_at"])[:16].replace("T"," ")}'
                    f' by {e(meta.get("uploaded_by","—"))}</p>', unsafe_allow_html=True)
        c = st.columns(3)
        c[0].metric("New weeks added", len(last_diff.get("new_weeks", [])))
        c[1].metric("Cells changed", last_diff.get("n_changed_cells", 0))
        c[2].metric("In-app edits live", len(overrides))
        if last_diff.get("new_weeks"):
            st.markdown(f'<div style="background:#ecfdf5;border:1px solid #a7f3d0;border-radius:8px;'
                        f'padding:9px 13px;font-size:.79rem;color:#065f46">'
                        f'<b>New:</b> {e(", ".join(ri.week_label(w) for w in last_diff["new_weeks"]))}</div>',
                        unsafe_allow_html=True)
        ch = last_diff.get("changed", [])
        if ch:
            st.markdown('<p class="sec">Cell changes</p>', unsafe_allow_html=True)
            st.dataframe(pd.DataFrame([{"Week": ri.week_range_text(x["week"]) if x.get("week") else "", "Person": x.get("name", ""),
                                        "Date": x.get("date", ""), "Was": x.get("old") or "—",
                                        "Now": x.get("new") or "—"} for x in ch]),
                         use_container_width=True, hide_index=True,
                         height=min(60 + 35 * len(ch), 460))
            if last_diff.get("n_changed_cells", 0) > len(ch):
                st.caption(f"Showing the first {len(ch)} of "
                           f"{last_diff['n_changed_cells']} changes.")
        elif not last_diff.get("new_weeks"):
            st.info("The last upload matched what was already stored.")

        if overrides:
            st.markdown('<p class="sec">In-app edits</p>', unsafe_allow_html=True)
            orows = []
            for k, v in sorted(overrides.items()):
                try: wk, name, dt = k.split("|", 2)
                except ValueError: continue
                orows.append({"Week": ri.week_range_text(wk), "Person": name, "Date": dt,
                              "Set to": v.get("value") or "(cleared)",
                              "By": v.get("by", ""), "When": str(v.get("at", ""))[:16].replace("T", " ")})
            st.dataframe(pd.DataFrame(orows), use_container_width=True, hide_index=True,
                         height=min(60 + 35 * len(orows), 320))

        # ── export ────────────────────────────────────────────────────────────
        st.markdown('<p class="sec">Export back to Excel</p>', unsafe_allow_html=True)
        file_info = db.staff_file_info()
        if not overrides:
            st.caption("No in-app edits to write back yet.")
        elif not (st.session_state.get("ri_raw_bytes") or file_info.get("size")):
            st.info("No workbook stored yet — upload it on the **Upload & sync** tab.")
        else:
            src = ("this session's upload" if st.session_state.get("ri_raw_bytes")
                   else f'the stored copy ({e(file_info.get("file_name","")) or "workbook"}, '
                        f'saved {str(file_info.get("saved_at",""))[:16].replace("T"," ")})')
            st.caption(f"Writes every in-app edit into a copy of {src}. Formulas and cell "
                       f"comments are preserved. Download it and put it back on OneDrive "
                       f"to bring Excel in line with the app.")
            if st.button("Build updated workbook", key="ri_export"):
                try:
                    raw = st.session_state.get("ri_raw_bytes")
                    if not raw:
                        with st.spinner("Fetching the stored workbook…"):
                            raw, _fi = db.load_staff_file()
                    if not raw:
                        st.error("The stored workbook could not be read.")
                        st.stop()
                    stored_weeks = db.load_staff_weeks()
                    out, n, skipped = ri.write_overrides_to_workbook(raw, stored_weeks, overrides)
                    st.session_state["ri_export_bytes"] = out
                    st.session_state["ri_export_note"] = (n, len(skipped))
                except Exception as ex:
                    st.error(f"Export failed: {ex}")
            if st.session_state.get("ri_export_bytes"):
                n, n_skip = st.session_state.get("ri_export_note", (0, 0))
                if n_skip:
                    st.warning(f"{n_skip} edit(s) could not be placed — their sheet or row "
                               f"is no longer in the stored workbook.")
                st.success(f"{n} edit(s) written.")
                st.download_button(
                    "Download updated Schedule.xlsx",
                    data=st.session_state["ri_export_bytes"],
                    file_name=f"Schedule_updated_{clock.today():%Y%m%d}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="ri_dl")

# ══════════════════════════════════════════════════════════════════════════════
#  APPLY TO ROSTER
# ══════════════════════════════════════════════════════════════════════════════
with tab_apply:
    if not week_keys:
        st.info("No weeks stored yet. Upload the workbook on the **Upload & sync** tab.")
    else:
        all_days = []
        for wk in week_keys:
            base = datetime.date.fromisoformat(wk)
            all_days += [base + datetime.timedelta(days=i) for i in range(7)]
        all_days = sorted(set(all_days))
        today = clock.today()
        default = today if today in all_days else min(all_days, key=lambda d: abs((d - today).days))
        c1, c2 = st.columns([1, 2])
        with c1:
            picked = st.date_input("Schedule date", value=default,
                                   min_value=all_days[0], max_value=all_days[-1], key="ri_day")
        wk_for = max([w for w in week_keys if w <= picked.isoformat()], default=None)
        week = db.load_staff_week(wk_for) if wk_for else None
        if not week or picked.isoformat() not in week["dates"]:
            st.warning(f"**{picked:%A, %B %d, %Y}** is not in any stored week.")
            st.stop()
        eff, _applied = ri.apply_overrides(week, overrides, wk_for)
        people = ri.week_to_people(eff, picked.isoformat())
        update = ri.build_roster_update(people, st.session_state.get("hk_roster", {}))
        with c2:
            st.markdown(f'<div style="padding-top:26px;font-size:.79rem;color:#5b6675">'
                        f'From sheet <b>{e(week["sheet"])}</b> · {e(ri.week_label(wk_for))}</div>',
                        unsafe_allow_html=True)

        n_hk = sum(1 for v in update["hk_roster"].values() if v["present"])
        n_ins = sum(1 for v in update["insp_roster"].values() if v)
        m = st.columns(4)
        m[0].metric("Housekeepers present", n_hk)
        m[1].metric("On Daily Service", len(update["ds_team"]))
        m[2].metric("Inspectors present", n_ins)
        m[3].metric("RQS 1 / RQS 2", f'{update["rqs1"] or "—"} / {update["rqs2"] or "—"}')

        # The sheet is the authority on buildings, but a person sitting under a
        # different building from the one they have always worked is worth a
        # second look — it is either a real transfer or a typo in the sheet.
        _bsug = ri.suggest_building(
            _history(f'{meta.get("uploaded_at","")}|{len(overrides)}|{len(week_keys)}'))
        _odd = []
        for p in people:
            if p["group"] != "hk" or not p["present"]:
                continue
            s = _bsug.get(ri.norm_name(p['name']))
            if s and s["settled"] != p["building"]:
                _odd.append((p["name"], p["building"], s["settled"], s["history"]))
        if _odd:
            rows_odd = "<br>".join(
                f'&nbsp;&nbsp;<b>{e(n)}</b>: on the sheet in Building {b}, '
                f'normally Building {s} <span style="opacity:.7">({h})</span>'
                for n, b, s, h in _odd)
            st.markdown(f'<div style="background:#f5f3ff;border:1px solid #ddd6fe;'
                        f'border-radius:8px;padding:10px 13px;font-size:.78rem;'
                        f'color:#5b21b6;margin-top:8px">'
                        f'<b>{len(_odd)} against their usual building</b> — the sheet wins, '
                        f'but check it is a real transfer:<br>{rows_odd}</div>',
                        unsafe_allow_html=True)

        if update.get("cover"):
            rows = "<br>".join(
                f'&nbsp;&nbsp;<b>{e(c["name"])}</b> ({e(c["from"])}) → Building {c["building"]} '
                f'<span style="opacity:.7">· "{e(c["raw"])}"</span>'
                for c in update["cover"])
            st.markdown(f'<div style="background:#ede9fe;border:1px solid #c4b5fd;border-radius:8px;'
                        f'padding:10px 13px;font-size:.78rem;color:#5b21b6;margin-top:8px">'
                        f'<b>{len(update["cover"])} person(s) drafted onto rooms</b> — the sheet '
                        f'puts them on housekeeping today, so they are added to the roster:'
                        f'<br>{rows}</div>', unsafe_allow_html=True)

        if update["building_changes"]:
            rows = "<br>".join(f'&nbsp;&nbsp;<b>{e(n)}</b>: Building {o} → Building {w}'
                               for n, o, w in update["building_changes"])
            st.markdown(f'<div style="background:#fffbeb;border:1px solid #fcd34d;border-radius:8px;'
                        f'padding:10px 13px;font-size:.78rem;color:#92400e;margin-top:8px">'
                        f'<b>{len(update["building_changes"])} building change(s)</b> — the sheet '
                        f'will overwrite the current roster:<br>{rows}</div>', unsafe_allow_html=True)
        if update["new_people"]:
            st.markdown(f'<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;'
                        f'padding:10px 13px;font-size:.78rem;color:#1e40af;margin-top:8px">'
                        f'<b>{len(update["new_people"])} new name(s)</b>: '
                        f'{e(", ".join(update["new_people"]))} — check these are not a '
                        f're-spelling of someone already on the roster.</div>', unsafe_allow_html=True)
        if update["unknown"]:
            st.markdown(f'<div style="background:#faf5ff;border:1px solid #e9d5ff;border-radius:8px;'
                        f'padding:10px 13px;font-size:.78rem;color:#7e22ce;margin-top:8px">'
                        f'<b>{len(update["unknown"])} duty we have no name for</b> — counted as '
                        f'a worked day, just not offered guest rooms: '
                        f'{", ".join(f"<b>{e(n)}</b> ({e(v)})" for n, v in update["unknown"])}'
                        f'</div>', unsafe_allow_html=True)

        st.markdown('<p class="sec">Detail</p>', unsafe_allow_html=True)
        det = pd.DataFrame([{"Name": p["name"], "Section": p["section"],
                             "Status": KIND_LABEL[p["kind"]], "Cell": p["raw"] or "—",
                             "Means": ri.legend_for(p, p["raw"]) or "—"}
                            for p in people])
        st.dataframe(det, use_container_width=True, hide_index=True,
                     height=min(60 + 33 * len(det), 460))

        st.markdown('<p class="sec">Apply</p>', unsafe_allow_html=True)
        o = st.columns(3)
        do_hk = o[0].checkbox("Housekeepers + buildings", value=True, key="ri_do_hk")
        do_rqs = o[1].checkbox("Inspectors + RQS 1/2", value=True, key="ri_do_rqs")
        do_ds = o[2].checkbox("Daily Service team", value=True, key="ri_do_ds")
        keep = st.checkbox("Keep people not on this sheet (mark absent instead of removing)",
                           value=False, key="ri_keep")
        st.caption(f"Then open the Schedule page and press **Generate** to build "
                   f"{picked:%b %d}'s charts.")

        if st.button("Apply to roster", type="primary", key="ri_apply"):
            applied_msg = []
            if do_hk:
                new_roster = dict(update["hk_roster"])
                if keep:
                    for name, v in st.session_state.get("hk_roster", {}).items():
                        if name not in new_roster:
                            new_roster[name] = {"building": v.get("building", 1), "present": False}
                st.session_state["hk_roster"] = new_roster
                for k in [k for k in list(st.session_state) if k.startswith("att_")]:
                    del st.session_state[k]
                applied_msg.append(f"{sum(1 for v in new_roster.values() if v['present'])} "
                                   f"housekeepers present of {len(new_roster)}")
            if do_rqs:
                new_insp = dict(update["insp_roster"])
                if keep:
                    for name, v in st.session_state.get("insp_roster", {}).items():
                        new_insp.setdefault(name, False)
                st.session_state["insp_roster"] = new_insp
                for k in [k for k in list(st.session_state) if k.startswith("insp_att_")]:
                    del st.session_state[k]
                st.session_state["rqs1"] = update["rqs1"]
                st.session_state["rqs2"] = update["rqs2"]
                # Point the Schedule page's RQS selectboxes at these people.
                # Clearing their keys instead would reset them to "no one" and
                # overwrite rqs1/rqs2 the next time that page runs.
                for k, v in (("rqs1_sel", update["rqs1"]), ("rqs2_sel", update["rqs2"])):
                    st.session_state[k] = v if (v and new_insp.get(v)) else RQS_NONE
                applied_msg.append(f"{sum(1 for v in new_insp.values() if v)} inspectors present")
            if do_ds:
                roster_now = st.session_state.get("hk_roster", {})
                team = ([n for n in update["ds_team"] if roster_now.get(n, {}).get("present")]
                        if do_hk else list(update["ds_team"]))
                st.session_state["ds_team"] = team
                applied_msg.append(f"{len(team)} on the Daily Service team")
            if not applied_msg:
                st.warning("Nothing selected to apply.")
            else:
                if do_hk or do_rqs:
                    try:
                        db.save_roster(st.session_state.get("hk_roster", {}),
                                       st.session_state.get("insp_roster", {}))
                    except Exception as ex:
                        st.warning(f"Applied to this session, but saving to the database "
                                   f"failed — it will not survive a reload. ({ex})")
                st.session_state["roster_import_note"] = (
                    f"{picked:%b %d, %Y} from sheet '{week['sheet']}'")
                st.success("Applied — " + "; ".join(applied_msg) + ".")
