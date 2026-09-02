"""
roster_import.py — Parse the weekly staff Schedule.xlsx into a day roster.

The workbook holds one sheet per week. Every sheet uses the same shape:

    col A            = section header, or a person's name
    cols B..H        = Sunday..Saturday
    a repeated row   = the real dates for those seven columns

Sheet names are unreliable (typos, duplicates, sheets copied without re-dating),
so a day is located by the DATES INSIDE each sheet, never by the sheet name.

Nothing here imports streamlit, so it can be unit-tested on its own.
"""
import io
import re
import datetime as _dt
import zoneinfo as _zi

#: The property's timezone. A UTC host rolls the date over at six in
#: the evening Mountain, which would age a week early.
_MTN = _zi.ZoneInfo("America/Denver")


def _local_today():
    return _dt.datetime.now(_MTN).date()

from collections import OrderedDict

#: Bump when a page starts relying on a helper added here. Pages compare it
#: against what they need and reload this module if the server handed them a
#: stale copy from a previous deploy — otherwise the first call to a new
#: function dies as an AttributeError inside a widget callback, which Streamlit
#: reports with the message redacted.
__version__ = 14

# ── Section headers (verified identical across sheets spanning a full year) ────
HK_SECTIONS = OrderedDict([
    ("housekeeper building 1", 1),
    ("housekeeper building 2", 2),
    ("housekeeper building 3", 3),
])
RQS_SECTION = "room quality supervisors"
OTHER_SECTIONS = OrderedDict([
    ("managers",        "Managers"),
    ("houseperson am",  "Houseperson AM"),
    ("am lead",         "AM Lead"),
    ("houseperson pm",  "Houseperson PM"),
    ("pm leads",        "PM Leads"),
    ("sales/ spa",      "Sales / Spa"),
    ("overnight team",  "Overnight Team"),
    ("ullr - ice rink", "ULLR - Ice Rink"),
])

# Rows that are structure or weekly summary metrics, never a person.
SKIP_LABELS = {
    "", ".", "events", "holds", "daily services", "check outs", "rqs needed",
    "hskp needed", "hskp actual #", "extras hskp",
}

def _norm(s) -> str:
    """Collapse whitespace and lowercase — for matching labels and names."""
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()

#: Lead rows carry the building in the name — "Jose M -BLD 1", "Willy bld 2",
#: "Luis R. - B2", "Willy - LEAD 2". Same person, different day. Stripping this
#: is safe because it is a suffix describing the shift, not part of the name.
_SUFFIX_RE = re.compile(
    r"[\s.,\-–]*\b(?:lead\s+)?(?:bld|bldg|building|b|lead)\s*[123]\s*$", re.I)

#: Column A sometimes holds something that is not a person at all — a stray
#: backtick, a lone digit, a filename someone pasted. A name needs at least two
#: letters and must not look like a file.
_NOT_A_NAME_RE = re.compile(r"\.(xlsx?|csv|pdf|docx?)\s*$", re.I)

#: Staff visiting from a sister property, written either way round:
#: "PEAK 7 - Adriana", "VICTOR S. - GL7", "Jaritza GL7".
_PROPERTY_RE = re.compile(r"\b(peak\s*7|gl\s*7)\b", re.I)

def split_property(label):
    """(person's name, home property). Property is "" for your own staff.

    The marker is stripped from the name so lists show the person, but it is
    kept alongside — confirmed as a different Cesar from the one in Building 3,
    so the property has to stay part of who they are.
    """
    s = re.sub(r"\s+", " ", str(label or "").strip())
    m = _PROPERTY_RE.search(s)
    if not m:
        return s, ""
    prop = "Peak 7" if "peak" in m.group(1).lower() else "GL7"
    bare = _PROPERTY_RE.sub("", s)
    bare = re.sub(r"\s{2,}", " ", bare).strip(" -–:,").strip()
    return (bare or s), prop

def looks_like_person(label) -> bool:
    s = str(label or "").strip()
    if not s or _NOT_A_NAME_RE.search(s):
        return False
    return sum(c.isalpha() for c in s) >= 2

def read_person(key, rec):
    """(name, home) for a stored person record, or (None, None) to skip it.

    Weeks saved by an earlier version kept the raw column-A text as the name
    and had no home property, so the same cleaning is applied on the way OUT of
    storage as on the way in. Without this, junk like a lone backtick keeps
    appearing in every list until someone re-uploads the workbook.
    """
    name = rec.get("name", key)
    if not looks_like_person(name):
        return None, None
    home = rec.get("home")
    if home is None:
        name, home = split_property(name)
    return name, home

#: Name variants confirmed by the manager as one person. Only what was actually
#: confirmed goes here — "Jhoselyn A"/"Jhoselyn M" are one person, while the
#: Daniels ("Daniela G", "Daniel T", "Daniel R", "Daniel Rdgz") are NOT and are
#: deliberately absent.
NAME_ALIASES = {
    "liz":               "liz salazar",
    "rigoberto":         "rigo garcia",
    "rigo g":            "rigo garcia",
    "josseling":         "josselin",
    "jenny caicedo":     "jenni caicedo",
    "jennifer c":        "jenni caicedo",
    "cecilia":           "cecilia angeles",
    "cecilia a":         "cecilia angeles",
    "elibeth":           "elibeth herrera",
    "elibeth h":         "elibeth herrera",
    "jhoselyn a":        "jhoselyn",
    "jhoselyn m":        "jhoselyn",
    # One person who has worked Building 1, Building 3, Houseperson AM and PM.
    # A different building in a different week never means a different person —
    # only the spelling of the name does.
    "junior d":          "junior torres",
    "junior t":          "junior torres",
    # Short form and full form of one person.
    "danny":             "danny r",
    "avelino":           "avelino rafael",
    "dianis chavez":     "dianis",
    "escarleth j":       "escarleth",
    "francys m":         "francys",
    "mauricio rivas":    "mauricio",
    "mauricio r":        "mauricio",
    "jose zendejas":     "jose z",       # Jose M is the AM Lead, a different person
    "nancy ramirez":     "nancy",
    # The Jennyfer spellings are one person, and NOT the same as Jenny Caicedo.
    "jennyfer c":        "jennyfer caicedo",
    "jennyfer":          "jennyfer caicedo",
    # Jorge Luis, however written. "JORGE" and "Jorge Lead" are other people.
    "jorge luis vides":  "jorge luis",
    "jorge l":           "jorge luis",
}

#: Preferred display name once variants are merged, where the fullest form is
#: not the one written most often. Without this the label would be whichever
#: spelling appears most, e.g. "Junior D" rather than "Junior Torres".
CANONICAL_LABELS = {
    "junior torres":  "Junior Torres",
    "liz salazar":    "Liz Salazar",
    "rigo garcia":    "Rigo Garcia",
    "cecilia angeles":"Cecilia Angeles",
    "elibeth herrera":"Elibeth Herrera",
    "jenni caicedo":  "Jenny Caicedo",
    "jennyfer caicedo": "Jennyfer Caicedo",
    "avelino rafael": "Avelino Rafael",
    "jorge luis":     "Jorge Luis",
}

def norm_name(s) -> str:
    """Key used to match a sheet name against an existing roster name."""
    n = _norm(s).rstrip(".").strip()
    prev = None
    while prev != n:                      # "Willy - LEAD 2 BLD 1" needs two passes
        prev = n
        n = _SUFFIX_RE.sub("", n).strip(" .,-–")
    return NAME_ALIASES.get(n, n)

# ── Cell status vocabulary ────────────────────────────────────────────────────
# Genuinely not at work. Note VTO is deliberately NOT here: it is paid, and a
# cell like "VTO - KEYSTONE" must not be read as ordinary time off.
_OFF_RE = re.compile(
    r"\b(r\s*[/-]?\s*off|off\s*/?\s*grant?ed|off|sick|fmla|fmlo|vacation|"
    r"call\s*out|p\s*-?\s*loa|ploa|loa)\b")
# Volunteered time off: paid, but not a worked day here. Keystone and FIRC are
# the sister properties — going there is volunteering away, however it is
# written: "VTO", "KEYSTONE VTO", "KEYSTONE", "FIRC", "VTO - FIRC".
_VTO_RE = re.compile(r"\b(vto|v\.t\.o\.?|keystone|firc)\b", re.I)
# At work but not on guest rooms. Used only to tell a recognised duty from an
# unfamiliar one — both count as a worked day either way.
_KNOWN_OTHER_RE = re.compile(
    r"\b(hsp|ullr|garages?|projects?|food|maps|stripping\s+linen|rollaways|"
    r"lavar\s+botes|cleaning\s+(windows|carpets)|breck\s*in|training|eng|"
    r"help\s+with\s+party|safety\s+meeting|stairs\s*only|baseboards|"
    r"listening\s+session|conference|meeting|spiff|touch\s*up|sales|"
    r"am\s+team|voluntario|deep\s+clean|inventory|inventario|windows)\b")
# Inspector role codes.
_RQS_RE = re.compile(r"^(rqs?\s*[12]|rq\s*[12])\b")
# A leading number is the person's chart/zone load — they are working.
_LEAD_NUM_RE = re.compile(r"^\d+(\.\d+)?\s*(\+|-|/|$|\s)")

KIND_WORKING = "working"
KIND_DAILY   = "daily_service"
KIND_OFF     = "off"
KIND_OTHER   = "other_duty"
KIND_VTO     = "vto"
KIND_NOCALL  = "no_call"
KIND_UNKNOWN = "unknown"

#: Available to be given guest rooms today.
PRESENT_KINDS = (KIND_WORKING, KIND_DAILY)

#: On the clock. Projects, garages, HSP cover and the like are a working day —
#: the person is at work and paid for it, they are simply not on guest rooms.
WORKED_KINDS = (KIND_WORKING, KIND_DAILY, KIND_OTHER)

#: Paid, but not a worked day. VTO is volunteered time off: still paid, and
#: deliberately counted apart from cleaning and other duty.
PAID_OFF_KINDS = (KIND_VTO,)

#: Cell text that says a no-call/no-show outright, independent of the red fill.
_NOCALL_RE = re.compile(r"\bn[c/\s-]*n[s/\s-]*h?\b|\bno\s*call\b|\bno\s*show\b", re.I)

#: Voluntary time off, when that is the whole story for the day. A cell that
#: leads with hours worked ("3 + VTO") is a worked day that ended early, and
#: "VTO - KEYSTONE" is time off spent at another property.
_VTO_ONLY_RE = re.compile(r"^\s*(vto|v\.t\.o\.?)\s*$", re.I)

def classify_full(raw):
    """(kind, recognised) for one schedule cell.

    The rule is deliberately inverted from where this started: a day is a
    WORKING day unless it is off, sick, a no-call, or volunteered away. Cooking
    for the staff, projects, garages, deep cleans — the person is at work and
    paid, so it counts. An unfamiliar code therefore counts as worked too, and
    is merely flagged as unrecognised rather than silently written off.

    Order matters. What the cell LEADS with wins: "3 + VTO" is a full shift
    that ended early and "VTO + Daily Service" is a daily-service shift, even
    though both mention VTO.
    """
    s = re.sub(r"\s+", " ", str(raw or "").strip())
    if not s:
        return KIND_OFF, True
    low = s.lower()
    if _NOCALL_RE.search(low) or "call sick" in low:
        return KIND_NOCALL, True
    if "daily service" in low:
        return KIND_DAILY, True
    if _LEAD_NUM_RE.match(s):
        return KIND_WORKING, True
    if _RQS_RE.match(low):
        return KIND_WORKING, True
    if low in ("on", "on am", "on pm", "hskp", "housekeeper") or re.match(r"^on\b", low):
        return KIND_WORKING, True
    if _OFF_RE.search(low):
        return KIND_OFF, True
    if _VTO_RE.search(low):
        return KIND_VTO, True
    if "lead" in low:
        return KIND_WORKING, True
    # Anything left is a duty of some sort — a worked day, recognised or not.
    return KIND_OTHER, bool(_KNOWN_OTHER_RE.search(low))

def classify(raw) -> str:
    return classify_full(raw)[0]

def is_recognised(raw) -> bool:
    """False for a duty we have no name for. It still counts as worked."""
    return classify_full(raw)[1]

# ── Controlled vocabulary for editing ─────────────────────────────────────────
# One canonical option list per role, ordered by how often each value actually
# appears in the workbook. Picking from these instead of typing is what stops
# the drift already in the data — "Daily service" / "Daily Service" /
# "Daily services" are three spellings of one thing, and "6 + 7" / "6+7" /
# "6 +7" three of another.
BLANK_LABEL = "— off / blank —"

_COMMON_OFF = ["R/OFF", "OFF GRANTED", "SICK", "FMLA", "PLOA", "VTO"]

ROLE_OPTIONS = {
    "hk": ["3", "Daily Service", "ON", "3 + VTO"] + _COMMON_OFF +
          ["HSP AM", "HSP PM", "KEYSTONE", "GARAGES", "PROJECTS", "BRECK IN",
           "FOOD", "ULLR", "Stripping linen"],
    "rqs": ["ON", "RQS 1", "RQS 2", "RQ2 + dust and vac inspect", "RQS 1 + RQS 2"]
           + _COMMON_OFF +
           ["ENG", "ULLR", "KEYSTONE", "Rollaways 8-9am and then, Cleaning carpets"],
    "Houseperson AM": ["1", "2", "3", "4", "5", "6", "7", "6 + 7"] + _COMMON_OFF +
                      ["Stripping linen", "Rollaways 8-9am and then, Cleaning carpets",
                       "KEYSTONE"],
    "Houseperson PM": ["1", "2", "3", "4", "5", "6", "7", "6 + 7"] + _COMMON_OFF +
                      ["Stripping linen", "KEYSTONE"],
    "Managers":       ["ON", "ON AM", "ON PM"] + _COMMON_OFF + ["Conference"],
    "AM Lead":        ["Lead", "Lead AM", "2 Lead -8:30 am"] + _COMMON_OFF,
    "PM Leads":       ["Lead", "Lead AM", "Lead - PM", "Lead PM"] + _COMMON_OFF,
    "Sales / Spa":    ["ON"] + _COMMON_OFF,
    "Overnight Team": ["ON"] + _COMMON_OFF,
    "ULLR - Ice Rink":["ON", "ULLR"] + _COMMON_OFF,
}

def role_key(rec) -> str:
    """Which option list a person's row uses."""
    g = rec.get("group")
    return "hk" if g == "hk" else ("rqs" if g == "rqs" else rec.get("section", ""))

def options_for(rec, present_values=(), extra=()) -> list:
    """Dropdown choices for one person's row.

    Anything already in the sheet is appended so an existing one-off never
    disappears when its cell is opened — the curated values simply come first.
    """
    opts = list(ROLE_OPTIONS.get(role_key(rec), []))
    seen = {o.casefold() for o in opts}
    for v in list(extra) + list(present_values):
        v = str(v or "").strip()
        if v and v.casefold() not in seen:
            opts.append(v); seen.add(v.casefold())
    return [BLANK_LABEL] + opts

# ── Legends, transcribed from the workbook's own side columns ────────────────
#: Housekeeper cells lead with a number: their 8-10am extra task. They clean
#: rooms the rest of the day, so the number never means "not available".
#: (Sheet "Jan 5 - 11" cols J-L; wording drifts slightly between weeks.)
HK_TASK_LEGEND = {
    "1": "Breakroom - Bld 2 or 3 (clean and tidy)",
    "2": "Linea vieja / mattress pad - check cart closet",
    "3": "Vacuum & cart - keep clean and serviceable",
    "4": "Amenities bottles - clean, refill, organise",
    "5": "Garages - pick up trash B1/B2/B3",
}
#: Houseperson cells are a zone number, not a task.
#: (Sheet "Dec 14 - 20": "Task for houseperson according to their number".)
HP_ZONE_LEGEND = {
    "1": "Runner Building 2",
    "2": "Lobby",
    "3": "Pool - Building 1",
    "4": "Runner Building 1",
    "5": "Runner Building 3",
    "6": "Pool - Building 3",
    "7": "Plaza - Building 3",
}

def legend_for(rec, raw):
    """Plain-English meaning of a leading number, by role. '' when none."""
    m = re.match(r"^(\d+)", str(raw or "").strip())
    if not m:
        return ""
    n = m.group(1)
    if rec.get("group") == "hk":
        return HK_TASK_LEGEND.get(n, "")
    if "Houseperson" in str(rec.get("section", "")):
        return HP_ZONE_LEGEND.get(n, "")
    return ""

# ── Non-housekeepers drafted onto rooms ───────────────────────────────────────
#: On short days a houseperson (or lead) is marked to clean rooms instead.
#: The sheet says so in words -- "HSKP", "HOUSEKEEPER", "cleaning Rooms",
#: "Stripping linen or HSKP (cleaning Rooms)" -- so those people are genuine
#: room cleaners for the day and belong in the housekeeper roster.
_COVER_RE = re.compile(r"\b(hskp|housekeep\w*|cleaning\s+rooms?)\b", re.I)
#: An RQS who "opens the office" and assigns daily services on Hotsos is doing
#: desk work, not cleaning -- see the RQS 1/RQS 2 + FLOATING legend. Without
#: this, "Opening the office + Daily services - Send the report" would draft an
#: inspector onto rooms they are not actually cleaning.
_ADMIN_RE = re.compile(
    r"\b(office|assign\w*|report\w*|reporte|enviar|hotsos|send|email|calendar)\b", re.I)

def is_room_cover(raw) -> bool:
    """True when a non-housekeeper's cell puts them on guest rooms today."""
    s = str(raw or "")
    if not s:
        return False
    if _COVER_RE.search(s):          # says HSKP / HOUSEKEEPER outright
        return True
    if _ADMIN_RE.search(s):          # desk work that merely mentions the words
        return False
    return "daily service" in s.lower()

def cover_building(raw, default=1) -> int:
    """Building named in a cover cell, else Building 1.

    Building 1 is the safe default: under the movement rule a B1 housekeeper
    may work any building, so a drafted-in helper stays assignable anywhere.
    """
    m = re.search(r"\bbld\s*([123])\b|\bbuilding\s*([123])\b", str(raw or ""), re.I)
    if m:
        return int(m.group(1) or m.group(2))
    return default

def rqs_role(raw):
    """Return 1 or 2 when a cell names an explicit RQS role, else None."""
    low = _norm(raw)
    if re.search(r"\b(rqs?\s*1|rq\s*1)\b", low): return 1
    if re.search(r"\b(rqs?\s*2|rq\s*2)\b", low): return 2
    return None

# ── Sheet scanning ────────────────────────────────────────────────────────────
def sheet_dates(ws, max_scan_rows=200):
    """Map date -> column index for one worksheet.

    The date header repeats several times down a sheet; every repeat carries the
    same seven dates, so later rows simply confirm the same mapping.
    """
    out = {}
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        for c in range(1, min(ws.max_column, 12) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, (_dt.datetime, _dt.date)):
                d = v.date() if isinstance(v, _dt.datetime) else v
                out.setdefault(d, c)
    return out

def index_workbook(wb):
    """Build {date: [(sheet_name, col), ...]} across the whole workbook."""
    idx = {}
    for ws in wb.worksheets:
        for d, c in sheet_dates(ws).items():
            idx.setdefault(d, []).append((ws.title, c))
    return idx

def available_dates(wb):
    return sorted(index_workbook(wb).keys())

def _walk_people_rows(ws):
    """Yield (row, name, group, building_or_label) for every person row.

    Column A carries both section headers and names, so the current section is
    tracked as we descend. Shared by the single-day and whole-week readers.
    """
    section = None          # ("hk", 1|2|3) | ("rqs", None) | ("other", label) | None
    for r in range(1, ws.max_row + 1):
        label_raw = ws.cell(r, 1).value
        label = _norm(label_raw)
        if label in SKIP_LABELS or label.startswith("schedule subject to change"):
            continue
        if label in HK_SECTIONS:
            section = ("hk", HK_SECTIONS[label]); continue
        if label == RQS_SECTION:
            section = ("rqs", None); continue
        # Sheet "Nov 2 - 8" has no "Room Quality Supervisors" header at all, so
        # every inspector fell under Building 3 and showed up as a duplicate of
        # themselves. This note only ever sits at the top of the RQS block, so
        # it doubles as a fallback marker; where the real header exists this
        # changes nothing.
        if label.startswith("please, assign to each"):
            if section != ("rqs", None):
                section = ("rqs", None)
            continue
        # A long sentence with no days against it is a note someone typed into
        # the name column, not a person -- e.g. "Stripping linen or HSKP
        # (cleaning Rooms)" sitting in column A of that same sheet.
        if len(label) > 30 and " " in label and all(
                ws.cell(r, c).value in (None, "") for c in range(2, 9)):
            continue
        matched_other = next((v for k, v in OTHER_SECTIONS.items() if label.startswith(k)), None)
        if matched_other:
            section = ("other", matched_other); continue
        if section is None or not label:
            continue
        # A date row in column A would already have been skipped as non-text.
        if isinstance(label_raw, (_dt.datetime, _dt.date)):
            continue
        name = re.sub(r"\s+", " ", str(label_raw).strip())
        if not looks_like_person(name):
            continue
        kind, bld = section
        yield r, name, kind, bld

def cell_kind(rec, iso):
    """Status of one person on one day, red fill included.

    The red no-call marks live beside the cells, not in them, and most such
    cells are blank — so classifying the text alone reports "off" and the
    no-call disappears everywhere except the history. Everything that turns a
    stored week into a status goes through here.
    """
    raw = (rec.get("cells") or {}).get(iso, "")
    if iso in (rec.get("nocall") or []):
        return KIND_NOCALL, True
    return classify_full(raw)

def _person_record(name, group, bld, raw, kind=None, known=None):
    raw = re.sub(r"\s+", " ", str(raw or "").strip())
    if kind is None:
        kind, known = classify_full(raw)
    return {
        "name":     name,
        "section":  {"hk": f"Building {bld}", "rqs": "RQS", "other": bld}[group],
        "group":    group,
        "building": bld if group == "hk" else None,
        "raw":      raw,
        "kind":     kind,
        "present":  kind in PRESENT_KINDS,
        "worked":   kind in WORKED_KINDS,
        "recognised": known,
        "daily_service": kind == KIND_DAILY,
    }

def parse_day(ws, col):
    """Read one day column and return the people found, by section.

    Returns a list of dicts: name, section, building (housekeepers only),
    raw cell text, kind, present, daily_service.
    """
    return [_person_record(name, group, bld, ws.cell(r, col).value)
            for r, name, group, bld in _walk_people_rows(ws)]

#: Pure red fill on a person's cell means they called off at the last minute
#: without notice. Confirmed against the workbook: of 454 such cells, 392 are
#: blank and the rest spell it out — "NC/NS", "NC/NSH", "call sick". The pale
#: pink FFF4CCCC is NOT this; it is how Lead rows are shaded.
NOCALL_FILLS = {"FFFF0000", "FFCC0000", "FFE06666"}

def _is_nocall_fill(cell) -> bool:
    try:
        fill = cell.fill
        if not fill or fill.patternType is None:
            return False
        rgb = getattr(fill.fgColor, "rgb", None)
        return isinstance(rgb, str) and rgb.upper() in NOCALL_FILLS
    except Exception:
        return False

def parse_week(ws, ws_styles=None):
    """Read a whole sheet in one pass.

    Returns {"sheet", "dates": [iso...], "people": {name: {...,"cells": {iso: raw}}}}
    — the compact shape that gets stored, diffed and rendered as a week grid.

    `ws_styles` is the same sheet from a workbook opened WITHOUT data_only, so
    cell fills are readable; openpyxl gives values or styles, never both.
    """
    dates = sheet_dates(ws)
    if not dates:
        return None
    ordered = sorted(dates)
    people = OrderedDict()
    for r, name, group, bld in _walk_people_rows(ws):
        section = {"hk": f"Building {bld}", "rqs": "RQS", "other": bld}[group]
        name, home = split_property(name)
        # The same first name can appear in two sections — the real sheet has a
        # Leonardo in Building 2 AND a Leonardo on the Overnight Team. Keying on
        # the name alone would merge them and let one person's shift decide the
        # other's attendance, so a repeat in a DIFFERENT section gets its own id.
        key = name
        if key in people and people[key]["section"] != section:
            key = f"{name} · {section}"
        rec = people.setdefault(key, {
            "name": name, "home": home,
            "group": group, "building": bld if group == "hk" else None,
            "section": section,
            "row": r, "cells": {}, "nocall": [],
        })
        for d in ordered:
            raw = re.sub(r"\s+", " ", str(ws.cell(r, dates[d]).value or "").strip())
            if raw:
                rec["cells"][d.isoformat()] = raw
            if ws_styles is not None and _is_nocall_fill(ws_styles.cell(r, dates[d])):
                rec["nocall"].append(d.isoformat())
    cols = {d.isoformat(): dates[d] for d in ordered}
    return {"sheet": ws.title,
            "dates": [d.isoformat() for d in ordered],
            # Column index per date, captured now so writing edits back never has
            # to re-derive it from a workbook opened without cached values.
            "cols": cols,
            # The hidden headcount block. ws_styles is the same sheet opened
            # without cached values, which is where the formulas -- and so the
            # labour minutes -- can be read.
            "metrics": parse_metrics(ws, ws_styles, cols),
            "people": people}

def needs_fill_backfill(week) -> bool:
    """True for a week stored before cell fills were read.

    A week planned in the app has no sheet behind it and never will, so it is
    not counted as missing anything.
    """
    if not week or week.get("planned"):
        return False
    return not any("nocall" in rec for rec in week.get("people", {}).values())

def merge_fill_data(stored, parsed):
    """Copy the red no-call marks from a fresh parse onto a stored week.

    Only the fill-derived fields are taken. The cells themselves are left
    exactly as stored, so nothing a manager typed is disturbed, and in-app
    edits live in a separate overrides record anyway.

    Returns (updated_week, marks_added) or (None, 0) when there is nothing to do.
    """
    if not stored or not parsed:
        return None, 0
    by_row, by_name = {}, {}
    for key, rec in parsed.get("people", {}).items():
        if rec.get("row"):
            by_row[rec["row"]] = rec
        by_name[norm_name(rec.get("name", key))] = rec
    added = 0
    for key, rec in stored.get("people", {}).items():
        src = by_row.get(rec.get("row")) or by_name.get(norm_name(rec.get("name", key)))
        marks = list((src or {}).get("nocall") or [])
        rec["nocall"] = marks
        if src is not None and "home" in src and "home" not in rec:
            rec["home"] = src["home"]
        added += len(marks)
    return stored, added

def week_key(week) -> str:
    """Stable id for a week — the ISO date of its first (Sunday) column."""
    return week["dates"][0]

def parse_all_weeks(wb, wb_styles=None):
    """Parse every dated sheet. Later sheets win a duplicated week key.

    `wb_styles` is the same workbook opened without data_only, used only to
    read the red no-call fills.
    """
    out = {}
    for ws in wb.worksheets:
        styled = None
        if wb_styles is not None and ws.title in wb_styles.sheetnames:
            styled = wb_styles[ws.title]
        wk = parse_week(ws, styled)
        if wk:
            out[week_key(wk)] = wk
    return out

# ── Comparing an upload against what is already stored ────────────────────────
def diff_week(old, new):
    """Cell-level differences between two parses of the same week."""
    changed, added_people, removed_people = [], [], []
    o_people = (old or {}).get("people", {})
    n_people = new.get("people", {})
    for name, rec in n_people.items():
        if name not in o_people:
            added_people.append(name); continue
        o_cells, n_cells = o_people[name].get("cells", {}), rec.get("cells", {})
        for iso in sorted(set(o_cells) | set(n_cells)):
            ov, nv = o_cells.get(iso, ""), n_cells.get(iso, "")
            if ov != nv:
                changed.append({"name": name, "date": iso, "old": ov, "new": nv})
    removed_people = [n for n in o_people if n not in n_people]
    return {"changed": changed, "added_people": added_people,
            "removed_people": removed_people}

def diff_all(stored, incoming):
    """Compare a whole upload against the stored weeks."""
    new_weeks = [k for k in incoming if k not in stored]
    gone_weeks = [k for k in stored if k not in incoming]
    per_week = {}
    for k in incoming:
        if k in stored:
            d = diff_week(stored[k], incoming[k])
            if d["changed"] or d["added_people"] or d["removed_people"]:
                per_week[k] = d
    return {"new_weeks": sorted(new_weeks), "gone_weeks": sorted(gone_weeks),
            "changed_weeks": per_week,
            "n_changed_cells": sum(len(d["changed"]) for d in per_week.values())}

# ── In-app edits ──────────────────────────────────────────────────────────────
def override_key(wk: str, name: str, iso: str) -> str:
    return f"{wk}|{name}|{iso}"

def apply_overrides(week, overrides, wk=None):
    """Overlay in-app edits onto a parsed week.

    Returns (week_copy, {(name, iso): {"value","excel"}}). The Excel value is
    kept alongside so the grid can show what was replaced and offer a revert.
    """
    wk = wk or week_key(week)
    out = {"sheet": week["sheet"], "dates": list(week["dates"]),
           "people": OrderedDict()}
    applied = {}
    for name, rec in week["people"].items():
        if read_person(name, rec)[0] is None:
            continue                      # junk row from an older parse
        cells = dict(rec.get("cells", {}))
        for iso in week["dates"]:
            ov = overrides.get(override_key(wk, name, iso))
            if ov is None:
                continue
            excel_val = cells.get(iso, "")
            new_val = str(ov.get("value", "") or "")
            if new_val:
                cells[iso] = new_val
            else:
                cells.pop(iso, None)
            applied[(name, iso)] = {"value": new_val, "excel": excel_val,
                                    "by": ov.get("by", ""), "at": ov.get("at", "")}
        out["people"][name] = {**rec, "cells": cells}
    return out, applied

def find_week_key(week_keys, iso):
    """The stored week containing this date, or None.

    A week key is its Sunday; a date belongs to it only if it lands inside the
    seven days that follow — otherwise a gap in the stored weeks would silently
    resolve to a much earlier week.
    """
    earlier = [w for w in sorted(week_keys) if w <= iso]
    if not earlier:
        return None
    wk = earlier[-1]
    try:
        delta = (_dt.date.fromisoformat(iso) - _dt.date.fromisoformat(wk)).days
    except ValueError:
        return None
    return wk if 0 <= delta <= 6 else None

def day_roster(week, overrides, wk, iso, existing_roster=None):
    """Everything the scheduler needs for one date, in-app edits included."""
    eff, _applied = apply_overrides(week, overrides or {}, wk)
    if iso not in eff["dates"]:
        return None
    return build_roster_update(week_to_people(eff, iso), existing_roster)

def merge_roster(update, existing_roster, keep_missing=True):
    """Combine a parsed day with the standing roster.

    keep_missing leaves anyone absent from the sheet on the roster, marked not
    present, rather than deleting them — the safe choice when this runs
    automatically and nobody is watching.
    """
    new_roster = dict(update["hk_roster"])
    if keep_missing:
        for name, v in (existing_roster or {}).items():
            if name not in new_roster:
                new_roster[name] = {"building": v.get("building", 1), "present": False}
    return new_roster

# ── Attendance history ────────────────────────────────────────────────────────
def history_rows(weeks, overrides=None, upto=None):
    """Flatten every stored week into one row per person per dated day.

    Rows carry the role, the raw cell, the classified kind and whether it
    counts as worked — enough to answer "how many days has this person worked
    this month, and doing what".
    """
    overrides = overrides or {}
    out = []
    for wk in sorted(weeks):
        week = weeks[wk]
        eff, _applied = apply_overrides(week, overrides, wk)
        for key, rec in eff["people"].items():
            name, home = read_person(key, rec)
            if name is None:
                continue
            for iso in eff["dates"]:
                if upto and iso > upto:
                    continue
                raw = rec.get("cells", {}).get(iso, "")
                kind, known = cell_kind(rec, iso)
                out.append({
                    "date": iso, "week": wk, "person": name,
                    # Identity across weeks. The sheet writes the same person
                    # several ways -- "Hari" in 39 weeks, "HARI" in 2 -- and an
                    # exact-key match would split their history in half.
                    # Confirmed by the manager: a name appearing in two teams is
                    # one person doing both jobs, so the team is NOT part of the
                    # id. The storage key inside a week still keeps them apart,
                    # which is what stops one row's cells overwriting another's.
                    "pid": norm_name(name) + (f"@{home}" if home else ""),
                    "home": home or "",
                    "key": key, "section": rec["section"], "group": rec["group"],
                    "building": rec.get("building"),
                    "raw": raw, "kind": kind,
                    "worked": kind in WORKED_KINDS,
                    "recognised": known,
                    "paid_off": kind in PAID_OFF_KINDS,
                    "no_call": kind == KIND_NOCALL,
                    "daily_service": kind == KIND_DAILY,
                    "cover": rec["group"] != "hk" and is_room_cover(raw),
                    "dow": _dt.date.fromisoformat(iso).strftime("%a"),
                })
    return out

def duplicate_candidates(index, threshold=0.86):
    """Names that look like the same person but are not merged.

    Building suffixes are handled in norm_name, so what is left needs a human:
    "Jhoselyn A" and "Jhoselyn M" are 90% similar and are two different people,
    while "JENNI CAICEDO" and "Jenny Caicedo" are one. This reports; it never
    merges on its own.
    """
    import difflib
    by_group = {}
    for pid, info in index.items():
        by_group.setdefault(info["group"], []).append((info["label"], info["n"]))
    out = []
    for items in by_group.values():
        items.sort()
        for i in range(len(items)):
            for j in range(i + 1, len(items)):
                (a, na), (b, nb) = items[i], items[j]
                ratio = difflib.SequenceMatcher(None, norm_name(a), norm_name(b)).ratio()
                if ratio >= threshold:
                    out.append((a, b, round(ratio, 3), na, nb))
    return sorted(out, key=lambda x: -x[2])

def people_index(rows):
    """{pid: {label, group, sections, n}} — one entry per real person.

    The label is the spelling that appears most often, so a name written in
    caps a couple of times does not become the display name.
    """
    spellings, groups, sections, counts, homes = {}, {}, {}, {}, {}
    for r in rows:
        pid = r["pid"]
        homes[pid] = r.get("home", "")
        spellings.setdefault(pid, {})
        spellings[pid][r["person"]] = spellings[pid].get(r["person"], 0) + 1
        groups[pid] = r["group"]
        sections.setdefault(pid, set()).add(r["section"])
        counts[pid] = counts.get(pid, 0) + 1
    return {pid: {"label": CANONICAL_LABELS.get(
                      pid.split("@")[0], max(sp.items(), key=lambda x: x[1])[0]),
                  "group": groups[pid], "sections": sorted(sections[pid]),
                  "home": homes.get(pid, ""), "n": counts[pid]}
            for pid, sp in spellings.items()}

def match_person(index, *candidates):
    """Best pid for a sign-in name, or None.

    Tries each candidate whole, then with trailing digits stripped — app
    usernames look like "hari8059" while the sheet just says "Hari" — then on
    first name alone.
    """
    names = []
    for c in candidates:
        # Usernames separate words with dots or underscores where the sheet
        # uses spaces: "jenny.caicedo" is "Jenny Caicedo".
        c = norm_name(re.sub(r"[._\-]+", " ", str(c or "")))
        if not c:
            continue
        names.append(c)
        stripped = re.sub(r"\d+\s*$", "", c).strip()
        if stripped and stripped != c:
            names.append(stripped)
    for want in names:
        for pid, info in index.items():
            if norm_name(info["label"]) == want:
                return pid
    for want in names:
        head = want.split(" ")[0]
        if len(head) < 3:
            continue
        for pid, info in index.items():
            if norm_name(info["label"]).split(" ")[0] == head:
                return pid
    return None

def summarise_person(rows, pid=None, person_key=None, person=None):
    """Totals and habits for one person, from history_rows output."""
    if pid is not None:
        mine = [r for r in rows if r["pid"] == pid]
    else:
        mine = [r for r in rows
                if (r["key"] == person_key if person_key else r["person"] == person)]
    worked = [r for r in mine if r["worked"]]
    no_calls = [r for r in mine if r.get("no_call")]
    paid_off = [r for r in mine if r.get("paid_off")]
    dow_counts = {}
    for r in worked:
        dow_counts[r["dow"]] = dow_counts.get(r["dow"], 0) + 1
    roles, shifts = {}, {}
    for r in worked:
        roles[r["section"]] = roles.get(r["section"], 0) + 1
        label = ("Daily Service" if r["daily_service"]
                 else "Room cover" if r["cover"]
                 else "Other duty" if r["kind"] == KIND_OTHER else "Regular")
        shifts[label] = shifts.get(label, 0) + 1
    order = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    return {
        "rows": mine, "worked": worked,
        "n_worked": len(worked), "n_days": len(mine),
        "off": len([r for r in mine if r["kind"] == KIND_OFF]),
        "no_calls": no_calls, "n_no_call": len(no_calls),
        "paid_off": paid_off, "n_paid_off": len(paid_off),
        "by_dow": {d: dow_counts.get(d, 0) for d in order},
        "roles": dict(sorted(roles.items(), key=lambda x: -x[1])),
        "shifts": dict(sorted(shifts.items(), key=lambda x: -x[1])),
        "usual_days": [d for d in order
                       if dow_counts.get(d, 0) >= max(1, max(dow_counts.values()) * 0.6)]
                      if dow_counts else [],
    }

def period_bounds(today=None):
    """(week_start, month_start, year_start) as ISO strings, weeks Sunday-based."""
    today = today or _local_today()
    week_start = today - _dt.timedelta(days=(today.weekday() + 1) % 7)
    return (week_start.isoformat(),
            today.replace(day=1).isoformat(),
            today.replace(month=1, day=1).isoformat())

# ── Planning a week that does not exist yet ───────────────────────────────────
#: How fast older weeks stop mattering. 0.85 per week back means a month ago
#: carries about half the weight of last week — recent enough to follow a
#: rota change, long enough to see a pattern.
_DECAY = 0.85

#: A houseperson cell that is purely a zone: "4", "6 + 7", "6+7".
_ZONE_ONLY_RE = re.compile(r"^[1-7](\s*\+\s*[1-7])*$")

def zone_of(raw):
    """The zone a houseperson cell names, or None if it is not a zone cell."""
    s = re.sub(r"\s+", " ", str(raw or "").strip())
    return s if _ZONE_ONLY_RE.match(s) else None

def suggest_building(rows, recent_weeks=8, today=None):
    """Which building each housekeeper belongs in.

    Measured over the whole workbook, buildings barely move: only 1.5% of
    week-to-week transitions are a switch and 106 of 121 housekeepers spend
    90%+ of their weeks in one building. So the answer is simply where they
    have settled — but a move, when it happens, is a permanent transfer rather
    than a rotation, so the recent window wins and the change is flagged.
    """
    today = today or _local_today()
    cut = (today - _dt.timedelta(weeks=recent_weeks)).isoformat()
    allb, recb, label = {}, {}, {}
    for r in rows:
        if r["group"] != "hk" or not r.get("building"):
            continue
        b = r["building"]
        allb.setdefault(r["pid"], {})
        allb[r["pid"]][b] = allb[r["pid"]].get(b, 0) + 1
        label[r["pid"]] = r["person"]
        if r["date"] >= cut:
            recb.setdefault(r["pid"], {})
            recb[r["pid"]][b] = recb[r["pid"]].get(b, 0) + 1
    out = {}
    for pid, counts in allb.items():
        settled = max(counts.items(), key=lambda x: x[1])[0]
        rec = recb.get(pid) or counts
        recent = max(rec.items(), key=lambda x: x[1])[0]
        out[pid] = {
            "person": label[pid],
            "building": recent,
            "settled": settled,
            "moved": recent != settled,
            "confidence": rec[recent] / sum(rec.values()),
            "history": dict(sorted(counts.items())),
        }
    return out

def suggest_week(rows, target_start, lookback_weeks=16, today=None):
    """Predict a week's cells from each person's own history.

    For every person and weekday, the values they have had on that weekday are
    tallied with an exponential recency weight; the heaviest wins. This models
    the thing that actually drives the sheet — people work the same days and do
    the same job on them — and it explains itself, which a language model
    guessing at names could not.

    Returns {pid: {iso: {"value", "confidence", "basis", "n"}}}.
    """
    start = _dt.date.fromisoformat(target_start)
    cutoff = (start - _dt.timedelta(weeks=lookback_weeks)).isoformat()
    horizon = (today or _local_today()).isoformat()
    # Only learn from days that have actually happened and predate the target.
    past = [r for r in rows
            if cutoff <= r["date"] < target_start and r["date"] <= horizon]

    tally = {}          # pid -> dow -> value -> weight
    seen = {}           # pid -> dow -> observation count
    for r in past:
        weeks_back = max((start - _dt.date.fromisoformat(r["date"])).days, 0) / 7.0
        w = _DECAY ** weeks_back
        d = tally.setdefault(r["pid"], {}).setdefault(r["dow"], {})
        d[r["raw"]] = d.get(r["raw"], 0.0) + w
        seen.setdefault(r["pid"], {})
        seen[r["pid"]][r["dow"]] = seen[r["pid"]].get(r["dow"], 0) + 1

    # Housepersons are a different problem. Their cell is a ZONE, and zones do
    # not follow the weekday: predicting from weekday scores 35.9%, while simply
    # carrying forward the zone they last worked scores 54.4%. They hold a zone
    # for a stretch (median 2 consecutive working days, mean 3.4) then move to
    # another, so persistence beats frequency. Whether they work at all is still
    # a weekday question, so only the zone itself is overridden.
    last_zone = {}
    for r in sorted(past, key=lambda x: x["date"]):
        if "Houseperson" in str(r.get("section", "")) and zone_of(r["raw"]):
            last_zone[r["pid"]] = r["raw"]

    out = {}
    for pid, bydow in tally.items():
        for i in range(7):
            d = start + _dt.timedelta(days=i)
            dow = d.strftime("%a")
            counts = bydow.get(dow)
            if not counts:
                continue
            total = sum(counts.values()) or 1.0
            ranked = sorted(counts.items(), key=lambda x: -x[1])
            value, weight = ranked[0]
            model = "weekday"
            if zone_of(value) and pid in last_zone:
                carried = last_zone[pid]
                if carried != value:
                    model = "last zone"
                    value = carried
            out.setdefault(pid, {})[d.isoformat()] = {
                "value": value,
                "confidence": weight / total,
                "n": seen[pid].get(dow, 0),
                "model": model,
                "basis": [(v, round(w / total, 3)) for v, w in ranked[:3]],
            }
    return out

def balance_zones(week, rows):
    """Stop a draft from putting two housepersons on the same zone.

    Carrying each person's last zone forward independently makes collisions
    that the real sheet does not have — zones are doubled on only 1% of actual
    shift-days. Where two people land on one zone, it stays with whoever has
    worked it most, and the other moves to an uncovered zone, preferring one
    they have actually worked before.

    Mutates and returns the week. Reports the moves it made.
    """
    hist = {}
    for r in rows:
        z = zone_of(r["raw"])
        if z and "Houseperson" in str(r.get("section", "")):
            hist.setdefault(r["pid"], {})
            hist[r["pid"]][z] = hist[r["pid"]].get(z, 0) + 1
    moves = []
    for section in ("Houseperson AM", "Houseperson PM"):
        members = [(k, rec) for k, rec in week["people"].items()
                   if rec.get("section") == section]
        if not members:
            continue
        for iso in week["dates"]:
            holders = {}
            for k, rec in members:
                z = zone_of(rec.get("cells", {}).get(iso, ""))
                if z:
                    holders.setdefault(z, []).append((k, rec))
            taken = set(holders)
            for z, who in list(holders.items()):
                if len(who) < 2:
                    continue
                pid_of = lambda rec, k: norm_name(rec.get("name", k))
                who.sort(key=lambda kr: -hist.get(pid_of(kr[1], kr[0]), {}).get(z, 0))
                for k, rec in who[1:]:
                    free = [c for c in "1234567" if c not in taken]
                    if not free:
                        break
                    pid = pid_of(rec, k)
                    free.sort(key=lambda c: -hist.get(pid, {}).get(c, 0))
                    new = free[0]
                    rec["cells"][iso] = new
                    taken.add(new)
                    moves.append({"person": rec.get("name", k), "date": iso,
                                  "section": section, "from": z, "to": new})
    week["zone_moves"] = moves
    return week

def zone_coverage(week):
    """Zones missing or doubled on each houseperson shift, per day.

    Full 1-7 cover happens on only 39% of shift-days in the real workbook, so
    this reports rather than enforces — but a doubled zone is rare (1%) and
    usually a mistake worth seeing.
    """
    report = {}
    for section in ("Houseperson AM", "Houseperson PM"):
        members = [rec for rec in week["people"].values()
                   if rec.get("section") == section]
        if not members:
            continue
        for iso in week["dates"]:
            zones = []
            for rec in members:
                z = zone_of(rec.get("cells", {}).get(iso, ""))
                if z:
                    zones.extend(p.strip() for p in z.split("+"))
            if not zones:
                continue
            missing = [z for z in "1234567" if z not in zones]
            dupes = sorted({z for z in zones if zones.count(z) > 1})
            if missing or dupes:
                report.setdefault(section, {})[iso] = {"missing": missing,
                                                       "doubled": dupes}
    return report

def _shift_cols(template, dates):
    """Re-key a template's column map onto a new week's dates.

    The map is {date: column}. Copied verbatim it still points at the dates it
    came from, so an exported plan keeps the old week's headers. Columns are
    positional, so the nth date takes the nth column.
    """
    src = template.get("cols") or {}
    ordered = [src[d] for d in sorted(src)]
    return {iso: ordered[i] for i, iso in enumerate(dates) if i < len(ordered)}

def suggestion_to_week(index, suggestions, target_start, template, rows=None):
    """Build a storable week dict from suggestions.

    `template` is the most recent real week — it supplies each person's row
    number, section and building so the result looks like a parsed sheet and
    can still be written back to Excel.
    """
    start = _dt.date.fromisoformat(target_start)
    dates = [(start + _dt.timedelta(days=i)).isoformat() for i in range(7)]
    people = OrderedDict()
    for key, rec in template["people"].items():
        pid = norm_name(rec.get("name", key))
        cells = {}
        for iso in dates:
            got = suggestions.get(pid, {}).get(iso)
            if got and got["value"]:
                cells[iso] = got["value"]
        people[key] = {**rec, "cells": cells, "nocall": []}
    week = {"sheet": f"(planned {target_start})", "dates": dates,
            "cols": _shift_cols(template, dates), "people": people,
            "planned": True, "from_sheet": template.get("sheet", "")}
    return balance_zones(week, rows) if rows else week

def copy_week(template, target_start):
    """Same shape, same values, shifted onto a new set of dates."""
    start = _dt.date.fromisoformat(target_start)
    dates = [(start + _dt.timedelta(days=i)).isoformat() for i in range(7)]
    src = list(template["dates"])
    people = OrderedDict()
    for key, rec in template["people"].items():
        cells = {}
        for i, iso in enumerate(dates):
            if i < len(src):
                v = rec.get("cells", {}).get(src[i], "")
                if v:
                    cells[iso] = v
        # nocall marks belong to the week they happened in, never to a copy.
        people[key] = {**rec, "cells": cells, "nocall": []}
    return {"sheet": f"(copied to {target_start})", "dates": dates,
            "cols": _shift_cols(template, dates), "people": people,
            "planned": True, "from_sheet": template.get("sheet", "")}

def week_range_text(iso, with_year=None, today=None) -> str:
    """'Aug 23 – 29', or 'Aug 30 – Sep 5' when the week straddles a month.

    The year is appended only when it is not the current one, so the common
    case stays short.
    """
    start = _dt.date.fromisoformat(iso)
    end = start + _dt.timedelta(days=6)
    today = today or _local_today()
    if start.month == end.month:
        text = f"{start:%b} {start.day} – {end.day}"
    else:
        text = f"{start:%b} {start.day} – {end:%b} {end.day}"
    show_year = with_year if with_year is not None else (start.year != today.year)
    return f"{text}, {start.year}" if show_year else text

def week_label(iso, today=None, note="") -> str:
    """Human label for a week key: 'This week · Aug 23 – 29'.

    Weeks near today get named rather than dated, because that is how people
    actually refer to them; anything further out falls back to the date range,
    which stays unambiguous.
    """
    today = today or _local_today()
    this_sun = today - _dt.timedelta(days=(today.weekday() + 1) % 7)
    start = _dt.date.fromisoformat(iso)
    delta = (start - this_sun).days // 7
    rel = {0: "This week", 1: "Next week", -1: "Last week"}.get(delta)
    if rel is None:
        if 2 <= delta <= 5:
            rel = f"In {delta} weeks"
        elif -5 <= delta <= -2:
            rel = f"{-delta} weeks ago"
    body = week_range_text(iso, today=today)
    label = f"{rel} · {body}" if rel else body
    return f"{label} · {note}" if note else label

def next_missing_weeks(week_keys, count=6, today=None):
    """Upcoming Sundays with no stored week yet."""
    today = today or _local_today()
    sun = today - _dt.timedelta(days=(today.weekday() + 1) % 7)
    have = set(week_keys)
    out, d = [], sun
    while len(out) < count:
        iso = d.isoformat()
        if iso not in have and iso >= sun.isoformat():
            out.append(iso)
        d += _dt.timedelta(days=7)
        if (d - sun).days > 370:
            break
    return out

def export_week_xlsx(raw_bytes, week, title=None):
    """One week as a workbook in the sheet's own layout.

    Rather than building a sheet from scratch, this keeps the real one: the
    stored workbook is opened WITHOUT data_only so formulas and formatting
    survive, every other sheet is dropped, and this week's values are written
    over the top. A week planned in the app has no sheet of its own, so it
    borrows the one it was built from and gets its dates rewritten.

    Returns (bytes, sheet_title).
    """
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=False)
    src = week.get("sheet") if week.get("sheet") in wb.sheetnames else None
    if src is None:
        src = week.get("from_sheet")
    if src not in wb.sheetnames:
        raise ValueError("The week's sheet is not in the stored workbook.")
    ws = wb[src]
    for name in list(wb.sheetnames):
        if name != src:
            del wb[name]

    # Cells swallowed by a merge are read-only in openpyxl; only the top-left
    # anchor can be written. Some rows really do merge days together (one note
    # spanning Thu-Sat), so those are left exactly as the sheet has them.
    merged_locked = set()
    for rng in ws.merged_cells.ranges:
        for row in ws[rng.coord]:
            for cell in row:
                if (cell.row, cell.column) != (rng.min_row, rng.min_col):
                    merged_locked.add((cell.row, cell.column))

    def writable(r, c):
        return None if (r, c) in merged_locked else ws.cell(r, c)

    dates = list(week.get("dates") or [])
    cols = dict(week.get("cols") or {})
    if not cols and dates:                       # planned week built off a template
        cols = {iso: 2 + i for i, iso in enumerate(dates)}
    col_to_iso = {c: iso for iso, c in cols.items()}

    # Rewrite every repeated date header so the sheet reads as this week.
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 220)):
        for cell in row:
            if isinstance(cell.value, (_dt.datetime, _dt.date)) and cell.column in col_to_iso:
                target = writable(cell.row, cell.column)
                if target is not None:
                    target.value = _dt.datetime.fromisoformat(col_to_iso[cell.column])

    red = PatternFill("solid", fgColor="FFFF0000")
    blank = PatternFill(fill_type=None)
    for key, rec in week.get("people", {}).items():
        row = rec.get("row")
        if not row:
            continue
        marks = set(rec.get("nocall") or [])
        for iso, col in cols.items():
            cell = writable(int(row), int(col))
            if cell is None:
                continue
            cell.value = rec.get("cells", {}).get(iso) or None
            # Carry the no-call marks, and clear a red left over from the week
            # this sheet used to hold — otherwise a plan inherits someone
            # else's absence.
            if iso in marks:
                cell.fill = red
            elif _is_nocall_fill(cell):
                cell.fill = blank

    if title:
        ws.title = title[:31]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), ws.title

def write_overrides_to_workbook(raw_bytes, weeks, overrides):
    """Write in-app edits into a copy of the uploaded workbook.

    Loaded WITHOUT data_only so the sheet's own formulas (HSKP Needed, Extras
    Hskp, ...) survive the round trip — opening with cached values and saving
    would replace every formula with a frozen number.

    Returns (new_bytes, written_count, skipped_keys).
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=False)
    written, skipped = 0, []
    for key, ov in (overrides or {}).items():
        try:
            wk, name, iso = key.split("|", 2)
        except ValueError:
            skipped.append(key); continue
        week = (weeks or {}).get(wk)
        if not week or week.get("sheet") not in wb.sheetnames:
            skipped.append(key); continue
        rec = week.get("people", {}).get(name)
        col = (week.get("cols") or {}).get(iso)
        if not rec or not rec.get("row") or not col:
            skipped.append(key); continue
        ws = wb[week["sheet"]]
        ws.cell(int(rec["row"]), int(col)).value = (ov.get("value") or None)
        written += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), written, skipped

def week_to_people(week, iso):
    """Materialise one day out of a week dict, in parse_day's shape.

    Uses each entry's real name, not its (possibly disambiguated) storage key,
    so roster matching sees the name as the sheet writes it.
    """
    out = []
    for key, rec in week["people"].items():
        name, _home = read_person(key, rec)
        if name is None:
            continue
        k, known = cell_kind(rec, iso)
        out.append(_person_record(
            name, rec["group"],
            rec["building"] if rec["group"] == "hk" else rec["section"],
            rec.get("cells", {}).get(iso, ""), k, known))
    return out

#: Weekly summary rows near the top of every sheet, keyed by their column-A label.
METRIC_LABELS = OrderedDict([
    ("hskp actual #", "HSKP on schedule"),
    ("hskp needed",   "HSKP needed"),
    ("daily services","Daily services"),
    ("check outs",    "Check outs"),
    ("rqs needed",    "RQS needed"),
])

#: The hidden rows 5-11 every sheet carries, keyed by their column-A label.
METRIC_ROWS = {
    "hskp needed":   "hskp_needed",
    "holds":         "holds",
    "daily services": "dailies",
    "check outs":    "checkouts",
    "rqs needed":    "rqs_needed",
    "hskp actual #": "hskp_actual",
    "extras hskp":   "extras",
}

#: "HSKP Needed" is written as =2070/360 -- the day's labour minutes live
#: inside that formula and nowhere else in the workbook, so they can only be
#: had by reading the formula itself.
_MINUTES_RE = re.compile(r"^=\s*\(?\s*([0-9]+(?:\.[0-9]+)?)\s*\)?\s*/\s*([0-9]+)")


def _labour_minutes(formula, value):
    """The day's cleaning minutes, from =N/360 or from the number it left."""
    if isinstance(formula, str):
        m = _MINUTES_RE.match(formula.replace(" ", ""))
        if m:
            return float(m.group(1))
    if isinstance(value, (int, float)):
        # No formula survived, so rebuild it from the headcount it produced.
        return round(float(value) * 360)
    return None


def parse_metrics(ws, ws_formulas, cols):
    """The sheet's own numbers for each day: labour, rooms, headcount.

    These sit in rows the workbook keeps hidden, so nobody sees them and
    nothing has ever read them. They are the only record of how much work a
    day actually held.
    """
    out = {}
    rows = {}
    for r in range(1, min(ws.max_row, 40) + 1):
        label = _norm(ws.cell(r, 1).value)
        if label in METRIC_ROWS and METRIC_ROWS[label] not in rows:
            rows[METRIC_ROWS[label]] = r
    for iso, col in (cols or {}).items():
        day = {}
        for key, r in rows.items():
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                day[key] = round(float(v), 4)
        r5 = rows.get("hskp_needed")
        if r5:
            mins = _labour_minutes(
                ws_formulas.cell(r5, col).value if ws_formulas is not None else None,
                ws.cell(r5, col).value)
            if mins:
                day["minutes"] = mins
        if day:
            out[iso] = day
    return out


def day_metrics(ws, col):
    """Read the sheet's own headcount figures for one day.

    These are the manager's numbers, not ours — showing them next to what we
    parsed makes a miscount obvious at a glance.
    """
    out = OrderedDict()
    for r in range(1, min(ws.max_row, 40) + 1):
        label = _norm(ws.cell(r, 1).value)
        if label in METRIC_LABELS and METRIC_LABELS[label] not in out:
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                out[METRIC_LABELS[label]] = round(float(v), 2)
    return out

def parse_sheet_for_date(wb, sheet_name, target_date):
    """Parse one named sheet for one date. Raises ValueError if absent."""
    ws = wb[sheet_name]
    col = sheet_dates(ws).get(target_date)
    if col is None:
        raise ValueError(f"{target_date} is not in sheet {sheet_name!r}")
    return parse_day(ws, col)

# ── Turning parsed people into scheduler state ────────────────────────────────
def build_roster_update(people, existing_roster=None):
    """Produce the hk_roster / inspector / ds_team / RQS values for a day.

    Existing roster names win on SPELLING when they match case-insensitively,
    so live room tracking keyed by housekeeper name keeps working. The sheet
    still wins on BUILDING, which is the whole point of importing it.
    """
    existing_roster = existing_roster or {}
    by_norm = {norm_name(n): n for n in existing_roster}

    hk_roster, ds_team, changes, new_people = {}, [], [], []
    inspectors, rqs = {}, {1: "", 2: ""}
    others, unknown, cover = [], [], []

    for p in people:
        if p["raw"] and not p.get("recognised", True):
            unknown.append((p["name"], p["raw"]))
        # A houseperson or lead put on rooms today counts as a housekeeper.
        if p["group"] != "hk" and is_room_cover(p["raw"]):
            canonical = by_norm.get(norm_name(p["name"]), p["name"])
            bld = cover_building(p["raw"])
            hk_roster[canonical] = {"building": bld, "present": True}
            cover.append({"name": canonical, "from": p["section"],
                          "raw": p["raw"], "building": bld})
            if "daily service" in p["raw"].lower():
                ds_team.append(canonical)
            others.append(p)
            continue
        if p["group"] == "hk":
            canonical = by_norm.get(norm_name(p["name"]), p["name"])
            old = existing_roster.get(canonical, {})
            if old and old.get("building") not in (None, p["building"]):
                changes.append((canonical, old.get("building"), p["building"]))
            if not old:
                new_people.append(canonical)
            hk_roster[canonical] = {"building": p["building"], "present": p["present"]}
            if p["daily_service"]:
                ds_team.append(canonical)
        elif p["group"] == "rqs":
            inspectors[p["name"]] = p["present"]
            role = rqs_role(p["raw"])
            if role and p["present"] and not rqs[role]:
                rqs[role] = p["name"]
        else:
            others.append(p)

    return {
        "hk_roster":   hk_roster,
        "ds_team":     ds_team,
        "insp_roster": inspectors,
        "rqs1":        rqs[1],
        "rqs2":        rqs[2],
        "building_changes": changes,
        "new_people":  new_people,
        "others":      others,
        "unknown":     unknown,
        "cover":       cover,
    }


# ── Daily service, shared out fairly ─────────────────────────────────────────
#: What the sheet writes in the cell. "Daily service" is far and away the most
#: common of the six spellings in the workbook, so a suggestion uses it.
DS_LABEL = "Daily service"

#: How far back a turn still counts against you.
WINDOW_DAYS = 28

#: Pseudo-days of smoothing applied to a thin record. See daily_service_record.
SMOOTH_DAYS = 10


def daily_service_record(rows, before=None, since=None):
    """Who has done daily service, how often, and how long ago.

    Counted against days actually worked, not calendar days: somebody part
    time is not owed fewer turns per shift than somebody full time.
    """
    rec = {}
    for r in rows:
        if r.get("group") != "hk":
            continue
        if before and r["date"] >= before:
            continue
        if since and r["date"] < since:
            continue
        p = r["person"]
        d = rec.setdefault(p, {"person": p, "label": r.get("label") or p,
                               "worked": 0, "ds": 0, "last": None})
        if r.get("kind") in WORKED_KINDS:
            d["worked"] += 1
        if r.get("daily_service"):
            d["ds"] += 1
            if not d["last"] or r["date"] > d["last"]:
                d["last"] = r["date"]
    # Smoothed, not raw. A raw ds/worked lets somebody with three worked days
    # and no turns outrank a veteran who has done twenty in three hundred, so
    # the rotation chases whoever is newest instead of whoever is owed a turn.
    # Ten pseudo-days at the house rate pulls thin records back toward average
    # until there is enough of a record to speak for itself.
    tot_ds = sum(d["ds"] for d in rec.values())
    tot_worked = sum(d["worked"] for d in rec.values())
    rate = (tot_ds / tot_worked) if tot_worked else 0.0
    for d in rec.values():
        d["raw_share"] = d["ds"] / d["worked"] if d["worked"] else 0.0
        d["share"] = ((d["ds"] + SMOOTH_DAYS * rate) /
                      (d["worked"] + SMOOTH_DAYS)) if (d["worked"] + SMOOTH_DAYS) else 0.0
        # How many turns this person is owed, against what their days imply.
        d["owed"] = d["worked"] * rate - d["ds"]
    return rec


def suggest_daily_service(rows, week, dates, per_day, before=None,
                          window_days=WINDOW_DAYS):
    """Pick who does daily service on each day of a planned week.

    Two rules, in that order:

    1.  Nobody takes a second turn in the week until everyone who could have
        taken a first has had one. That is the "no repeats" rule, and the
        "unless there is nobody left" exception falls out of it: when the pool
        of people on nought is empty the count simply starts again.

    2.  Between people on equal footing for the week, the turn goes to
        whoever has done it least often across their worked days, and then to
        whoever has waited longest. This is the part that matters. Within a
        week repeats are already rare; it is across the weeks that the job
        piles up on the same few people -- in this workbook one housekeeper
        spends 18% of her days on it while another has worked 77 days and
        never been asked once.

    `per_day` maps each date to how many people that day needs. Returns
    {date: [names]} plus the reasoning, so the suggestion can be argued with.
    """
    # A rolling window, not the whole record. Balancing on everything ever
    # worked makes the rotation spend the next two months repaying old debts:
    # it picks the same owed people every day they are in and skips everybody
    # else entirely. Replayed over ten real weeks that was measurably WORSE
    # than what the sheet does by hand. Fairness people actually feel is "have
    # I had it lately", so only the last few weeks count.
    edge = None
    if window_days and dates:
        edge = (_dt.date.fromisoformat(min(dates))
                - _dt.timedelta(days=window_days)).isoformat()
    hist = daily_service_record(rows, before=before, since=edge)
    this_week = {p: 0 for p in hist}
    picked, why = {}, {}

    for iso in dates:
        want = int((per_day or {}).get(iso) or 0)
        # Only housekeepers, and only the ones the plan has working that day.
        avail = []
        for name, r in (week.get("people") or {}).items():
            if r.get("group") != "hk":
                continue
            raw = (r.get("cells") or {}).get(iso, "")
            if classify(raw) not in WORKED_KINDS:
                continue
            avail.append(name)

        chosen = []
        for _ in range(min(want, len(avail))):
            pool = [n for n in avail if n not in chosen]
            if not pool:
                break
            fewest = min(this_week.get(n, 0) for n in pool)
            # The hard rule: only people on the lowest count for this week.
            tier = [n for n in pool if this_week.get(n, 0) == fewest]
            tier.sort(key=lambda n: (
                hist.get(n, {}).get("share", 0.0),          # least often, first
                hist.get(n, {}).get("last") or "",          # longest wait, first
                n.lower()))
            pick = tier[0]
            chosen.append(pick)
            this_week[pick] = this_week.get(pick, 0) + 1
            h = hist.get(pick, {})
            why[(iso, pick)] = {
                "share": h.get("share", 0.0), "ds": h.get("ds", 0),
                "worked": h.get("worked", 0), "last": h.get("last"),
                "second_turn": this_week[pick] > 1,
            }
        picked[iso] = chosen
    return {"picks": picked, "why": why, "history": hist,
            "week_counts": {p: n for p, n in this_week.items() if n}}
