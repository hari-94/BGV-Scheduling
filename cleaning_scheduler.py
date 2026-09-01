"""
Cleaning Schedule Grouper v10
"""
import re, html as _html
import hashlib as _hashlib
import pandas as pd
import sys as _sys2, os as _os2
_sys2.path.insert(0, _os2.path.dirname(__file__))

import streamlit as st
import streamlit.components.v1 as components

st.set_page_config(
    page_title="Cleaning Schedule",
    page_icon="GC8",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Import local modules after set_page_config
import auth, db
import ui

# ── Hide Streamlit's auto-generated page navigation IMMEDIATELY ───────────────
# This must be the first markdown call so the nav never flashes. We use every
# known selector variant across Streamlit versions to be bulletproof.
st.markdown("""<style>
/* Hide ONLY Streamlit's auto-generated page nav. The config.toml setting
   showSidebarNavigation=false is the primary fix; this is a safe backup that
   targets only the nav testid, never the sidebar content itself. */
[data-testid="stSidebarNav"]{
  display: none !important;
}
</style>""", unsafe_allow_html=True)

# ── Theme: locked to a single formal "light"theme for office use ─────────────
st.session_state["theme"] = "light"
_THEME = "light"
_IS_GLASS = False
_IS_LIGHT = True

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════
def e(s): return _html.escape(str(s) if s else "")

# ── Shared time helpers (Mountain time) ──────────────────────────────────────
# Consolidated here so the timezone and "now" / formatting logic live in one
# place instead of being rebuilt inside the HK view and Live tab.
import zoneinfo as _zoneinfo
from datetime import datetime as _datetime
_MTN_TZ = _zoneinfo.ZoneInfo("America/Denver")
def _now_iso():
    """Current Mountain-time timestamp as ISO string (for saved statuses)."""
    return _datetime.now(_MTN_TZ).isoformat()
def _fmt_mtn(ts):
    """Format a stored ISO timestamp as 'HH:MM AM/PM' in Mountain time."""
    if not ts: return ""
    try:
        dt = _datetime.fromisoformat(str(ts).replace("Z", "+00:00"))
        return dt.astimezone(_MTN_TZ).strftime("%I:%M %p")
    except Exception:
        return ""

def make_labels(prefix: str, n: int) -> list:
    alpha = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    out = []
    for c in alpha:
        out.append(f"{prefix}-{c}")
        if len(out) >= n: return out
    for c1 in alpha:
        for c2 in alpha:
            out.append(f"{prefix}-{c1}{c2}")
            if len(out) >= n: return out
    return out

LOW_MIN = 330
MAX_FC = 380
MAX_DS = 560
DS_OVER = 700
LOW_FILL = 350

# Daily Service: HARD cap per housekeeper. Fill up to DS_CAP, never exceed;
# leftover rooms form new charts (blank housekeeper if none available, filled in
# manually later based on how the day goes).
DS_CAP = 460
# IH charts below this total are treated as scraps and spill to Daily Service
# (kept charts at/above this are inspected by RQS 2).
IH_KEEP_MIN = 310

SVC_FC = "Full Clean"
SVC_IH = "Full Clean (IH)"
SVC_DS = "Daily Service"
SVC_DV = "Dust n Vac"

# Placeholder shown when a chart has no real person on it. Numbered per chart
# (Need Housekeeper 1, 2, ...) so the manager sees how many bodies are short.
NEED_HK_PREFIX = "Need Housekeeper"
NO_HK_LABEL    = "No HK available"
#: "nobody" entry in the RQS selectboxes. Named because the auto-import has to
#: write this exact value into the widget's state to clear a role.
RQS_NONE       = "— none —"

def is_unassigned_hk(name) -> bool:
    """True for empty / placeholder housekeeper values (not a real person)."""
    s = str(name or "")
    return (not s) or s.startswith(NO_HK_LABEL) or s.startswith(NEED_HK_PREFIX)

# RQS assigned to inspect all IH charts.
IH_RQS = "RQS 2"

DEFAULT_TIMES = {
    SVC_FC: {"A":120,"B":70,"C":70,"D":120,"E":140,"F":70,"G":70,"H":70,"I":70},
    SVC_IH: {"A":120,"B":70,"C":70,"D":120,"E":140,"F":70,"G":70,"H":70,"I":70},
    SVC_DS: {"A":35, "B":20,"C":20,"D":35, "E":40, "F":20,"G":20,"H":20,"I":20},
    SVC_DV: {},
}
DV_DEFAULT_TIME = 0

def default_time_for(room: str, svc: str) -> int:
    room_type = ""
    for ch in reversed(str(room).strip().upper()):
        if ch.isalpha(): room_type = ch; break
    return DEFAULT_TIMES.get(svc, {}).get(room_type, 70)

PALETTE = [
    ("#5B4FE9","#EEEEFF"),("#0D9488","#ECFDF5"),("#2563EB","#EFF6FF"),
    ("#D97706","#FFFBEB"),("#DC2626","#FFF5F5"),("#7C3AED","#F5F3FF"),
    ("#0891B2","#ECFEFF"),("#EA580C","#FFF7ED"),("#DB2777","#FDF2F8"),
    ("#059669","#F0FDF4"),("#65A30D","#F7FEE7"),("#4F46E5","#EEF2FF"),
    ("#9333EA","#FAF5FF"),("#16A34A","#F0FDF4"),("#E11D48","#FFF1F2"),
    ("#0284C7","#F0F9FF"),("#B45309","#FFFBEB"),("#6D28D9","#F5F3FF"),
    ("#047857","#ECFDF5"),("#B91C1C","#FEF2F2"),
]
SVC_PALETTE = {
    SVC_FC: ("#2563EB","#EFF6FF"),
    SVC_IH: ("#7C3AED","#F5F3FF"),
    SVC_DS: ("#0D9488","#ECFDF5"),
    SVC_DV: ("#D97706","#FFFBEB"),
}
IC = ["#5B4FE9","#0D9488","#D97706","#DC2626","#7C3AED","#0891B2","#EA580C","#DB2777"]
BLD_COLORS = {1:("#2563EB","#EFF6FF"), 2:("#0D9488","#ECFDF5"), 3:("#D97706","#FFFBEB")}

def pal(i): return PALETTE[i % len(PALETTE)]

DEFAULT_HK = {
    1: ["Maricruz","Melissa","ARACELI","Anaberta","Darling","Adrian","Leonardo","Rosibel"],
    2: ["Liliana L","DIANIS","Cecilia Angeles","Elibeth Herrera","Jenifer S.",
        "Norma E","Santos","Yadira","Senia","Gloria R.","Ana Centeno","Camila O","DANIA","Andres"],
    3: ["JENNI CAICEDO","Federico","Nancy","Minoska","Lourdes","AMALIA",
        "Elizabeth","Jorge Luis","Lorena","Nury","Lilliam","Janiris",
        "Luis Urbina","Ana Hernandez","JOSSELIN"],
}
DEFAULT_INSPECTORS = [
    "Sayra","Oralia","Alejandro","Claudia P.","Castor","Hari",
    "Danny R.","Mayra B.","Ana Casia","Bryan","Gustavo","Edward","David S.","Grace",
]

# ══════════════════════════════════════════════════════════════════════════════
# CSS
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;500;600;700;800&family=DM+Mono:ital,wght@0,300;0,400;0,500;1,400&family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,600&display=swap');

:root {
  --bg:#080810; --bg1:#0e0e1a; --bg2:#13131f; --bg3:#1a1a2e;
  --border:rgba(99,102,241,.18); --border-hi:rgba(99,102,241,.45);
  --indigo:#6366f1; --indigo-lo:rgba(99,102,241,.12);
  --cyan:#22d3ee; --cyan-lo:rgba(34,211,238,.1);
  --teal:#14b8a6; --amber:#f59e0b; --rose:#f43f5e;
  --txt:#e2e8f0; --txt2:#94a3b8; --txt3:#475569;
  --glow-i:0 0 24px rgba(99,102,241,.35),0 0 60px rgba(99,102,241,.12);
  --glow-c:0 0 24px rgba(34,211,238,.3),0 0 60px rgba(34,211,238,.1);
  --glow-sm:0 0 12px rgba(99,102,241,.25);
  --radius:14px; --radius-sm:8px;
}
*,*::before,*::after{box-sizing:border-box;}
html,body,[class*="css"]{
  font-family:'DM Sans',sans-serif!important;
  -webkit-font-smoothing:antialiased;
  color:var(--txt)!important;
}
.stApp{
  background:var(--bg)!important;
  background-image:
    radial-gradient(ellipse 80% 50% at 20% -10%,rgba(99,102,241,.12) 0%,transparent 60%),
    radial-gradient(ellipse 60% 40% at 80% 100%,rgba(34,211,238,.07) 0%,transparent 55%)!important;
}
.block-container{padding-top:1.6rem!important;max-width:1440px!important;background:transparent!important;}
::-webkit-scrollbar{width:4px;height:4px;}
::-webkit-scrollbar-track{background:var(--bg1);}
::-webkit-scrollbar-thumb{background:var(--indigo);border-radius:99px;}

/* ── Typography ── */
.pg-title{
  font-family:'Syne',sans-serif!important;font-size:2rem;font-weight:800;letter-spacing:-.04em;
  background:linear-gradient(135deg,#fff 0%,var(--indigo) 45%,var(--cyan) 100%);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;
  margin:0 0 4px;line-height:1.1;animation:titleReveal .7s cubic-bezier(.16,1,.3,1) both;
}
@keyframes titleReveal{from{opacity:0;transform:translateY(12px)}to{opacity:1;transform:translateY(0)}}
.pg-sub{font-size:.82rem;color:var(--txt2);margin:0 0 1rem;font-weight:400;animation:titleReveal .7s .1s cubic-bezier(.16,1,.3,1) both;}
.sec{font-family:'DM Mono',monospace!important;font-size:.6rem;font-weight:500;text-transform:uppercase;
  letter-spacing:.16em;color:var(--indigo);padding-bottom:6px;border-bottom:1px solid var(--border);margin:1.2rem 0 .6rem;}

/* ── Sidebar ── */
section[data-testid="stSidebar"]{
  background:rgba(13,13,26,.88)!important;backdrop-filter:blur(20px)!important;
  -webkit-backdrop-filter:blur(20px)!important;border-right:1px solid var(--border)!important;
  box-shadow:4px 0 40px rgba(0,0,0,.5)!important;
  min-width:340px!important;max-width:380px!important;
}
/* Only when EXPLICITLY collapsed (aria-expanded="false") do we zero it out and
   strip the shadow/border/blur so no ghost strip remains. The default rule
   above always gives the sidebar a width, so it can never vanish on load even
   if Streamlit doesn't set aria-expanded="true". */
section[data-testid="stSidebar"][aria-expanded="false"]{
  min-width:0!important;max-width:0!important;width:0!important;
  box-shadow:none!important;
  border-right:none!important;
  backdrop-filter:none!important;
  -webkit-backdrop-filter:none!important;
  background:transparent!important;
  overflow:hidden!important;
}
section[data-testid="stSidebar"][aria-expanded="false"]::before{display:none!important;}
section[data-testid="stSidebar"]::before{
  content:'';position:absolute;top:0;left:0;right:0;height:2px;
  background:linear-gradient(90deg,var(--indigo),var(--cyan));z-index:10;
}
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p{font-size:.8rem;color:var(--txt2);}
section[data-testid="stSidebar"] h2{
  font-family:'Syne',sans-serif!important;font-size:1rem!important;font-weight:700!important;
  color:var(--txt)!important;letter-spacing:-.01em;padding-bottom:8px;margin-bottom:8px;
  border-bottom:1px solid var(--border)!important;
}
section[data-testid="stSidebar"] h3{
  font-family:'DM Mono',monospace!important;font-size:.62rem!important;font-weight:500!important;
  text-transform:uppercase;letter-spacing:.14em;color:var(--indigo)!important;margin:12px 0 6px!important;
}
section[data-testid="stSidebar"] hr{border-color:var(--border)!important;}

/* ── Stat cards ── */
.stat-row{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:1.2rem;}
.sc{
  flex:1 1 100px;background:rgba(255,255,255,.03)!important;border:1px solid var(--border)!important;
  border-radius:var(--radius)!important;padding:16px 14px;text-align:center;
  backdrop-filter:blur(12px);transition:transform .2s,box-shadow .2s,border-color .2s;
  animation:cardIn .5s cubic-bezier(.16,1,.3,1) both;
}
.sc:hover{transform:translateY(-3px) scale(1.02);border-color:var(--border-hi)!important;box-shadow:var(--glow-sm);}
.sc.hi{background:linear-gradient(135deg,rgba(99,102,241,.25),rgba(99,102,241,.1))!important;border-color:rgba(99,102,241,.4)!important;box-shadow:var(--glow-sm);}
.sc.ds{background:linear-gradient(135deg,rgba(20,184,166,.2),rgba(20,184,166,.06))!important;border-color:rgba(20,184,166,.35)!important;}
.sc.dv{background:linear-gradient(135deg,rgba(245,158,11,.2),rgba(245,158,11,.06))!important;border-color:rgba(245,158,11,.35)!important;}
.sc .n{font-family:'Syne',sans-serif!important;font-size:1.8rem;font-weight:800;color:#fff;line-height:1;margin-bottom:3px;text-shadow:0 0 20px rgba(99,102,241,.5);}
.sc.ds .n{text-shadow:0 0 20px rgba(20,184,166,.5);}
.sc.dv .n{text-shadow:0 0 20px rgba(245,158,11,.5);}
.sc .l{font-family:'DM Mono',monospace!important;font-size:.58rem;color:var(--txt2);text-transform:uppercase;letter-spacing:.1em;font-weight:500;}
@keyframes cardIn{from{opacity:0;transform:translateY(16px) scale(.97)}to{opacity:1;transform:translateY(0) scale(1)}}
.sc:nth-child(1){animation-delay:.05s}.sc:nth-child(2){animation-delay:.1s}.sc:nth-child(3){animation-delay:.15s}
.sc:nth-child(4){animation-delay:.2s}.sc:nth-child(5){animation-delay:.25s}.sc:nth-child(6){animation-delay:.3s}
.sc:nth-child(7){animation-delay:.35s}.sc:nth-child(8){animation-delay:.4s}

/* ── Rules box ── */
.rules-box{background:rgba(99,102,241,.06);border:1px solid var(--border);border-radius:var(--radius);padding:14px 18px;font-size:.81rem;color:var(--txt2);}
.rules-box li{margin-bottom:5px;line-height:1.6;}
.rules-box strong{color:var(--cyan);}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"]{gap:4px;background:rgba(255,255,255,.03)!important;border:1px solid var(--border)!important;border-radius:var(--radius-sm);padding:4px!important;}
.stTabs [data-baseweb="tab"]{border-radius:6px!important;padding:7px 18px!important;font-family:'DM Sans',sans-serif!important;font-size:.78rem!important;font-weight:600!important;color:var(--txt2)!important;border:none!important;background:transparent!important;transition:all .2s!important;}
.stTabs [data-baseweb="tab"]:hover{color:var(--txt)!important;background:rgba(99,102,241,.1)!important;}
.stTabs [aria-selected="true"]{background:rgba(99,102,241,.2)!important;color:var(--cyan)!important;box-shadow:0 0 0 1px rgba(99,102,241,.4),var(--glow-sm)!important;}

/* ── Buttons ── */
.stButton>button{border-radius:var(--radius-sm)!important;font-family:'DM Sans',sans-serif!important;font-weight:600!important;font-size:.82rem!important;border:1px solid var(--border)!important;background:rgba(255,255,255,.04)!important;color:var(--txt)!important;transition:all .2s!important;}
.stButton>button:hover{border-color:var(--border-hi)!important;background:rgba(99,102,241,.1)!important;transform:translateY(-1px);}
.stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--indigo),#818cf8)!important;border:none!important;color:#fff!important;box-shadow:var(--glow-sm),0 4px 20px rgba(99,102,241,.3)!important;font-size:.88rem!important;animation:pulseCTA 2.5s ease-in-out infinite;}
.stButton>button[kind="primary"]:hover{animation:none!important;transform:translateY(-2px)!important;box-shadow:var(--glow-i)!important;}
@keyframes pulseCTA{0%,100%{box-shadow:0 0 12px rgba(99,102,241,.35),0 4px 20px rgba(99,102,241,.2);}50%{box-shadow:0 0 28px rgba(99,102,241,.6),0 4px 30px rgba(99,102,241,.35);}}

/* ── Inputs ── */
.stTextArea textarea,.stTextInput input{background:rgba(255,255,255,.03)!important;border:1px solid var(--border)!important;border-radius:var(--radius-sm)!important;color:var(--txt)!important;font-family:'DM Mono',monospace!important;font-size:.78rem!important;}
.stTextArea textarea:focus,.stTextInput input:focus{border-color:var(--indigo)!important;box-shadow:0 0 0 2px rgba(99,102,241,.2)!important;}
.stTextArea textarea::placeholder{color:var(--txt3)!important;}

/* ── Selects / dropdowns ── */
.stSelectbox [data-baseweb="select"]>div,.stMultiSelect [data-baseweb="select"]>div{background:rgba(255,255,255,.03)!important;border:1px solid var(--border)!important;border-radius:var(--radius-sm)!important;color:var(--txt)!important;}
[data-baseweb="popover"] [role="listbox"],[data-baseweb="menu"]{background:var(--bg2)!important;border:1px solid var(--border)!important;border-radius:var(--radius-sm)!important;}
[data-baseweb="option"]:hover,[aria-selected="true"]{background:rgba(99,102,241,.15)!important;}
[data-baseweb="tag"]{background:rgba(99,102,241,.2)!important;border:1px solid rgba(99,102,241,.4)!important;border-radius:6px!important;color:var(--cyan)!important;}

/* ── Checkboxes / sliders ── */
.stCheckbox label{font-size:.8rem!important;color:var(--txt2)!important;}
.stCheckbox label:hover{color:var(--txt)!important;}
[data-testid="stSlider"] div[role="slider"]{background:var(--indigo)!important;box-shadow:var(--glow-sm)!important;}

/* ── Expander ── */
.streamlit-expanderHeader{background:rgba(255,255,255,.03)!important;border:1px solid var(--border)!important;border-radius:var(--radius-sm)!important;color:var(--txt2)!important;font-size:.8rem!important;font-weight:600!important;}
.streamlit-expanderHeader:hover{border-color:var(--border-hi)!important;color:var(--txt)!important;}
.streamlit-expanderContent{background:rgba(255,255,255,.02)!important;border:1px solid var(--border)!important;border-top:none!important;border-radius:0 0 var(--radius-sm) var(--radius-sm)!important;}

/* ── Alerts ── */
.stAlert{border-radius:var(--radius-sm)!important;border:none!important;}
.stSuccess{background:rgba(20,184,166,.1)!important;border-left:3px solid var(--teal)!important;}
.stWarning{background:rgba(245,158,11,.08)!important;border-left:3px solid var(--amber)!important;}
.stError{background:rgba(244,63,94,.08)!important;border-left:3px solid var(--rose)!important;}
.stInfo{background:rgba(99,102,241,.08)!important;border-left:3px solid var(--indigo)!important;}

/* ── Misc ── */
hr{border:none!important;height:1px!important;background:var(--border)!important;margin:1rem 0!important;}
[data-testid="stSpinner"]>div{border-top-color:var(--indigo)!important;}
footer{visibility:hidden!important;}
#MainMenu{visibility:hidden!important;}
/* Hide the top header bar (leaves a white strip otherwise) */
header[data-testid="stHeader"]{
  background:transparent !important;
  height:0 !important;
  min-height:0 !important;
}
[data-testid="stToolbar"]{display:none !important;}
[data-testid="stDecoration"]{display:none !important;}
/* Hide Streamlit's auto-generated page nav — we use a custom role-based nav */
[data-testid="stSidebarNav"]{display:none !important;}

/* ── Mobile ── */
@media(max-width:768px){
  /* Tighter page padding, full width */
  .block-container{padding:.6rem .4rem!important;max-width:100%!important;}
  .pg-title{font-size:1.35rem!important;line-height:1.1!important;}
  .pg-sub{font-size:.72rem!important;}

  /* Stat cards: 2 per row */
  .stat-row{gap:5px!important;}
  .sc{flex:1 1 calc(50% - 5px)!important;min-width:calc(50% - 5px)!important;padding:9px 7px!important;}
  .sc .n{font-size:1.25rem!important;}
  .sc .l{font-size:.5rem!important;letter-spacing:.04em!important;}

  /* Tabs wrap and shrink */
  .stTabs [data-baseweb="tab-list"]{flex-wrap:wrap!important;gap:3px!important;}
  .stTabs [data-baseweb="tab"]{padding:5px 9px!important;font-size:.68rem!important;}

  /* Sidebar wider on mobile when open; fully off-screen when collapsed
     (identical to the Dashboard page, which collapses correctly) */
  section[data-testid="stSidebar"]{
    min-width:85vw!important;max-width:90vw!important;
  }
  section[data-testid="stSidebar"][aria-expanded="false"]{
    min-width:0!important;max-width:0!important;width:0!important;margin-left:-92vw!important;
  }

  /* Bigger tap targets for buttons */
  .stButton>button{min-height:42px!important;padding:8px 10px!important;}

  /* Inputs full width, easier to tap */
  .stTextInput input,.stTextArea textarea,
  .stSelectbox [data-baseweb="select"]>div,
  .stMultiSelect [data-baseweb="select"]>div{font-size:16px!important;}
  /* ^ 16px prevents iOS auto-zoom on focus */

  /* Expander headers bigger */
  .streamlit-expanderHeader{font-size:.82rem!important;padding:10px!important;}
}

/* ── BIG layout columns: phone-first behavior ── */
@media(max-width:768px){
  /* Default: every multi-column row may wrap instead of squishing */
  [data-testid="stHorizontalBlock"]{flex-wrap:wrap!important;gap:6px!important;}
  [data-testid="stHorizontalBlock"]>[data-testid="column"]{min-width:0!important;}

  /* Input areas (room data / front-desk email / add-staff) stack FULL width */
  [data-testid="stHorizontalBlock"]:has(.stTextArea),
  [data-testid="stHorizontalBlock"]:has(.stTextInput){
    flex-direction:column!important;
  }
  [data-testid="stHorizontalBlock"]:has(.stTextArea)>[data-testid="column"],
  [data-testid="stHorizontalBlock"]:has(.stTextInput)>[data-testid="column"]{
    width:100%!important;min-width:100%!important;flex:1 1 100%!important;
  }

  /* Filter bars (Groups tab, Live tab): selectbox / multiselect / checkbox
     columns flow as a 2-up grid instead of 5 squeezed slivers */
  [data-testid="stHorizontalBlock"]:has(.stSelectbox)>[data-testid="column"],
  [data-testid="stHorizontalBlock"]:has(.stMultiSelect)>[data-testid="column"],
  [data-testid="stHorizontalBlock"]:has(.stCheckbox)>[data-testid="column"]{
    flex:1 1 47%!important;min-width:47%!important;width:auto!important;
  }

  /* Button-only rows (status actions) STAY horizontal & equal width —
     thumb-friendly action bars under each room */
  [data-testid="stHorizontalBlock"]:has(.stButton):not(:has(.stTextArea)):not(:has(.stTextInput)):not(:has(.stSelectbox)):not(:has(.stCheckbox)){
    flex-direction:row!important;flex-wrap:nowrap!important;gap:4px!important;
  }
  [data-testid="stHorizontalBlock"]:has(.stButton):not(:has(.stTextArea)):not(:has(.stTextInput)):not(:has(.stSelectbox)):not(:has(.stCheckbox))>[data-testid="column"]{
    flex:1 1 0!important;min-width:0!important;
  }

  /* Sticky tab bar — keep Housekeepers/Inspectors/Groups/Live reachable
     while scrolling long schedules */
  .stTabs [data-baseweb="tab-list"]{
    position:sticky!important;top:0!important;z-index:99!important;
    backdrop-filter:blur(18px)!important;-webkit-backdrop-filter:blur(18px)!important;
  }

  /* Group / staff / inspector cards (iframes) hug the screen edge */
  iframe{width:100%!important;}

  /* Expander internals breathe less */
  .streamlit-expanderContent{padding:8px 6px!important;}

  /* Download + Generate buttons: full-width thumb targets */
  .stDownloadButton>button{width:100%!important;min-height:46px!important;}
}

/* ── Narrow phones ── */
@media(max-width:480px){
  .pg-title{font-size:1.1rem!important;}
  .sc .n{font-size:1.05rem!important;}
  .sc{padding:8px 5px!important;}
  /* 4 stat cards per two rows is still a lot — drop label tracking further */
  .sc .l{font-size:.46rem!important;}
  /* Tab pills shrink to fit all 4 on one line */
  .stTabs [data-baseweb="tab"]{padding:4px 7px!important;font-size:.64rem!important;}
  /* Status action buttons: compact but still ≥40px tall */
  .stButton>button{min-height:40px!important;padding:6px 4px!important;font-size:.72rem!important;}
  /* Section headers tighter */
  .sec{font-size:.55rem!important;margin:.8rem 0 .4rem!important;}
  /* Hide the vertical gap Streamlit adds between stacked widgets */
  [data-testid="stVerticalBlock"]{gap:.45rem!important;}
}
</style>""", unsafe_allow_html=True)

# ── LIGHT THEME OVERRIDE ──────────────────────────────────────────────────────
# When the user picks light mode, override the CSS variables + backgrounds.
# Because every component reads from var(--…), this single block reskins the
# entire app without touching the individual style rules above.
if _THEME == "light":
    st.markdown("""
<style>
/* ════════════════════════════════════════════════════════════════════════
   FORMAL / OFFICE THEME
   A clean corporate look: neutral slate surfaces, one restrained blue accent,
   crisp hairline borders, and soft shadows instead of neon glows. Overrides the
   base design-system variables so every var(--…) consumer reskins at once.
   ════════════════════════════════════════════════════════════════════════ */
:root {
  --bg:#f4f5f7; --bg1:#eceef1; --bg2:#ffffff; --bg3:#f7f8fa;
  --border:#e2e5ea; --border-hi:#c3c9d4;
  --indigo:#2563a8; --indigo-lo:rgba(37,99,168,.07);
  --cyan:#3b7fb8; --cyan-lo:rgba(59,127,184,.07);
  --teal:#0f766e; --amber:#b45309; --rose:#be123c;
  --txt:#1f2733; --txt2:#5b6675; --txt3:#8a93a1;
  --glow-i:0 1px 3px rgba(20,32,54,.08),0 1px 2px rgba(20,32,54,.04);
  --glow-c:0 1px 3px rgba(20,32,54,.08);
  --glow-sm:0 1px 2px rgba(20,32,54,.07);
  --radius:10px; --radius-sm:7px;
}
/* Neutral, distraction-free background — no radial neon washes. */
.stApp{
  background:#f4f5f7!important;
  background-image:none!important;
}
.block-container{padding-top:1.8rem!important;max-width:1360px!important;}

/* Page title: solid, confident slate — no gradient text, no animation flourish. */
.pg-title{
  font-family:'Syne',sans-serif!important;font-weight:700!important;
  font-size:1.7rem!important;letter-spacing:-.02em!important;
  color:#16202e!important;
  -webkit-text-fill-color:#16202e!important;background:none!important;
  animation:none!important;
}
.pg-sub{color:var(--txt2)!important;}
.sec{color:var(--txt2)!important;border-bottom:1px solid var(--border)!important;
  letter-spacing:.14em!important;font-weight:600!important;}

/* Cards & surfaces: white with a hairline border and a soft shadow. */
.sc{
  background:#ffffff!important;
  border:1px solid var(--border)!important;
  box-shadow:0 1px 2px rgba(20,32,54,.06)!important;
}
.sc.hi{background:#ffffff!important;border-left:3px solid var(--indigo)!important;}
.sc.ds{background:#ffffff!important;border-left:3px solid var(--teal)!important;}
.sc.dv{background:#ffffff!important;border-left:3px solid var(--amber)!important;}
.sc .n{color:#16202e!important;text-shadow:none!important;}
.sc .l{color:var(--txt2)!important;}
.rules-box{background:#ffffff!important;border:1px solid var(--border)!important;}

/* Accent top-bars on cards flat, single tone (no indigocyan gradient). */
.sc::before,.rules-box::before{background:var(--indigo)!important;}

/* Sidebar: solid white panel with a clean divider (no blur/glow). */
section[data-testid="stSidebar"]{
  background:#ffffff!important;
  backdrop-filter:none!important;-webkit-backdrop-filter:none!important;
  border-right:1px solid var(--border)!important;
  box-shadow:1px 0 0 rgba(20,32,54,.03)!important;
}
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p{color:var(--txt)!important;}

/* Tabs: understated, with a solid accent for the active tab. */
.stTabs [data-baseweb="tab-list"]{background:#ffffff!important;border:1px solid var(--border)!important;}
.stTabs [data-baseweb="tab"]{color:var(--txt2)!important;}
.stTabs [aria-selected="true"]{background:var(--indigo)!important;color:#ffffff!important;}

/* Buttons: clean white default, solid accent on hover/primary. */
.stButton>button{
  background:#ffffff!important;color:var(--txt)!important;
  border:1px solid var(--border-hi)!important;box-shadow:none!important;
}
.stButton>button:hover{
  background:var(--indigo)!important;color:#ffffff!important;
  border-color:var(--indigo)!important;
}

/* Inputs: white fields, subtle border, accent focus ring. */
.stTextArea textarea,.stTextInput input,
.stSelectbox [data-baseweb="select"]>div,
.stMultiSelect [data-baseweb="select"]>div{
  background:#ffffff!important;color:var(--txt)!important;border:1px solid var(--border-hi)!important;
}
.stTextArea textarea:focus,.stTextInput input:focus{
  border-color:var(--indigo)!important;box-shadow:0 0 0 3px var(--indigo-lo)!important;
}
.streamlit-expanderHeader{background:#ffffff!important;border:1px solid var(--border)!important;}
.streamlit-expanderContent{background:#ffffff!important;}
[data-baseweb="popover"] [role="listbox"],[data-baseweb="menu"]{background:#ffffff!important;}

/* Metrics / dataframes read cleanly on white. */
[data-testid="stMetricValue"]{color:#16202e!important;}
[data-testid="stMetricLabel"]{color:var(--txt2)!important;}

/* Scrollbar: neutral, not accent-colored. */
::-webkit-scrollbar-thumb{background:#c3c9d4!important;}
::-webkit-scrollbar-track{background:#eceef1!important;}

/* Primary button: solid accent, flat, no pulsing glow (formal). */
.stButton>button[kind="primary"]{
  background:var(--indigo)!important;border:1px solid var(--indigo)!important;color:#ffffff!important;
  box-shadow:0 1px 2px rgba(20,32,54,.10)!important;animation:none!important;
}
.stButton>button[kind="primary"]:hover{
  background:#1f5697!important;transform:none!important;
  box-shadow:0 2px 6px rgba(20,32,54,.14)!important;
}
/* Card hover: gentle lift, no scale-pop or glow. */
.sc:hover{transform:translateY(-1px)!important;box-shadow:0 2px 6px rgba(20,32,54,.10)!important;}
/* Sidebar collapse control and misc glows flattened. */
[data-testid="stSlider"] div[role="slider"]{box-shadow:none!important;}
</style>""", unsafe_allow_html=True)

# ── LIQUID GLASS THEMES (iOS-style frosted translucency) ─────────────────────
# Two variants: glass-light (bright frosted) and glass-dark (smoked glass).
if _THEME == "glass-light":
    st.markdown("""
<style>
:root {
  --bg:#eef1f7; --bg1:#e8ecf4; --bg2:rgba(255,255,255,.62); --bg3:rgba(255,255,255,.45);
  --border:rgba(255,255,255,.75); --border-hi:rgba(94,92,230,.45);
  --indigo:#5e5ce6; --indigo-lo:rgba(94,92,230,.10);
  --cyan:#0a84c1; --cyan-lo:rgba(10,132,193,.10);
  --teal:#0d9488; --amber:#c77700; --rose:#d6275d;
  --txt:#1c1c1e; --txt2:#5b5b60; --txt3:#9a9aa2;
  --glow-i:0 8px 32px rgba(31,38,135,.10);
  --glow-c:0 8px 32px rgba(31,38,135,.08);
  --glow-sm:0 2px 10px rgba(31,38,135,.06);
}
.stApp{
  background:#eef1f7!important;
  background-image:
    radial-gradient(ellipse 90% 60% at 15% -10%,rgba(94,92,230,.14) 0%,transparent 60%),
    radial-gradient(ellipse 70% 50% at 90% 0%,rgba(10,132,193,.10) 0%,transparent 55%),
    radial-gradient(ellipse 60% 45% at 70% 110%,rgba(255,150,160,.10) 0%,transparent 55%)!important;
  background-attachment:fixed!important;
}
.sc, .rules-box{
  background:rgba(255,255,255,.55)!important;
  backdrop-filter:blur(28px) saturate(180%)!important;
  -webkit-backdrop-filter:blur(28px) saturate(180%)!important;
  border:1px solid rgba(255,255,255,.78)!important;
  border-radius:18px!important;
  box-shadow:0 8px 32px rgba(31,38,135,.10), inset 0 1px 0 rgba(255,255,255,.95)!important;
}
.sc.hi{background:linear-gradient(135deg,rgba(94,92,230,.16),rgba(255,255,255,.5))!important;}
.sc.ds{background:linear-gradient(135deg,rgba(13,148,136,.13),rgba(255,255,255,.5))!important;}
.sc.dv{background:linear-gradient(135deg,rgba(199,119,0,.13),rgba(255,255,255,.5))!important;}
.sc .n{text-shadow:none!important;color:#1c1c1e!important;}
section[data-testid="stSidebar"]{
  background:rgba(255,255,255,.55)!important;
  backdrop-filter:blur(34px) saturate(180%)!important;
  -webkit-backdrop-filter:blur(34px) saturate(180%)!important;
  border-right:1px solid rgba(255,255,255,.7)!important;
  box-shadow:8px 0 32px rgba(31,38,135,.07)!important;
}
section[data-testid="stSidebar"]::before{background:linear-gradient(90deg,#5e5ce6,#0a84c1)!important;}
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p{color:var(--txt)!important;}
.stTabs [data-baseweb="tab-list"]{
  background:rgba(255,255,255,.5)!important;border:1px solid rgba(255,255,255,.75)!important;
  backdrop-filter:blur(20px) saturate(170%)!important;-webkit-backdrop-filter:blur(20px) saturate(170%)!important;
  border-radius:14px!important;box-shadow:0 4px 18px rgba(31,38,135,.07)!important;
}
.stTabs [aria-selected="true"]{
  background:rgba(255,255,255,.92)!important;color:var(--indigo)!important;
  box-shadow:0 2px 10px rgba(31,38,135,.12)!important;border-radius:10px!important;
}
.stButton>button{
  background:rgba(255,255,255,.6)!important;color:var(--txt)!important;
  border:1px solid rgba(255,255,255,.8)!important;border-radius:13px!important;
  backdrop-filter:blur(16px) saturate(160%)!important;-webkit-backdrop-filter:blur(16px) saturate(160%)!important;
  box-shadow:0 2px 12px rgba(31,38,135,.07), inset 0 1px 0 rgba(255,255,255,.95)!important;
}
.stButton>button:hover{background:rgba(255,255,255,.85)!important;border-color:rgba(94,92,230,.4)!important;}
.stButton>button[kind="primary"]{
  background:linear-gradient(135deg,rgba(94,92,230,.92),rgba(122,120,255,.88))!important;
  border:1px solid rgba(255,255,255,.5)!important;color:#fff!important;
  box-shadow:0 6px 22px rgba(94,92,230,.35), inset 0 1px 0 rgba(255,255,255,.45)!important;
}
.stTextArea textarea,.stTextInput input,
.stSelectbox [data-baseweb="select"]>div,
.stMultiSelect [data-baseweb="select"]>div{
  background:rgba(255,255,255,.62)!important;color:var(--txt)!important;
  border:1px solid rgba(255,255,255,.85)!important;border-radius:13px!important;
  backdrop-filter:blur(14px)!important;-webkit-backdrop-filter:blur(14px)!important;
}
.streamlit-expanderHeader{background:rgba(255,255,255,.55)!important;border-radius:14px!important;}
.streamlit-expanderContent{background:rgba(255,255,255,.4)!important;}
[data-baseweb="popover"] [role="listbox"],[data-baseweb="menu"]{
  background:rgba(255,255,255,.92)!important;backdrop-filter:blur(28px)!important;
  border:1px solid rgba(255,255,255,.85)!important;border-radius:14px!important;
}
hr{background:rgba(28,28,30,.12)!important;}
</style>""", unsafe_allow_html=True)

elif _THEME == "glass-dark":
    st.markdown("""
<style>
:root {
  --bg:#0b0b0f; --bg1:#101016; --bg2:rgba(38,38,48,.5); --bg3:rgba(48,48,60,.4);
  --border:rgba(255,255,255,.14); --border-hi:rgba(125,122,255,.5);
  --indigo:#7d7aff; --indigo-lo:rgba(125,122,255,.12);
  --cyan:#64d2ff; --cyan-lo:rgba(100,210,255,.10);
  --teal:#2dd4bf; --amber:#ffb340; --rose:#ff6482;
  --txt:#f2f2f7; --txt2:#aeaeb6; --txt3:#6c6c75;
  --glow-i:0 8px 32px rgba(0,0,0,.45);
  --glow-c:0 8px 32px rgba(0,0,0,.4);
  --glow-sm:0 2px 12px rgba(0,0,0,.35);
}
.stApp{
  background:#0b0b0f!important;
  background-image:
    radial-gradient(ellipse 90% 60% at 15% -10%,rgba(125,122,255,.16) 0%,transparent 60%),
    radial-gradient(ellipse 70% 50% at 90% 0%,rgba(100,210,255,.09) 0%,transparent 55%),
    radial-gradient(ellipse 60% 45% at 70% 110%,rgba(255,100,130,.08) 0%,transparent 55%)!important;
  background-attachment:fixed!important;
}
.sc, .rules-box{
  background:rgba(36,36,46,.45)!important;
  backdrop-filter:blur(28px) saturate(160%)!important;
  -webkit-backdrop-filter:blur(28px) saturate(160%)!important;
  border:1px solid rgba(255,255,255,.14)!important;
  border-radius:18px!important;
  box-shadow:0 8px 32px rgba(0,0,0,.45), inset 0 1px 0 rgba(255,255,255,.10)!important;
}
.sc.hi{background:linear-gradient(135deg,rgba(125,122,255,.2),rgba(36,36,46,.4))!important;}
.sc.ds{background:linear-gradient(135deg,rgba(45,212,191,.14),rgba(36,36,46,.4))!important;}
.sc.dv{background:linear-gradient(135deg,rgba(255,179,64,.13),rgba(36,36,46,.4))!important;}
.sc .n{color:#f2f2f7!important;text-shadow:0 0 22px rgba(125,122,255,.45)!important;}
section[data-testid="stSidebar"]{
  background:rgba(22,22,30,.55)!important;
  backdrop-filter:blur(36px) saturate(160%)!important;
  -webkit-backdrop-filter:blur(36px) saturate(160%)!important;
  border-right:1px solid rgba(255,255,255,.12)!important;
  box-shadow:8px 0 40px rgba(0,0,0,.5)!important;
}
section[data-testid="stSidebar"]::before{background:linear-gradient(90deg,#7d7aff,#64d2ff)!important;}
.stTabs [data-baseweb="tab-list"]{
  background:rgba(36,36,46,.45)!important;border:1px solid rgba(255,255,255,.13)!important;
  backdrop-filter:blur(22px) saturate(150%)!important;-webkit-backdrop-filter:blur(22px) saturate(150%)!important;
  border-radius:14px!important;
}
.stTabs [aria-selected="true"]{
  background:rgba(125,122,255,.25)!important;color:#cfcdff!important;
  box-shadow:0 0 0 1px rgba(125,122,255,.45), inset 0 1px 0 rgba(255,255,255,.16)!important;
  border-radius:10px!important;
}
.stButton>button{
  background:rgba(46,46,58,.5)!important;color:var(--txt)!important;
  border:1px solid rgba(255,255,255,.15)!important;border-radius:13px!important;
  backdrop-filter:blur(18px) saturate(150%)!important;-webkit-backdrop-filter:blur(18px) saturate(150%)!important;
  box-shadow:0 2px 14px rgba(0,0,0,.4), inset 0 1px 0 rgba(255,255,255,.10)!important;
}
.stButton>button:hover{background:rgba(125,122,255,.18)!important;border-color:rgba(125,122,255,.45)!important;}
.stButton>button[kind="primary"]{
  background:linear-gradient(135deg,rgba(125,122,255,.85),rgba(100,210,255,.7))!important;
  border:1px solid rgba(255,255,255,.3)!important;color:#fff!important;
  box-shadow:0 6px 26px rgba(125,122,255,.4), inset 0 1px 0 rgba(255,255,255,.35)!important;
}
.stTextArea textarea,.stTextInput input,
.stSelectbox [data-baseweb="select"]>div,
.stMultiSelect [data-baseweb="select"]>div{
  background:rgba(36,36,46,.5)!important;color:var(--txt)!important;
  border:1px solid rgba(255,255,255,.15)!important;border-radius:13px!important;
  backdrop-filter:blur(14px)!important;-webkit-backdrop-filter:blur(14px)!important;
}
.streamlit-expanderHeader{background:rgba(36,36,46,.5)!important;border-radius:14px!important;}
.streamlit-expanderContent{background:rgba(28,28,36,.4)!important;}
[data-baseweb="popover"] [role="listbox"],[data-baseweb="menu"]{
  background:rgba(30,30,40,.92)!important;backdrop-filter:blur(30px)!important;
  border:1px solid rgba(255,255,255,.16)!important;border-radius:14px!important;
}
hr{background:rgba(255,255,255,.1)!important;}
</style>""", unsafe_allow_html=True)

# ── Theme-aware constants for components.html cards ───────────────────────────
# Cards render inside iframes, which DON'T inherit the parent page's CSS vars,
# so we pass concrete colors based on the active theme. (backdrop-filter can't
# blur the parent page from inside an iframe, so glass cards here use soft
# translucent fills + hairline highlights that read as glass.)
if _THEME == "glass-light":
    _C = {
        "card_bg": "linear-gradient(135deg,rgba(255,255,255,.94),rgba(247,249,255,.9))",
        "card_br": "rgba(255,255,255,.95)",
        "card_sh": "0 8px 28px rgba(31,38,135,.10), inset 0 1px 0 rgba(255,255,255,.98)",
        "txt": "#1c1c1e",
        "txt2": "#5b5b60",
        "txt3": "#9a9aa2",
        "row_br": "rgba(94,92,230,.10)",
        "th_bg": "rgba(94,92,230,.05)",
        "tbl_bg": "rgba(255,255,255,.92)",
        "foot_bg": "rgba(94,92,230,.04)",
    }
    _BODY_TXT = "#1c1c1e"
elif _THEME == "glass-dark":
    _C = {
        "card_bg": "linear-gradient(135deg,rgba(40,40,52,.92),rgba(30,30,40,.94))",
        "card_br": "rgba(255,255,255,.16)",
        "card_sh": "0 8px 28px rgba(0,0,0,.5), inset 0 1px 0 rgba(255,255,255,.12)",
        "txt": "#f2f2f7",
        "txt2": "#aeaeb6",
        "txt3": "#6c6c75",
        "row_br": "rgba(255,255,255,.08)",
        "th_bg": "rgba(255,255,255,.04)",
        "tbl_bg": "rgba(32,32,42,.92)",
        "foot_bg": "rgba(255,255,255,.03)",
    }
    _BODY_TXT = "#f2f2f7"
elif _THEME == "light":
    _C = {
        "card_bg": "#ffffff",
        "card_br": "#e2e5ea",
        "card_sh": "0 1px 2px rgba(20,32,54,.06),0 0 0 1px rgba(20,32,54,.02)",
        "txt": "#16202e",
        "txt2": "#5b6675",
        "txt3": "#8a93a1",
        "row_br": "#eef0f3",
        "th_bg": "#f4f5f7",
        "tbl_bg": "#ffffff",
        "foot_bg": "#f7f8fa",
    }
    _BODY_TXT = "#16202e"
else:
    _C = {
        "card_bg": "linear-gradient(135deg,rgba(14,14,26,.95),rgba(19,19,31,.98))",
        "card_br": "rgba(99,102,241,.25)",
        "card_sh": "0 0 0 1px rgba(255,255,255,.04),0 8px 32px rgba(0,0,0,.4)",
        "txt": "#e2e8f0",
        "txt2": "#94a3b8",
        "txt3": "#475569",
        "row_br": "rgba(99,102,241,.07)",
        "th_bg": "rgba(255,255,255,.02)",
        "tbl_bg": "rgba(13,13,26,.9)",
        "foot_bg": "rgba(255,255,255,.015)",
    }
    _BODY_TXT = "#e2e8f0"

# ── Micro-interactions: animated buttons & status transitions (all themes) ───
st.markdown("""
<style>
.stButton>button{
  transition:transform .12s cubic-bezier(.34,1.56,.64,1), box-shadow .25s ease,
             background .2s ease, border-color .2s ease, color .2s ease!important;
  will-change:transform;
}
.stButton>button:hover{transform:translateY(-1px);}
.stButton>button:active{
  transform:scale(.93)!important;
  box-shadow:0 0 0 5px rgba(99,102,241,.18)!important;
}
@keyframes statusPop{
  0%{transform:scale(.55);opacity:0}
  60%{transform:scale(1.1)}
  100%{transform:scale(1);opacity:1}
}
@keyframes pulseDot{
  0%,100%{opacity:1;transform:scale(1)}
  50%{opacity:.4;transform:scale(.8)}
}
@keyframes ringPulse{
  0%{box-shadow:0 0 0 0 rgba(34,211,238,.45)}
  70%{box-shadow:0 0 0 7px rgba(34,211,238,0)}
  100%{box-shadow:0 0 0 0 rgba(34,211,238,0)}
}
</style>""", unsafe_allow_html=True)

SHARED_CSS = """<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@600;700;800&family=DM+Mono:wght@400;500&family=DM+Sans:wght@400;500;600&display=swap');
*{box-sizing:border-box;margin:0;padding:0;}
body{
  font-family:'DM Sans',sans-serif;
  background:transparent;
  color:""" + _BODY_TXT + """;
  overflow-x:auto;
  -webkit-overflow-scrolling:touch;
}
@keyframes rowIn{from{opacity:0;transform:translateX(-6px)}to{opacity:1;transform:translateX(0)}}
@keyframes glassIn{from{opacity:0;transform:translateY(10px)}to{opacity:1;transform:translateY(0)}}
table{min-width:0;width:100%;border-collapse:collapse;}
@media(max-width:600px){
  table{font-size:.72rem!important;}
  th,td{padding:5px 7px!important;}
  /* Hide low-value columns on phones — Bld + Service already shown in the
     card header, so the table keeps Room/Guest/Time/Pet/LateOut */
  .m-hide{display:none!important;}
}
</style>"""

# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
def _save_reassignment(fg):
    """Persist a manual housekeeper/RQS change.

    inspectors_data is rebuilt from the charts rather than patched, so the
    Inspectors view can never drift from what the charts actually say.
    """
    order, entries = [], {}
    for g in fg:
        name = g.get("inspector") or ""
        if not name:
            continue
        if name not in entries:
            order.append(name)
            entries[name] = {"id": len(order), "name": name,
                             "role": "RQS2" if name == st.session_state.get("rqs2")
                                     else "FC",
                             "groups": [], "buildings": set()}
        entries[name]["groups"].append(g["label"])
        entries[name]["buildings"] |= set(g.get("blds") or set())
    rebuilt = []
    for name in order:
        ent = entries[name]
        ent["buildings"] = sorted(ent["buildings"])
        rebuilt.append(ent)
    st.session_state["groups_data"] = fg
    st.session_state["inspectors_data"] = rebuilt
    used = {g.get("housekeeper", "") for g in fg
            if g.get("housekeeper") and not is_unassigned_hk(g.get("housekeeper", ""))
            and g.get("housekeeper") != "Manager"}
    st.session_state["used_hk_set"] = used
    try:
        db.save_full_schedule({
            "groups_data": fg, "total_rooms": st.session_state.get("total_rooms", 0),
            "inspectors_data": rebuilt, "used_hk_set": list(used),
            "hk_roster": dict(st.session_state.get("hk_roster", {})),
            "generated_by": st.session_state.get("username", "unknown"),
        })
    except Exception as ex:
        print(f"[app] _save_reassignment failed: {ex}")

def _persist_roster():
    """Save the current housekeeper + inspector rosters to the database so all
    changes (add, remove, present-toggle, building move, bulk set) survive page
    reloads, new sessions, and redeploys. Silent no-op if the DB is unreachable."""
    try:
        db.save_roster(st.session_state.get("hk_roster", {}),
                       st.session_state.get("insp_roster", {}))
    except Exception as _ex:
        print(f"[app] _persist_roster failed: {_ex}")

def _init_state():
    if "hk_roster" not in st.session_state:
        # Load the standing roster from the database first — this is what people
        # actually see, so add/remove of staff must persist. Only if nothing has
        # ever been saved do we seed from the hard-coded DEFAULT_HK.
        saved_roster = None
        try:
            saved_roster = db.load_roster()
        except Exception:
            saved_roster = None
        if saved_roster and saved_roster.get("hk_roster"):
            st.session_state["hk_roster"] = saved_roster["hk_roster"]
            st.session_state["insp_roster"] = saved_roster.get("insp_roster") \
                or {n: True for n in DEFAULT_INSPECTORS}
        else:
            roster = {}
            for bld, names in DEFAULT_HK.items():
                for n in names:
                    roster[n] = {"building": bld, "present": True}
            st.session_state["hk_roster"] = roster
            st.session_state["insp_roster"] = {n: True for n in DEFAULT_INSPECTORS}
    if "insp_roster" not in st.session_state:
        st.session_state["insp_roster"] = {n: True for n in DEFAULT_INSPECTORS}
    for k, default in [("groups_data",None),("total_rooms",None),
                        ("inspectors_data",None),("used_hk_set",None),
                        ("last_email",None),("rqs1",""),("rqs2",""),
                        ("priority_hks",[]),("ds_team",[])]:
        if k not in st.session_state:
            st.session_state[k] = default

def _auto_apply_today(force=False):
    """Set today's attendance from the stored staff schedule, once per day.

    Without this the roster keeps whatever presence flags were last applied, so
    a new day opens showing yesterday's crew. Guarded by a marker in the
    database rather than session state, so it happens once for the property —
    not once per person who opens the app, which would wipe out any manual
    correction made after the first run.

    Returns a short note describing what happened, or None if it did nothing.
    """
    import roster_import as _ri
    # A deploy can leave the previous copy in sys.modules; reload if it is old.
    if getattr(_ri, "__version__", 0) < 12:
        import importlib
        _ri = importlib.reload(_ri)
    # Property-local date, NOT the server's. On a UTC host date.today() rolls
    # over around 5-6pm Mountain, which would pull tomorrow's crew mid-shift.
    today = _datetime.now(_MTN_TZ).date().isoformat()
    if not force and st.session_state.get("_autoapply_done") == today:
        return None
    try:
        marker = db.load_autoapply() or {}
        if not force and marker.get("date") == today:
            st.session_state["_autoapply_done"] = today
            st.session_state["_autoapply_note"] = marker.get("note", "")
            return None
        wk = _ri.find_week_key(db.staff_week_keys(), today)
        if not wk:
            st.session_state["_autoapply_done"] = today   # nothing to do; don't retry
            return None
        week = db.load_staff_week(wk)
        if not week:
            st.session_state["_autoapply_done"] = today
            return None
        update = _ri.day_roster(week, db.load_staff_overrides(), wk, today,
                                st.session_state.get("hk_roster", {}))
        if not update:
            st.session_state["_autoapply_done"] = today
            return None
        # keep_missing: anyone not on the sheet stays on the roster, marked
        # absent, so nobody silently disappears when this runs unattended.
        st.session_state["hk_roster"] = _ri.merge_roster(
            update, st.session_state.get("hk_roster", {}), keep_missing=True)
        new_insp = dict(update["insp_roster"])
        for name in st.session_state.get("insp_roster", {}):
            new_insp.setdefault(name, False)
        st.session_state["insp_roster"] = new_insp
        st.session_state["rqs1"] = update["rqs1"]
        st.session_state["rqs2"] = update["rqs2"]
        st.session_state["ds_team"] = [
            n for n in update["ds_team"]
            if st.session_state["hk_roster"].get(n, {}).get("present")]
        # Attendance checkboxes redraw from the roster once their keys are gone.
        for _k in [k for k in list(st.session_state) if k.startswith(("att_", "insp_att_"))]:
            st.session_state.pop(_k, None)
        # The RQS selectboxes are different: they are keyed widgets that write
        # rqs1/rqs2 back on every run. Clearing their keys resets them to
        # "no one" and immediately overwrites what was just set, so point them
        # at the people the sheet names instead of deleting them.
        for _sel, _val in (("rqs1_sel", update["rqs1"]), ("rqs2_sel", update["rqs2"])):
            st.session_state[_sel] = _val if (_val and new_insp.get(_val)) else RQS_NONE
        n_hk = sum(1 for v in st.session_state["hk_roster"].values() if v["present"])
        note = (f"{n_hk} HK · {len(st.session_state['ds_team'])} on daily service · "
                f"RQS {update['rqs1'] or '—'}/{update['rqs2'] or '—'} (sheet {week['sheet']})")
        try:
            db.save_roster(st.session_state["hk_roster"], st.session_state["insp_roster"])
        except Exception:
            pass
        db.save_autoapply({"date": today, "note": note, "week": wk,
                           "at": _datetime.now().isoformat(timespec="seconds")})
        st.session_state["_autoapply_done"] = today
        st.session_state["_autoapply_note"] = note
        return note
    except Exception as _ex:
        print(f"[app] auto-apply failed: {_ex}")
        st.session_state["_autoapply_done"] = today
        return None

_init_state()
auth.init_auth()
if st.session_state.get("logged_in"):
    _auto_apply_today()

# ══════════════════════════════════════════════════════════════════════════════
# LOGIN GATE
# ══════════════════════════════════════════════════════════════════════════════
if not st.session_state.get("logged_in"):
    st.markdown("""<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400&display=swap');
/* Hide the entire sidebar + nav on the login screen */
section[data-testid="stSidebar"]{display:none !important;}
[data-testid="stSidebarNav"]{display:none !important;}
[data-testid="collapsedControl"]{display:none !important;}
button[kind="header"]{display:none !important;}
.stApp{
  background:#f4f5f7 !important;
  background-image:none !important;
}
.block-container{
  padding-top:0 !important;
  max-width:420px !important;
  margin:0 auto;
}
/* Center vertically */
.block-container > div:first-child {
  min-height:100vh;
  display:flex;
  flex-direction:column;
  justify-content:center;
  padding: 2rem 0;
}
.stButton>button{
  width:100%!important;border-radius:8px!important;font-weight:600!important;
  background:#2563a8!important;border:1px solid #2563a8!important;
  color:#fff!important;padding:12px!important;font-size:.88rem!important;letter-spacing:.01em;
  box-shadow:0 1px 2px rgba(20,32,54,.10)!important;
  transition:all .15s!important;
}
.stButton>button:hover{
  background:#1f5697!important;
  box-shadow:0 2px 6px rgba(20,32,54,.16)!important;
}
.stTextInput input {
  background:#ffffff!important;
  border:1px solid #c3c9d4!important;
  border-radius:8px!important;color:#16202e!important;
  font-family:'DM Sans',sans-serif!important;font-size:.88rem!important;
  padding:12px 14px!important;
  -webkit-text-fill-color:#16202e!important;
}
.stTextInput input:focus{
  border-color:#2563a8!important;
  box-shadow:0 0 0 3px rgba(37,99,168,.12)!important;
  -webkit-text-fill-color:#16202e!important;
}
.stTextInput input::placeholder{color:#8a93a1!important;-webkit-text-fill-color:#8a93a1!important;}
/* Stop browser autofill from forcing an off-color box */
.stTextInput input:-webkit-autofill,
.stTextInput input:-webkit-autofill:hover,
.stTextInput input:-webkit-autofill:focus{
  -webkit-text-fill-color:#16202e!important;
  caret-color:#16202e!important;
  -webkit-box-shadow:0 0 0 1000px #ffffff inset!important;
  box-shadow:0 0 0 1000px #ffffff inset!important;
  transition:background-color 9999s ease-in-out 0s!important;
}
label{color:#5b6675!important;font-size:.78rem!important;font-weight:500!important;font-family:'DM Sans',sans-serif!important;}
footer{visibility:hidden!important;}
</style>""", unsafe_allow_html=True)

    # Logo + title
    st.markdown("""
<div style="text-align:center;margin-bottom:28px">
  <div style="width:52px;height:52px;margin:0 auto 14px;border-radius:12px;
              background:#2563a8;display:flex;align-items:center;justify-content:center;
              box-shadow:0 2px 8px rgba(37,99,168,.25);font-size:1.05rem;font-weight:700;font-family:'Syne',sans-serif;color:#ffffff;letter-spacing:.02em">GC8</div>
  <div style="font-family:'Syne',sans-serif;font-size:1.5rem;font-weight:700;letter-spacing:-.02em;
              color:#16202e;margin-bottom:5px">Grand Timber GC8</div>
  <div style="font-family:'DM Sans',sans-serif;font-size:.76rem;color:#5b6675;letter-spacing:.06em;
              text-transform:uppercase;font-weight:500">Housekeeping · Scheduling · Live Tracking</div>
</div>
""", unsafe_allow_html=True)

    # Card wrapper
    st.markdown("""
<div style="background:#ffffff;border:1px solid #e2e5ea;
            border-radius:14px;padding:30px 28px 26px;
            box-shadow:0 1px 3px rgba(20,32,54,.08),0 8px 24px rgba(20,32,54,.06)">
  <div style="font-family:'Syne',sans-serif;font-size:1.02rem;font-weight:700;color:#16202e;
              margin-bottom:4px">Welcome back</div>
  <div style="font-family:'DM Sans',sans-serif;font-size:.8rem;color:#5b6675;margin-bottom:24px">
    Sign in with your Grand Timber email
  </div>
""", unsafe_allow_html=True)
    _db_ok = True; _db_msg = ""
    try:
        db.ensure_admin_exists()
    except Exception as _ex:
        _db_ok = False; _db_msg = str(_ex)
    if not _db_ok:
        st.error("Cannot connect to database.")
        st.markdown(f"**Error:** `{_db_msg}`\n\n**Fix:** Add `SUPABASE_URL` and `SUPABASE_KEY` to Streamlit Secrets.")
        st.stop()
    with st.form("login_form"):
        _uname = st.text_input("Username")
        _pw = st.text_input("Password", type="password")
        _remember = st.checkbox("Keep me signed in on this device", value=True)
        _sub = st.form_submit_button("Sign In", type="primary", use_container_width=True)
    if _sub:
        if not _uname or not _pw:
            st.error("Please enter both username and password.")
        else:
            _user = db.authenticate(_uname.strip(), _pw)
            if _user:
                auth.login(_user)
                if _remember:
                    try:
                        import session as _session
                        _session.remember(_user)
                    except Exception as _ex:
                        print(f"[app] could not persist session: {_ex}")
                # Record who signed in and when (best-effort; never blocks login).
                try:
                    db.log_login(_user.get("username",""),
                                 _user.get("display_name") or _user.get("username",""),
                                 _user.get("role",""))
                except Exception:
                    pass
                # Everyone lands on their own schedule first. Managers navigate
                # on to the scheduler from there.
                try:
                    st.switch_page("pages/4_My_Home.py")
                except Exception:
                    st.rerun()
            else:
                st.error("Invalid username or password.")

    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("""
<div style="text-align:center;margin-top:20px;font-family:'DM Mono',monospace;
            font-size:.65rem;color:#1e293b;letter-spacing:.06em">
  GRAND TIMBER GC8 · CONFIDENTIAL
</div>""", unsafe_allow_html=True)
    st.stop()

# ── Restore today's shared schedule — ONCE per browser session ────────────────
# Runs only on the first script execution of a session. After that the flag
# stays True for the whole session, so generating a new schedule (which sets
# groups_data directly) is never overwritten by a stale DB load.
if not st.session_state.get("_did_initial_restore", False):
    st.session_state["_did_initial_restore"] = True
    try:
        saved = db.load_full_schedule()
        # NOTE: we deliberately do NOT restore hk_roster / insp_roster from the
        # daily schedule here. The authoritative, persistent roster is loaded in
        # _init_state() from db.load_roster(). A day's saved schedule can contain
        # a stale roster (e.g. someone you removed yesterday), and restoring it
        # here would bring removed people back — the exact bug we're fixing.
        if saved and saved.get("groups_data"):
            _loaded_groups = saved.get("groups_data") or []
            # JSON serialization turns sets into lists (or strings). Restore them
            # to sets so sorted(g["blds"]) and set operations work correctly.
            for _g in _loaded_groups:
                _b = _g.get("blds", [])
                if isinstance(_b, str):
                    # e.g. "{1, 2}"extract digits
                    _b = [int(x) for x in re.findall(r'\d+', _b)]
                _g["blds"] = set(_b) if not isinstance(_b, set) else _b
                _f = _g.get("floors", [])
                if isinstance(_f, str):
                    _f = [int(x) for x in re.findall(r'\d+', _f)]
                _g["floors"] = set(_f) if not isinstance(_f, set) else _f
            st.session_state["groups_data"] = _loaded_groups
            st.session_state["total_rooms"] = saved.get("total_rooms", 0)
            st.session_state["inspectors_data"] = saved.get("inspectors_data", [])
            st.session_state["used_hk_set"] = set(saved.get("used_hk_set", []))
    except Exception:
        pass
# ══════════════════════════════════════════════════════════════════════════════
SKIP_SERVICES = {"p/u models","pu models","p/u model","showcase","model unit","p/u"}

def normalize_service(raw: str) -> str:
    s = re.sub(r'\s+', ' ', str(raw).strip().lower())
    if s in SKIP_SERVICES or "p/u" in s or (s.startswith("p") and "model" in s):
        return "__SKIP__"
    if "daily" in s: return SVC_DS
    # Sheet Exchange is a light turn-down service handled by the Daily Service
    # crew, so it's grouped with Daily Service.
    if "sheet" in s: return SVC_DS
    # "Full Clean (IH)" / "Full Clean( IH)" / "... IH" -> separate IH stream,
    # packed apart from regular Full Clean and inspected by RQS 2.
    if ("full clean" in s or s.startswith("fc")) and "ih" in s:
        return SVC_IH
    if s.startswith("full clean") or s.startswith("fc"): return SVC_FC
    if "dust" in s or "d&v" in s or "dnv" in s: return SVC_DV
    if "vac" in s: return SVC_DV
    return SVC_FC

def parse_room_code(room: str) -> dict:
    s = str(room).strip()
    try:
        bld = int(s[0])
        floor = int(s[1]) if len(s)>1 and s[1].isdigit() else 0
        digits = "".join(c for c in s[2:] if c.isdigit())
        num = int(digits) if digits else 0
    except Exception:
        bld, floor, num = 0, 0, 0
    return {"bld": bld, "floor": floor, "num": num}

def get_building(room): return parse_room_code(room)["bld"]

def expand_compound_room(room_str: str) -> list:
    m = re.match(r'^([1-9]\d{3})([A-Z]{2,4})$', room_str.upper())
    if not m: return [room_str.upper()]
    base, suffix = m.group(1), m.group(2)
    if len(suffix) == 2 and ord(suffix[1]) == ord(suffix[0]) + 1:
        return [f"{base}{suffix[0]}", f"{base}{suffix[1]}"]
    return [room_str.upper()]

def expand_rooms(raw_list: list) -> list:
    return [r for s in raw_list for r in expand_compound_room(s)]

def parse_email_notes(text: str) -> dict:
    late_co: dict = {}
    notes: dict = {}
    if not text or not text.strip():
        return {"late_checkout": late_co, "notes": notes}
    ROOM_RE = re.compile(r'\b([1-9]\d{3}[A-Z]{1,4})\b')
    TIME_RE = re.compile(r'\b(\d{1,2}:\d{2}\s*(?:am|pm))\b', re.IGNORECASE)
    SECTION_RE = re.compile(r'^([A-Za-z][A-Za-z &\'/]+):\s*$')
    MOVE_RE = re.compile(r'([1-9]\d{3}[A-Z]{1,4})\s*[-\u2013>\u2192]+\s*([1-9]\d{3}[A-Z]{1,4})')
    CELEB_RE = re.compile(r'^(Birthday|Anniversary|Misc\.?)$', re.IGNORECASE)
    DEBULLET = re.compile(r'^[\s\t]*[*\u2022\u25e6\u2023\u2043\-]?\s*')
    NOTE_LABELS = {
        "vip inspections":"VIP","room moves":"Room Move","stayovers":"Stayover",
        "robes":"Robes","pack n play":"Pack n Play","highchairs":"Highchair",
        "rollaway":"Rollaway","special requests":"Special Request",
        "dogs arriving":"Dog arriving","celebrations":"Celebration",
        "early ins":"Early In","late arrival":"Late Arrival",
    }
    LATE_KEY = "late checkouts"
    section = None; sub_label = None; late_time = None
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped: continue
        hdr_m = SECTION_RE.match(stripped)
        if hdr_m and not ROOM_RE.search(stripped):
            hdr_key = hdr_m.group(1).strip().lower()
            if hdr_key in NOTE_LABELS or hdr_key == LATE_KEY:
                section = hdr_key; sub_label = None; late_time = None; continue
        content = DEBULLET.sub('', line).strip()
        if not content or content.lower() == "n/a": continue
        if section == LATE_KEY:
            time_m = TIME_RE.search(content)
            rooms = expand_rooms(ROOM_RE.findall(content.upper()))
            if time_m:
                late_time = re.sub(r'\s+', ' ', time_m.group(1).strip())
                for rm in rooms: late_co[rm] = f"Late Out: {late_time}"
            elif rooms and late_time:
                for rm in rooms: late_co[rm] = f"Late Out: {late_time}"
            continue
        if not section or section not in NOTE_LABELS: continue
        label = NOTE_LABELS[section]
        if section == "room moves":
            for mv in MOVE_RE.finditer(content.upper()):
                for rf in expand_compound_room(mv.group(1)):
                    for rt in expand_compound_room(mv.group(2)):
                        # Tag BOTH rooms with the full move so origin and
                        # destination are both visible, e.g. "Room Move 2234A>2234B".
                        move_txt = f"Room Move {rf}>{rt}"
                        notes.setdefault(rf, []).append(move_txt)
                        notes.setdefault(rt, []).append(move_txt)
            continue
        if section == "celebrations":
            cm = CELEB_RE.match(content)
            if cm:
                t = cm.group(1).strip()
                sub_label = None if t.lower().startswith("misc") else t; continue
            if sub_label: label = f"Celebration ({sub_label})"
        rooms = expand_rooms(ROOM_RE.findall(content.upper()))
        if not rooms: continue
        qty = re.search(r'x(\d+)', content, re.IGNORECASE)
        qty_s = f"x{qty.group(1)}" if qty else ""
        if section == "special requests":
            # Capture the actual request detail (e.g. "2135A - Humidifier" ->
            # "Special Request: Humidifier"). Skip "n/a". Strip the room codes and
            # surrounding punctuation to leave just the request text.
            detail = ROOM_RE.sub('', content.upper() if False else content)
            detail = re.sub(r'\b[1-9]\d{3}[A-Z]{1,4}\b', '', detail)
            detail = detail.strip(" -\u2013:;,\t").strip()
            if detail.lower() in ("n/a","na",""):
                for rm in rooms: notes.setdefault(rm, []).append(f"{label}{qty_s}")
            else:
                for rm in rooms:
                    notes.setdefault(rm, []).append(f"{label}: {detail}{qty_s}")
            continue
        for rm in rooms: notes.setdefault(rm, []).append(f"{label}{qty_s}")
    return {"late_checkout": late_co, "notes": notes}

def excel_to_room_text(file_obj):
    """Convert an uploaded 'Housekeeping Dashboard' .xlsx into the same
    tab-separated text the paste box accepts, so it can flow through the exact
    same parse_rooms() pipeline. Reads only the cleaning-services table and
    stops before the footer/summary sections. Column positions are located by
    header label (not fixed indices) so minor export changes don't misread.
    Returns (tsv_text, n_rooms, sheet_name)."""
    xls = pd.ExcelFile(file_obj)
    sheet = None
    for cand in xls.sheet_names:
        if "housekeeping dashboard" in str(cand).strip().lower():
            sheet = cand; break
    if sheet is None:
        sheet = xls.sheet_names[1] if len(xls.sheet_names) > 1 else xls.sheet_names[0]
    df = pd.read_excel(xls, sheet_name=sheet, header=None)

    # Find the header row (contains both "Room" and "Service").
    hdr = None
    for i in range(min(30, len(df))):
        vals = [str(v).strip().lower() for v in df.iloc[i] if pd.notna(v)]
        if "room" in vals and "service" in vals:
            hdr = i; break
    if hdr is None:
        raise ValueError("Couldn't find a 'Room'/'Service' header row in the sheet.")

    header_cells = {j: str(v).strip().lower()
                    for j, v in enumerate(df.iloc[hdr]) if pd.notna(v)}
    def find_col(*needles):
        # Prefer an exact header match, then fall back to substring.
        for n in needles:
            for j, name in header_cells.items():
                if name == n: return j
        for n in needles:
            for j, name in header_cells.items():
                if n in name: return j
        return None
    c_room = find_col("room")
    c_svc = find_col("service")
    # Prefer the canonical minutes column ("time (min)" / "sum of time (min)" /
    # "time") over any other column that merely contains the word "time" (e.g. a
    # "clean time start"). This prevents reading the wrong number for a room.
    c_time = find_col("time (min)", "sum of time (min)", "time(min)", "time")
    c_pet = find_col("pet")
    c_guest = find_col("current guest or status", "current guest", "guest")
    c_late = find_col("late checkout", "late check")
    c_status= find_col("status") # exact -> real Status col, not "guest or status"
    c_notes = find_col("notes")
    c_arr = find_col("arriving guest", "arriving")
    c_rtype = find_col("res type")

    _ROOM_RE = re.compile(r'^[1-9]\d{2,3}[A-Z]{1,4}$')
    order = [("Room", c_room), ("Service", c_svc), ("Time", c_time), ("Pet", c_pet),
             ("Current Guest or Status", c_guest), ("Res Type", c_rtype),
             ("Late Checkout", c_late), ("Status", c_status), ("Notes", c_notes),
             ("Arriving Guest", c_arr)]
    lines = ["\t".join(name for name, _ in order)]
    n_rooms = 0
    for i in range(hdr + 1, len(df)):
        rv = df.iloc[i, c_room] if c_room is not None else None
        if pd.isna(rv): continue
        room = str(rv).strip()
        if not _ROOM_RE.match(room.upper()):
            low = room.lower()
            if any(k in low for k in ("non-clean", "range totals", "daily labor",
                                       "daily housekeeper", "housekeeping hold", "room type")):
                break # reached the footer — stop reading rooms
            continue # stray non-room line — skip
        def cell(idx):
            if idx is None: return ""
            v = df.iloc[i, idx]
            if pd.isna(v): return ""
            if isinstance(v, float) and v == int(v): return str(int(v))
            return str(v).strip()
        lines.append("\t".join(cell(idx) for _, idx in order))
        n_rooms += 1
    return "\n".join(lines), n_rooms, sheet

def parse_rooms(text: str) -> pd.DataFrame:
    lines = [l for l in text.strip().splitlines() if l.strip()]
    if not lines: return pd.DataFrame()
    rows = [re.split(r"\t", l) for l in lines]
    header = [c.strip().lower() for c in rows[0]]

    def col(*names):
        # Prefer an EXACT header match before falling back to substring, so a
        # query like "status"doesn't accidentally match "current guest or
        # status". Names are tried in priority order.
        for n in names:
            for i, h in enumerate(header):
                if h.strip().lower() == n: return i
        for n in names:
            for i, h in enumerate(header):
                h_clean = h.strip().lower()
                if not h_clean: continue
                if n in h_clean: return i
        return None

    has_header = any(h in ("room","service","time","guest","current guest") for h in header)
    if has_header:
        data_rows = rows[1:]
        i_room = col("room"); i_svc = col("service")
        i_time = col("time (min)","sum of time (min)","time(min)","time")
        i_pet = col("pet"); i_guest = col("current guest","guest")
        i_late = col("late checkout","late check")
        i_status = col("status"); i_notes = col("notes")
        i_arriving = col("arriving guest","arriving"); i_restype = col("res type")
    else:
        data_rows = rows
        i_room,i_svc,i_time,i_pet,i_guest = 0,1,2,3,4
        i_late=i_status=i_notes=i_arriving=i_restype = None

    def get(row, idx, default=""):
        try: return str(row[idx]).strip() if idx is not None and idx<len(row) else default
        except: return default

    records = []
    for row in data_rows:
        room = get(row, i_room); svc = get(row, i_svc)
        ts = get(row, i_time)
        try: ti = int(float(ts))
        except: ti = 0
        if not room: continue
        norm_svc = normalize_service(svc)
        # P/U Models used to be dropped (__SKIP__). Now we keep them as a
        # "verify"room (no HK/RQS, pushed to the bottom for manual review).
        is_pu_skip = (norm_svc == "__SKIP__")
        if is_pu_skip:
            norm_svc = SVC_FC # give it a real service type so it can render
        if ti <= 0:
            if norm_svc == SVC_DV:
                ti = DV_DEFAULT_TIME
            else:
                ti = default_time_for(room, norm_svc)
                if ti <= 0:
                    suffix = room[-1].upper() if room else ""
                    if suffix == "E": ti = 140
                    elif suffix in ("D","A"):ti = 120
                    else: ti = 70
        import re as _re
        raw_guest = get(row, i_guest)
        norm_guest = _re.sub(r'\s+', ' ', raw_guest.strip())
        status_raw = get(row, i_status).strip().lower()
        guest_raw = norm_guest.lower().strip()
        notes_raw_val = get(row, i_notes).strip().lower()
        svc_raw_lower = str(svc).strip().lower()
        has_stayover_excel = "stayover" in notes_raw_val or "stay over" in notes_raw_val
        # P/U Models can appear in the service column OR the notes column
        has_pu_models = ("p/u model" in notes_raw_val or "pu model" in notes_raw_val
                         or "p/u model" in svc_raw_lower or "pu model" in svc_raw_lower)
        # "verify"rooms: stayover or P/U models — never auto-assign HK/RQS,
        # pushed to the bottom of the schedule for manual review.
        needs_verify = has_stayover_excel or has_pu_models or is_pu_skip
        row_text = " ".join(str(c).strip().lower() for c in row if c)
        has_pending_anywhere = "pending" in row_text
        is_uncertain = (
            (guest_raw in ("unallocated","---","room, walk","","deposit, deposit") and
             ("pending" in status_raw or has_pending_anywhere))
            or has_stayover_excel
        )
        records.append({
            "Room":get(row,i_room),"Service":norm_svc,"ServiceRaw":svc,
            "Time":ti,"Pet":get(row,i_pet),"Guest":norm_guest,
            "LateCheckout":get(row,i_late),"Status":get(row,i_status),
            "NotesRaw":get(row,i_notes),"ArrivingGuest":get(row,i_arriving),
            "ResType":get(row,i_restype),"uncertain":is_uncertain,
            "verify":needs_verify,
        })
    if not records: return pd.DataFrame()
    df = pd.DataFrame(records)
    pc = df["Room"].apply(parse_room_code)
    df["bld"] = pc.apply(lambda x: x["bld"])
    df["floor"] = pc.apply(lambda x: x["floor"])
    df["num"] = pc.apply(lambda x: x["num"])
    return df

# ══════════════════════════════════════════════════════════════════════════════
# GROUPING LOGIC
# ══════════════════════════════════════════════════════════════════════════════
def bld_ok(blds, b):
    for x in blds:
        if (x==2 and b==3) or (x==3 and b==2): return False
    return True

def same_bld(blds, ublds): return len(blds | ublds) == 1

def proximity_score(group_rooms, unit_rooms):
    if not group_rooms or not unit_rooms: return 999
    total, count = 0, 0
    for gr in group_rooms:
        for ur in unit_rooms:
            gb, ub = gr.get("bld",0), ur.get("bld",0)
            gf, uf = gr.get("floor",0), ur.get("floor",0)
            if gb != ub: total += 300
            elif gf != uf: total += 30
            else: total += min(abs(gr.get("num",0)-ur.get("num",0))//10, 9)
            count += 1
    return total // max(count,1)

def can_add_fc(g, unit):
    if g.get("service_type") != SVC_FC: return False
    ut = sum(r["time"] for r in unit)
    if g["time"]+ut > MAX_FC: return False
    for r in unit:
        if not bld_ok(g["blds"], r["bld"]): return False
    new_blds = g["blds"] | set(r["bld"] for r in unit)
    if 2 in new_blds and 3 in new_blds: return False
    u140 = sum(1 for r in unit if r["time"]==140)
    u120 = sum(1 for r in unit if r["time"]==120)
    new_c140 = g["c140"] + u140
    new_c120 = g["c120"] + u120
    if new_c140 > 1: return False
    # A group may not hold a 140 together with MORE THAN ONE 120.
    # 120+140+70 · 120+120+120 · 120+120+140 
    if new_c140 >= 1 and new_c120 > 1: return False
    return True

def can_add_ds(g, unit, allow_overflow=False):
    if g.get("service_type") != SVC_DS: return False
    ut = sum(r["time"] for r in unit)
    cap = DS_OVER if allow_overflow else MAX_DS
    if g["time"]+ut > cap: return False
    # Daily Service MAY span all three buildings — housekeepers wheel carts
    # between towers. (Full Clean keeps the stricter B2<->B3 block; DS does not.)
    return True

def unit_ok_fc(unit):
    # A same-guest cluster is only "OK to keep as one chart" if it can actually
    # fit one chart under ALL hard rules. If it can't, pack_rooms splits it into
    # individual rooms so they can be spread across housekeepers.
    if sum(r["time"] for r in unit) > MAX_FC: # over the 380 cap -> must split
        return False
    blds = set(r["bld"] for r in unit)
    ba = list(blds)
    for i in range(len(ba)):
        for j in range(i+1,len(ba)):
            if (ba[i]==2 and ba[j]==3) or (ba[i]==3 and ba[j]==2): return False
    n140 = sum(1 for r in unit if r["time"]==140)
    n120 = sum(1 for r in unit if r["time"]==120)
    if n140 > 1: return False
    # Same rule as can_add_fc: a 140 may coexist with at most ONE 120.
    if n140 >= 1 and n120 > 1: return False
    return True

def mk(unit, svc):
    return {"rooms":list(unit),"time":sum(r["time"] for r in unit),
            "blds":set(r["bld"] for r in unit),"floors":set(r.get("floor",0) for r in unit),
            "c140":sum(1 for r in unit if r["time"]==140),
            "c120":sum(1 for r in unit if r["time"]==120),
            "service_type":svc}

def absorb(g, unit):
    for r in unit:
        g["rooms"].append(r); g["blds"].add(r["bld"]); g["floors"].add(r.get("floor",0))
    g["time"] += sum(r["time"] for r in unit)
    g["c140"] += sum(1 for r in unit if r["time"]==140)
    g["c120"] += sum(1 for r in unit if r["time"]==120)

def _mix_penalty(g, unit):
    """Steer Full-Clean charts toward the preferred shapes:
        120+120+120 (360)
        140+120+70 (330)
        120+120+70+70(380)
        70x5 (350)
    A chart that already matches (or is a clean prefix of) one of these gets no
    penalty; the further its room-count of each size is from the nearest target
    shape, the higher the penalty. This is a soft preference used to break ties
    in placement, never a hard rule — capacity/guest/building rules still win.
    """
    times = [r["time"] for r in g["rooms"]] + [r["time"] for r in unit]
    if not times: return 0
    n140 = sum(1 for t in times if t == 140)
    n120 = sum(1 for t in times if t == 120)
    n70 = sum(1 for t in times if t <= 70) # 70s and anything smaller

    # (n140, n120, n70) signatures of the four preferred end-states
    TARGETS = [
        (0,3,0), # 120+120+120
        (1,1,1), # 140+120+70
        (0,2,2), # 120+120+70+70
        (0,0,5), # 70 x5
    ]
    # Distance to the closest target = sum of size-count mismatches. A chart
    # that is "on the way"to a target (fewer rooms than the target, matching
    # composition so far) is treated as a clean prefix low penalty.
    best = 99
    for t140,t120,t70 in TARGETS:
        # Over-shooting a size beyond its target is worse than under-shooting.
        over = max(0,n140-t140)+max(0,n120-t120)+max(0,n70-t70)
        under = max(0,t140-n140)+max(0,t120-n120)+max(0,t70-n70)
        best = min(best, over*3 + under) # overshoot weighted heavier
    return best * 25 # scale into "minutes-equivalent"tie-shaping range

def best_fit_generic(groups, unit, can_add_fn, same_bld_only, same_floor_only):
    ub = set(r["bld"] for r in unit)
    uf = set(r.get("floor",0) for r in unit)
    u_t = sum(r["time"] for r in unit)
    cap = MAX_DS if (unit and unit[0].get("service")==SVC_DS) else MAX_FC
    bi, best_score = -1, float("inf")
    for i, g in enumerate(groups):
        if not can_add_fn(g, unit): continue
        if same_bld_only and not same_bld(g["blds"], ub): continue
        if same_floor_only and not (g["floors"] & uf): continue
        prx = proximity_score(g["rooms"], unit)
        rem = cap - (g["time"] + u_t)
        # mix penalty only matters for Full Clean (DS rooms are all light)
        mix = _mix_penalty(g, unit) if (unit and unit[0].get("service")!=SVC_DS) else 0
        score = prx * 10000 + rem + mix
        if score < best_score: best_score, bi = score, i
    return bi

def _fc_feasible(units_list):
    """True if a set of same-guest units can share one Full-Clean chart under
    all hard rules: 380 cap, no B2+B3 together (B1 is the bridge), at most one
    140, and a 140 may coexist with at most one 120."""
    t = sum(u["time"] for u in units_list)
    if t > MAX_FC: return False
    blds = set()
    for u in units_list: blds |= u["blds"]
    if 2 in blds and 3 in blds: return False
    n140 = sum(u["n140"] for u in units_list)
    n120 = sum(u["n120"] for u in units_list)
    if n140 > 1: return False
    if n140 >= 1 and n120 > 1: return False
    return True

def _fc_spread(units_list):
    """Travel cost of one chart: building span (heaviest), floor span within each
    building, and number of distinct floor-stops. Lower = tighter / less walking."""
    if not units_list: return 0
    blds = set(); floors_by_bld = {}
    for u in units_list:
        blds |= u["blds"]
        for r in u["rooms"]:
            b = r.get("bld", 0)
            floors_by_bld.setdefault(b, set()).add(r.get("floor", 0))
    fspread = sum(max(fs) - min(fs) for fs in floors_by_bld.values())
    nstops = sum(len(fs) for fs in floors_by_bld.values())
    return (len(blds) - 1) * 100 + fspread * 10 + nstops * 3

def _fc_travel(charts):
    return sum(_fc_spread(c) for c in charts if c)

def _fc_optimize(units, seed=20240601, restarts=14):
    """Two-phase optimizer for Full Clean.
    Phase 1 — minimize the number of housekeepers: best-fit from several seed
      orders + random restarts, each refined by a move/swap local search that
      escapes greedy local optima (this is what matches the manual HK count).
    Phase 2 — with the housekeeper count fixed, reduce room travel: only moves/
      swaps that lower total building+floor spread WITHOUT increasing the chart
      count are accepted, so charts become tight (same building / nearby floors)
      like the manual schedule. Deterministic for a given input (fixed seed)."""
    import random as _rnd
    rng = _rnd.Random(seed)

    def pack_bestfit(order):
        charts = []
        for u in order:
            best, best_rem = None, 10**9
            for ch in charts:
                if _fc_feasible(ch + [u]):
                    rem = MAX_FC - (sum(x["time"] for x in ch) + u["time"])
                    if rem < best_rem: best_rem, best = rem, ch
            if best is None: charts.append([u])
            else: best.append(u)
        return charts

    # ── Phase 1: fewest charts ───────────────────────────────────────────────
    def _spans_two_bld(chart):
        return len(set(b for u in chart for b in u["blds"])) > 1
    def ls_mincount(charts, iters):
        charts = [c[:] for c in charts if c]
        for _ in range(iters):
            ne = [c for c in charts if c]
            if len(ne) < 2: break
            ne.sort(key=lambda c: sum(u["time"] for u in c))
            src = ne[0] if rng.random() < 0.5 else rng.choice(ne)
            if not src: continue
            u = rng.choice(src); order = ne[:]; rng.shuffle(order); moved = False
            for dst in order:
                if dst is src: continue
                if _fc_feasible(dst + [u]):
                    src.remove(u); dst.append(u); moved = True; break
            if not moved:
                for dst in order:
                    if dst is src: continue
                    for v in dst:
                        if _fc_feasible([x for x in src if x is not u] + [v]) and \
                           _fc_feasible([x for x in dst if x is not v] + [u]):
                            src.remove(u); dst.remove(v); src.append(v); dst.append(u)
                            moved = True; break
                    if moved: break
            charts = [c for c in charts if c]
        return [c for c in charts if c]

    orders = [
        sorted(units, key=lambda u: -u["time"]),
        sorted(units, key=lambda u: (-u["n140"], -u["n120"], -u["time"])),
        sorted(units, key=lambda u: ({3:0,1:1,2:2}.get(min(u["blds"]) if u["blds"] else 1, 9), -u["time"])),
    ]
    candidates = [pack_bestfit(o) for o in orders]
    for _ in range(restarts):
        us = units[:]; rng.shuffle(us)
        candidates.append(pack_bestfit(us))

    best_charts, best_n = None, None
    for c in candidates:
        refined = ls_mincount(c, 2500)
        n = len([ch for ch in refined if ch])
        if best_n is None or n < best_n:
            best_n, best_charts = n, refined
    charts = [c for c in best_charts if c]

    # ── Phase 1.5: un-cross cross-building charts (same count) ────────────────
    # Phase 1 minimizes housekeepers but may leave a chart spanning two buildings
    # (a B1 unit packed into a B2 chart, etc.). Removing such a crossing usually
    # needs a rotation: move the foreign unit into a same-building chart and push
    # that chart's displaced unit onward. We try 2-way swaps first, then 3-way
    # rotations, accepting only changes that reduce the number of crossings while
    # keeping the chart count fixed. Deterministic.
    def _nblds(chart):
        return len(set(b for u in chart for b in u["blds"]))
    def _decross(charts):
        charts = [c[:] for c in charts if c]
        for _ in range(40):
            crossed = [c for c in charts if _nblds(c) > 1]
            if not crossed: break
            improved = False
            for a in crossed:
                # foreign = the units belonging to a's minority building
                cnt = {}
                for u in a:
                    for bd in u["blds"]: cnt[bd] = cnt.get(bd, 0) + 1
                minority = min(set().union(*[u["blds"] for u in a]), key=lambda bd: cnt[bd])
                foreign = [u for u in a if minority in u["blds"]]
                for fu in foreign:
                    # 2-way: swap fu with a unit bu from another chart b
                    for b in charts:
                        if b is a: continue
                        for bu in b:
                            na = [x for x in a if x is not fu] + [bu]
                            nb = [x for x in b if x is not bu] + [fu]
                            if not (_fc_feasible(na) and _fc_feasible(nb)): continue
                            if (_nblds(na) + _nblds(nb)) < (_nblds(a) + _nblds(b)):
                                a.remove(fu); a.append(bu); b.remove(bu); b.append(fu)
                                improved = True; break
                        if improved: break
                    if improved: break
                    # 3-way rotation: a gives fu to b, b gives bu to c, c gives cu to a
                    for b in charts:
                        if b is a: continue
                        if not _fc_feasible(b + [fu]): continue
                        b2 = b + [fu]
                        for bu in b:
                            for c in charts:
                                if c is a or c is b: continue
                                if not _fc_feasible(c + [bu]): continue
                                c2 = c + [bu]
                                for cu in c:
                                    na = [x for x in a if x is not fu] + [cu]
                                    nb = [x for x in b2 if x is not bu]
                                    nc = [x for x in c2 if x is not cu]
                                    if not (_fc_feasible(na) and _fc_feasible(nb) and _fc_feasible(nc)):
                                        continue
                                    before = _nblds(a)+_nblds(b)+_nblds(c)
                                    after = _nblds(na)+_nblds(nb)+_nblds(nc)
                                    if after < before:
                                        a.remove(fu); a.append(cu)
                                        b.remove(bu); b.append(fu)
                                        c.remove(cu); c.append(bu)
                                        improved = True; break
                                if improved: break
                            if improved: break
                        if improved: break
                    if improved: break
                if improved: break
            if not improved: break
        return [c for c in charts if c]
    charts = [c for c in charts if c]

    # ── Phase 2: same chart count, minimize travel ──────────────────────────
    def ls_travel(charts, iters):
        charts = [c[:] for c in charts if c]
        for _ in range(iters):
            ne = [c for c in charts if c]
            if len(ne) < 2: break
            a, b = rng.sample(ne, 2)
            # move a unit a->b if it lowers combined spread and doesn't empty a
            if len(a) > 1:
                u = rng.choice(a)
                if _fc_feasible(b + [u]):
                    a2 = [x for x in a if x is not u]
                    if a2 and _fc_spread(a2) + _fc_spread(b + [u]) < _fc_spread(a) + _fc_spread(b):
                        a.remove(u); b.append(u); continue
            # swap u<->v if it lowers combined spread
            u = rng.choice(a); v = rng.choice(b)
            na = [x for x in a if x is not u] + [v]
            nb = [x for x in b if x is not v] + [u]
            if _fc_feasible(na) and _fc_feasible(nb) and \
               _fc_spread(na) + _fc_spread(nb) < _fc_spread(a) + _fc_spread(b):
                a.remove(u); a.append(v); b.remove(v); b.append(u)
        return [c for c in charts if c]

    charts = _decross(charts) # remove gratuitous building-crossings
    charts = ls_travel(charts, 15000)
    return [c for c in charts if c]


def pack_rooms(room_list, svc, can_add_fn, unit_ok_fn):
    if not room_list: return []
    import re as _re2
    for r in room_list:
        r["guest"] = _re2.sub(r'\s+', ' ', r.get("guest","").strip())
    gmap = {}
    for r in room_list: gmap.setdefault(r["guest"],[]).append(r)
    seen, unit_lists = set(), []
    for r in room_list:
        if r["guest"] not in seen:
            seen.add(r["guest"]); unit_lists.append(gmap[r["guest"]])

    # ── DAILY SERVICE (and any non-FC): keep the simple best-fit fill ────────
    # DS has no heavy-room / building constraints and a different cap, so the
    # FC optimizer doesn't apply. Behavior here is unchanged.
    if svc != SVC_FC:
        unit_lists.sort(key=lambda u:(
            u[0].get("bld",0), u[0].get("floor",0),
            -sum(r["time"] for r in u), -len(u),
        ))
        groups = []
        for unit in unit_lists:
            i=best_fit_generic(groups,unit,can_add_fn,True,True)
            if i==-1: i=best_fit_generic(groups,unit,can_add_fn,True,False)
            if i==-1: i=best_fit_generic(groups,unit,can_add_fn,False,False)
            if i>=0: absorb(groups[i],unit)
            else: groups.append(mk(unit,svc))
        return [g for g in groups if g["rooms"]]

    # ── FULL CLEAN: global bin-packing optimizer ────────────────────────────
    # A same-guest cluster that itself breaks the rules (too big for one chart,
    # spans B2+B3, too many heavy rooms) can't sit in one chart. Rather than
    # atomizing it into single rooms — which lets the optimizer scatter a guest's
    # rooms across many housekeepers — we split it into the FEWEST legal
    # sub-clusters, keeping as many of the guest's rooms together as the rules
    # allow (same building, <=380, <=one 140, a 140 with <=one 120). This keeps
    # e.g. a guest's two same-floor rooms on one cart instead of two.
    def _split_into_legal_subclusters(unit):
        # Sort rooms so heavy rooms seed sub-clusters and same building/floor
        # rooms stay adjacent, then greedily fill sub-clusters under the rules.
        rooms_sorted = sorted(unit, key=lambda r: (r.get("bld",0), r.get("floor",0), -r["time"]))
        subs = [] # each: {"rooms":[...],"time","blds","n140","n120"}
        for r in rooms_sorted:
            placed = False
            for s in subs:
                nb = s["blds"] | {r["bld"]}
                if 2 in nb and 3 in nb: continue
                n140 = s["n140"] + (1 if r["time"]==140 else 0)
                n120 = s["n120"] + (1 if r["time"]==120 else 0)
                if s["time"] + r["time"] > MAX_FC: continue
                if n140 > 1: continue
                if n140 >= 1 and n120 > 1: continue
                s["rooms"].append(r); s["time"] += r["time"]
                s["blds"] = nb; s["n140"] = n140; s["n120"] = n120
                placed = True; break
            if not placed:
                subs.append({"rooms":[r],"time":r["time"],"blds":{r["bld"]},
                             "n140":1 if r["time"]==140 else 0,
                             "n120":1 if r["time"]==120 else 0})
        return [s["rooms"] for s in subs]

    placeable = [] # each is a list-of-rooms (a locked cluster OR a legal sub-cluster)
    for unit in unit_lists:
        if unit_ok_fn(unit):
            placeable.append(unit)
        else:
            placeable.extend(_split_into_legal_subclusters(unit))

    # Wrap each placeable cluster with packing metadata.
    units = []
    for rooms in placeable:
        units.append({
            "rooms": rooms,
            "time": sum(r["time"] for r in rooms),
            "blds": set(r["bld"] for r in rooms),
            "n140": sum(1 for r in rooms if r["time"]==140),
            "n120": sum(1 for r in rooms if r["time"]==120),
        })

    charts = _fc_optimize(units)

    # Materialize charts into the standard group dict via mk().
    groups = []
    for chart in charts:
        chart_rooms = [r for u in chart for r in u["rooms"]]
        if chart_rooms: groups.append(mk(chart_rooms, svc))
    return groups

def _bld(r): return r.get("bld", 0)
def _flr(r): return r.get("floor", 0)
def _rnum(r): return r.get("num", 0)

def _norm_guest(r):
    g = re.sub(r'\s+', ' ', str(r.get("guest","")).strip())
    return g if g.lower() not in {"","unallocated","---","deposit, deposit",
                                  "room, walk","p/u models"} else ""

def _cluster_adjacent_same_guest(rooms):
    """HARD same-guest adjacency rule: rooms that share the SAME building, floor
    AND room-number and belong to the same (real) guest are the same physical
    unit and must never be split across housekeepers. Returns a list of "bundles"
    (each a list of room dicts) — a bundle of >1 is locked together; singles pass
    through as one-room bundles. Bundles are only kept together when they can
    still form a legal chart (<=380, <=one 140, 140+<=one 120); if a guest's
    adjacent rooms are collectively too big for one chart, they're split minimally
    so each piece is chart-legal (a genuine no-option case)."""
    from collections import defaultdict
    by_key = defaultdict(list)
    singles = []
    for r in rooms:
        g = _norm_guest(r)
        if g:
            # key on same physical room: building + floor + room number + guest
            by_key[(_bld(r), _flr(r), _rnum(r), g)].append(r)
        else:
            singles.append([r])
    bundles = []
    for key, grp in by_key.items():
        if len(grp) == 1:
            bundles.append(grp); continue
        # keep the adjacent same-guest rooms together, but if they can't legally
        # share one chart, split into the fewest legal sub-bundles.
        if _chart_feasible(grp):
            bundles.append(grp)
        else:
            cur = []
            for r in sorted(grp, key=lambda x: -x["time"]):
                if cur and _chart_feasible(cur + [r]):
                    cur.append(r)
                elif cur:
                    bundles.append(cur); cur = [r]
                else:
                    cur = [r]
            if cur: bundles.append(cur)
    return bundles + singles

def _unit_times(chart):
    """Expand any composite bundle units into their real per-room times, so the
    140/120 rules are checked against ACTUAL rooms, not a bundle's summed time."""
    times = []
    for r in chart:
        members = r.get("_members")
        if members:
            times.extend(m["time"] for m in members)
        else:
            times.append(r["time"])
    return times

def _chart_feasible(chart, single_building=False):
    """Hard rules for a Full Clean / IH chart (list of room dicts or bundle units)."""
    times = _unit_times(chart)
    if sum(times) > MAX_FC: return False
    if times.count(140) > 1: return False # never two 140s
    if times.count(140) >= 1 and times.count(120) > 1: return False
    blds = set(_bld(r) for r in chart)
    if 2 in blds and 3 in blds: return False # B2 and B3 never share
    if single_building and len(blds) > 1: return False
    return True

def _chart_nbld(chart): return len(set(_bld(r) for r in chart))
def _chart_nflr(chart): return len(set((_bld(r), _flr(r)) for r in chart))
def _chart_floor_gap(chart):
    """Largest vertical gap between floors used within a single building. 0 for a
    single-floor chart, 1 for adjacent floors (3+4), 2+ for distant floors (0+3).
    Distant-floor charts are what we want to avoid."""
    by_bld = {}
    for r in chart:
        by_bld.setdefault(_bld(r), set()).add(_flr(r))
    gap = 0
    for fs in by_bld.values():
        if len(fs) > 1:
            gap = max(gap, max(fs) - min(fs))
    return gap
def _chart_spread(chart):
    if not chart: return 0
    fl = [_flr(r) for r in chart]; nm = [_rnum(r) for r in chart]
    # Travel cost, in priority order:
    # • crossing buildings is worst,
    # • then how many DISTINCT floors the housekeeper must visit,
    # • then — critically — how FAR APART those floors are: a chart that must
    # span two floors should use ADJACENT floors (3+4), never distant ones
    # (0+3). The gap term grows fast so the solver pairs neighbouring floors.
    # • then room-number distance along the hall.
    n_floors = _chart_nflr(chart)
    gap = _chart_floor_gap(chart)
    return ((_chart_nbld(chart)-1)*400
            + (n_floors-1)*120
            + (gap*gap)*40
            + (max(fl)-min(fl))*10
            + (max(nm)-min(nm))*2)
def _charts_xb(charts): return sum(1 for c in charts if _chart_nbld(c) > 1)
def _charts_spread(charts): return sum(_chart_spread(c) for c in charts if c)

def _fc_bestfit(order):
    """Best-fit placement that prefers, in order: not adding a new building, not
    adding a new floor (keep the housekeeper on one floor), then tightest fit."""
    charts = []
    for r in order:
        best = None; bkey = None
        for c in charts:
            if _chart_feasible(c + [r]):
                rem = MAX_FC - (sum(x["time"] for x in c) + r["time"])
                intro_bld = _chart_nbld(c + [r]) > _chart_nbld(c)
                intro_flr = _chart_nflr(c + [r]) > _chart_nflr(c)
                key = (1 if intro_bld else 0, 1 if intro_flr else 0, rem)
                if bkey is None or key < bkey: bkey = key; best = c
        if best is None: charts.append([r])
        else: best.append(r)
    return charts

def _fc_ls_count(charts, iters, rng, avoid_xb=True):
    """Local search to reduce housekeeper count (move/swap). When avoid_xb, don't
    create a new building-crossing unless the move frees a whole housekeeper."""
    charts = [c[:] for c in charts if c]
    for _ in range(iters):
        ne = [c for c in charts if c]
        if len(ne) < 2: break
        ne.sort(key=lambda c: sum(x["time"] for x in c))
        src = ne[0] if rng.random() < 0.5 else rng.choice(ne)
        if not src: continue
        u = rng.choice(src); order = ne[:]; rng.shuffle(order); moved = False
        for dst in order:
            if dst is src: continue
            if _chart_feasible(dst + [u]):
                frees = (len(src) == 1)
                makes = avoid_xb and _chart_nbld(dst + [u]) > _chart_nbld(dst)
                if frees or not makes:
                    src.remove(u); dst.append(u); moved = True; break
        if not moved:
            for dst in order:
                if dst is src: continue
                for v in dst:
                    if _chart_feasible([x for x in src if x is not u] + [v]) and \
                       _chart_feasible([x for x in dst if x is not v] + [u]):
                        ns = [x for x in src if x is not u] + [v]
                        nd = [x for x in dst if x is not v] + [u]
                        if avoid_xb and ((_chart_nbld(ns) > _chart_nbld(src)) or
                                         (_chart_nbld(nd) > _chart_nbld(dst))): continue
                        src.remove(u); dst.remove(v); src.append(v); dst.append(u)
                        moved = True; break
                if moved: break
        charts = [c for c in charts if c]
    return [c for c in charts if c]

def _fc_ls_tidy(charts, iters, rng):
    """Within a fixed housekeeper count, move/swap to reduce travel (spread)."""
    charts = [c[:] for c in charts if c]
    for _ in range(iters):
        ne = [c for c in charts if c]
        if len(ne) < 2: break
        a, b = rng.sample(ne, 2)
        if len(a) > 1:
            u = rng.choice(a)
            if _chart_feasible(b + [u]):
                a2 = [x for x in a if x is not u]
                if a2 and _chart_spread(a2) + _chart_spread(b + [u]) < _chart_spread(a) + _chart_spread(b):
                    a.remove(u); b.append(u); continue
        u = rng.choice(a); v = rng.choice(b)
        na = [x for x in a if x is not u] + [v]; nb = [x for x in b if x is not v] + [u]
        if _chart_feasible(na) and _chart_feasible(nb) and \
           _chart_spread(na) + _chart_spread(nb) < _chart_spread(a) + _chart_spread(b):
            a.remove(u); a.append(v); b.remove(v); b.append(u)
    return [c for c in charts if c]

def _fc_floor_consolidate(charts, rounds=25):
    """Deterministic best-improvement pass dedicated to reducing floor travel:
    first the number of distinct floors a housekeeper visits, and — when a chart
    must span floors — keeping those floors ADJACENT (small gap) rather than
    distant. Never changes the housekeeper count or breaks a hard rule."""
    charts = [c[:] for c in charts if c]
    def cost(c):
        # floor-count dominates; floor-gap is the secondary term.
        return _chart_nflr(c) * 100 + _chart_floor_gap(c) * 10
    for _ in range(rounds):
        best_gain = 0; best_op = None
        n = len(charts)
        for ai in range(n):
            a = charts[ai]
            if not a: continue
            for bi in range(n):
                if bi == ai: continue
                b = charts[bi]
                if not b: continue
                base = cost(a) + cost(b)
                for u in a:
                    if len(a) == 1: break
                    if _chart_feasible(b + [u]):
                        na = [x for x in a if x is not u]; nb = b + [u]
                        gain = base - (cost(na) + cost(nb))
                        if gain > best_gain:
                            best_gain = gain; best_op = ("move", ai, bi, u, None)
                for u in a:
                    for v in b:
                        na = [x for x in a if x is not u] + [v]
                        nb = [x for x in b if x is not v] + [u]
                        if _chart_feasible(na) and _chart_feasible(nb):
                            gain = base - (cost(na) + cost(nb))
                            if gain > best_gain:
                                best_gain = gain; best_op = ("swap", ai, bi, u, v)
        if best_op is None: break
        kind, ai, bi, u, v = best_op
        if kind == "move":
            charts[ai].remove(u); charts[bi].append(u)
        else:
            charts[ai].remove(u); charts[ai].append(v)
            charts[bi].remove(v); charts[bi].append(u)
        charts = [c for c in charts if c]
    return [c for c in charts if c]

def _bundle_to_unit(bundle):
    """Collapse a locked bundle of adjacent same-guest rooms into one composite
    'unit' the packer treats atomically. Carries the sub-rooms in _members."""
    if len(bundle) == 1:
        u = dict(bundle[0]); u["_members"] = bundle; return u
    return {
        "room": bundle[0]["room"],
        "time": sum(r["time"] for r in bundle),
        "bld": bundle[0].get("bld", 0),
        "floor": bundle[0].get("floor", 0),
        "num": bundle[0].get("num", 0),
        "guest": bundle[0].get("guest", ""),
        "_members": bundle,
    }

def _unpack_units(charts):
    """Expand composite units back into their real room dicts."""
    out = []
    for c in charts:
        rooms = []
        for u in c:
            rooms.extend(u.get("_members", [u]))
        out.append(rooms)
    return out

def solve_full_clean(fc_rooms, seed=20240601, restarts=16):
    """Tidy-first, min-count Full Clean solver. Priority: (1) fewest housekeepers,
    (2) tidiest charts (single-building, contiguous) within that count. Adjacent
    same-guest rooms (same building+floor+room#) are locked together as one unit
    and never split. Returns a list of charts (each a list of room dicts).
    Deterministic (fixed seed)."""
    if not fc_rooms: return []
    import random as _rnd
    rng = _rnd.Random(seed)
    # HARD adjacency rule: bundle adjacent same-guest rooms, then pack bundles.
    bundles = _cluster_adjacent_same_guest(fc_rooms)
    units = [_bundle_to_unit(b) for b in bundles]
    BLD = {3:0, 1:1, 2:2}
    orders = [
        sorted(units, key=lambda r: -r["time"]),
        sorted(units, key=lambda r: (-(r["time"]==140), -(r["time"]==120), -r["time"])),
        sorted(units, key=lambda r: (BLD.get(_bld(r),1), _flr(r), _rnum(r))),
        # Strict floor-sequential: fill one (building, floor) completely, heavy
        # units first within the floor, before moving to the next floor — this is
        # how the manual builds charts, and it maximizes single-floor housekeepers.
        sorted(units, key=lambda r: (BLD.get(_bld(r),1), _flr(r), -r["time"], _rnum(r))),
    ]
    cands = [_fc_bestfit(o) for o in orders]
    for _ in range(restarts):
        rr = list(units); rng.shuffle(rr); cands.append(_fc_bestfit(rr))
    # Phase 1: building-aware count minimization
    best = None
    for c in cands:
        r = _fc_ls_count(c, 1500, rng, avoid_xb=True)
        nb = len([x for x in r if x])
        key = (nb, _charts_xb(r), _charts_spread(r))
        if best is None or key < best[0]: best = (key, r)
    nb0 = best[0][0]
    # Phase 2: if crossings were forbidden at a cost, retry allowing them to see
    # whether a lower count is reachable (manual bridges B1 only when it saves a HK)
    for c in cands:
        r = _fc_ls_count(c, 1500, rng, avoid_xb=False)
        nb = len([x for x in r if x])
        if nb < nb0:
            key = (nb, _charts_xb(r), _charts_spread(r))
            if key < best[0]: best = (key, r); nb0 = nb
    charts = [c[:] for c in best[1] if c]
    # Fill charts toward the 330-380 band FIRST so nobody is left light, then let
    # the floor-tidy passes run to restore the single-floor layout...
    charts = _fc_greedy_finish(charts)
    charts = _fc_ls_tidy(charts, 12000, rng)
    charts = _fc_floor_consolidate(charts)     # dedicated floor-reduction pass
    charts = _fc_ls_tidy(charts, 4000, rng)    # final hall-tightening polish
    # ...then a final top-up: the tidy passes above optimize only floor spread and
    # can pull a chart back under 330. This pass lifts any under-330 chart back up
    # toward the cap by pulling in rooms — preferring same-floor rooms so tidiness
    # is preserved, and only crossing to another floor if that's the only way to
    # clear 330. Never exceeds the cap or breaks a rule.
    charts = _fc_topup_light(charts)
    # Finally, reduce the "easy" 120+70+70+70 = 330 charts: try to swap one of the
    # 70s for a bigger room (120/140) from a heavier chart so the chart becomes a
    # tougher combination, but only when a valid swap exists that keeps both charts
    # legal and neither drops below 330. Only fires when an alternative exists.
    charts = _fc_reduce_easy_330(charts)
    return _unpack_units(charts)

def _fc_reduce_easy_330(charts, rounds=120):
    """Discourage the exact 120+70+70+70 = 330 shape. For each such chart, try to
    swap one of its 70s with a 120/140 from another chart, provided both charts
    stay <=380 and >=LOW_MIN and no rule breaks. Keeps room counts and building
    coherence; only changes which rooms sit where. No-op when no valid swap."""
    charts = [c[:] for c in charts if c]
    if len(charts) < 2: return charts
    def _times(c):
        out = []
        for u in c:
            m = u.get("_members")
            out.extend(x["time"] for x in m) if m else out.append(u["time"])
        return out
    def _t(c): return sum(_times(c))
    def _is_easy(c): return sorted(_times(c)) == [70, 70, 70, 120]
    for _ in range(rounds):
        moved = False
        for i, ci in enumerate(charts):
            if not _is_easy(ci): continue
            # a 70-min unit in this chart to give away
            for a in ci:
                if _unit_time_sum([a]) != 70: continue
                # find a donor chart with a 120/140 unit we can take in exchange
                for j, cj in enumerate(charts):
                    if j == i: continue
                    for b in cj:
                        bt = _unit_time_sum([b])
                        if bt < 120: continue          # want a bigger room
                        new_ci = [x for x in ci if x is not a] + [b]
                        new_cj = [x for x in cj if x is not b] + [a]
                        if not (_chart_feasible(new_ci) and _chart_feasible(new_cj)): continue
                        # both must stay in the valid band after the swap
                        if not (LOW_MIN <= _t(new_ci) <= MAX_FC): continue
                        if not (LOW_MIN <= _t(new_cj) <= MAX_FC): continue
                        # and the target chart must no longer be the easy shape
                        if _is_easy(new_ci): continue
                        charts[i] = new_ci; charts[j] = new_cj; moved = True; break
                    if moved: break
                if moved: break
            if moved: break
        if not moved: break
    return charts

def _fc_topup_light(charts, rounds=200):
    """Lift under-LOW_MIN charts toward the cap, and — higher value — try to
    dissolve a light chart entirely by redistributing its rooms into others, which
    both removes a light chart and frees a housekeeper.

    Two strategies each round:
      A. DISSOLVE: if the lightest chart's rooms can all be absorbed by other
         charts (within cap + rules), do it — one fewer housekeeper, no light chart.
      B. TOP-UP: otherwise pull a room from a donor into the lightest chart to lift
         it toward LOW_MIN, preferring same-floor donors and never dropping a donor
         that's at/above LOW_MIN below it."""
    charts = [c[:] for c in charts if c]
    if len(charts) < 2: return charts
    def _tt(u):
        m = u.get("_members")
        return sum(x["time"] for x in m) if m else u["time"]
    def t(c): return sum(_tt(u) for u in c)
    for _ in range(rounds):
        order = sorted(range(len(charts)), key=lambda i: t(charts[i]))
        light_i = order[0]
        if t(charts[light_i]) >= LOW_MIN: break

        # ── Strategy A: can we place EVERY unit of the light chart elsewhere? ──
        others = [i for i in range(len(charts)) if i != light_i]
        trial = {i: list(charts[i]) for i in others}
        ok = True
        for u in charts[light_i]:
            placed = False
            ub, uf = _bld(u), _flr(u)
            # Rank candidate charts so the unit lands close by: same floor first,
            # then same building, then (as a last resort) fullest chart. This keeps
            # a dissolved single-room housekeeper's rooms with a neighbour rather
            # than scattering them across buildings.
            def _rank(k):
                c = trial[k]
                same_floor = any((_bld(x), _flr(x)) == (ub, uf) for x in c)
                same_bld   = any(_bld(x) == ub for x in c)
                return (0 if same_floor else 1, 0 if same_bld else 1, -t(c))
            for i in sorted(others, key=_rank):
                if _chart_feasible(trial[i] + [u]):
                    trial[i].append(u); placed = True; break
            if not placed: ok = False; break
        if ok:
            for i in others: charts[i] = trial[i]
            charts[light_i] = []
            charts = [c for c in charts if c]
            continue

        # ── Strategy B: pull one donor room into the light chart ──────────────
        cur = t(charts[light_i]); gap = MAX_FC - cur
        cur_flrs = set((_bld(x), _flr(x)) for x in charts[light_i])
        best = None
        for di in range(len(charts)):
            if di == light_i or len(charts[di]) <= 1: continue
            dt = t(charts[di])
            for u in charts[di]:
                ut = _tt(u)
                if ut > gap: continue
                if dt - ut < LOW_MIN and dt >= LOW_MIN: continue
                if not _chart_feasible(charts[light_i] + [u]): continue
                same_floor = 1 if (_bld(u), _flr(u)) in cur_flrs else 0
                cand = (same_floor, ut)
                if best is None or cand > best[0]:
                    best = (cand, di, u)
        if best is None: break
        _, di, u = best
        charts[di].remove(u); charts[light_i].append(u)
        charts = [c for c in charts if c]
    return charts

def _fc_greedy_finish(charts):
    """Re-pack units so each chart is FILLED before the next is opened, keeping the
    same units (same-guest bundles intact), never breaking a rule, never adding a
    housekeeper.

    Filling to the cap depends on which unit seeds each chart, so for each chart we
    try a handful of promising seeds, fill the rest best-fit, and keep the seed
    that fills the chart most. This reaches the cap when a full combination exists
    (e.g. 190+120+70=380) instead of stalling on a big seed."""
    units = [u for c in charts for u in c]
    n_target = len([c for c in charts if c])
    # Process in building/floor/room order so consecutive charts follow the floors
    # (keeps a housekeeper's chart on one floor when the fill allows it).
    BLD = {3:0, 1:1, 2:2}
    pool = sorted(units, key=lambda u: (BLD.get(_bld(u),1), _flr(u), _rnum(u)))

    def _fill_from(seed_i, avail):
        chart = [avail[seed_i]]; used = {seed_i}
        while True:
            cur = _unit_time_sum(chart)
            gap = MAX_FC - cur
            if gap <= 0: break
            cur_flrs = set((_bld(x), _flr(x)) for x in chart)
            # Only step onto a NEW floor if this chart is still under the 330 floor
            # (i.e. it genuinely needs more work to avoid being light). Once at/above
            # 330 we stop reaching to other floors — keeping travel minimal.
            allow_new_floor = cur < LOW_MIN
            best_j = None; best_key = None
            for j, u in enumerate(avail):
                if j in used: continue
                ut = _unit_time_sum([u])
                if ut > gap or not _chart_feasible(chart + [u]): continue
                same_floor = (_bld(u), _flr(u)) in cur_flrs
                if not same_floor and not allow_new_floor:
                    continue                      # don't wander off-floor once full enough
                # Primary: stay on the current floor. Secondary: for off-floor
                # options, pick the physically nearest floor (small vertical gap).
                # Tertiary: bigger fill. This keeps each housekeeper on as few
                # floors as possible while still reaching the target band.
                if same_floor:
                    floor_pen = 0
                else:
                    # distance to the nearest floor already in the chart. Crossing
                    # into a DIFFERENT building is treated as much more expensive
                    # than moving floors within the same building, so the fill only
                    # crosses buildings when there is genuinely no same-building
                    # room left to reach LOW_MIN — this stops a lone B1 room from
                    # being stranded onto an otherwise-B3 chart (and vice-versa).
                    floor_pen = min(abs(_flr(u)-_flr(x)) + (0 if _bld(u)==_bld(x) else 100)
                                    for x in chart)
                key = (0 if same_floor else 1, floor_pen, -ut)   # lower is better
                if best_key is None or key < best_key:
                    best_key = key; best_j = j
            if best_j is None: break
            used.add(best_j); chart.append(avail[best_j])
        return used, _unit_time_sum(chart)

    result = []
    while pool and len(result) < n_target:
        # Try a few seeds; keep the packing that best (1) clears the 330 floor,
        # (2) uses the FEWEST distinct floors (minimal travel), then (3) fills most.
        best = None
        n_try = min(len(pool), 6)
        for i in range(n_try):
            used, total = _fill_from(i, pool)
            u_chart = [pool[k] for k in used]
            nflr = len(set((_bld(x), _flr(x)) for x in u_chart))
            rank = (1 if total >= LOW_MIN else 0, -nflr, total)
            if best is None or rank > best[0]:
                best = (rank, used, total)
                if total >= MAX_FC and nflr == 1: break
        used = best[1]
        chart = [pool[i] for i in sorted(used)]
        pool = [u for i, u in enumerate(pool) if i not in used]
        result.append(chart)

    # Leftover units → the lightest chart that can LEGALLY take them (feasibility
    # enforced, so the cap is never exceeded). If none can, open nothing new —
    # append to the lightest chart only if it stays within the cap; otherwise keep
    # the unit as its own chart (should not happen given n_target math).
    for u in pool:
        placed = False
        for c in sorted(result, key=_unit_time_sum):
            if _chart_feasible(c + [u]): c.append(u); placed = True; break
        if not placed:
            result.append([u])
    return [c for c in result if c]

def _unit_time_sum(chart):
    total = 0
    for u in chart:
        m = u.get("_members")
        total += sum(x["time"] for x in m) if m else u["time"]
    return total

def solve_ih(ih_rooms, seed=20240601):
    """Pack IH rooms into SINGLE-building charts <=380, keeping a guest's rooms
    together where possible. Returns (kept_charts, leftover_rooms). Charts totalling
    >= IH_KEEP_MIN are kept (inspected by RQS 2); genuine scraps spill to Daily
    Service."""
    if not ih_rooms: return [], []
    BLD = {3:0, 1:1, 2:2}
    pool = sorted(ih_rooms, key=lambda r: (BLD.get(_bld(r),1), r.get("guest",""), _flr(r), _rnum(r)))
    used = [False]*len(pool); charts = []
    for i in range(len(pool)):
        if used[i]: continue
        chart = [pool[i]]; used[i] = True
        while sum(r["time"] for r in chart) < MAX_FC:
            best = None
            for j in range(len(pool)):
                if used[j]: continue
                if not _chart_feasible(chart + [pool[j]], single_building=True): continue
                rj = pool[j]
                gj = rj.get("guest","")
                same_g = 0 if any(r.get("guest","")==gj and gj not in ("","Unallocated")
                                  for r in chart) else 1
                d = min(abs(_flr(r)-_flr(rj))*30 + abs(_rnum(r)-_rnum(rj)) for r in chart)
                key = (same_g, d)
                if best is None or key < best[0]: best = (key, j)
            if best is None: break
            used[best[1]] = True; chart.append(pool[best[1]])
        charts.append(chart)
    keep = []; leftover = []
    for c in charts:
        if sum(r["time"] for r in c) >= IH_KEEP_MIN: keep.append(c)
        else: leftover.extend(c)
    return keep, leftover

def split_daily_service(ds_rooms, extra_rooms=None, cap=DS_CAP):
    """Split Daily Service rooms into charts, packing each housekeeper as FULL as
    possible (up to the hard cap) before opening a new one, so we use the fewest
    housekeepers. Rooms are kept physically close (building/floor/room order) and
    no chart ever exceeds the cap.

    Approach:
      1. Sort rooms by building/floor/room so each chart stays contiguous.
      2. Greedily fill the current chart to the cap; only open a new chart when the
         next room genuinely won't fit.
      3. Run a tightening pass that pulls early rooms forward to top up
         under-filled charts, so the last chart carries the small remainder rather
         than leaving several half-full charts."""
    import math
    rooms = list(ds_rooms) + list(extra_rooms or [])
    if not rooms: return []
    BLD = {3:0, 1:1, 2:2}
    rooms = sorted(rooms, key=lambda r: (BLD.get(_bld(r),1), _flr(r), _rnum(r)))
    total = sum(r["time"] for r in rooms)
    # Fewest housekeepers this could possibly need at the hard cap.
    n_min = max(1, math.ceil(total / cap))

    # Greedy fill-to-cap: keep loading the current chart until the next room won't
    # fit, then start a new one. This packs each housekeeper tight.
    charts = []; cur = []; cur_t = 0
    for r in rooms:
        if cur and cur_t + r["time"] > cap:
            charts.append(cur); cur = []; cur_t = 0
        cur.append(r); cur_t += r["time"]
    if cur: charts.append(cur)

    # Tightening pass: if we ended up with more charts than the theoretical
    # minimum, try to top up earlier charts by pulling the first room of a later
    # chart forward whenever it fits under the cap. This concentrates the load and
    # can drop the trailing (nearly empty) chart entirely, cutting a housekeeper.
    _ds_pack_tight(charts, cap, n_min)
    return charts

def _ds_pack_tight(charts, cap, n_min, rounds=400):
    """Move boundary rooms EARLIER to fill charts to the cap, so the fewest
    housekeepers are used. Only the first room of a later chart moves to the end of
    an earlier one (keeps contiguity), and only when it fits under the cap."""
    def t(c): return sum(x["time"] for x in c)
    for _ in range(rounds):
        moved = False
        # try to top up each chart from the one after it
        for i in range(len(charts)-1):
            a = charts[i]
            if not a: continue
            ta = t(a)
            # pull rooms from later charts forward while they fit
            for j in range(i+1, len(charts)):
                b = charts[j]
                while b and ta + b[0]["time"] <= cap:
                    r = b.pop(0); a.append(r); ta += r["time"]; moved = True
                if ta >= cap: break
        # drop any charts we emptied
        before = len(charts)
        charts[:] = [c for c in charts if c]
        if len(charts) != before: moved = True
        if not moved: break

def build_all_groups(rooms, priority_hks=None):
    verify_rooms = [r for r in rooms if r.get("verify")]
    rooms = [r for r in rooms if not r.get("verify")]

    fc_rooms = [r for r in rooms if r.get("service")==SVC_FC]
    ih_rooms = [r for r in rooms if r.get("service")==SVC_IH]
    ds_rooms = [r for r in rooms if r.get("service")==SVC_DS]
    dv_rooms = [r for r in rooms if r.get("service")==SVC_DV]

    # ── Stage 1: regular Full Clean — tidy-first, minimum housekeepers ────────
    # (priority_hks preserved: any rooms pre-claimed by a named housekeeper are
    # packed first via the legacy path, the rest go through the new solver.)
    priority_groups = []
    remaining_fc = list(fc_rooms)
    priority_hks = priority_hks or []
    if priority_hks:
        try: roster = st.session_state.get("hk_roster", {})
        except: roster = {}
        for hk_name in priority_hks:
            home_bld = roster.get(hk_name, {}).get("building", 0)
            pool = sorted(remaining_fc, key=lambda r:(
                0 if r.get("bld",0)==home_bld else 1,
                r.get("floor",0), -r.get("time",0)
            ))
            state = {"rooms":[], "time":0, "c140":0, "c120":0, "used":set()}
            def can_add_p(rooms_to_add, s=state):
                t = sum(r["time"] for r in rooms_to_add)
                if s["time"]+t > MAX_FC: return False
                has140 = any(r["time"]==140 for r in rooms_to_add)
                nc140 = s["c140"]+(1 if has140 else 0)
                nc120 = s["c120"]+sum(1 for r in rooms_to_add if r["time"]==120)
                if nc140 > 1: return False
                if nc140>=1 and nc120>=1 and any(r["time"]>70 for r in rooms_to_add): return False
                cur_blds = set(r.get("bld",0) for r in s["rooms"])
                new_blds = cur_blds | set(r.get("bld",0) for r in rooms_to_add)
                if 2 in new_blds and 3 in new_blds: return False
                return True
            def add_p(rooms_to_add, s=state):
                for r in rooms_to_add:
                    s["rooms"].append(r); s["used"].add(id(r))
                    s["time"] += r["time"]
                    s["c140"] += 1 if r["time"]==140 else 0
                    s["c120"] += 1 if r["time"]==120 else 0
            import re as _re3
            guest_map = {}
            for r in pool:
                g = _re3.sub(r'\s+', ' ', r.get("guest","").strip())
                if g.lower() not in {"","unallocated","---","deposit, deposit","room, walk","p/u models"}:
                    guest_map.setdefault(g,[]).append(r)
            seen2 = set()
            for r in pool:
                if id(r) in state["used"]: continue
                g = _re3.sub(r'\s+', ' ', r.get("guest","").strip())
                if g in seen2: continue
                seen2.add(g)
                unit = guest_map.get(g,[r])
                if any(id(u) in state["used"] for u in unit): continue
                if can_add_p(unit): add_p(unit)
                if state["time"] >= MAX_FC: break
            if state["time"] < MAX_FC:
                for r in sorted(pool, key=lambda r:-r["time"]):
                    if id(r) in state["used"]: continue
                    if can_add_p([r]): add_p([r])
                    if state["time"] >= MAX_FC: break
            if state["time"] < 380:
                gap = MAX_FC - state["time"]
                fill_pool = sorted([r for r in remaining_fc if id(r) not in state["used"]],
                                   key=lambda r:abs(r["time"]-gap))
                for r in fill_pool:
                    if id(r) in state["used"]: continue
                    if can_add_p([r]): add_p([r])
                    if state["time"] >= MAX_FC: break
            if state["time"] < 330:
                for r in sorted(remaining_fc, key=lambda r:-r["time"]):
                    if id(r) in state["used"]: continue
                    if can_add_p([r]): add_p([r])
                    if state["time"] >= MAX_FC: break
            if state["rooms"]:
                grp = mk(state["rooms"], SVC_FC)
                grp["priority_hk"] = hk_name
                priority_groups.append(grp)
                used_ids = {id(r) for r in state["rooms"]}
                remaining_fc = [r for r in remaining_fc if id(r) not in used_ids]

    fc_charts = solve_full_clean(remaining_fc)
    fc_groups_normal = [mk(c, SVC_FC) for c in fc_charts]
    fc_groups = priority_groups + fc_groups_normal

    # ── Stage 2: Full Clean (IH) — packed separately, inspected by RQS 2 ──────
    ih_leftover = []
    ih_groups = []
    if ih_rooms:
        ih_charts, ih_leftover = solve_ih(ih_rooms)
        for c in ih_charts:
            g = mk(c, SVC_IH)
            g["ih_group"] = True
            g["rqs"] = IH_RQS # RQS 2 inspects all IH charts
            ih_groups.append(g)

    # ── Stage 3: Daily Service — HARD 460-min cap per chart, +leftover IH ──────
    if ds_rooms or ih_leftover:
        ds_charts = split_daily_service(ds_rooms, extra_rooms=ih_leftover)
        ds_groups = [mk(c, SVC_DS) for c in ds_charts]
        for g in ds_groups:
            g["ds_overflow"] = g["time"] > DS_CAP # shouldn't happen with hard cap
    else:
        ds_groups = []

    if dv_rooms:
        # All Dust n Vac goes to RQS 2 (alongside Daily Service + IH inspection),
        # not the Manager.
        dv_groups = [{"rooms":list(dv_rooms),"time":0,"blds":set(r["bld"] for r in dv_rooms),
                      "floors":set(r.get("floor",0) for r in dv_rooms),"c140":0,"c120":0,
                      "service_type":SVC_DV,"dv_rqs2":True}]
    else:
        dv_groups = []
    # Verify group(s): stayover / P-U Models — no HK, no RQS, flagged for review.
    if verify_rooms:
        verify_groups = [{
            "rooms":list(verify_rooms),
            "time":sum(r.get("time",0) for r in verify_rooms),
            "blds":set(r["bld"] for r in verify_rooms),
            "floors":set(r.get("floor",0) for r in verify_rooms),
            "c140":0,"c120":0,"service_type":SVC_FC,
            "verify_group":True,
        }]
    else:
        verify_groups = []
    # Sequencing: Full Clean (regular + IH) first, then Daily Service, then DV.
    return fc_groups + ih_groups + ds_groups + dv_groups + verify_groups

# ══════════════════════════════════════════════════════════════════════════════
# STAFF ASSIGNMENT
# ══════════════════════════════════════════════════════════════════════════════
def assign_hk_building_aware(groups, present_hk, roster, ds_team=None):
    # Housekeepers dedicated to Daily Service today. They are RESERVED: only DS
    # charts draw from them, and they are held out of the general pool so no
    # Full Clean / IH / Dust n Vac chart can take them. Empty => old behaviour.
    ds_set = {n for n in (ds_team or []) if n in present_hk}

    pool = {1:[], 2:[], 3:[]}; ds_pool = {1:[], 2:[], 3:[]}
    for n in present_hk:
        b = roster.get(n,{}).get("building",0)
        if b not in pool: continue
        (ds_pool if n in ds_set else pool)[b].append(n)
    available    = {b: list(v) for b, v in pool.items()}
    ds_available = {b: list(v) for b, v in ds_pool.items()}
    assignment = {}; used = set()
    def hk_can_take(hk_bld, group_blds, is_ds=False):
        # Daily Service: housekeepers may wheel carts across all buildings,
        # so the B2<->B3 block does NOT apply. Full Clean keeps the block.
        if is_ds: return True
        for gb in group_blds:
            if hk_bld==2 and gb==3: return False
            if hk_bld==3 and gb==2: return False
        return True
    # Which buildings a housekeeper from building X is allowed to cover (FC).
    # Movement rule: B1<->B2 ok, B1<->B3 ok, B2<->B3 blocked.
    ADJ = {1:[1,2,3], 2:[2,1], 3:[3,1]}
    def find_hk(group_blds, is_ds=False, avail=None):
        """Pull the best legal housekeeper out of `avail` (default: the general
        pool). Returns None when that pool cannot cover the chart."""
        if avail is None: avail = available
        # A Full Clean group spanning BOTH B2 and B3 is structurally impossible
        # for one housekeeper (B2<->B3 is blocked) -- never assign it.
        if not is_ds and (2 in group_blds and 3 in group_blds):
            return None
        primary = min(group_blds) if group_blds else 1
        # 1) Try a housekeeper whose home building IS the group's primary building.
        for hk in list(avail.get(primary,[])):
            if hk_can_take(roster.get(hk,{}).get("building",0), group_blds, is_ds):
                avail[primary].remove(hk); return hk
        # 2) Borrow from an ADJACENT allowed building (or any, for DS).
        adj_order = [1,2,3] if is_ds else ADJ.get(primary, [1,2,3])
        for b in adj_order:
            for hk in list(avail.get(b,[])):
                if hk_can_take(roster.get(hk,{}).get("building",0), group_blds, is_ds):
                    avail[b].remove(hk); return hk
        # 3) Any remaining present housekeeper who can legally take the group.
        for b in [1,2,3]:
            for hk in list(avail.get(b,[])):
                if hk_can_take(roster.get(hk,{}).get("building",0), group_blds, is_ds):
                    avail[b].remove(hk); return hk
        # 3.5) SHORTAGE fallback. When home/adjacent housekeepers are exhausted,
        # a housekeeper may be moved cross-building to cover a group, as long
        # as the GROUP itself spans only an allowed combination -- B1&2 or B1&3
        # (or a single building), NEVER B2&3. This relaxes the per-housekeeper
        # home-building rule on busy days while keeping the hard B2<->B3 block.
        group_spans_b2b3 = (2 in group_blds and 3 in group_blds)
        if not group_spans_b2b3:
            for b in [1,2,3]:
                if avail.get(b):
                    return avail[b].pop(0)
        # 4) Truly exhausted -- caller substitutes a placeholder.
        return None
    def drop(name):
        """Remove a named housekeeper from whichever pool still holds them."""
        for av in (available, ds_available):
            for b in [1,2,3]:
                if name in av.get(b,[]):
                    av[b].remove(name); return
    # Priority HKs first
    for g in groups:
        if g.get("verify_group"): # never assign verify groups
            assignment[g["label"]] = ""; continue
        phk = g.get("priority_hk","")
        if not phk: continue
        if g.get("dv_rqs2"): assignment[g["label"]]=""; continue # DV -> RQS2 inspector
        drop(phk)
        assignment[g["label"]] = phk; used.add(phk)
    # Daily Service goes ONLY to the dedicated team when one is selected. Once
    # the team is exhausted the remaining DS charts get a numbered placeholder
    # (Need Housekeeper 1, 2, ...) instead of a name, so the gap is visible.
    if ds_set:
        need_n = 0
        for g in groups:
            if g["label"] in assignment: continue
            if g.get("service_type") != SVC_DS: continue
            if g.get("dv_rqs2"): assignment[g["label"]]=""; continue
            matched = find_hk(g.get("blds",{1}), True, ds_available)
            if matched:
                assignment[g["label"]] = matched; used.add(matched)
            else:
                need_n += 1
                assignment[g["label"]] = f"{NEED_HK_PREFIX} {need_n}"
    # Everyone else — assign the most CONSTRAINED groups first so they get the
    # scarce building-specific housekeepers. Order: Full Clean (strict building
    # rule) Dust n Vac Daily Service (flexible, can take any HK).
    def _order(g):
        st_ = g.get("service_type","")
        return {SVC_FC:0, SVC_IH:1, SVC_DV:2, SVC_DS:3}.get(st_, 4)
    for g in sorted(groups, key=_order):
        if g["label"] in assignment: continue
        if g.get("verify_group"):
            assignment[g["label"]] = ""; continue
        # Dust n Vac is carried by RQS 2 (inspector), not a housekeeper — leave the
        # housekeeper field blank here; assign_inspectors puts it on RQS 2.
        if g.get("dv_rqs2"): assignment[g["label"]]=""; continue
        is_ds = (g.get("service_type") == SVC_DS)
        matched = find_hk(g.get("blds",{1}), is_ds) or NO_HK_LABEL
        assignment[g["label"]] = matched
        if not is_unassigned_hk(matched): used.add(matched)
    return assignment, used

def _primary_bld(g): return min(g["blds"]) if g["blds"] else 0
def _group_complexity(g): return sum(r.get("time",70)/70 for r in g.get("rooms",[]))
def _batch_complexity(batch): return sum(_group_complexity(g) for g in batch)
def _insp_travel_score(batch):
    # Travel cost = buildings visited + floors visited within those buildings,
    # plus the vertical spread (how many floors apart the highest and lowest
    # rooms are). Keeping an inspector on one floor — or a tight floor band —
    # scores lowest.
    blds=set(); cross=0
    floor_keys=set() # distinct (building, floor) stops
    floors_by_bld={}
    for g in batch:
        blds |= g["blds"]
        if len(g["blds"])>1: cross += len(g["blds"])-1
        for b in g["blds"]:
            for f in (g.get("floors") or {0}):
                floor_keys.add((b,f))
                floors_by_bld.setdefault(b,set()).add(f)
    # Vertical spread within each building (max floor - min floor)
    spread=0
    for b,fs in floors_by_bld.items():
        if fs: spread += (max(fs)-min(fs))
    # Non-sequential penalty: an inspector should cover a CONTIGUOUS band of
    # floors (e.g. 2-3-4), not hop over gaps (e.g. 2 then 5). For each building we
    # count the missing floors between the lowest and highest floor the inspector
    # visits — every gap floor is extra elevator/stair travel — and penalize it
    # heavily so the optimizer keeps each inspector on sequential floors.
    gap=0
    for b,fs in floors_by_bld.items():
        if len(fs) > 1:
            lo, hi = min(fs), max(fs)
            gap += (hi - lo + 1) - len(fs)   # count of skipped floors in the band
    # Horizontal spread: how far apart the room numbers are within each building
    # (walking down a long hall). Grouping physically close rooms cuts RQS travel.
    nums_by_bld={}
    for g in batch:
        for b in g["blds"]:
            for r in g.get("rooms",[]):
                if r.get("bld")==b:
                    nums_by_bld.setdefault(b,[]).append(r.get("num",0))
    hspread=0
    for b,ns in nums_by_bld.items():
        if ns: hspread += (max(ns)-min(ns))
    # Weights: building hops are most expensive, then number of distinct
    # floor-stops, then how far apart those floors are. A batch that spans all
    # THREE buildings is very hard to inspect, so it gets a steep extra penalty
    # on top of the linear building cost — we want the optimizer to treat a
    # 3-building inspector as nearly forbidden.
    nbld = len(blds)
    three_bld_penalty = 400 if nbld >= 3 else 0
    return (nbld*12 + three_bld_penalty + cross*4 + len(floor_keys)*3
            + spread*2 + gap*8 + hspread)
def _batch_heavy(batch):
    """Count of big rooms (120/140 min) across a batch — used to spread the
    hard-to-inspect rooms evenly across inspectors."""
    return sum(1 for g in batch for r in g.get("rooms",[]) if r.get("time",70) >= 120)
def _is_heavy_chart(g):
    """A housekeeper's chart is 'heavy' (hard to inspect) when it's loaded with
    big rooms — e.g. 120+120+120 or 140+120+70. Defined as 2+ big rooms AND a
    total at/above 330 min."""
    big = sum(1 for r in g.get("rooms",[]) if r.get("time",70) >= 120)
    return big >= 2 and g.get("time",0) >= 330
def _batch_heavy_charts(batch):
    return sum(1 for g in batch if _is_heavy_chart(g))
def _insp_combined_score(batches):
    tt = sum(_insp_travel_score(b) for b in batches)
    nonempty = [b for b in batches if b]
    if len(nonempty) < 2: return tt
    cx = [_batch_complexity(b) for b in nonempty]
    hv = [_batch_heavy(b) for b in nonempty]
    hc = [_batch_heavy_charts(b) for b in nonempty]
    # Worst single inspector's travel: minimizing the SUM alone can still leave
    # one inspector hauling all over the property while the rest are tight. Adding
    # the maximum individual travel makes the optimizer actively pull work off the
    # worst-travelling RQS, evening out travel rather than just totalling it.
    travels = [_insp_travel_score(b) for b in nonempty]
    worst_travel = max(travels) if travels else 0
    # Count inspectors forced across all 3 buildings — this is the worst case
    # for an inspector and must be avoided even at the cost of heavy-work
    # balance. Each such batch adds an overriding penalty that dwarfs the
    # balance terms, so the optimizer will almost always restructure to remove
    # a 3-building inspector before worrying about heavy-chart fairness.
    three_bld_batches = sum(1 for b in nonempty
                            if len(set().union(*[g["blds"] for g in b])) >= 3)
    THREE_BLD = three_bld_batches * 10000
    # An inspector must never be sent between Building 2 and Building 3 — they are
    # at opposite ends with Building 1 as the only bridge, so covering both means
    # walking the entire property. Penalize any batch spanning B2 and B3 as hard
    # as a 3-building batch.
    b2b3_batches = sum(1 for b in nonempty
                       if {2,3}.issubset(set().union(*[g["blds"] for g in b])))
    B2B3 = b2b3_batches * 10000
    # Balance terms, in priority order (below the building overrides):
    # 1) heavy-CHART spread — no inspector gets all the 120+120+120 charts.
    # 2) heavy-ROOM spread — even count of individual 120/140 rooms.
    # 3) travel (building + floor aware).
    # 4) overall complexity spread — tie-breaker.
    return (THREE_BLD + B2B3
            + (max(hc)-min(hc))*14
            + (max(hv)-min(hv))*8
            + tt*2
            + worst_travel*3
            + (max(cx)-min(cx)))

def assign_inspectors(groups, present_insp, per, rqs1, rqs2):
    # Verify groups (stayover / P-U Models) never get an inspector.
    verify_groups = [g for g in groups if g.get("verify_group")]
    for g in verify_groups: g["inspector"] = ""
    groups = [g for g in groups if not g.get("verify_group")]
    fc_groups = [g for g in groups if g.get("service_type")==SVC_FC]
    ih_groups = [g for g in groups if g.get("service_type")==SVC_IH]
    ds_groups = [g for g in groups if g.get("service_type")==SVC_DS]
    dv_groups = [g for g in groups if g.get("service_type")==SVC_DV]
    inspectors=[]; assigned_names=set()

    # IH groups are always inspected by RQS 2 (regardless of FC/DS load).
    for g in ih_groups:
        g["inspector"] = rqs2 or IH_RQS

    def units(grp_list): # total rooms across a list of groups
        return sum(len(g["rooms"]) for g in grp_list)

    # ── Dedicated FC inspectors = present inspectors who are NOT rqs1/rqs2 ────
    # These are used FIRST for Full Clean. RQS1/RQS2 only step in for FC when
    # the dedicated inspectors can't cover everything.
    fc_inspectors = [n for n in present_insp if n not in (rqs1, rqs2)]

    # ── Assign dedicated FC inspectors to FC groups first ────────────────────
    fc_sorted=sorted(fc_groups, key=lambda g:(
        _primary_bld(g),
        min(g.get("floors",{0})) if g.get("floors") else 0,
        min(r.get("num",0) for r in g["rooms"]) if g["rooms"] else 0
    ))
    # How many FC charts can the dedicated inspectors cover? With tight room-based
    # packing an inspector holds up to ~13 rooms (often 4-5 small charts), so the
    # dedicated pool can absorb more charts than n_dedicated*per. Estimate capacity
    # by rooms and only push the genuine overflow to RQS1/RQS2.
    n_dedicated = len(fc_inspectors)
    coverable = n_dedicated * max(per, 5)  # generous chart budget; room cap binds
    dedicated_fc = fc_sorted[:coverable]
    leftover_fc = fc_sorted[coverable:] # only these need RQS1/RQS2 help

    # Pack each inspector as FULL as possible (up to a room ceiling) before opening
    # the next, so we use the fewest inspectors and each gets as many complete
    # charts as will fit — rather than spreading charts thinly across everyone.
    # The ROOM ceiling is the real limit (an inspector can carry ~12-13 rooms);
    # we allow more than `per` charts as long as rooms stay under the ceiling, so
    # inspectors fill up to capacity instead of stopping at 3 thin charts.
    INSP_ROOM_MAX = 13          # an inspector can carry up to ~12-13 rooms
    INSP_CHART_MAX = max(per, 5)  # allow filling toward the room ceiling
    def _grooms(g): return len(g["rooms"])
    batches = []
    cur = []; cur_rooms = 0
    for g in dedicated_fc:
        gr = _grooms(g)
        if cur and (cur_rooms + gr > INSP_ROOM_MAX or len(cur) >= INSP_CHART_MAX):
            batches.append(cur); cur = []; cur_rooms = 0
        cur.append(g); cur_rooms += gr
    if cur: batches.append(cur)
    # Balance batches to reduce cross-building and floor travel. Run several full
    # passes; each pass keeps swapping charts between batches while any swap lowers
    # the combined score (which now also penalizes the single worst-travelling
    # inspector), so we don't settle for a solution where one RQS roams. Swaps are
    # rejected if they would push an inspector over the room ceiling.
    def _brooms(b): return sum(len(g["rooms"]) for g in b)
    def _within_cap(b): return _brooms(b) <= INSP_ROOM_MAX
    def _optimize_batches(batches):
        improved=True; max_iter=(len(batches)*per*4 if batches else 0); iters=0
        while improved and iters<max_iter:
            improved=False; iters+=1
            for bi in range(len(batches)):
                for bj in range(bi+1,len(batches)):
                    for gi,ga in enumerate(batches[bi]):
                        for gj,gb in enumerate(batches[bj]):
                            new_bi=batches[bi][:gi]+[gb]+batches[bi][gi+1:]
                            new_bj=batches[bj][:gj]+[ga]+batches[bj][gj+1:]
                            # don't let a travel-swap push an inspector over the ceiling
                            if not (_within_cap(new_bi) and _within_cap(new_bj)): continue
                            if _insp_combined_score([new_bi,new_bj])<_insp_combined_score([batches[bi],batches[bj]]):
                                batches[bi],batches[bj]=new_bi,new_bj; improved=True; break
                        if improved: break
                    if improved: break
        return batches
    batches=_optimize_batches(batches)
    # Extra pass: explicitly target the worst-travelling batch and try to hand its
    # farthest-out chart to a batch that can absorb it more cheaply.
    for _ in range(len(batches)):
        if len(batches)<2: break
        tscore=[(_insp_travel_score(b),i) for i,b in enumerate(batches) if b]
        if not tscore: break
        _, wi = max(tscore)
        moved=False
        for gi,ga in enumerate(batches[wi]):
            for bj in range(len(batches)):
                if bj==wi or not batches[bj]: continue
                for gj,gb in enumerate(batches[bj]):
                    new_wi=batches[wi][:gi]+[gb]+batches[wi][gi+1:]
                    new_bj=batches[bj][:gj]+[ga]+batches[bj][gj+1:]
                    if not (_within_cap(new_wi) and _within_cap(new_bj)): continue
                    if _insp_combined_score([new_wi,new_bj])<_insp_combined_score([batches[wi],batches[bj]]):
                        batches[wi],batches[bj]=new_wi,new_bj; moved=True; break
                if moved: break
            if moved: break
        if not moved: break
        batches=_optimize_batches(batches)
    fc_inspectors_q = list(fc_inspectors)
    for batch in batches:
        if not batch: continue
        name=fc_inspectors_q.pop(0) if fc_inspectors_q else f"Inspector {len(inspectors)+1}"
        blds=sorted(set(b for g in batch for b in g["blds"]))
        cx=_batch_complexity(batch)
        entry={"id":len(inspectors)+1,"name":name,"role":"FC",
               "groups":[g["label"] for g in batch],"buildings":blds,
               "travel_warning":len(blds)>2,"heavy_warning":cx>15,"complexity":round(cx,1)}
        for g in batch: g["inspector"]=name
        inspectors.append(entry); assigned_names.add(name)

    # ── RQS role assignment (per operations policy) ──────────────────────────
    # • RQS 2 carries Daily Service + ALL Dust n Vac, and inspects IH.
    # • RQS 1 is NOT auto-assigned any rooms — left free for manual assignment
    # later based on how the day goes.
    # • Leftover Full Clean that dedicated inspectors can't cover opens extra
    # inspector slots (never dumped on RQS 1).
    fc_shortage = len(leftover_fc) > 0

    # RQS 2 shares leftover FC only if there genuinely aren't enough inspectors,
    # but its primary load is DS + DV (+ IH already assigned above).
    rqs2_fc_groups = []
    if fc_shortage and rqs2:
        budget = 6
        for g in sorted(leftover_fc, key=lambda g:(len(g["rooms"]), _primary_bld(g))):
            if len(g["rooms"]) <= budget:
                rqs2_fc_groups.append(g); budget -= len(g["rooms"])
            if budget <= 0: break
        leftover_fc = [g for g in leftover_fc if g not in rqs2_fc_groups]

    if (ds_groups or dv_groups or rqs2_fc_groups) and rqs2:
        r2_groups = ds_groups + dv_groups + rqs2_fc_groups
        blds=sorted(set(b for g in r2_groups for b in g["blds"]))
        entry={"id":len(inspectors)+1,"name":rqs2,"role":"RQS2",
               "groups":[g["label"] for g in r2_groups],"buildings":blds}
        for g in r2_groups: g["inspector"]=rqs2
        inspectors.append(entry); assigned_names.add(rqs2)

    # RQS 1: intentionally left with NO auto-assigned rooms.

    # Any leftover FC still unassigned extra inspector slots (not RQS 1).
    while leftover_fc:
        batch = leftover_fc[:per]; leftover_fc = leftover_fc[per:]
        name = f"Inspector {len(inspectors)+1}"
        blds=sorted(set(b for g in batch for b in g["blds"]))
        cx=_batch_complexity(batch)
        entry={"id":len(inspectors)+1,"name":name,"role":"FC",
               "groups":[g["label"] for g in batch],"buildings":blds,
               "travel_warning":len(blds)>2,"heavy_warning":cx>15,"complexity":round(cx,1)}
        for g in batch: g["inspector"]=name
        inspectors.append(entry)

    # Safety net: if RQS 2 wasn't present, any DV without an inspector falls back
    # to RQS 2's label so it's never silently dropped.
    for g in dv_groups:
        if not g.get("inspector"): g["inspector"] = rqs2 or IH_RQS
    return inspectors

# ══════════════════════════════════════════════════════════════════════════════
# HTML BUILDERS
# ══════════════════════════════════════════════════════════════════════════════
def _rgba_a(rgba: str, a) -> str:
    """Return the same rgba color with a different alpha. Safe on any input."""
    m = re.match(r"rgba\((\d+),\s*(\d+),\s*(\d+)", str(rgba))
    if not m: return rgba
    return f"rgba({m.group(1)},{m.group(2)},{m.group(3)},{a})"

def group_card_html(g, idx):
    svc = g.get("service_type", SVC_FC)
    cap = MAX_DS if svc==SVC_DS else MAX_FC
    pct = min(int(g["time"]/max(cap,1)*100), 100)
    is_verify = g.get("verify_group", False)

    # Color per service type. In the formal light theme these are muted,
    # professional tones with soft (near-flat) accent bars and no neon glow.
    if _IS_LIGHT:
        SVC_COLORS = {
            SVC_FC: {"accent":"#2563a8","glow":"rgba(37,99,168,.14)","bar":"#2563a8","badge_bg":"rgba(37,99,168,.10)","badge_txt":"#1d4e86"},
            SVC_IH: {"accent":"#6d5bb5","glow":"rgba(109,91,181,.14)","bar":"#6d5bb5","badge_bg":"rgba(109,91,181,.10)","badge_txt":"#574699"},
            SVC_DS: {"accent":"#0f766e","glow":"rgba(15,118,110,.14)","bar":"#0f766e","badge_bg":"rgba(15,118,110,.10)","badge_txt":"#0c5d56"},
            SVC_DV: {"accent":"#b45309","glow":"rgba(180,83,9,.14)","bar":"#b45309","badge_bg":"rgba(180,83,9,.10)","badge_txt":"#8f420a"},
        }
    else:
        SVC_COLORS = {
            SVC_FC: {"accent":"#6366f1","glow":"rgba(99,102,241,.45)","bar":"linear-gradient(90deg,#6366f1,#818cf8)","badge_bg":"rgba(99,102,241,.18)","badge_txt":"#a5b4fc"},
            SVC_IH: {"accent":"#8b5cf6","glow":"rgba(139,92,246,.45)","bar":"linear-gradient(90deg,#8b5cf6,#a78bfa)","badge_bg":"rgba(139,92,246,.18)","badge_txt":"#c4b5fd"},
            SVC_DS: {"accent":"#14b8a6","glow":"rgba(20,184,166,.4)","bar":"linear-gradient(90deg,#14b8a6,#2dd4bf)","badge_bg":"rgba(20,184,166,.15)","badge_txt":"#5eead4"},
            SVC_DV: {"accent":"#f59e0b","glow":"rgba(245,158,11,.4)","bar":"linear-gradient(90deg,#f59e0b,#fbbf24)","badge_bg":"rgba(245,158,11,.15)","badge_txt":"#fcd34d"},
        }
    c = SVC_COLORS.get(svc, SVC_COLORS[SVC_FC])
    # Verify groups use a distinct warning palette (muted in light mode).
    if is_verify:
        if _IS_LIGHT:
            c = {"accent":"#be123c","glow":"rgba(190,18,60,.14)","bar":"#be123c",
                 "badge_bg":"rgba(190,18,60,.10)","badge_txt":"#9f1239"}
        else:
            c = {"accent":"#f43f5e","glow":"rgba(244,63,94,.45)",
                 "bar":"linear-gradient(90deg,#f43f5e,#fb7185)",
                 "badge_bg":"rgba(244,63,94,.18)","badge_txt":"#fda4af"}
    ac = c["accent"]; glow = c["glow"]; bar = c["bar"]

    hk_raw = g.get("housekeeper","") or ""
    need_hk = hk_raw.startswith(NEED_HK_PREFIX)
    no_hk = not hk_raw or hk_raw.startswith("No HK available")
    if is_verify:
        unassigned_badge = "" # verify groups intentionally have no HK badge
        hk_raw = ""
    elif need_hk:
        # Short-staffed Daily Service: keep the numbered placeholder as the name
        # so the manager can see exactly how many bodies are missing.
        unassigned_badge = f'<span style="background:rgba(245,158,11,.2);color:#fcd34d;border-radius:5px;padding:1px 8px;font-size:.66rem;font-weight:700;border:1px solid rgba(245,158,11,.4);letter-spacing:.03em"> NEED HK</span>'
    elif no_hk:
        unassigned_badge = f'<span style="background:rgba(244,63,94,.2);color:#fb7185;border-radius:5px;padding:1px 8px;font-size:.66rem;font-weight:700;border:1px solid rgba(244,63,94,.35);letter-spacing:.03em"> NO HK</span>'
        hk_raw = hk_raw.replace("No HK available","Unassigned") if hk_raw else "Unassigned"
    else:
        unassigned_badge = ""

    hk = e(hk_raw or "—")
    insp = e(g.get("inspector","") or "—")
    _blds_raw = g.get("blds", set())
    if isinstance(_blds_raw, str):
        _blds_raw = [int(x) for x in re.findall(r'\d+', _blds_raw)]
    bld_str = " · ".join(f"Bldg {b}" for b in sorted(set(_blds_raw)))

    def badge(txt, bg, clr, border="transparent"):
        return f'<span style="background:{bg};color:{clr};border:1px solid {border};border-radius:5px;padding:1px 8px;font-size:.66rem;font-weight:600;letter-spacing:.02em">{txt}</span>'

    if is_verify:
        svc_badge = badge("VERIFY & ASSIGN","rgba(244,63,94,.2)","#fda4af","rgba(244,63,94,.4)")
    else:
        svc_badge = badge(svc, c["badge_bg"], c["badge_txt"], _rgba_a(c["glow"], ".25"))
    overflow_badge = badge("DS Overflow","rgba(245,158,11,.15)","#fcd34d","rgba(245,158,11,.3)") if g.get("ds_overflow") else ""
    priority_badge = badge("Priority","rgba(234,179,8,.15)","#fde047","rgba(234,179,8,.3)") if g.get("priority_hk") else ""
    cross_badge = badge("Cross-bld","rgba(168,85,247,.15)","#d8b4fe","rgba(168,85,247,.3)") if (g.get("cross_bld") and not is_verify) else ""

    t_col = "#4ade80" if pct<=87 else ("#fbbf24" if pct<=95 else "#f87171")

    rows = ""
    for i, r in enumerate(g["rooms"]):
        notes_lower = r.get("notes","").lower()
        is_stayover = "stayover" in notes_lower or "stay over" in notes_lower
        row_bg = "rgba(34,211,238,.06)" if (r.get("uncertain") and is_stayover) else "rgba(245,158,11,.06)" if r.get("uncertain") else "transparent"
        pet_badge = '<span style="background:rgba(244,63,94,.15);color:#fb7185;border-radius:4px;padding:1px 6px;font-size:.64rem;font-weight:600"></span>' if r.get("pet") else ""
        late_co = e(r.get("late_checkout",""))
        late_badge2= f'<span style="background:rgba(245,158,11,.15);color:#fcd34d;border-radius:4px;padding:1px 6px;font-size:.64rem;font-weight:600"> {late_co}</span>' if late_co else ""
        delay = f"{i*0.04:.2f}s"
        rows += f"""<tr style="background:{row_bg};border-bottom:1px solid {_C["row_br"]};animation:rowIn .3s {delay} both">
          <td style="font-family:'DM Mono',monospace;font-size:.76rem;font-weight:500;color:{ac};padding:8px 10px;white-space:nowrap">{e(r.get("room",""))}</td>
          <td class="m-hide"style="padding:8px 10px;color:#64748b;font-size:.75rem">B{r.get("bld","")}</td>
          <td style="padding:8px 10px;color:{_C["txt"]};font-size:.78rem;font-weight:500;max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{e(r.get("guest",""))}</td>
          <td class="m-hide"style="padding:8px 10px;color:{_C["txt2"]};font-size:.73rem">{e(r.get("service",""))}</td>
          <td style="padding:8px 10px;font-family:'DM Mono',monospace;font-weight:500;color:{_C["txt"]};font-size:.76rem">{"—" if r.get("time",0)==0 else str(r.get("time",""))+"m"}</td>
          <td style="padding:8px 10px">{pet_badge}</td>
          <td style="padding:8px 10px">{late_badge2}</td>
        </tr>"""

    lbl = e(g.get("label",""))
    th = f"padding:6px 10px;text-align:left;font-family:'DM Mono',monospace;font-size:.6rem;font-weight:500;text-transform:uppercase;letter-spacing:.1em;color:{_C["txt3"]};background:{_C["th_bg"]};border-bottom:1px solid {_C["row_br"]}"

    # Verify groups: distinct title, pill text, hidden HK/RQS meta + time meter
    if is_verify:
        title_text = "Unsure — verify &amp; assign"
        pill_text = "VERIFY"
        header_bg = "linear-gradient(90deg,rgba(244,63,94,.1),transparent)"
        meta_line = (f'<div style="font-size:.71rem;color:{_C["txt2"]};margin-top:3px;'
                        f'font-family:\'DM Mono\',monospace">{bld_str} &nbsp;·&nbsp; '
                        f'<span style="color:#fda4af">no housekeeper / RQS — assign manually</span></div>')
        time_meter = ""
    else:
        title_text = f"Group {lbl}"
        pill_text = lbl
        header_bg = "linear-gradient(90deg,rgba(99,102,241,.06),transparent)"
        meta_line = (f'<div style="font-size:.71rem;color:{_C["txt2"]};margin-top:3px;font-family:\'DM Mono\',monospace">'
                        f'{bld_str} &nbsp;·&nbsp; <span style="color:{_C["txt3"]}"></span> {hk} &nbsp;·&nbsp; '
                        f'<span style="color:{_C["txt3"]}"></span> {insp}</div>')
        time_meter = (f'<div style="display:flex;align-items:center;gap:8px;flex-shrink:0">'
                        f'<span style="font-family:\'DM Mono\',monospace;font-size:.82rem;font-weight:500;color:{ac};'
                        f'text-shadow:0 0 8px {glow}">{g.get("time","")} <span style="color:#475569;font-size:.7rem">/ {cap}m</span></span>'
                        f'<div style="background:rgba(255,255,255,.06);border-radius:99px;height:5px;width:72px;overflow:hidden;border:1px solid rgba(255,255,255,.06)">'
                        f'<div style="background:{bar};width:{pct}%;height:5px;border-radius:99px;box-shadow:0 0 6px {glow};transition:width .4s"></div>'
                        f'</div></div>')

    return f"""<!DOCTYPE html><html><head>{SHARED_CSS}</head><body>
<div style="border-radius:12px;overflow:hidden;border:1px solid {glow};
            background:{_C["card_bg"]};
            backdrop-filter:blur(16px);margin-bottom:4px;
            box-shadow:{_C["card_sh"]};
            animation:glassIn .35s cubic-bezier(.16,1,.3,1) both">
  <!-- Card header -->
  <div style="padding:10px 14px;display:flex;align-items:center;gap:10px;flex-wrap:wrap;
              border-bottom:1px solid rgba(99,102,241,.1);
              background:{header_bg}">
    <!-- Label pill with glow -->
    <div style="background:{ac};color:#fff;border-radius:6px;padding:4px 10px;
                font-family:'Syne',sans-serif;font-weight:800;font-size:.78rem;
                white-space:nowrap;flex-shrink:0;letter-spacing:.04em;
                box-shadow:0 0 12px {glow},0 0 24px {_rgba_a(glow, ".18")}">
      {pill_text}
    </div>
    <!-- Title + badges -->
    <div style="flex:1;min-width:0">
      <div style="display:flex;align-items:center;gap:5px;flex-wrap:wrap">
        <span style="font-family:'Syne',sans-serif;font-weight:700;font-size:.88rem;color:{_C["txt"]}">{title_text}</span>
        {svc_badge} {overflow_badge} {priority_badge} {cross_badge} {unassigned_badge}
      </div>
      {meta_line}
    </div>
    <!-- Time meter -->
    {time_meter}
  </div>
  <!-- Room table -->
  <table>
    <thead><tr>
      <th style="{th}">Room</th><th class="m-hide"style="{th}">Bld</th><th style="{th}">Guest</th>
      <th class="m-hide"style="{th}">Service</th><th style="{th}">Time</th>
      <th style="{th}">Pet</th><th style="{th}">Late Out</th>
    </tr></thead>
    <tbody>{rows}</tbody>
  </table>
  <!-- Footer -->
  <div style="background:{_C["foot_bg"]};padding:6px 12px;display:flex;justify-content:space-between;
              border-top:1px solid {_C["row_br"]};font-family:'DM Mono',monospace;font-size:.68rem;color:{_C["txt3"]}">
    <span>{len(g["rooms"])} rooms &nbsp;·&nbsp; {g.get("c120",0)}×120 &nbsp;·&nbsp; {g.get("c140",0)}×140</span>
    <span style="color:{t_col};font-weight:500">{g.get("time","")}m used</span>
  </div>
</div></body></html>"""

def staff_table_html(rows, cols, cell_fns, row_bg_fn):
    th_s = ("padding:8px 12px;text-align:left;font-family:'DM Mono',monospace;"
            "font-size:.6rem;font-weight:500;text-transform:uppercase;letter-spacing:.1em;"
            f"color:{_C['txt3']};background:{_C['th_bg']};border-bottom:1px solid {_C['row_br']}")
    ths = "".join(f'<th style="{th_s}">{e(c)}</th>' for c in cols)
    body = ""
    for i, row in enumerate(rows):
        bg = row_bg_fn(row)
        # Map known status backgrounds to theme-appropriate tints
        if bg == "#f0fdf4": bg = "rgba(20,184,166,.07)" if _IS_LIGHT else "rgba(20,184,166,.08)"
        elif bg == "#fefce8": bg = "rgba(245,158,11,.06)"
        elif bg not in ("transparent","","#fff"): bg = "rgba(99,102,241,.05)"
        elif bg in ("#fff","") : bg = "transparent"
        delay = f"{i*0.03:.2f}s"
        tds = "".join(
            f'<td style="padding:8px 12px;border-bottom:1px solid {_C["row_br"]};'
            f'vertical-align:middle;background:{bg};animation:rowIn .3s {delay} both">{fn(row)}</td>'
            for fn in cell_fns)
        body += f"<tr>{tds}</tr>"
    return f"""<!DOCTYPE html><html><head>{SHARED_CSS}</head><body>
<div style="border-radius:12px;overflow:hidden;border:1px solid {_C['card_br']};
            background:{_C['tbl_bg']};backdrop-filter:blur(12px);
            box-shadow:{_C['card_sh']}">
<table style="width:100%;border-collapse:collapse;font-size:.8rem">
  <thead><tr>{ths}</tr></thead><tbody>{body}</tbody>
</table></div></body></html>"""

def insp_card_html(insp, fg, color):
    name = e(insp.get("name",""))
    role = insp.get("role","FC")
    blds = insp.get("buildings",[])
    BLD_NEON = {
        1: ("rgba(99,102,241,.2)","#a5b4fc","rgba(99,102,241,.4)"),
        2: ("rgba(20,184,166,.18)","#5eead4","rgba(20,184,166,.35)"),
        3: ("rgba(245,158,11,.18)","#fcd34d","rgba(245,158,11,.35)"),
    }
    bld_tags_parts = []
    for b in blds:
        bg3, txt3, bdr3 = BLD_NEON.get(b, ("rgba(99,99,99,.15)","#94a3b8","rgba(99,99,99,.3)"))
        bld_tags_parts.append(
            f'<span style="background:{bg3};color:{txt3};border-radius:5px;padding:2px 8px;'
            f'font-family:\'DM Mono\',monospace;font-size:.65rem;font-weight:500;margin-right:4px;'
            f'border:1px solid {bdr3}">Bldg {b}</span>'
        )
    bld_tags = "".join(bld_tags_parts)
    role_map = {
        "RQS1":("rgba(245,158,11,.18)","#fcd34d","RQS1 · DV"),
        "RQS2":("rgba(20,184,166,.18)","#5eead4","RQS2 · DS"),
        "FC": ("rgba(99,102,241,.18)","#a5b4fc","Full Clean"),
    }
    rbg,rtxt,rlbl = role_map.get(role,("rgba(99,99,99,.1)","#94a3b8",role))
    role_badge = (f'<span style="background:{rbg};color:{rtxt};border-radius:5px;'
                  f'padding:2px 8px;font-family:\'DM Mono\',monospace;font-size:.65rem;font-weight:500;margin-left:6px">{rlbl}</span>')
    heavy_warn = (f'<span style="background:rgba(244,63,94,.15);color:#fb7185;border-radius:5px;'
                  f'padding:2px 8px;font-size:.66rem;font-weight:600;margin-left:5px">'
                  f' Heavy {insp.get("complexity",0)}pts</span>') if insp.get("heavy_warning") else ""
    pills = ""; total_t = 0
    for gl in insp["groups"]:
        gobj = next((g for g in fg if g["label"]==gl), None)
        if not gobj: continue
        ac2 = "#6366f1" if gobj.get("service_type")==SVC_FC else ("#14b8a6" if gobj.get("service_type")==SVC_DS else "#f59e0b")
        hk = e(gobj.get("housekeeper","") or f"Grp {gl}")
        total_t += gobj.get("time",0)
        pills += (f'<span style="display:inline-block;background:rgba(99,102,241,.08);'
                  f'border:1px solid rgba(99,102,241,.25);border-radius:20px;'
                  f'padding:3px 10px;font-size:.73rem;margin:2px 2px;">'
                  f'<span style="font-family:\'DM Mono\',monospace;font-weight:500;color:{ac2};font-size:.72rem">{gl}</span>'
                  f' <span style="color:#64748b">·</span>'
                  f' <span style="color:{_C["txt"]}">{hk}</span></span>')
    return f"""<!DOCTYPE html><html><head>{SHARED_CSS}</head><body>
<div style="border-radius:12px;padding:14px 16px;
            border:1px solid {color}44;
            background:{_C['card_bg']};
            backdrop-filter:blur(16px);
            box-shadow:{_C['card_sh']},0 0 20px {color}18">
  <div style="display:flex;align-items:center;flex-wrap:wrap;gap:5px;margin-bottom:8px">
    <span style="font-family:'Syne',sans-serif;font-weight:700;font-size:.9rem;
                 color:{color}"> {name}</span>
    {role_badge}{heavy_warn}
  </div>
  <div style="margin-bottom:8px;display:flex;flex-wrap:wrap;gap:4px">{bld_tags}</div>
  <div style="line-height:1.8">{pills or f'<span style="color:{_C["txt3"]};font-family:\'DM Mono\',monospace;font-size:.77rem">— no groups —</span>'}</div>
  <div style="margin-top:8px;font-family:'DM Mono',monospace;font-size:.67rem;color:{_C['txt3']};
              border-top:1px solid {_C['row_br']};padding-top:6px">
    {len(insp["groups"])} groups &nbsp;·&nbsp; {total_t} min
  </div>
</div></body></html>"""

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    _cu = auth.current_user()
    _ac, _bg = auth.ROLE_COLORS.get(_cu["role"], ("#6366f1","rgba(99,102,241,.15)"))
    st.markdown(f"""
<div style="background:rgba(99,102,241,.1);border:1px solid rgba(99,102,241,.25);
            border-radius:10px;padding:10px 14px;display:flex;justify-content:space-between;
            align-items:center;margin-bottom:10px;
            box-shadow:0 0 16px rgba(99,102,241,.1)">
  <div>
    <div style="font-family:'Syne',sans-serif;font-weight:700;font-size:.85rem;
                color:#a5b4fc">{_cu["username"]}</div>
    <div style="font-family:'DM Mono',monospace;font-size:.65rem;color:#475569;
                text-transform:uppercase;letter-spacing:.08em">{_cu["role"].title()}</div>
  </div>
  <span style="font-size:1.3rem;opacity:.8">{"" if _cu["role"]=="admin" else "" if _cu["role"]=="rqs" else ""}</span>
</div>""", unsafe_allow_html=True)
    # Sign out lives in the top bar now, the same place as on every page.

    # ── Role-based navigation ──────────────────────────────────────────────
    # admin Schedule + Dashboard + Admin
    # rqs Schedule + Dashboard
    # hk nothing (just name/role/signout above)
    _role = _cu["role"]

    # ── Housekeepers: NO attendance/config sidebar ─────────────────────────
    # They only see their name, role, and sign-out above. Everything below
    # (attendance, roster, RQS roles, priority HKs) is admin/rqs only.
    if auth.is_housekeeper():
        present_hk = [n for n,v in st.session_state["hk_roster"].items() if v["present"]]
        present_insp = [n for n,v in st.session_state["insp_roster"].items() if v]
        rqs1 = st.session_state.get("rqs1","")
        rqs2 = st.session_state.get("rqs2","")
        priority_hks = st.session_state.get("priority_hks",[])
        ds_team      = st.session_state.get("ds_team",[])
        groups_per_insp = 3
    else:
        st.markdown("## Daily Attendance")
        _ri_note = st.session_state.get("roster_import_note")
        if _ri_note:
            st.markdown(f'<div style="background:#eff6ff;border:1px solid #bfdbfe;'
                        f'border-radius:8px;padding:6px 10px;font-size:.72rem;color:#1e40af">'
                        f'Imported {_html.escape(str(_ri_note))}</div>', unsafe_allow_html=True)
        # Today is pulled from the staff schedule automatically, once per day.
        # This re-runs it after a mid-day re-upload or an accidental change.
        if st.button("↻ Reload today from schedule", key="btn_reload_sched",
                     use_container_width=True,
                     help="Re-applies today's attendance, buildings, RQS roles and "
                          "daily-service team from the stored staff schedule."):
            _n = _auto_apply_today(force=True)
            if _n:
                st.success(f"Reloaded — {_n}")
                st.rerun()
            else:
                st.warning("No stored staff schedule covers today.")
        st.markdown("---")
        with st.expander("Add / Remove Housekeeper"):
            col_a, col_b = st.columns([2,1])
            with col_a: new_hk_name = st.text_input("Name", key="new_hk_inp")
            with col_b: new_hk_bld = st.selectbox("Bldg", [1,2,3], key="new_hk_bld")
            if st.button("Add HK", key="btn_add_hk"):
                n = new_hk_name.strip()
                if n and n not in st.session_state["hk_roster"]:
                    st.session_state["hk_roster"][n] = {"building":new_hk_bld,"present":True}
                    _persist_roster(); st.rerun()
            rm_hk = st.selectbox("Remove", ["—"]+list(st.session_state["hk_roster"].keys()), key="rm_hk_sel")
            if st.button("Remove", key="btn_rm_hk") and rm_hk != "—":
                del st.session_state["hk_roster"][rm_hk]
                _persist_roster(); st.rerun()

        st.markdown("### Housekeepers")
        st.caption("Check to mark present. Use the arrow buttons to move between buildings.")
        roster = st.session_state["hk_roster"]
        present_hk = []
        import copy as _copy
        _hk_before = _copy.deepcopy(roster)
        for bld in [1,2,3]:
            bld_hks = [n for n,v in roster.items() if v["building"]==bld]
            # Note: we no longer skip empty buildings — the bulk-paste field
            # below lets you populate an empty building from scratch.
            # Solid, high-contrast building colors that read clearly on the white
            # sidebar (formal theme): each building a distinct professional tone,
            # white text, and a readable count badge.
            BLD_SB = {
                1: ("#2563a8", "#1c4e86"),   # blue   (bar bg, badge bg)
                2: ("#0f766e", "#0c5d56"),   # teal
                3: ("#b45309", "#8f420a"),   # amber
            }
            bar_bg, badge_bg = BLD_SB.get(bld, ("#475569", "#334155"))
            n_present = sum(1 for n in bld_hks if roster[n]["present"])
            st.markdown(
                f'<div style="background:{bar_bg};color:#ffffff;border-radius:8px;'
                f'padding:7px 12px;font-family:\'DM Mono\',monospace;font-size:.7rem;font-weight:600;'
                f'display:flex;justify-content:space-between;align-items:center;'
                f'margin:10px 0 4px;'
                f'letter-spacing:.05em;text-transform:uppercase">'
                f'<span>Building {bld}</span>'
                f'<span style="background:{badge_bg};color:#ffffff;border-radius:20px;padding:2px 10px;'
                f'font-size:.66rem;font-weight:700">{n_present}/{len(bld_hks)}</span>'
                f'</div>',
                unsafe_allow_html=True)

            # ── Select all / none for this building ──────────────────────────
            c_all, c_none = st.columns(2)
            with c_all:
                if st.button("Select all", key=f"selall_b{bld}", use_container_width=True):
                    for _n in bld_hks:
                        roster[_n]["present"] = True
                        # also update the checkbox widget's own state so the tick
                        # reflects the change after rerun
                        st.session_state[f"att_{_n}"] = True
                    _persist_roster(); st.rerun()
            with c_none:
                if st.button("Unselect all", key=f"selnone_b{bld}", use_container_width=True):
                    for _n in bld_hks:
                        roster[_n]["present"] = False
                        st.session_state[f"att_{_n}"] = False
                    _persist_roster(); st.rerun()

            # ── Bulk replace: paste a list of names (e.g. from Excel) to replace
            # ALL housekeepers currently in this building ────────────────────
            with st.expander(f"Bulk set Building {bld} names"):
                _bulk = st.text_area(
                    "Paste names (one per line or comma-separated)",
                    key=f"bulk_hk_b{bld}", height=90,
                    placeholder="Maria Lopez\nAna Garcia\nRosa Diaz",
                    label_visibility="collapsed")
                if st.button(f"Replace Building {bld} list", key=f"bulk_apply_b{bld}",
                             use_container_width=True):
                    # Parse: split on newlines and commas, trim, drop blanks/dupes,
                    # PRESERVING the pasted order.
                    raw = re.split(r'[\n,]+', _bulk or "")
                    new_names = []
                    seen = set()
                    for nm in raw:
                        nm = re.sub(r'\s+', ' ', nm).strip()
                        if nm and nm.lower() not in seen:
                            seen.add(nm.lower()); new_names.append(nm)
                    if new_names:
                        # 1) Remove everyone currently in THIS building.
                        for _n in [n for n,v in roster.items() if v["building"]==bld]:
                            del roster[_n]
                        # 2) Remove any pasted name that already exists in ANOTHER
                        #    building, so the name lives ONLY in the new building
                        #    (no duplicates across buildings).
                        _new_lower = {nm.lower() for nm in new_names}
                        for _n in [n for n in roster if n.lower() in _new_lower]:
                            del roster[_n]
                        # 3) Rebuild this building in the exact pasted order. Because
                        #    dict preserves insertion order, appending them now keeps
                        #    the order you pasted.
                        for _n in new_names:
                            roster[_n] = {"building":bld, "present":True}
                        st.session_state["hk_roster"] = roster
                        _persist_roster()
                        st.toast(f"Building {bld}: set {len(new_names)} housekeepers")
                        st.rerun()
                    else:
                        st.warning("No valid names found to set.")

            for name in bld_hks:
                c_chk, c_name, c_left, c_right = st.columns([0.4,3.2,0.6,0.6])
                with c_chk:
                    # Seed the widget's state from the roster once; thereafter the
                    # widget (and Select/Unselect all, which write the same key)
                    # drives the value. Passing both value= and an existing key
                    # would conflict, so we only set the key.
                    _ck = f"att_{name}"
                    if _ck not in st.session_state:
                        st.session_state[_ck] = roster[name]["present"]
                    checked = st.checkbox("", key=_ck, label_visibility="collapsed")
                    if roster[name]["present"] != checked:
                        roster[name]["present"] = checked
                        _persist_roster()
                    else:
                        roster[name]["present"] = checked
                with c_name:
                    col2 = "inherit" if checked else "#94a3b8"
                    td = "none" if checked else "line-through"
                    st.markdown(f'<div style="font-size:.8rem;color:{col2};padding:4px 0;text-decoration:{td}">{e(name)}</div>', unsafe_allow_html=True)
                with c_left:
                    if bld > 1:
                        if st.button("◀", key=f"ml_{name}", use_container_width=True):
                            roster[name]["building"] = bld-1; _persist_roster(); st.rerun()
                with c_right:
                    if bld < 3:
                        if st.button("▶", key=f"mr_{name}", use_container_width=True):
                            roster[name]["building"] = bld+1; _persist_roster(); st.rerun()
                if roster[name]["present"]: present_hk.append(name)

        # Persist HK attendance immediately so it survives logout/login. The
        # authoritative store is the 'roster' record (via _persist_roster); we no
        # longer double-write to save_full_schedule, which used a different key and
        # could resurrect deleted names.
        if roster != _hk_before:
            _persist_roster()

        st.markdown("---")
        with st.expander("Add / Remove Inspector"):
            new_insp = st.text_input("Name", key="new_insp_inp")
            if st.button("Add Inspector", key="btn_add_insp"):
                n = new_insp.strip()
                if n and n not in st.session_state["insp_roster"]:
                    st.session_state["insp_roster"][n]=True; _persist_roster(); st.rerun()
            rm_insp = st.selectbox("Remove",["—"]+list(st.session_state["insp_roster"].keys()),key="rm_insp_sel")
            if st.button("Remove", key="btn_rm_insp") and rm_insp != "—":
                del st.session_state["insp_roster"][rm_insp]; _persist_roster(); st.rerun()

        st.markdown("### Inspectors")
        insp_roster = st.session_state["insp_roster"]
        # ── Shuffle RQS order so they pair with different housekeepers ───────
        if st.button("Shuffle RQS order", key="btn_shuffle_rqs", use_container_width=True,
                     help="Randomize inspector order so RQS get paired with different housekeepers each run"):
            import random as _rnd
            _items = list(insp_roster.items())
            _rnd.shuffle(_items)
            st.session_state["insp_roster"] = dict(_items)
            insp_roster = st.session_state["insp_roster"]
            # clear cached RQS role picks so they re-pick from the new order
            for _k in ("rqs1","rqs2"):
                if _k in st.session_state: st.session_state[_k] = ""
            _persist_roster()
            st.toast("RQS order shuffled")
            st.rerun()
        present_insp = []
        _insp_before = dict(insp_roster)
        for name in list(insp_roster.keys()):
            insp_roster[name] = st.checkbox(name, value=insp_roster[name], key=f"insp_att_{name}")
            if insp_roster[name]: present_insp.append(name)
        # Persist attendance immediately so it survives logout/login even without
        # regenerating the schedule.
        if insp_roster != _insp_before:
            try:
                _existing = db.load_full_schedule() or {}
                _existing["insp_roster"] = dict(insp_roster)
                _existing["hk_roster"] = dict(st.session_state.get("hk_roster",{}))
                db.save_full_schedule(_existing)
            except Exception:
                pass

        st.markdown("---")
        st.markdown("### RQS Roles Today")
        rqs_opts = [RQS_NONE] + present_insp
        rqs1_sel = st.selectbox("RQS 1 (Dust & Vac)", rqs_opts, key="rqs1_sel")
        rqs2_sel = st.selectbox("RQS 2 (Daily Service)", rqs_opts, key="rqs2_sel")
        rqs1 = "" if rqs1_sel==RQS_NONE else rqs1_sel
        rqs2 = "" if rqs2_sel==RQS_NONE else rqs2_sel
        st.session_state["rqs1"] = rqs1; st.session_state["rqs2"] = rqs2

        st.markdown("---")
        groups_per_insp = st.select_slider("Groups / FC inspector", options=[3,4], value=3)

        st.markdown("---")
        st.markdown("### Priority HKs")
        st.caption("Select HKs who need a full 380-min group (productivity recovery).")
        if "priority_hks" not in st.session_state: st.session_state["priority_hks"] = []
        saved_priority = [n for n in st.session_state["priority_hks"] if n in present_hk]
        if saved_priority != st.session_state["priority_hks"]: st.session_state["priority_hks"] = saved_priority
        st.multiselect("Priority HKs", options=present_hk, key="priority_hks",
                       label_visibility="collapsed", placeholder="Choose HKs for priority full groups…")
        priority_hks = st.session_state["priority_hks"]
        if priority_hks:
            st.markdown(f'<div style="background:#fef3c7;border:1px solid #fcd34d;border-radius:8px;'
                        f'padding:7px 11px;font-size:.75rem;color:#92400e">'
                        f' <b>{len(priority_hks)}</b> HK(s): {", ".join(priority_hks)}</div>',
                        unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("### Daily Service Team")
        st.caption("HKs dedicated to Daily Service today. Only they get DS charts — "
                   "and they are held back from Full Clean / Dust n Vac.")
        if "ds_team" not in st.session_state: st.session_state["ds_team"] = []
        saved_ds_team = [n for n in st.session_state["ds_team"] if n in present_hk]
        if saved_ds_team != st.session_state["ds_team"]: st.session_state["ds_team"] = saved_ds_team
        st.multiselect("Daily Service Team", options=present_hk, key="ds_team",
                       label_visibility="collapsed", placeholder="Choose dedicated DS housekeepers…")
        ds_team = st.session_state["ds_team"]
        if ds_team:
            _ds_last = [g for g in (st.session_state.get("groups_data") or [])
                        if g.get("service_type")==SVC_DS and not g.get("verify_group")]
            _short   = max(len(_ds_last) - len(ds_team), 0)
            st.markdown(f'<div style="background:#ccfbf1;border:1px solid #5eead4;border-radius:8px;'
                        f'padding:7px 11px;font-size:.75rem;color:#115e59">'
                        f'<b>{len(ds_team)}</b> on Daily Service: {", ".join(ds_team)}</div>',
                        unsafe_allow_html=True)
            if _short:
                st.markdown(f'<div style="background:#fee2e2;border:1px solid #fca5a5;border-radius:8px;'
                            f'padding:7px 11px;font-size:.75rem;color:#991b1b;margin-top:6px">'
                            f'Last run was short <b>{_short}</b> — those charts showed '
                            f'"{NEED_HK_PREFIX} 1…{_short}".</div>',
                            unsafe_allow_html=True)
        else:
            st.caption("None selected — Daily Service is assigned from the full roster as usual.")

        st.markdown(f'<div style="background:#f1f5f9;border-radius:8px;padding:9px 11px;font-size:.77rem;color:#475569;margin-top:8px">'
                    f' <b>{len(present_hk)}</b> HKs &nbsp;·&nbsp; <b>{len(present_insp)}</b> inspectors<br>'
                    f'RQS1: <b>{rqs1 or "—"}</b> &nbsp;·&nbsp; RQS2: <b>{rqs2 or "—"}</b></div>',
                    unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# MAIN INPUT
# ══════════════════════════════════════════════════════════════════════════════
_cu = auth.current_user()
_disp_name = _cu.get("display_name") or _cu.get("username","")
_first_name = _disp_name.split()[0].title() if _disp_name else "there"
_welcome_msg = auth.get_welcome_msg(_cu["role"])

ui.topnav("Schedule", hide_sidebar=False)

st.markdown(f'<p class="pg-title">Good morning, {_first_name}! </p>', unsafe_allow_html=True)
st.markdown(f'<p class="pg-sub">{_welcome_msg}</p>', unsafe_allow_html=True)

# ── Staff schedule stamp: when Schedule.xlsx was last loaded into the app ─────
@st.cache_data(ttl=120, show_spinner=False)
def _staff_stamp():
    """Cached so the header does not hit the database on every rerun."""
    try:
        return db.load_staff_meta() or {}
    except Exception:
        return {}

def _ago(iso):
    try:
        _t = _datetime.fromisoformat(str(iso))
    except Exception:
        return ""
    _n = _datetime.now(_t.tzinfo) if _t.tzinfo else _datetime.now()
    _s = (_n - _t).total_seconds()
    if _s < 90:     return "just now"
    if _s < 5400:   return f"{int(_s//60)} min ago"
    if _s < 172800: return f"{int(_s//3600)} hours ago"
    return f"{int(_s//86400)} days ago"

_stamp = _staff_stamp()
if _stamp.get("uploaded_at"):
    _sw = str(_stamp["uploaded_at"])[:16].replace("T", " ")
    _stale = _ago(_stamp["uploaded_at"]).endswith("days ago") and \
             int(_ago(_stamp["uploaded_at"]).split()[0] or 0) >= 8
    _c = ("#b45309", "#fffbeb", "#fcd34d") if _stale else ("#5b6675", "#f8fafc", "#e2e5ea")
    st.markdown(
        f'<div style="display:inline-flex;align-items:center;gap:9px;background:{_c[1]};'
        f'border:1px solid {_c[2]};border-radius:8px;padding:5px 12px;margin:-4px 0 10px;'
        f'font-size:.74rem;color:{_c[0]}">'
        f'<span style="font-family:\'DM Mono\',monospace;font-size:.6rem;'
        f'text-transform:uppercase;letter-spacing:.11em;opacity:.75">Staff schedule loaded</span>'
        f'<b>{_sw}</b><span style="opacity:.7">({_ago(_stamp["uploaded_at"])})</span>'
        f'<span style="opacity:.55">· {e(_stamp.get("file_name",""))}'
        f' · {_stamp.get("n_weeks","?")} weeks</span>'
        f'{" · <b>update due</b>" if _stale else ""}</div>', unsafe_allow_html=True)
elif auth.can("can_edit_roster"):
    st.markdown(
        '<div style="display:inline-flex;align-items:center;gap:8px;background:#fffbeb;'
        'border:1px solid #fcd34d;border-radius:8px;padding:5px 12px;margin:-4px 0 10px;'
        'font-size:.74rem;color:#b45309">No staff schedule loaded yet — '
        'import it from the <b>Roster Import</b> page.</div>', unsafe_allow_html=True)

_aa_note = st.session_state.get("_autoapply_note")
if _aa_note:
    st.markdown(
        f'<div style="display:inline-flex;align-items:center;gap:9px;background:#ecfdf5;'
        f'border:1px solid #a7f3d0;border-radius:8px;padding:5px 12px;margin:-4px 0 10px;'
        f'font-size:.74rem;color:#065f46">'
        f'<span style="font-family:\'DM Mono\',monospace;font-size:.6rem;text-transform:uppercase;'
        f'letter-spacing:.11em;opacity:.75">Today loaded automatically</span>'
        f'{e(_aa_note)}</div>', unsafe_allow_html=True)

st.markdown("---")
# Once a schedule has been generated, collapse the upload/paste inputs so the
# dashboard (metrics + table) is front and center. The section can be reopened
# any time to re-upload or edit and regenerate.
_has_sched = bool(st.session_state.get("groups_data"))
_inp_exp = st.expander("Room Data + Front-Desk Email"
                       + ("  —  tap to edit / regenerate" if _has_sched else ""),
                       expanded=not _has_sched)
with _inp_exp:
    col_data, col_cfg = st.columns([5,1], gap="medium")
    with col_data:
        inp_a, inp_b = st.columns([3,2], gap="small")
        with inp_a:
            # ── Upload the Housekeeping Dashboard .xlsx directly (optional) ────────
            # Alternative to copy-paste: an uploaded file is converted to the same
            # tab-separated text the paste box uses and run through the exact same
            # parser, so nothing downstream changes. We push the converted text into
            # the text-area's session_state (tracking which file we've already read,
            # so re-runs don't clobber manual edits) rather than passing value=,
            # which avoids Streamlit's value/key conflict warning.
            if auth.can("can_paste_input"):
                _xl = st.file_uploader("Upload Housekeeping Dashboard (.xlsx)",
                                       type=["xlsx"], key="room_xlsx",
                                       help="Optional — upload the exported dashboard instead of pasting.")
                if _xl is not None:
                    _sig = f"{_xl.name}:{getattr(_xl,'size','?')}"
                    if st.session_state.get("_room_xlsx_sig") != _sig:
                        try:
                            _txt, _n_up, _sheet_up = excel_to_room_text(_xl)
                            st.session_state["room_input"] = _txt
                            st.session_state["_room_xlsx_sig"] = _sig
                            st.session_state["_room_xlsx_msg"] = (
                                "ok", f"Read {_n_up} rooms from '{_sheet_up}'. "
                                      f"Click Generate to build the schedule.")
                        except ImportError:
                            st.session_state["_room_xlsx_sig"] = _sig
                            st.session_state["_room_xlsx_msg"] = (
                                "err", "The .xlsx reader isn't installed on the server yet "
                                       "(openpyxl). Once the app redeploys it'll work — for "
                                       "now you can copy-paste the data below.")
                        except Exception as _e:
                            st.session_state["_room_xlsx_sig"] = _sig
                            st.session_state["_room_xlsx_msg"] = (
                                "err", f"Couldn't read that file: {_e}. "
                                       f"You can still copy-paste the data below.")
                    _msg = st.session_state.get("_room_xlsx_msg")
                    if _msg:
                        if _msg[0] == "ok":
                            st.caption(_msg[1])
                        else:
                            st.error(_msg[1])
            raw_input = st.text_area("rooms", label_visibility="collapsed", height=230,
                disabled=not auth.can("can_paste_input"),
                placeholder="Room\tService\tTime\tPet\tCurrent Guest or Status\n1020D\tFull Clean\t120\t\tSmith, John",
                key="room_input")
            st.caption("Upload the .xlsx above, or copy-paste from Excel (include header row).")
        with inp_b:
            email_text = st.text_area("email", label_visibility="collapsed", height=230,
                disabled=not auth.can("can_paste_input"),
                placeholder="Paste today's front-desk email...\n\nLate Checkouts:\n* 10:30 am\n * 1234A",
                key="email_input")
            st.caption("Late checkouts, room moves, notes auto-matched.")
    with col_cfg:
        st.markdown('<p class="sec"></p>', unsafe_allow_html=True)
        _today_bg = ("rgba(255,255,255,.7)" if _IS_GLASS else "#ffffff") if _IS_LIGHT else ("rgba(40,40,52,.5)" if _IS_GLASS else "rgba(255,255,255,.03)")
        _today_br = "rgba(99,102,241,.18)"
        _today_hd = "#0f172a" if _IS_LIGHT else "#e2e8f0"
        _today_tx = "#475569" if _IS_LIGHT else "#94a3b8"
        st.markdown(f'<div style="background:{_today_bg};border:1px solid {_today_br};border-radius:10px;'
                    f'padding:11px 13px;font-size:.78rem;color:{_today_tx};margin-bottom:10px">'
                    f'<div style="font-weight:700;color:{_today_hd};margin-bottom:5px">Today</div>'
                    f'<div> <b>{len(present_hk)}</b> HKs present</div>'
                    f'<div> <b>{len(present_insp)}</b> inspectors</div></div>', unsafe_allow_html=True)
        _can_gen = auth.can("can_generate")
        run = st.button("Generate", type="primary", use_container_width=True,
                        disabled=not _can_gen,
                        help="" if _can_gen else "Housekeeper role — view only")

# ══════════════════════════════════════════════════════════════════════════════
# SNAPSHOT
# ══════════════════════════════════════════════════════════════════════════════
def _build_snapshot(fg, total_rooms, inspectors):
    from datetime import date
    hk_snap = {}
    for g in fg:
        hk = g.get("housekeeper","")
        if hk and hk != "Manager" and not is_unassigned_hk(hk):
            if hk not in hk_snap:
                hk_snap[hk] = {"time":0,"rooms":0,"rooms_fc":0,"rooms_ds":0,"rooms_dv":0}
            n = len(g.get("rooms",[]))
            hk_snap[hk]["time"] += g.get("time",0)
            hk_snap[hk]["rooms"] += n
            svc = g.get("service_type","")
            if svc==SVC_FC: hk_snap[hk]["rooms_fc"] += n
            elif svc==SVC_DS:hk_snap[hk]["rooms_ds"] += n
            elif svc==SVC_DV:hk_snap[hk]["rooms_dv"] += n
    insp_snap = {}
    for insp in inspectors:
        nm = insp.get("name","")
        if nm:
            labels = set(insp.get("groups",[]))
            n_rooms = sum(len(g.get("rooms",[])) for g in fg if g.get("label") in labels)
            insp_snap[nm] = {"rooms":n_rooms,"groups":len(labels),"role":insp.get("role","FC")}
    return {"date":_datetime.now(_MTN_TZ).date().isoformat(),"total_rooms":total_rooms,"n_groups":len(fg),
            "hk":hk_snap,"inspectors":insp_snap,
            "saved_by":st.session_state.get("username","unknown"),"schema_v":2}

# ══════════════════════════════════════════════════════════════════════════════
# GENERATE
# ══════════════════════════════════════════════════════════════════════════════
if run:
    st.session_state["last_email"] = email_text
    if not raw_input.strip():
        st.warning("Paste room data first.")
    elif not present_hk:
        st.warning("No housekeepers marked as present.")
    else:
        # ── Animated loading overlay (formal theme) ──────────────────────────
        # A polished animation shown while the schedule is built. It cycles
        # through the real pipeline stages so the wait feels purposeful.
        _loader = st.empty()
        _loader.markdown("""
<style>
@keyframes gc8spin { to { transform: rotate(360deg); } }
@keyframes gc8pulse { 0%,100% { opacity:.35; } 50% { opacity:1; } }
@keyframes gc8bar {
  0% { left:-40%; width:40%; }
  50% { width:55%; }
  100% { left:100%; width:40%; }
}
@keyframes gc8fade { from { opacity:0; transform:translateY(4px);} to { opacity:1; transform:translateY(0);} }
.gc8-wrap{
  display:flex;flex-direction:column;align-items:center;justify-content:center;
  padding:44px 24px;margin:8px 0 4px;
  background:#ffffff;border:1px solid #e2e5ea;border-radius:16px;
  box-shadow:0 1px 3px rgba(20,32,54,.06),0 12px 30px rgba(20,32,54,.05);
  animation:gc8fade .4s ease both;
}
.gc8-ring{
  width:54px;height:54px;border-radius:50%;
  border:4px solid #e6eaf1;border-top-color:#2563a8;
  animation:gc8spin .8s linear infinite;margin-bottom:20px;
}
.gc8-title{
  font-family:'Syne',sans-serif;font-weight:700;font-size:1.12rem;
  color:#16202e;letter-spacing:-.01em;margin-bottom:6px;
}
.gc8-sub{
  font-family:'DM Sans',sans-serif;font-size:.82rem;color:#5b6675;
  margin-bottom:20px;min-height:1.1em;
}
.gc8-track{
  position:relative;width:min(340px,80%);height:6px;border-radius:99px;
  background:#eceef1;overflow:hidden;
}
.gc8-fill{
  position:absolute;top:0;height:100%;border-radius:99px;
  background:linear-gradient(90deg,#2563a8,#4b8bd0);
  animation:gc8bar 1.35s cubic-bezier(.65,0,.35,1) infinite;
}
.gc8-steps{
  display:flex;gap:18px;margin-top:22px;flex-wrap:wrap;justify-content:center;
  font-family:'DM Mono',monospace;font-size:.68rem;letter-spacing:.03em;
  text-transform:uppercase;color:#8a93a1;
}
.gc8-steps span{animation:gc8pulse 1.4s ease-in-out infinite;}
.gc8-steps span:nth-child(2){animation-delay:.2s;}
.gc8-steps span:nth-child(3){animation-delay:.4s;}
.gc8-steps span:nth-child(4){animation-delay:.6s;}
.gc8-steps span:nth-child(5){animation-delay:.8s;}
</style>
<div class="gc8-wrap">
  <div class="gc8-ring"></div>
  <div class="gc8-title">Building your schedule</div>
  <div class="gc8-sub">Packing rooms into the fewest, tidiest charts…</div>
  <div class="gc8-track"><div class="gc8-fill"></div></div>
  <div class="gc8-steps">
    <span>Parsing</span><span>Grouping</span><span>Floors</span><span>Assigning</span><span>Inspectors</span>
  </div>
</div>
""", unsafe_allow_html=True)
        if True:
            try:
                df = parse_rooms(raw_input)
                if df.empty:
                    _loader.empty()
                    st.error("No valid rows — check tab-separated data with a header row.")
                else:
                    email_data = parse_email_notes(email_text)
                    late_co_map = email_data["late_checkout"]
                    email_notes_map = email_data["notes"]
                    if email_text.strip():
                        n_late = len(late_co_map)
                        n_notes= sum(len(v) for v in email_notes_map.values())
                        # Show the email-parse summary quietly in the sidebar rather
                        # than as a large banner in the middle of the page.
                        if n_late > 0:
                            late_rooms = ", ".join(f"{rm} ({t.replace('Late Out: ','')})" for rm,t in sorted(late_co_map.items()))
                            st.sidebar.caption(
                                f"Email parsed — {n_late} late checkout(s), {n_notes} note(s).")
                            st.sidebar.caption(f"Late rooms: {late_rooms}")
                        else:
                            st.sidebar.caption("Email parsed — no late checkouts found.")
                    records_raw = df.to_dict("records")
                    rds = []
                    for r in records_raw:
                        rm_upper = str(r["Room"]).strip().upper()
                        excel_late = r.get("LateCheckout","").strip()
                        email_late = late_co_map.get(rm_upper,"")
                        if email_late: late_co = email_late
                        elif excel_late and excel_late.lower() not in ("","late check out","late checkout","late check-out"): late_co = excel_late
                        elif excel_late: late_co = "Late Out"
                        else: late_co = ""
                        notes_parts = []
                        if r.get("NotesRaw","").strip(): notes_parts.append(r["NotesRaw"].strip())
                        if rm_upper in email_notes_map: notes_parts += email_notes_map[rm_upper]
                        has_stayover = (
                            r.get("uncertain",False) or
                            any("stayover" in n.lower() or "stay over" in n.lower() for n in notes_parts)
                        )
                        # verify rooms (stayover / P-U Models): no auto HK/RQS, sent to bottom
                        needs_verify = (
                            r.get("verify",False) or
                            any("stayover" in n.lower() or "stay over" in n.lower()
                                or "p/u model" in n.lower() or "pu model" in n.lower()
                                for n in notes_parts)
                        )
                        rds.append({
                            "room":r["Room"],"service":r["Service"],"time":r["Time"],
                            "pet":r["Pet"],"guest":r["Guest"],
                            "bld":r.get("bld",get_building(r["Room"])),
                            "floor":r.get("floor",0),"num":r.get("num",0),
                            "late_checkout":late_co,"status":r.get("Status",""),
                            "notes":"; ".join(notes_parts),"arriving":r.get("ArrivingGuest",""),
                            "res_type":r.get("ResType",""),"uncertain":has_stayover,
                            "verify":needs_verify,
                        })
                    fg = build_all_groups(rds, priority_hks=st.session_state.get("priority_hks",[]))
                    fc_gs=[g for g in fg if g.get("service_type")==SVC_FC and not g.get("verify_group")]
                    ih_gs=[g for g in fg if g.get("service_type")==SVC_IH and not g.get("verify_group")]
                    ds_gs=[g for g in fg if g.get("service_type")==SVC_DS and not g.get("verify_group")]
                    dv_gs=[g for g in fg if g.get("service_type")==SVC_DV and not g.get("verify_group")]
                    vr_gs=[g for g in fg if g.get("verify_group")]
                    for g,lbl in zip(fc_gs, make_labels("FC",len(fc_gs))): g["label"]=lbl
                    for g,lbl in zip(ih_gs, make_labels("IH",len(ih_gs))): g["label"]=lbl
                    for g,lbl in zip(ds_gs, make_labels("DS",len(ds_gs))): g["label"]=lbl
                    for g,lbl in zip(dv_gs, make_labels("DV",len(dv_gs))): g["label"]=lbl
                    for g,lbl in zip(vr_gs, make_labels("VERIFY",len(vr_gs))): g["label"]=lbl
                    for g in fg: g["cross_bld"] = len(g["blds"])>1

                    # (The new solve_full_clean already consolidates to the fewest
                    # charts, so the legacy tiny-merge pass below is disabled to
                    # avoid undoing its tidy, building-coherent packing.)
                    fg=[g for g in fg if g["rooms"]]
                    p_fc=[g for g in fg if g.get("service_type")==SVC_FC and g.get("priority_hk") and not g.get("verify_group")]
                    n_fc=[g for g in fg if g.get("service_type")==SVC_FC and not g.get("priority_hk") and not g.get("verify_group")]
                    ih2=[g for g in fg if g.get("service_type")==SVC_IH and not g.get("verify_group")]
                    ds2=[g for g in fg if g.get("service_type")==SVC_DS and not g.get("verify_group")]
                    dv2=[g for g in fg if g.get("service_type")==SVC_DV and not g.get("verify_group")]
                    vr2=[g for g in fg if g.get("verify_group")]
                    for g,lbl in zip(p_fc+n_fc, make_labels("FC",len(p_fc)+len(n_fc))): g["label"]=lbl
                    for g,lbl in zip(ih2, make_labels("IH",len(ih2))): g["label"]=lbl
                    for g,lbl in zip(ds2, make_labels("DS",len(ds2))): g["label"]=lbl
                    for g,lbl in zip(dv2, make_labels("DV",len(dv2))): g["label"]=lbl
                    for g,lbl in zip(vr2, make_labels("VERIFY",len(vr2))): g["label"]=lbl
                    for g in fg: g["cross_bld"]=len(g["blds"])>1

                    hk_asgn, used_hk_set = assign_hk_building_aware(
                        fg, present_hk, roster, ds_team=st.session_state.get("ds_team",[]))
                    for g in fg: g["housekeeper"] = hk_asgn.get(g["label"],"")
                    inspectors = assign_inspectors(fg, present_insp, groups_per_insp, rqs1, rqs2)

                    # Store fresh result in session state
                    st.session_state["groups_data"] = fg
                    st.session_state["total_rooms"] = len(df)
                    st.session_state["inspectors_data"] = inspectors
                    st.session_state["used_hk_set"] = used_hk_set

                    # Save to DB for sharing + dashboard (non-blocking)
                    try:
                        db.save_full_schedule({
                            "groups_data": fg, "total_rooms": len(df),
                            "inspectors_data": inspectors,
                            "used_hk_set": list(used_hk_set),
                            "hk_roster": dict(st.session_state.get("hk_roster",{})),
                            "insp_roster": dict(st.session_state.get("insp_roster",{})),
                            "generated_by": st.session_state.get("username","unknown"),
                        })
                    except Exception:
                        pass
                    try:
                        db.save_snapshot(_build_snapshot(fg,len(df),inspectors))
                    except Exception:
                        pass

                    _loader.empty()
                    st.success(f"Schedule generated — {len(fg)} groups from {len(df)} rooms.")
            except Exception as ex:
                _loader.empty()
                st.error(f"Error: {ex}")
                import traceback; st.code(traceback.format_exc())

# ══════════════════════════════════════════════════════════════════════════════
# RESULTS
# ══════════════════════════════════════════════════════════════════════════════
if not st.session_state.get("groups_data"): st.stop()

fg = st.session_state["groups_data"]
total_rooms = st.session_state["total_rooms"]
inspectors = st.session_state["inspectors_data"]
used_hk_set = st.session_state.get("used_hk_set") or set()
present_hk = [n for n,v in st.session_state["hk_roster"].items() if v["present"]]
present_insp= [n for n,v in st.session_state["insp_roster"].items() if v]

st.markdown("---")
fc_g=[g for g in fg if g.get("service_type")==SVC_FC]
ih_g=[g for g in fg if g.get("service_type")==SVC_IH]
ds_g=[g for g in fg if g.get("service_type")==SVC_DS]
dv_g=[g for g in fg if g.get("service_type")==SVC_DV]
# Room counts per service (what actually matters day-to-day), not group counts.
fc_rooms_n = sum(len(g["rooms"]) for g in fc_g)
ih_rooms_n = sum(len(g["rooms"]) for g in ih_g)
ds_rooms_n = sum(len(g["rooms"]) for g in ds_g)
dv_rooms_n = sum(len(g["rooms"]) for g in dv_g)
n_free_hk=sum(1 for n in present_hk if n not in used_hk_set)
n_low_hk =sum(1 for g in fg if g.get("housekeeper") and g.get("housekeeper")!="Manager"
              and not is_unassigned_hk(g.get("housekeeper")) and g["time"]<LOW_MIN)
n_need_hk=sum(1 for g in fg if str(g.get("housekeeper","")).startswith(NEED_HK_PREFIX))

st.markdown(f"""<div class="stat-row">
  <div class="sc hi"><div class="n">{total_rooms}</div><div class="l">Total Rooms</div></div>
  <div class="sc"><div class="n" style="color:#2563EB">{fc_rooms_n}</div><div class="l">Full Clean</div></div>
  <div class="sc"><div class="n" style="color:#7C3AED">{ih_rooms_n}</div><div class="l">Full Clean (IH)</div></div>
  <div class="sc ds"><div class="n">{ds_rooms_n}</div><div class="l">Daily Service</div></div>
  <div class="sc dv"><div class="n">{dv_rooms_n}</div><div class="l">Dust &amp; Vac</div></div>
  <div class="sc"><div class="n" style="color:{'#059669' if n_free_hk==0 else '#d97706'}">{n_free_hk}</div>
    <div class="l">Free HKs</div></div>
  <div class="sc"><div class="n" style="color:{'#059669' if n_low_hk==0 else '#dc2626'}">{n_low_hk}</div>
    <div class="l">Low-Hour HKs</div></div>
</div>""", unsafe_allow_html=True)

_is_hk = auth.is_housekeeper()
_my_name = auth.my_display_name()

if _is_hk:
    # ══════════════════════════════════════════════════════════════════════
    # HOUSEKEEPER VIEW — single "My Schedule"tab, own rooms only
    # ══════════════════════════════════════════════════════════════════════
    _NOW = _now_iso # shared Mountain-time timestamp helper

    _tmsg_hk = st.session_state.pop("_live_toast", None)
    if _tmsg_hk:
        try: st.toast(_tmsg_hk)
        except Exception: pass

    STATUS_META_HK = {
        "pending": {"icon":"","label":"Pending","color":"#475569","bg":"rgba(71,85,105,.2)","border":"rgba(71,85,105,.35)"},
        "already_clean": {"icon":"","label":"Already Clean", "color":"#34d399","bg":"rgba(52,211,153,.12)","border":"rgba(52,211,153,.35)"},
        "cleaning_started": {"icon":"","label":"In Progress","color":"#fbbf24","bg":"rgba(251,191,36,.15)","border":"rgba(251,191,36,.4)"},
        "cleaning_done": {"icon":"","label":"Done","color":"#60a5fa","bg":"rgba(96,165,250,.12)","border":"rgba(96,165,250,.35)"},
        "inspected": {"icon":"","label":"Inspected ","color":"#a78bfa","bg":"rgba(167,139,250,.15)","border":"rgba(167,139,250,.4)"},
    }

    # Init room statuses
    if "room_statuses" not in st.session_state:
        st.session_state["room_statuses"] = {}
    if not st.session_state.get("_live_loaded"):
        st.session_state["_live_loaded"] = True
        try:
            st.session_state["room_statuses"] = db.get_room_statuses()
        except Exception:
            pass
    rs = st.session_state["room_statuses"]

    def _save_status_hk(room, fields):
        if room not in rs: rs[room] = {"room":room,"status":"pending"}
        rs[room].update(fields)
        rs[room]["updated_by"] = _my_name
        try:
            db.upsert_room_status(room, fields | {"updated_by": _my_name})
        except Exception:
            pass

    # Collect this HK's groups
    my_groups = [g for g in fg if g.get("housekeeper","") == _my_name]

    # Stats
    my_rooms_all = [r for g in my_groups for r in g["rooms"]]
    n_total = len(my_rooms_all)
    n_done = sum(1 for r in my_rooms_all if rs.get(r["room"],{}).get("status") in ("already_clean","cleaning_done","inspected"))
    n_active = sum(1 for r in my_rooms_all if rs.get(r["room"],{}).get("status") == "cleaning_started")
    n_insp = sum(1 for r in my_rooms_all if rs.get(r["room"],{}).get("status") == "inspected")
    pct = int(n_done / max(n_total,1) * 100)

    # Progress header
    st.markdown(f"""
<div style="background:rgba(99,102,241,.08);border:1px solid rgba(99,102,241,.2);
            border-radius:14px;padding:16px 20px;margin-bottom:20px">
  <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px">
    <div>
      <div style="font-family:'Syne',sans-serif;font-size:1.1rem;font-weight:700;color:#e2e8f0">
        Your Rooms Today
      </div>
      <div style="font-family:'DM Mono',monospace;font-size:.7rem;color:#475569;margin-top:2px">
        {n_total} rooms &nbsp;·&nbsp; {len(my_groups)} group{"s" if len(my_groups)!=1 else ""}
      </div>
    </div>
    <div style="display:flex;gap:10px">
      <div style="text-align:center;background:rgba(251,191,36,.1);border:1px solid rgba(251,191,36,.25);border-radius:8px;padding:8px 14px">
        <div style="font-family:'Syne',sans-serif;font-size:1.3rem;font-weight:800;color:#fbbf24">{n_active}</div>
        <div style="font-family:'DM Mono',monospace;font-size:.58rem;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em">Active</div>
      </div>
      <div style="text-align:center;background:rgba(96,165,250,.1);border:1px solid rgba(96,165,250,.25);border-radius:8px;padding:8px 14px">
        <div style="font-family:'Syne',sans-serif;font-size:1.3rem;font-weight:800;color:#60a5fa">{n_done}</div>
        <div style="font-family:'DM Mono',monospace;font-size:.58rem;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em">Done</div>
      </div>
      <div style="text-align:center;background:rgba(167,139,250,.1);border:1px solid rgba(167,139,250,.25);border-radius:8px;padding:8px 14px">
        <div style="font-family:'Syne',sans-serif;font-size:1.3rem;font-weight:800;color:#a78bfa">{n_insp}</div>
        <div style="font-family:'DM Mono',monospace;font-size:.58rem;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em">Inspected</div>
      </div>
    </div>
  </div>
  <div style="margin-top:12px">
    <div style="background:rgba(255,255,255,.05);border-radius:99px;height:7px;overflow:hidden;border:1px solid rgba(255,255,255,.04)">
      <div style="background:linear-gradient(90deg,#6366f1,#22d3ee);width:{pct}%;height:7px;
                  border-radius:99px;box-shadow:0 0 8px rgba(99,102,241,.5);transition:width .5s"></div>
    </div>
    <div style="font-family:'DM Mono',monospace;font-size:.65rem;color:#334155;margin-top:4px">
      {pct}% complete
    </div>
  </div>
</div>""", unsafe_allow_html=True)

    if not my_groups:
        st.markdown("""
<div style="text-align:center;padding:60px 20px;color:#334155;font-family:'DM Mono',monospace">
  <div style="font-size:2rem;margin-bottom:12px"></div>
  <div style="font-size:.9rem">No rooms assigned yet for today.</div>
  <div style="font-size:.75rem;margin-top:6px;color:#1e293b">Check back once the schedule is generated.</div>
</div>""", unsafe_allow_html=True)
    else:
        # Render each group
        for _hgidx, g in enumerate(my_groups):
            g_label = g.get("label","")
            insp_name = g.get("inspector","—")
            g_rooms = g["rooms"]
            g_done = sum(1 for r in g_rooms if rs.get(r["room"],{}).get("status") in ("already_clean","cleaning_done","inspected"))
            g_pct = int(g_done / max(len(g_rooms),1) * 100)
            g_color = "#6366f1" if g.get("service_type")==SVC_FC else ("#14b8a6" if g.get("service_type")==SVC_DS else "#f59e0b")

            # Group header
            st.markdown(f"""
<div style="background:linear-gradient(135deg,rgba(14,14,26,.95),rgba(19,19,31,.98));
            border:1px solid rgba(99,102,241,.25);border-radius:12px;
            padding:12px 16px;margin-bottom:8px;
            box-shadow:0 0 0 1px rgba(255,255,255,.03),0 4px 20px rgba(0,0,0,.3)">
  <div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px">
    <div style="display:flex;align-items:center;gap:10px">
      <div style="background:{g_color};color:#fff;border-radius:6px;padding:4px 10px;
                  font-family:'Syne',sans-serif;font-weight:800;font-size:.78rem;
                  box-shadow:0 0 10px {g_color}66">{g_label}</div>
      <div>
        <div style="font-family:'DM Sans',sans-serif;font-size:.82rem;font-weight:600;color:#e2e8f0">
          {g.get("service_type","")} &nbsp;<span style="color:#334155">·</span>&nbsp;
          <span style="font-family:'DM Mono',monospace;font-size:.72rem;color:#475569">
            Inspector: {insp_name}
          </span>
        </div>
        <div style="font-family:'DM Mono',monospace;font-size:.65rem;color:#334155;margin-top:2px">
          {g_done}/{len(g_rooms)} done
        </div>
      </div>
    </div>
    <div style="display:flex;align-items:center;gap:8px">
      <div style="background:rgba(255,255,255,.04);border-radius:99px;height:5px;width:80px;overflow:hidden;border:1px solid rgba(255,255,255,.05)">
        <div style="background:linear-gradient(90deg,{g_color},{g_color}aa);width:{g_pct}%;height:5px;border-radius:99px"></div>
      </div>
      <span style="font-family:'DM Mono',monospace;font-size:.7rem;color:{g_color}">{g_pct}%</span>
    </div>
  </div>
</div>""", unsafe_allow_html=True)

            # Room rows (dedupe so a room repeated in a group can't collide)
            _seen_hr = set(); _dedup_g_rooms = []
            for r in g_rooms:
                _rc = r.get("room")
                if _rc in _seen_hr: continue
                _seen_hr.add(_rc); _dedup_g_rooms.append(r)
            for _hridx, r in enumerate(_dedup_g_rooms):
                rm = r["room"]
                _hrk = f"{_hgidx}_{_hridx}_{rm}" # fully unique row key
                # Ensure room is initialised in rs with this HK's info
                if rm not in rs:
                    rs[rm] = {"room":rm,"status":"pending","housekeeper":_my_name,
                              "group_label":g_label,"inspector":insp_name}
                r_state = rs.get(rm, {"status":"pending"})
                cur = r_state.get("status","pending")
                sm = STATUS_META_HK.get(cur, STATUS_META_HK["pending"])

                _fmt = _fmt_mtn

                pet_icon = " " if r.get("pet") else ""
                late_icon = " " if r.get("late_checkout") else ""
                late_html = (f'<span style="font-family:\'DM Mono\',monospace;font-size:.62rem;'
                             f'color:#f59e0b;background:rgba(245,158,11,.12);border-radius:4px;'
                             f'padding:1px 5px;margin-left:4px"> {r.get("late_checkout","")}</span>'
                             if r.get("late_checkout") else "")
                guest_disp = r.get("guest","")[:22]

                # ── Info line: room + guest + animated status badge ──
                _hk_active = (cur == "cleaning_started")
                _hk_dot = (f'<span style="display:inline-block;width:6px;height:6px;'
                           f'border-radius:50%;background:{sm["color"]};margin-right:5px;'
                           f'vertical-align:middle;'
                           + ('animation:pulseDot 1.1s ease-in-out infinite;' if _hk_active else '')
                           + '"></span>')
                _hk_ring = 'animation:statusPop .35s cubic-bezier(.34,1.56,.64,1) both' \
                           + (',ringPulse 2s ease-out infinite' if _hk_active else '')
                st.markdown(
                    f'<div style="display:flex;align-items:center;justify-content:space-between;'
                    f'gap:8px;flex-wrap:wrap;padding:8px 10px;background:rgba(255,255,255,.02);'
                    f'border:1px solid rgba(99,102,241,.1);border-radius:8px;margin-bottom:4px">'
                    f' <div style="display:flex;align-items:center;gap:8px;min-width:0;flex:1">'
                    f' <span style="font-family:\'DM Mono\',monospace;font-size:.85rem;'
                    f'font-weight:600;color:#6366f1;white-space:nowrap">{rm}</span>'
                    f' <span style="font-size:.78rem;color:#94a3b8;overflow:hidden;'
                    f'text-overflow:ellipsis;white-space:nowrap">{guest_disp}{pet_icon}</span>'
                    f' {late_html}'
                    f' </div>'
                    f' <div style="background:{sm["bg"]};border:1px solid {sm["border"]};'
                    f'border-radius:6px;padding:3px 9px;font-size:.7rem;font-weight:600;'
                    f'color:{sm["color"]};white-space:nowrap;flex-shrink:0;{_hk_ring}">'
                    f'{_hk_dot}{sm["icon"]} {sm["label"]}</div>'
                    f'</div>', unsafe_allow_html=True)

                # ── Action buttons row — compact, equal width, stays horizontal ──
                b1,b2,b3 = st.columns(3)
                with b1:
                    if cur == "pending":
                        if st.button("Clean", key=f"hk_ac_{_hrk}", use_container_width=True):
                            _save_status_hk(rm, {"status":"already_clean","marked_clean_at":_NOW()})
                            st.session_state["_live_toast"] = f" {rm} marked Already Clean"
                            st.rerun()
                    elif cur == "already_clean":
                        if st.button("Undo", key=f"hk_uac_{_hrk}", use_container_width=True):
                            _save_status_hk(rm, {"status":"pending"})
                            st.rerun()
                with b2:
                    if cur == "pending":
                        if st.button("Start", key=f"hk_s_{_hrk}", use_container_width=True):
                            _save_status_hk(rm, {"status":"cleaning_started","started_at":_NOW()})
                            st.session_state["_live_toast"] = f" {rm} — cleaning started"
                            st.rerun()
                    elif cur == "cleaning_started":
                        if st.button("Done", key=f"hk_d_{_hrk}", use_container_width=True):
                            _save_status_hk(rm, {"status":"cleaning_done","cleaned_at":_NOW()})
                            st.session_state["_live_toast"] = f" {rm} — done, awaiting inspection"
                            st.rerun()
                with b3:
                    if cur not in ("pending",):
                        if st.button("Reset", key=f"hk_r_{_hrk}", use_container_width=True):
                            _save_status_hk(rm, {"status":"pending","started_at":None,"cleaned_at":None,"inspected_at":None,"marked_clean_at":None})
                            st.rerun()

                # Timestamp trail
                ts_parts = []
                if _fmt(r_state.get("marked_clean_at","")): ts_parts.append(f' {_fmt(r_state.get("marked_clean_at",""))}')
                if _fmt(r_state.get("started_at","")): ts_parts.append(f' {_fmt(r_state.get("started_at",""))}')
                if _fmt(r_state.get("cleaned_at","")): ts_parts.append(f' {_fmt(r_state.get("cleaned_at",""))}')
                if _fmt(r_state.get("inspected_at","")): ts_parts.append(f' {_fmt(r_state.get("inspected_at",""))}')
                if ts_parts:
                    st.markdown(
                        f'<div style="font-family:\'DM Mono\',monospace;font-size:.6rem;color:#1e293b;'
                        f'padding:0 0 6px 4px">{" · ".join(ts_parts)}</div>', unsafe_allow_html=True)

            st.markdown("<div style='margin-bottom:12px'></div>", unsafe_allow_html=True)

    # Refresh button
    if st.button("Refresh Status", key="hk_refresh", use_container_width=False):
        try:
            st.session_state["room_statuses"] = db.get_room_statuses()
        except Exception:
            pass
        st.rerun()

else:
    # ══════════════════════════════════════════════════════════════════════
    # ADMIN / RQS VIEW — one comprehensive schedule table + Live tracking
    # ══════════════════════════════════════════════════════════════════════
    tab_sched, tab_reassign, tab_live = st.tabs(["Schedule", "Reassign", "Live"])

    with tab_sched:
        # Optional filter bar
        f1, f2, f3 = st.columns([2,2,2])
        _all_hk  = sorted(set(g.get("housekeeper","") for g in fg if g.get("housekeeper","")))
        _all_rqs = sorted(set(g.get("inspector","")   for g in fg if g.get("inspector","")))
        with f1: _fhk  = st.selectbox("Housekeeper", ["All"]+_all_hk,  key="sch_hk_filter")
        with f2: _frqs = st.selectbox("Inspector (RQS)", ["All"]+_all_rqs, key="sch_rqs_filter")
        with f3: _fsvc = st.selectbox("Service", ["All",SVC_FC,SVC_IH,SVC_DS,SVC_DV], key="sch_svc_filter")

        # ── Build one row per housekeeper, grouped under their RQS (like the
        #    pivot). Verify/blank-housekeeper groups collect under "Unassigned". ─
        by_hk = {}   # hk_name -> {"insp":str, "bld":set, "rooms":[...], "time":int}
        def _room_is_unalloc(r):
            return str(r.get("guest","")).strip().lower() in ("unallocated","---","")
        for g in fg:
            hk   = g.get("housekeeper","") or ("Unassigned" if g.get("verify_group") else "")
            insp = g.get("inspector","")
            svc  = g.get("service_type","")
            if not hk: hk = "Unassigned"
            # Unallocated Full Clean / IH rooms are left for manual assignment (same
            # as the Excel download): show them under "Unassigned" with no RQS,
            # rather than pre-assigning a housekeeper. Dust n Vac keeps RQS 2 and
            # Daily Service keeps its assignment, so only FC/IH are rerouted.
            unalloc_here = svc in (SVC_FC, SVC_IH) and not g.get("verify_group")
            alloc_rooms   = [r for r in g["rooms"] if not (unalloc_here and _room_is_unalloc(r))]
            unalloc_rooms = [r for r in g["rooms"] if      unalloc_here and _room_is_unalloc(r)]

            if alloc_rooms:
                rec = by_hk.setdefault(hk, {"insp":set(), "bld":set(), "rooms":[], "time":0})
                if insp: rec["insp"].add(insp)
                for r in alloc_rooms:
                    rec["bld"].add(r.get("bld",0))
                    rec["rooms"].append({**r, "_svc":svc})
                    rec["time"] += r.get("time",0)
            if unalloc_rooms:
                rec = by_hk.setdefault("Unassigned", {"insp":set(), "bld":set(), "rooms":[], "time":0})
                for r in unalloc_rooms:
                    rec["bld"].add(r.get("bld",0))
                    rec["rooms"].append({**r, "_svc":svc})
                    rec["time"] += r.get("time",0)

        def _keep(hk, rec):
            if _fhk  != "All" and hk != _fhk: return False
            if _frqs != "All" and _frqs not in rec["insp"]: return False
            if _fsvc != "All" and not any(x["_svc"]==_fsvc for x in rec["rooms"]): return False
            return True

        SVC_SHORT = {SVC_FC:"FC", SVC_IH:"IH", SVC_DS:"DS", SVC_DV:"DV"}
        SVC_COL   = {SVC_FC:"#2563a8", SVC_IH:"#7c3aed", SVC_DS:"#0f766e", SVC_DV:"#b45309"}

        def _primary_rqs(rec):
            return sorted(rec["insp"])[0] if rec["insp"] else "Unassigned"

        # Service ranking used to order both the RQS groups and rooms: Full Clean
        # first, then IH, Daily Service, Dust n Vac.
        _SVC_RANK = {SVC_FC:0, SVC_IH:1, SVC_DS:2, SVC_DV:3}
        def _rec_primary_svc_rank(rec):
            # an RQS/HK's "primary" service = the lowest-ranked (most senior)
            # service among the rooms they carry.
            ranks = [_SVC_RANK.get(x["_svc"], 4) for x in rec["rooms"]]
            return min(ranks) if ranks else 4

        # Group housekeepers under their (primary) RQS.
        from collections import OrderedDict as _OD
        rqs_groups = {}
        for hk, rec in by_hk.items():
            if not _keep(hk, rec): continue
            rqs_groups.setdefault(_primary_rqs(rec), []).append((hk, rec))
        # Each RQS group's service rank = the best (lowest) rank among its HKs, so
        # RQS groups that lead Full Clean sort to the top and Daily Service to the
        # bottom. "Unassigned" always sorts last.
        def _rqs_svc_rank(rqs):
            return min((_rec_primary_svc_rank(rec) for _,rec in rqs_groups[rqs]), default=4)
        ordered_rqs = sorted(rqs_groups.keys(),
                             key=lambda r:(r=="Unassigned", _rqs_svc_rank(r), r.lower()))
        # Within each RQS, sort housekeepers by their service (FC first), then
        # building, then name.
        for r in rqs_groups:
            rqs_groups[r].sort(key=lambda kv:(kv[0]=="Unassigned",
                                              _rec_primary_svc_rank(kv[1]),
                                              min(kv[1]["bld"]) if kv[1]["bld"] else 9,
                                              kv[0].lower()))

        def _rooms_cell(rec):
            rooms = rec["rooms"]
            if _fsvc != "All": rooms = [x for x in rooms if x["_svc"]==_fsvc]
            rooms = sorted(rooms, key=lambda x:(x.get("bld",0), x.get("floor",0), x.get("num",0)))
            out = []
            for x in rooms:
                col = SVC_COL.get(x["_svc"], "#475569")
                tm  = x.get("time",0)
                tm_s = f' <span style="color:#9aa4b2;font-size:.66rem">{tm}m</span>' if tm else ""
                out.append(
                    f'<span style="display:inline-block;background:{col}14;color:{col};'
                    f'border:1px solid {col}33;border-radius:6px;padding:2px 8px;margin:2px 3px 2px 0;'
                    f'font-size:.72rem;font-weight:600;white-space:nowrap">'
                    f'{e(x.get("room",""))}{tm_s}</span>')
            return "".join(out) or '<span style="color:#9aa4b2">—</span>'

        def _svc_cell(rec):
            svcs = []
            for s in [SVC_FC,SVC_IH,SVC_DS,SVC_DV]:
                if any(x["_svc"]==s for x in rec["rooms"]):
                    if _fsvc!="All" and s!=_fsvc: continue
                    c=SVC_COL[s]
                    svcs.append(f'<span style="background:{c}14;color:{c};border-radius:5px;'
                                f'padding:1px 7px;font-size:.68rem;font-weight:700;margin-right:3px;'
                                f'white-space:nowrap">{SVC_SHORT[s]}</span>')
            return "".join(svcs) or '<span style="color:#9aa4b2">—</span>'

        def _notes_cell(rec):
            parts = []; seen = set()
            for x in rec["rooms"]:
                n = (x.get("notes","") or "").strip()
                if n and n not in seen:
                    seen.add(n)
                    parts.append(f'<div style="font-size:.7rem;color:#5b6675;line-height:1.5">'
                                 f'<span style="color:#8a93a1">{e(x.get("room",""))}:</span> {e(n)}</div>')
            return "".join(parts) or '<span style="color:#9aa4b2">—</span>'

        def _late_cell(rec):
            parts = []
            for x in sorted(rec["rooms"], key=lambda x:x.get("room","")):
                lc = (x.get("late_checkout","") or "").strip()
                if lc:
                    t = lc.replace("Late Out: ","").replace("Late Out","").strip() or "Late"
                    parts.append(f'<div style="font-size:.7rem;color:#b45309;font-weight:600;'
                                 f'line-height:1.5">{e(x.get("room",""))} · {e(t)}</div>')
            return "".join(parts) or '<span style="color:#9aa4b2">—</span>'

        def _hk_cell(hk, rec):
            blds = "".join(
                f'<span style="background:{BLD_COLORS.get(b,("#888","#eee"))[1]};'
                f'color:{BLD_COLORS.get(b,("#888","#eee"))[0]};border-radius:4px;'
                f'padding:0 6px;font-size:.62rem;font-weight:700;margin-left:5px">B{b}</span>'
                for b in sorted(rec["bld"]) if b)
            low = ""
            if hk!="Unassigned" and rec["time"] and rec["time"]<LOW_MIN:
                low = ('<span style="background:#fff4e5;color:#b45309;border-radius:4px;'
                       'padding:0 6px;font-size:.6rem;font-weight:700;margin-left:5px">LOW</span>')
            nm_col = "#9aa4b2" if hk=="Unassigned" else "#16202e"
            return (f'<span style="font-weight:700;color:{nm_col};font-size:.85rem">{e(hk)}</span>'
                    f'{blds}{low}'
                    f'<div style="font-size:.66rem;color:#8a93a1;margin-top:2px">{rec["time"]}m</div>')

        def _rqs_cell(rqs, span_first):
            if not span_first: return ""   # blank on repeat rows within the RQS group
            col = "#9aa4b2" if rqs=="Unassigned" else "#16202e"
            return f'<span style="font-weight:700;color:{col};font-size:.85rem">{e(rqs)}</span>'

        # ── Render: columns lead with RQS, then Housekeeper, Rooms, Service,
        #    Notes, Late Out (same order as the pivot). ─────────────────────────
        th = ("padding:10px 12px;text-align:left;font-family:'DM Mono',monospace;"
              "font-size:.6rem;font-weight:600;text-transform:uppercase;letter-spacing:.09em;"
              f"color:{_C['txt3']};background:{_C['th_bg']};border-bottom:1px solid {_C['row_br']};"
              "position:sticky;top:0;z-index:2")
        cols   = ["RQS","Housekeeper","Rooms","Service","Notes","Late Out"]
        widths = ["11%","15%","33%","8%","23%","10%"]
        head = "".join(f'<th style="{th};width:{w}">{c}</th>' for c,w in zip(cols,widths))
        body = ""; ri = 0
        for rqs in ordered_rqs:
            members = rqs_groups[rqs]
            for j,(hk,rec) in enumerate(members):
                delay=f"{min(ri*0.02,0.6):.2f}s"; ri += 1
                # top border between RQS groups for visual separation
                grp_top = (f"border-top:2px solid {_C['row_br']};" if j==0 and body else "")
                cells = [_rqs_cell(rqs, j==0), _hk_cell(hk,rec), _rooms_cell(rec),
                         _svc_cell(rec), _notes_cell(rec), _late_cell(rec)]
                tds = "".join(
                    f'<td style="padding:9px 12px;border-bottom:1px solid {_C["row_br"]};{grp_top}'
                    f'vertical-align:top">{c}</td>' for c in cells)
                body += f'<tr style="animation:rowIn .3s {delay} both">{tds}</tr>'
        table_html = f"""<!DOCTYPE html><html><head>{SHARED_CSS}
<style>
tr{{transition:background .15s ease}}
tr:hover td{{background:{_C['th_bg']}!important}}
td{{transition:background .15s ease}}
</style></head><body>
<div style="border-radius:12px;overflow:hidden;border:1px solid {_C['card_br']};
            background:{_C['tbl_bg']};box-shadow:{_C['card_sh']}">
<table style="width:100%;border-collapse:collapse;font-size:.8rem">
  <thead><tr>{head}</tr></thead><tbody>{body}</tbody>
</table></div></body></html>"""
        _row_h = 0
        for rqs in ordered_rqs:
            for hk,rec in rqs_groups[rqs]:
                nrooms = len(rec["rooms"]) if _fsvc=="All" else len([x for x in rec["rooms"] if x["_svc"]==_fsvc])
                _row_h += max(52, 30 + ((nrooms+5)//6)*30)
        components.html(table_html, height=min(max(_row_h+70, 160), 4000), scrolling=True)

        # ── Free / low summary at the END of the table ────────────────────────
        # Housekeepers: low = under LOW_MIN minutes. RQS/inspectors: low = fewer
        # than 8 rooms. Free (present but assigned nothing) shown for both. Each
        # row also shows the service type(s) that person is handling today.
        SVC_SHORT = {SVC_FC:"FC", SVC_IH:"IH", SVC_DS:"DS", SVC_DV:"DV"}
        RQS_LOW_ROOMS = 8

        def _svc_summary(recs_rooms):
            svs = []
            for s in [SVC_FC,SVC_IH,SVC_DS,SVC_DV]:
                if any(x["_svc"]==s for x in recs_rooms):
                    c = SVC_COL[s]
                    svs.append(f'<span style="background:{c}14;color:{c};border-radius:5px;'
                               f'padding:1px 7px;font-size:.66rem;font-weight:700;margin-right:3px">'
                               f'{SVC_SHORT[s]}</span>')
            return "".join(svs) or '<span style="color:#9aa4b2">—</span>'

        # Housekeeper rows
        _hk_low = []   # (name, load_str, load_col, nrooms, svc_html, rqs_str, rooms_names, kind)
        def _room_names(recs_rooms):
            # sorted room codes for display next to the count
            names = [str(x.get("room","")) for x in recs_rooms if x.get("room")]
            return ", ".join(sorted(names))
        for hk, rec in by_hk.items():
            if hk == "Unassigned" or is_unassigned_hk(hk): continue
            if rec["time"] and rec["time"] < LOW_MIN:
                _hk_low.append((hk, f'{rec["time"]}m', "#b45309", len(rec["rooms"]),
                                _svc_summary(rec["rooms"]), " · ".join(sorted(rec["insp"])),
                                _room_names(rec["rooms"]), rec["time"]))
        _hk_low.sort(key=lambda t:t[-1])   # lightest first
        _hk_free = sorted(n for n in present_hk if n not in used_hk_set)

        # RQS/inspector rows — build room counts per inspector from the schedule.
        rqs_rooms = {}   # rqs -> {"n":int, "rooms":[...]}
        for g in fg:
            insp = g.get("inspector","")
            if not insp: continue
            r = rqs_rooms.setdefault(insp, {"n":0, "rooms":[]})
            for rm in g["rooms"]:
                r["n"] += 1
                r["rooms"].append({**rm, "_svc":g.get("service_type","")})
        _rqs_low = []
        for insp, info in rqs_rooms.items():
            if info["n"] < RQS_LOW_ROOMS:
                _rqs_low.append((insp, f'{info["n"]} rm', "#b45309", info["n"],
                                 _svc_summary(info["rooms"]), "",
                                 _room_names(info["rooms"]), info["n"]))
        _rqs_low.sort(key=lambda t:t[-1])
        _rqs_free = sorted(n for n in present_insp if n not in rqs_rooms)

        if n_need_hk:
            st.error(f"**{n_need_hk}** more housekeeper(s) needed for Daily Service — charts "
                     f"labelled *{NEED_HK_PREFIX} 1…{n_need_hk}* have nobody on them. "
                     f"Add more HKs to the Daily Service Team in the sidebar.")

        _any = _hk_low or _hk_free or _rqs_low or _rqs_free
        if _any:
            def _row(name, load, load_col, rooms_txt, svc_html, who, role_tag):
                return (f'<tr>'
                        f'<td style="padding:7px 12px;border-bottom:1px solid {_C["row_br"]};'
                        f'font-weight:600;color:#16202e">{e(name)}'
                        f'<span style="background:#eef0f3;color:#5b6675;border-radius:4px;'
                        f'padding:0 6px;font-size:.58rem;font-weight:700;margin-left:6px;'
                        f'text-transform:uppercase">{role_tag}</span></td>'
                        f'<td style="padding:7px 12px;border-bottom:1px solid {_C["row_br"]};'
                        f'color:{load_col};font-weight:700">{load}</td>'
                        f'<td style="padding:7px 12px;border-bottom:1px solid {_C["row_br"]};'
                        f'color:#5b6675">{rooms_txt}</td>'
                        f'<td style="padding:7px 12px;border-bottom:1px solid {_C["row_br"]}">{svc_html}</td>'
                        f'<td style="padding:7px 12px;border-bottom:1px solid {_C["row_br"]};'
                        f'color:#5b6675">{who}</td></tr>')
            # rooms_txt shows the count and, when present, the actual room codes
            def _rooms_cell(nrm, names):
                base = f'{nrm} room{"s" if nrm!=1 else ""}'
                if names:
                    return (f'{base}<span style="color:#9aa4b2"> — </span>'
                            f'<span style="color:#5b6675">{e(names)}</span>')
                return base
            srows = ""
            for hk, load, col, nrm, svc, insp, names, _ in _hk_low:
                srows += _row(hk, load, col, _rooms_cell(nrm, names), svc,
                              e(insp) or "—", "HK")
            for hk in _hk_free:
                srows += _row(hk, "Free", "#059669", "0 rooms",
                              '<span style="color:#9aa4b2">—</span>', "—", "HK")
            for insp, load, col, nrm, svc, _, names, _n in _rqs_low:
                srows += _row(insp, load, col, _rooms_cell(nrm, names), svc,
                              "—", "RQS")
            for insp in _rqs_free:
                srows += _row(insp, "Free", "#059669", "0 rooms",
                              '<span style="color:#9aa4b2">—</span>', "—", "RQS")
            sth = ("padding:8px 12px;text-align:left;font-family:'DM Mono',monospace;"
                   "font-size:.58rem;font-weight:600;text-transform:uppercase;letter-spacing:.09em;"
                   f"color:{_C['txt3']};background:{_C['th_bg']};border-bottom:1px solid {_C['row_br']}")
            summary_html = f"""<!DOCTYPE html><html><head>{SHARED_CSS}</head><body>
<div style="margin-top:6px;font-family:'DM Mono',monospace;font-size:.62rem;font-weight:600;
            letter-spacing:.09em;text-transform:uppercase;color:{_C['txt3']};margin-bottom:6px">
  Free &amp; Low-Load Staff &nbsp;(HK under {LOW_MIN} min · RQS under {RQS_LOW_ROOMS} rooms)</div>
<div style="border-radius:10px;overflow:hidden;border:1px solid {_C['card_br']};
            background:{_C['tbl_bg']};box-shadow:{_C['card_sh']}">
<table style="width:100%;border-collapse:collapse;font-size:.78rem">
  <thead><tr><th style="{sth};width:26%">Staff</th>
    <th style="{sth};width:14%">Load</th>
    <th style="{sth};width:16%">Rooms</th>
    <th style="{sth};width:22%">Service</th>
    <th style="{sth};width:22%">RQS</th></tr></thead>
  <tbody>{srows}</tbody>
</table></div></body></html>"""
            _nrows = len(_hk_low)+len(_hk_free)+len(_rqs_low)+len(_rqs_free)
            components.html(summary_html, height=min(100 + _nrows*40, 1100), scrolling=True)
        else:
            st.caption("All staff are assigned and at or above their thresholds.")

    # ══════════════════════════════════════════════════════════════════════════════
    # LIVE TAB — Real-time cleaning & inspection tracking
    # ══════════════════════════════════════════════════════════════════════════════
    with tab_reassign:
        st.markdown('<p class="sec">Board — drag a housekeeper to another RQS</p>',
                    unsafe_allow_html=True)
        # Sorting moves the CARDS, never the columns. An inspector's column
        # staying put is what makes the board readable -- when the columns
        # reordered too, the same place on screen held a different person and
        # every number appeared to have changed.
        KB_SORTS = ["Busiest first", "Lightest first", "Most rooms first",
                    "Building", "Name (A-Z)"]
        _sc1, _sc2 = st.columns([2, 3])
        with _sc1:
            kb_sort = st.selectbox("Sort housekeepers by", KB_SORTS,
                                   key="kb_sort")
        st.caption("One column per inspector, always in the same order. Sorting "
                   "reorders the cards inside each column, not the columns. "
                   "Dragging a card takes all of that housekeeper's charts. "
                   "Nothing is saved until you press Apply.")

        # A housekeeper's charts can already be split across inspectors; the card
        # sits under whoever has the bulk and the move takes the rest with it.
        KB_TARGET = 12          # rooms an RQS is meant to inspect in a day
        RQS_KB_LOW = 8          # below this a column is flagged light
        SVC_SHORT = {SVC_FC: "FC", SVC_IH: "IH", SVC_DS: "DS", SVC_DV: "DV"}
        hk_charts, hk_insp = {}, {}
        for _g in fg:
            _hk = _g.get("housekeeper", "")
            if not _hk or _hk == "Manager" or is_unassigned_hk(_hk):
                continue
            hk_charts.setdefault(_hk, []).append(_g)
            if _g.get("inspector"):
                hk_insp.setdefault(_hk, {})
                hk_insp[_hk][_g["inspector"]] = hk_insp[_hk].get(_g["inspector"], 0) + 1

        UNASSIGNED = "— no RQS —"
        _rqs_names = sorted({g.get("inspector", "") for g in fg if g.get("inspector")}
                            | set(present_insp))

        # Per-service fill targets for the load bars. A chart is judged against
        # its own service, not one blanket number: a 420-minute Daily Service is
        # thin, while a 380-minute Full Clean is exactly full.
        KB_FILL = {
            SVC_DS: (460, 400),      # (target, below-this-is-red)
            SVC_FC: (MAX_FC, LOW_MIN),
            SVC_IH: (MAX_FC, LOW_MIN),
            SVC_DV: (MAX_FC, LOW_MIN),
        }

        def _bar(mins, svc, width=6):
            """Load bar for one chart, coloured green through red.

            Drag cards are plain text, so a CSS gradient cannot reach inside
            them — the colour comes from emoji blocks instead. That gives four
            bands rather than a smooth ramp, which is enough to see at a glance
            whether a chart is full, thin, or badly short.
            """
            target, red = KB_FILL.get(svc, (MAX_FC, LOW_MIN))
            if mins >= target:      block = "🟩"      # full
            elif mins >= red:       block = "🟨"      # a little light
            elif mins >= red * 0.85: block = "🟧"     # well short
            else:                   block = "🟥"      # badly short
            filled = max(0, min(width, round(width * mins / max(target, 1))))
            return block * filled + "⬜" * (width - filled)

        def _rooms_of(g, cap=6):
            names = [str(r.get("room", "")).strip() for r in (g.get("rooms") or [])
                     if str(r.get("room", "")).strip()]
            shown = " ".join(names[:cap])
            return shown + (f" +{len(names) - cap}" if len(names) > cap else "")

        def _hk_card(hk, gs):
            blds = sorted({b for g in gs for b in g.get("blds") or set()})
            mins = sum(g["time"] for g in gs)
            rooms = sum(len(g.get("rooms") or []) for g in gs)
            head = (f"{hk}\n"
                    f"B{'/'.join(str(b) for b in blds) or '?'} · {len(gs)} "
                    f"chart{'s' if len(gs) != 1 else ''} · {rooms} rm · {mins}m")
            lines = []
            for g in sorted(gs, key=lambda x: x.get("label", "")):
                svc = g.get("service_type", "")
                # Rooms get their own line. Appended to the bar they wrapped
                # awkwardly beneath it and the card lost its shape.
                lines.append(f"{_bar(g['time'], svc)} {SVC_SHORT.get(svc, '?')} "
                             f"{g['time']}m")
                rooms_txt = _rooms_of(g)
                if rooms_txt:
                    lines.append(f"   {rooms_txt}")
            return head + "\n" + "\n".join(lines)

        def _col_header(name, hks):
            gs = [g for hk in hks for g in hk_charts.get(hk, [])]
            rooms = sum(len(g.get("rooms") or []) for g in gs)
            mins = sum(g["time"] for g in gs)
            role = ""
            if name == st.session_state.get("rqs1"): role = " · RQS 1"
            elif name == st.session_state.get("rqs2"): role = " · RQS 2"
            warn = ""
            if name != UNASSIGNED and rooms and rooms < RQS_KB_LOW:
                warn = "  ⚠ light"
            elif rooms > KB_TARGET + 4:
                warn = "  ⚠ heavy"
            return (f"{name}{role}\n"
                    f"{len(hks)} HK · {len(gs)} chart{'s' if len(gs) != 1 else ''} · {rooms} rm · {mins}m{warn}")

        home_of = {}
        for _hk, _gs in hk_charts.items():
            here = hk_insp.get(_hk) or {}
            home_of[_hk] = max(here, key=here.get) if here else UNASSIGNED

        def _hk_mins(hk):
            return sum(g["time"] for g in hk_charts.get(hk, []))

        def _hk_bld(hk):
            blds = sorted({b for g in hk_charts.get(hk, []) for b in g.get("blds") or set()})
            return blds[0] if blds else 99

        def _is_ds_hk(hk):
            return any(g.get("service_type") == SVC_DS
                       for g in hk_charts.get(hk, []))

        def _hk_rooms(hk):
            return sum(len(g.get("rooms") or []) for g in hk_charts.get(hk, []))

        def _card_order(hks, ds_last=False):
            """Order housekeepers by whichever sort is showing.

            `ds_last` pushes Daily Service rounds to the end regardless: on the
            room board those columns are far wider than the rest, so they sit
            at the bottom out of the way. The main board honours the sort
            literally, because a sort that quietly overrules itself is exactly
            what makes the numbers look wrong.
            """
            if kb_sort == "Building":
                key = lambda h: (_hk_bld(h), h.lower())         # noqa: E731
            elif kb_sort == "Name (A-Z)":
                key = lambda h: (h.lower(),)                    # noqa: E731
            elif kb_sort == "Lightest first":
                key = lambda h: (_hk_mins(h), h.lower())        # noqa: E731
            elif kb_sort == "Most rooms first":
                key = lambda h: (-_hk_rooms(h), h.lower())      # noqa: E731
            else:
                key = lambda h: (-_hk_mins(h), h.lower())       # noqa: E731
            if ds_last:
                return sorted(hks, key=lambda h: (_is_ds_hk(h),) + key(h))
            return sorted(hks, key=key)

        def _col_order(names):
            """A fixed column order: the two RQS on duty, then everyone else
            by name, then the holding pen.

            Deliberately independent of the sort. Columns that reshuffle when
            you change how the cards are ordered make the board unreadable --
            you look at the third column expecting the person who was there a
            moment ago.
            """
            r1 = st.session_state.get("rqs1")
            r2 = st.session_state.get("rqs2")
            lead = [n for n in (r1, r2) if n and n in names]
            rest = sorted((n for n in names
                           if n != UNASSIGNED and n not in lead), key=str.lower)
            return lead + rest + [UNASSIGNED]

        board, card_to_hk = {}, {}
        for name in _col_order(_rqs_names + [UNASSIGNED]):
            board[name] = _card_order([hk for hk, h in home_of.items() if h == name])
        containers = []
        for name, hks in board.items():
            items = []
            for hk in hks:
                card = _hk_card(hk, hk_charts[hk])
                card_to_hk[card] = hk
                items.append(card)
            containers.append({"header": _col_header(name, hks), "items": items})
        header_to_name = {c["header"]: n for c, n in zip(containers, board.keys())}

        def _board_key(prefix, conts):
            """A widget key that changes only when the board's contents do.

            streamlit-sortables seeds its React state from the items prop on
            mount and never syncs it again, so a board re-rendered with new
            data keeps showing the old cards -- a room moved away downstairs
            still appeared under its old housekeeper up here. Folding the
            contents into the key remounts the component when, and only when,
            the charts actually changed. Dragging alone leaves the charts
            untouched, so a drag in progress is not thrown away.
            """
            sig = repr([(c["header"], c["items"]) for c in conts]).encode("utf-8")
            return f"{prefix}_{_hashlib.sha1(sig).hexdigest()[:10]}"

        KANBAN_CSS = """
        /* The component only ships a flex layout for its "vertical" variant —
           the horizontal one has no layout rule at all, so the columns fall
           back to block and stack. Lay them out here instead: a grid that
           wraps, so a dozen inspectors fill the width rather than running off
           the side. */
        .sortable-component{display:grid !important;
            grid-template-columns:repeat(auto-fill,minmax(232px,1fr));
            gap:12px;align-items:start;background:transparent;
            font-family:'DM Sans',sans-serif;padding-bottom:4px}
        .sortable-container{background:#eef2f7;border:1px solid #e0e6ee;
            border-radius:16px;padding:9px;margin:0 !important;
            min-width:0 !important;max-width:none;width:100%}
        .sortable-container-header{
            background:linear-gradient(135deg,#1b4a80 0%,#2d72b8 55%,#4b93d1 100%);
            color:#fff;border-radius:11px;padding:9px 12px;
            font-size:.73rem;font-weight:700;line-height:1.5;
            white-space:pre-line;letter-spacing:.01em;
            box-shadow:0 3px 10px rgba(27,74,128,.24)}
        .sortable-container-body{min-height:76px;padding:9px 0 0 !important;
            background:transparent !important}
        .sortable-item{background:#fff !important;border:1px solid #e0e6ee;
            border-radius:12px;height:auto !important;
            padding:10px 12px !important;margin:0 0 8px !important;
            font-size:.74rem;line-height:1.55;
            color:#26313f !important;text-align:left;white-space:pre-line;
            font-weight:500;
            box-shadow:0 1px 2px rgba(16,26,42,.05);cursor:grab;width:100%;
            transition:box-shadow .15s ease,border-color .15s ease,transform .1s ease}
        .sortable-item:first-line{font-weight:800;font-size:.86rem;color:#16202e}
        /* The component paints .sortable-item AND .sortable-item:hover with
           color:#fff. Without !important the card text turns white on hover
           and disappears against the white card, leaving only the emoji bar. */
        .sortable-item:hover{background:#fff !important;color:#26313f !important;
            border-color:#9fc0e0;transform:translateY(-2px);
            box-shadow:0 8px 20px rgba(37,99,168,.16)}
        .sortable-item:active,.sortable-item.dragging{
            background:#fff !important;color:#26313f !important}
        .sortable-item.dragging{opacity:.9;cursor:grabbing;
            box-shadow:0 14px 30px rgba(16,26,42,.22);border-color:#2d72b8}
        """

        moved = None
        try:
            from streamlit_sortables import sort_items
            after = sort_items(containers, multi_containers=True,
                               direction="horizontal", custom_style=KANBAN_CSS,
                               key=_board_key("reassign_board", containers))
            moved = {header_to_name.get(c["header"], c["header"]): c["items"]
                     for c in after}
        except Exception as _ex:
            st.info("Drag-and-drop could not load, so use the pickers below — they do "
                    f"the same thing. ({_ex})")
            moved = {}
            mc = st.columns(4)
            for i, (_hk, _gs) in enumerate(sorted(hk_charts.items())):
                opts = [UNASSIGNED] + _rqs_names
                home = home_of.get(_hk, UNASSIGNED)
                with mc[i % 4]:
                    pick = st.selectbox(
                        f'{_hk} · {len(_gs)} charts',
                        opts, index=opts.index(home) if home in opts else 0,
                        key=f"reass_{_hk}")
                card = _hk_card(_hk, _gs)
                card_to_hk[card] = _hk
                moved.setdefault(pick, []).append(card)

        # What would change if this were applied?
        pending = {}
        for target, cards in (moved or {}).items():
            for card in cards:
                _hk = card_to_hk.get(card)
                if not _hk:
                    continue
                new_insp = "" if target == UNASSIGNED else target
                for _g in hk_charts.get(_hk, []):
                    if (_g.get("inspector") or "") != new_insp:
                        pending.setdefault(_hk, (set(), new_insp))
                        pending[_hk][0].add(_g.get("inspector") or UNASSIGNED)

        if pending:
            rows = "<br>".join(
                f'&nbsp;&nbsp;<b>{e(hk)}</b>: {e(" / ".join(sorted(froms)))} '
                f'<span style="opacity:.6">→</span> {e(to or UNASSIGNED)}'
                for hk, (froms, to) in sorted(pending.items()))
            st.markdown(f'<div style="background:#eef4fb;border:1px solid #cddff0;'
                        f'border-radius:8px;padding:10px 13px;font-size:.79rem;'
                        f'color:#1c4a78">{len(pending)} change(s) waiting:<br>{rows}</div>',
                        unsafe_allow_html=True)
            if st.button("Apply these moves", type="primary", key="btn_apply_moves"):
                n = 0
                for hk, (_froms, to) in pending.items():
                    for _g in hk_charts.get(hk, []):
                        _g["inspector"] = to
                        n += 1
                _save_reassignment(fg)
                st.success(f"Moved {len(pending)} housekeeper(s), {n} chart(s) reassigned.")
                st.rerun()
        else:
            st.caption("No moves pending.")

        # ── Rooms between housekeepers ────────────────────────────────────────
        st.markdown('<p class="sec">Move single rooms</p>', unsafe_allow_html=True)
        st.caption("Drag a room to any housekeeper on the board, including one "
                   "under a different RQS. Times, buildings and floors are "
                   "recalculated for both charts. Nothing is saved until you "
                   "press Apply.")

        def _rechart(g):
            """Recompute a chart's totals after its rooms changed."""
            g["time"] = sum(r.get("time", 0) for r in g["rooms"])
            g["blds"] = {r.get("bld") for r in g["rooms"] if r.get("bld")} or set()
            g["floors"] = {r.get("floor", 0) for r in g["rooms"]}
            g["c140"] = sum(1 for r in g["rooms"] if r.get("time") == 140)
            g["c120"] = sum(1 for r in g["rooms"] if r.get("time") == 120)
            g["cross_bld"] = len(g["blds"]) > 1

        _SVC_PREFIX = {SVC_FC: "FC", SVC_IH: "IH", SVC_DS: "DS", SVC_DV: "DV"}

        def _free_label(svc):
            """A label of the right family that no chart is using yet."""
            used = {g.get("label") for g in fg}
            pre = _SVC_PREFIX.get(svc, "FC")
            for lbl in make_labels(pre, len(fg) + 27):
                if lbl not in used:
                    return lbl
            return f"{pre}-{len(fg) + 1}"

        def _move_room(code, src_hk, dst_hk):
            """Move one room across housekeepers, or return why it could not."""
            src = next((g for g in fg if g.get("housekeeper") == src_hk
                        and any(str(r.get("room")) == code for r in g["rooms"])), None)
            if src is None:
                return f"{code}: no longer on {src_hk}"
            room = next(r for r in src["rooms"] if str(r.get("room")) == code)
            svc = src.get("service_type")
            cands = [g for g in fg if g.get("housekeeper") == dst_hk
                     and not g.get("verify_group")]
            dst = next((g for g in cands if g.get("service_type") == svc), None)
            if dst is None:
                # The service belongs to the room, not to whoever cleans it.
                # Dropping a Full Clean into the only chart a Daily Service
                # person has would relabel the room as a Daily Service and
                # cost it three hours of its time. Open a second chart of the
                # right service for them instead -- someone can carry a DS
                # round and a Full Clean room, and the board should say so.
                dst = mk([], svc)
                dst.update({"label": _free_label(svc), "housekeeper": dst_hk,
                            "inspector": (cands[0].get("inspector")
                                          if cands else "")})
                fg.append(dst)
            src["rooms"].remove(room)
            dst["rooms"].append(room)
            _rechart(src); _rechart(dst)
            return None

        # Every housekeeper on the property is a column, whoever inspects
        # them - a room is just as likely to belong next door under another
        # RQS as under its own. The filter narrows a crowded board; it does
        # not fence the move.
        rb1, rb2 = st.columns([3, 2])
        with rb1:
            room_scope = st.multiselect(
                "Inspectors shown", _rqs_names + [UNASSIGNED],
                default=_rqs_names + [UNASSIGNED], key="room_scope")
        with rb2:
            room_hide_empty = st.checkbox("Hide empty housekeepers", value=True,
                                          key="room_hide_empty")
        scope = set(room_scope) or set(_rqs_names) | {UNASSIGNED}

        scope_hks = _card_order([hk for hk in hk_charts
                                 if home_of.get(hk, UNASSIGNED) in scope
                                 and not is_unassigned_hk(hk)], ds_last=True)
        if not scope_hks:
            st.caption("No housekeepers under the inspectors you picked.")
        else:
            room_home, room_cards = {}, {}
            for _hk in scope_hks:
                items = []
                for g in sorted(hk_charts.get(_hk, []),
                                key=lambda x: x.get("label", "")):
                    for r in sorted(g["rooms"],
                                    key=lambda x: str(x.get("room", ""))):
                        code = str(r.get("room", ""))
                        if not code:
                            continue
                        card = (f"{code}\n"
                                f"{SVC_SHORT.get(g.get('service_type', ''), '?')} · "
                                f"{r.get('time', 0)}m"
                                + (" · 🐾" if str(r.get("pet", "")).strip() else ""))
                        room_cards[card] = (code, _hk)
                        items.append(card)
                if items or not room_hide_empty:
                    room_home[_hk] = items

            def _room_header(hk, items):
                mins = sum(g["time"] for g in hk_charts.get(hk, []))
                insp = home_of.get(hk, UNASSIGNED)
                blds = sorted({b for g in hk_charts.get(hk, [])
                               for b in g.get("blds") or set()})
                return (f"{hk}\n{insp if insp != UNASSIGNED else 'no RQS'} · "
                        f"B{'/'.join(str(b) for b in blds) or '?'}\n"
                        f"{len(items)} rm · {mins}m")

            # Written out rather than derived from KANBAN_CSS by a string
            # replace: that would silently no-op if the other block's wording
            # ever changed, and these columns would quietly stack again.
            ROOM_CSS = """
            .sortable-component{display:grid !important;
                grid-template-columns:repeat(auto-fill,minmax(166px,1fr));
                gap:10px;align-items:start;background:transparent;
                font-family:'DM Sans',sans-serif;padding-bottom:4px}
            .sortable-container{background:#eef2f7;border:1px solid #e0e6ee;
                border-radius:14px;padding:7px;margin:0 !important;
                min-width:0 !important;max-width:none;width:100%}
            .sortable-container-header{
                background:linear-gradient(135deg,#245c8f 0%,#3a83c4 60%,#5aa0d8 100%);
                color:#fff;border-radius:10px;padding:7px 10px;
                font-size:.68rem;font-weight:700;line-height:1.45;
                white-space:pre-line;letter-spacing:.01em;
                box-shadow:0 3px 9px rgba(27,74,128,.22)}
            .sortable-container-body{min-height:56px;padding:8px 0 0 !important;
                background:transparent !important}
            .sortable-item{background:#fff !important;border:1px solid #e0e6ee;
                border-radius:9px;height:auto !important;
                padding:7px 9px !important;margin:0 0 6px !important;
                font-size:.71rem;line-height:1.45;
                color:#26313f !important;text-align:left;white-space:pre-line;
                font-weight:500;box-shadow:0 1px 2px rgba(16,26,42,.05);
                cursor:grab;width:100%;
                transition:box-shadow .15s ease,border-color .15s ease,transform .1s ease}
            .sortable-item:first-line{font-weight:800;font-size:.79rem;color:#16202e}
            .sortable-item:hover{border-color:#9fc0e0;transform:translateY(-1px);
                color:#26313f !important;
                box-shadow:0 6px 16px rgba(37,99,168,.15)}
            .sortable-item:active,.sortable-item.dragging{opacity:.95;
                cursor:grabbing;color:#26313f !important;border-color:#2d72b8;
                box-shadow:0 12px 26px rgba(16,26,42,.2)}
            """

            # Those wide Daily Service columns get a full-width box with the
            # rooms flowing across it, rather than one 166px column running
            # far below everything else. The component gives containers no
            # per-column class, so they are addressed by position -- as
            # div:nth-of-type, which skips the <style> tag the component
            # renders as its own first child.
            ds_cols = [i + 1 for i, hk in enumerate(room_home) if _is_ds_hk(hk)]
            if ds_cols:
                def _sel(suffix=""):
                    return ",".join(
                        f".sortable-component > div:nth-of-type({i}){suffix}"
                        for i in ds_cols)
                ROOM_CSS += f"""
                {_sel()}{{grid-column:1/-1;width:100% !important;
                    background:#e8f0f9;border-color:#cbdff2}}
                {_sel(" .sortable-container-header")}{{
                    background:linear-gradient(135deg,#1d6b52 0%,#2f9169 60%,#46b184 100%);
                    box-shadow:0 3px 9px rgba(29,107,82,.22)}}
                {_sel(" .sortable-container-body")}{{display:flex !important;
                    flex-wrap:wrap;gap:6px;min-height:52px;align-content:flex-start}}
                {_sel(" .sortable-item")}{{width:auto !important;
                    min-width:104px;margin:0 !important;flex:0 0 auto}}
                """

            r_moved = None
            try:
                from streamlit_sortables import sort_items
                _room_conts = [{"header": _room_header(hk, v), "items": v}
                               for hk, v in room_home.items()]
                r_after = sort_items(
                    _room_conts,
                    multi_containers=True, direction="horizontal",
                    custom_style=ROOM_CSS,
                    # Keyed on the contents, not merely the column set: two
                    # housekeepers that both keep rooms leave the columns
                    # unchanged, and the board would otherwise not notice the
                    # room that moved between them.
                    key=_board_key("room_board", _room_conts))
                r_moved = {c["header"].split("\n")[0]: c["items"] for c in r_after}
            except Exception as _ex:
                st.info(f"Drag-and-drop unavailable: {_ex}")
                r_moved = None

            if r_moved:
                room_pending = []
                for dst_hk, cards in r_moved.items():
                    for card in cards:
                        got = room_cards.get(card)
                        if got and got[1] != dst_hk:
                            room_pending.append((got[0], got[1], dst_hk))
                if room_pending:
                    rows = "<br>".join(
                        f'&nbsp;&nbsp;<b>{e(c)}</b>: {e(a)} '
                        f'<span style="opacity:.6">→</span> {e(b)}'
                        + ('  <span style="opacity:.6">(across RQS)</span>'
                           if home_of.get(a) != home_of.get(b) else '')
                        for c, a, b in room_pending)
                    st.markdown(f'<div style="background:#eef4fb;border:1px solid #cddff0;'
                                f'border-radius:8px;padding:10px 13px;font-size:.79rem;'
                                f'color:#1c4a78">{len(room_pending)} room move(s) waiting:'
                                f'<br>{rows}</div>', unsafe_allow_html=True)
                    if st.button("Apply room moves", type="primary", key="btn_apply_rooms"):
                        problems = [p for p in
                                    (_move_room(c, a, b) for c, a, b in room_pending) if p]
                        # A chart emptied by the move is dropped, the same as
                        # when the schedule is first built.
                        st.session_state["groups_data"] = [g for g in fg if g["rooms"]]
                        _save_reassignment(st.session_state["groups_data"])
                        if problems:
                            st.warning("Some rooms could not move:\n\n"
                                       + "\n\n".join(f"- {p}" for p in problems))
                        st.success(f"Moved {len(room_pending) - len(problems)} room(s).")
                        st.rerun()
                else:
                    st.caption("No room moves pending.")

        # ── Gaps: charts nobody is on ─────────────────────────────────────────
        st.markdown('<p class="sec">Charts still needing someone</p>',
                    unsafe_allow_html=True)
        need_hk_charts = [g for g in fg
                          if is_unassigned_hk(g.get("housekeeper", ""))
                          and not g.get("verify_group") and not g.get("dv_rqs2")]
        need_rqs_charts = [g for g in fg
                           if not g.get("inspector") and not g.get("verify_group")
                           and not g.get("dv_rqs2")]

        if not need_hk_charts and not need_rqs_charts:
            st.success("Every chart has a housekeeper and an RQS.")
        else:
            free_hk = [n for n in present_hk if n not in used_hk_set]
            gc1, gc2 = st.columns(2)
            with gc1:
                st.markdown(f'<div class="mono">NEEDS A HOUSEKEEPER · '
                            f'{len(need_hk_charts)}</div>', unsafe_allow_html=True)
                if not need_hk_charts:
                    st.caption("None.")
                for _g in need_hk_charts:
                    lbl = _g["label"]
                    opts = ["— leave unassigned —"] + (free_hk or []) + \
                           [n for n in present_hk if n not in free_hk]
                    pick = st.selectbox(
                        f'{lbl} · {_g.get("service_type","")} · '
                        f'Bldg {",".join(str(b) for b in sorted(_g["blds"]))} · {_g["time"]}m',
                        opts, key=f"fillhk_{lbl}")
                    if pick != opts[0] and st.button("Assign", key=f"fillhk_btn_{lbl}"):
                        _g["housekeeper"] = pick
                        st.session_state["used_hk_set"] = (
                            set(st.session_state.get("used_hk_set") or set()) | {pick})
                        _save_reassignment(fg)
                        st.success(f"{lbl} → {pick}")
                        st.rerun()
            with gc2:
                st.markdown(f'<div class="mono">NEEDS AN RQS · '
                            f'{len(need_rqs_charts)}</div>', unsafe_allow_html=True)
                if not need_rqs_charts:
                    st.caption("None.")
                for _g in need_rqs_charts:
                    lbl = _g["label"]
                    opts = ["— leave unassigned —"] + sorted(set(present_insp) | set(_rqs_names))
                    pick = st.selectbox(
                        f'{lbl} · {_g.get("service_type","")} · '
                        f'{_g.get("housekeeper","") or "no housekeeper"}',
                        opts, key=f"fillrqs_{lbl}")
                    if pick != opts[0] and st.button("Assign", key=f"fillrqs_btn_{lbl}"):
                        _g["inspector"] = pick
                        _save_reassignment(fg)
                        st.success(f"{lbl} → {pick}")
                        st.rerun()

    with tab_live:
        _tmsg = st.session_state.pop("_live_toast", None)
        if _tmsg:
            try: st.toast(_tmsg)
            except Exception: pass
        import json as _json
        _NOW = _now_iso # shared Mountain-time timestamp helper

        # ── Status pipeline definition ─────────────────────────────────────────
        STATUS_FLOW = ["pending","already_clean","cleaning_started","cleaning_done","inspected"]
        STATUS_META = {
            "pending": {"icon":"","label":"Pending","color":"#334155","bg":"rgba(51,65,85,.25)","border":"rgba(71,85,105,.4)"},
            "already_clean": {"icon":"","label":"Already Clean","color":"#34d399","bg":"rgba(52,211,153,.12)","border":"rgba(52,211,153,.35)"},
            "cleaning_started":{"icon":"","label":"Cleaning...","color":"#fbbf24","bg":"rgba(251,191,36,.12)","border":"rgba(251,191,36,.35)"},
            "cleaning_done": {"icon":"","label":"Cleaned","color":"#60a5fa","bg":"rgba(96,165,250,.12)","border":"rgba(96,165,250,.35)"},
            "inspected": {"icon":"","label":"Inspected ","color":"#a78bfa","bg":"rgba(167,139,250,.15)","border":"rgba(167,139,250,.4)"},
        }

        # ── Load / init room statuses ──────────────────────────────────────────
        if "room_statuses" not in st.session_state:
            st.session_state["room_statuses"] = {}

        def _load_statuses():
            try:
                st.session_state["room_statuses"] = db.get_room_statuses()
            except Exception:
                pass # table may not exist yet — use session-only tracking

        def _save_status(room, fields):
            rs = st.session_state["room_statuses"]
            if room not in rs:
                rs[room] = {"room": room, "status": "pending"}
            rs[room].update(fields)
            rs[room]["updated_by"] = st.session_state.get("username","?")
            try:
                db.upsert_room_status(room, fields | {
                    "updated_by": st.session_state.get("username","?")
                })
            except Exception:
                pass # persist in session even if DB fails

        def _init_statuses_from_schedule():
            """Pre-populate room_statuses from the schedule if not in DB."""
            rs = st.session_state["room_statuses"]
            for g in fg:
                for r in g["rooms"]:
                    rm = r["room"]
                    if rm not in rs:
                        rs[rm] = {
                            "room": rm,
                            "status": "pending",
                            "group_label": g.get("label",""),
                            "housekeeper": g.get("housekeeper",""),
                            "inspector": g.get("inspector",""),
                            "svc": r.get("service",""),
                            "guest": r.get("guest",""),
                            "pet": r.get("pet",""),
                            "late": r.get("late_checkout",""),
                            "bld": r.get("bld",""),
                        }
                    else:
                        # Always keep schedule info in sync, EXCEPT housekeeper
                        # if the room was swapped (swapped_from set) — keep DB HK.
                        rs[rm]["group_label"] = g.get("label","")
                        if not rs[rm].get("swapped_from"):
                            rs[rm]["housekeeper"] = g.get("housekeeper","")
                        rs[rm]["inspector"] = g.get("inspector","")
                        rs[rm]["svc"] = r.get("service","")
                        rs[rm]["guest"] = r.get("guest","")
                        rs[rm]["pet"] = r.get("pet","")
                        rs[rm]["late"] = r.get("late_checkout","")
                        rs[rm]["bld"] = r.get("bld","")

        # Load from DB on first visit to this tab, then init from schedule
        if not st.session_state.get("_live_loaded"):
            st.session_state["_live_loaded"] = True
            _load_statuses()
        _init_statuses_from_schedule()

        rs = st.session_state["room_statuses"]

        # ── Header bar ────────────────────────────────────────────────────────
        # Progress summary — only count non-DV rooms consistently
        _live_rooms = [r for r in rs.values() if r.get("svc","") != "Dust n Vac"]
        total_rooms_live = len(_live_rooms)
        n_clean = sum(1 for r in _live_rooms if r.get("status") in ("already_clean","cleaning_done","inspected"))
        n_insp = sum(1 for r in _live_rooms if r.get("status") == "inspected")
        n_active = sum(1 for r in _live_rooms if r.get("status") == "cleaning_started")
        pct_done = int(n_clean / max(total_rooms_live,1) * 100)

        st.markdown(f"""
    <div style="background:rgba(99,102,241,.08);border:1px solid rgba(99,102,241,.2);
                border-radius:12px;padding:14px 18px;margin-bottom:16px;
                display:flex;align-items:center;gap:20px;flex-wrap:wrap">
      <div style="flex:1;min-width:200px">
        <div style="font-family:'DM Mono',monospace;font-size:.6rem;color:#475569;
                    text-transform:uppercase;letter-spacing:.1em;margin-bottom:6px">Overall Progress</div>
        <div style="background:rgba(255,255,255,.06);border-radius:99px;height:8px;overflow:hidden;border:1px solid rgba(255,255,255,.05)">
          <div style="background:linear-gradient(90deg,#6366f1,#22d3ee);width:{pct_done}%;height:8px;
                      border-radius:99px;box-shadow:0 0 8px rgba(99,102,241,.5);transition:width .4s"></div>
        </div>
        <div style="font-family:'DM Mono',monospace;font-size:.68rem;color:#94a3b8;margin-top:4px">
          {n_clean}/{total_rooms_live} rooms done &nbsp;·&nbsp; {pct_done}%
        </div>
      </div>
      <div style="display:flex;gap:12px;flex-wrap:wrap">
        <div style="text-align:center;background:rgba(251,191,36,.1);border:1px solid rgba(251,191,36,.25);
                    border-radius:8px;padding:8px 14px">
          <div style="font-family:'Syne',sans-serif;font-size:1.4rem;font-weight:800;color:#fbbf24;
                      text-shadow:0 0 10px rgba(251,191,36,.4)">{n_active}</div>
          <div style="font-size:.6rem;color:#94a3b8;font-family:'DM Mono',monospace;text-transform:uppercase;letter-spacing:.06em">Active</div>
        </div>
        <div style="text-align:center;background:rgba(96,165,250,.1);border:1px solid rgba(96,165,250,.25);
                    border-radius:8px;padding:8px 14px">
          <div style="font-family:'Syne',sans-serif;font-size:1.4rem;font-weight:800;color:#60a5fa;
                      text-shadow:0 0 10px rgba(96,165,250,.4)">{n_clean - n_insp}</div>
          <div style="font-size:.6rem;color:#94a3b8;font-family:'DM Mono',monospace;text-transform:uppercase;letter-spacing:.06em">Awaiting Insp</div>
        </div>
        <div style="text-align:center;background:rgba(167,139,250,.1);border:1px solid rgba(167,139,250,.25);
                    border-radius:8px;padding:8px 14px">
          <div style="font-family:'Syne',sans-serif;font-size:1.4rem;font-weight:800;color:#a78bfa;
                      text-shadow:0 0 10px rgba(167,139,250,.4)">{n_insp}</div>
          <div style="font-size:.6rem;color:#94a3b8;font-family:'DM Mono',monospace;text-transform:uppercase;letter-spacing:.06em">Inspected</div>
        </div>
      </div>
      <div style="display:flex;gap:8px;align-items:center">
    """, unsafe_allow_html=True)

        lv1, lv2 = st.columns([1,1])
        with lv1:
            if st.button("Refresh", key="live_refresh", use_container_width=True):
                # Pull fresh status records from DB and merge onto schedule
                try:
                    fresh = db.get_room_statuses()
                    cur_rs = st.session_state.get("room_statuses", {})
                    for rm_code, rec in fresh.items():
                        if rm_code in cur_rs:
                            # keep schedule metadata, take DB status + timestamps
                            cur_rs[rm_code].update({
                                k: rec.get(k) for k in
                                ("status","started_at","cleaned_at","inspected_at",
                                 "marked_clean_at","housekeeper","swapped_from")
                                if rec.get(k) is not None
                            })
                        else:
                            cur_rs[rm_code] = rec
                    st.session_state["room_statuses"] = cur_rs
                except Exception:
                    pass
                _init_statuses_from_schedule()
                st.rerun()
        with lv2:
            if st.button("Reset All", key="live_reset", use_container_width=True):
                for rm in rs:
                    if rs[rm].get("status") != "pending":
                        rs[rm]["status"] = "pending"
                        for f in ["started_at","cleaned_at","inspected_at","marked_clean_at"]:
                            rs[rm].pop(f, None)
                st.rerun()

        st.markdown("</div></div>", unsafe_allow_html=True)

        # ── View filter ───────────────────────────────────────────────────────
        lf1, lf2, lf3, lf4 = st.columns([2,2,2,1])
        with lf1:
            live_view_insp = st.selectbox("Inspector",
                ["All"] + sorted(set(r.get("inspector","") for r in rs.values() if r.get("inspector","")), ),
                key="live_insp_filter")
        with lf2:
            live_view_hk = st.selectbox("Housekeeper",
                ["All"] + sorted(set(r.get("housekeeper","") for r in rs.values() if r.get("housekeeper","")), ),
                key="live_hk_filter")
        with lf3:
            live_status_filter = st.selectbox("Status",
                ["All"] + list(STATUS_META.keys()),
                key="live_status_filter",
                format_func=lambda s: f"{STATUS_META[s]['icon']} {STATUS_META[s]['label']}" if s != "All" else "All Statuses")
        with lf4:
            show_dv = st.checkbox("Show DV", value=False, key="live_show_dv")

        # ── Group rooms by inspector batch ────────────────────────────────────
        # Build inspector groups rooms mapping
        insp_groups: dict = {}
        for g in fg:
            if g.get("service_type") == "Dust n Vac" and not show_dv:
                continue
            insp_name = g.get("inspector","—")
            if live_view_insp != "All" and insp_name != live_view_insp: continue
            if insp_name not in insp_groups:
                insp_groups[insp_name] = []
            insp_groups[insp_name].append(g)

        if not insp_groups:
            st.markdown('<div style="text-align:center;padding:40px;color:#334155;font-family:\'DM Mono\',monospace">No groups match the current filters</div>', unsafe_allow_html=True)
        else:
            # Render each inspector section
            for insp_name, groups in sorted(insp_groups.items()):
                # Inspector header
                insp_rooms_all = [r for g in groups for r in g["rooms"]]
                insp_done = sum(1 for r in insp_rooms_all if rs.get(r["room"],{}).get("status") in ("already_clean","cleaning_done","inspected"))
                insp_pct = int(insp_done / max(len(insp_rooms_all),1) * 100)

                st.markdown(f"""
    <div style="margin:16px 0 8px;display:flex;align-items:center;gap:10px">
      <div style="font-family:'Syne',sans-serif;font-weight:700;font-size:.95rem;color:#a5b4fc">{insp_name}</div>
      <div style="flex:1;background:rgba(255,255,255,.05);border-radius:99px;height:4px;overflow:hidden">
        <div style="background:linear-gradient(90deg,#6366f1,#22d3ee);width:{insp_pct}%;height:4px;border-radius:99px"></div>
      </div>
      <div style="font-family:'DM Mono',monospace;font-size:.65rem;color:#475569">{insp_done}/{len(insp_rooms_all)}</div>
    </div>""", unsafe_allow_html=True)

                # Render each group as a compact card
                for _gidx, g in enumerate(groups):
                    hk_name = g.get("housekeeper","—")
                    g_label = g.get("label","—")
                    _gk = f"{g_label}_{_gidx}" # unique per-group key prefix

                    # Filter by HK
                    rooms_in_g = g["rooms"]
                    if live_view_hk != "All":
                        rooms_in_g = [r for r in rooms_in_g if rs.get(r["room"],{}).get("housekeeper","") == live_view_hk or g.get("housekeeper","") == live_view_hk]
                    if not rooms_in_g: continue

                    # Filter by status
                    if live_status_filter != "All":
                        rooms_in_g = [r for r in rooms_in_g if rs.get(r["room"],{}).get("status","pending") == live_status_filter]
                    if not rooms_in_g: continue

                    g_done = sum(1 for r in g["rooms"] if rs.get(r["room"],{}).get("status") in ("already_clean","cleaning_done","inspected"))
                    g_pct = int(g_done / max(len(g["rooms"]),1) * 100)
                    g_color = "#6366f1" if g.get("service_type")==SVC_FC else ("#14b8a6" if g.get("service_type")==SVC_DS else "#f59e0b")

                    with st.expander(f"{g_label} · {hk_name} · {g_done}/{len(g['rooms'])} done", expanded=(g_pct < 100)):

                        # ── Room swap UI ─────────────────────────────────────
                        swap_col, _ = st.columns([3,1])
                        with swap_col:
                            # Collect all HKs in same inspector batch for swap target
                            batch_hks = sorted(set(
                                gg.get("housekeeper","") for gg in fg
                                if gg.get("inspector","") == insp_name
                                and not is_unassigned_hk(gg.get("housekeeper",""))
                            ))
                            if len(batch_hks) > 1:
                                st.markdown(f'<div style="font-family:\'DM Mono\',monospace;font-size:.62rem;color:#475569;text-transform:uppercase;letter-spacing:.08em;margin-bottom:4px"> Swap rooms to:</div>', unsafe_allow_html=True)
                                swap_target = st.selectbox(
                                    "Swap to HK", ["— select HK —"] + [h for h in batch_hks if h != hk_name],
                                    key=f"swap_target_{_gk}", label_visibility="collapsed"
                                )

                                if swap_target and swap_target != "— select HK —":
                                    # Show swappable rooms (pending or cleaning_started only)
                                    swappable = [r for r in g["rooms"] if rs.get(r["room"],{}).get("status","pending") in ("pending","cleaning_started")]
                                    if swappable:
                                        swap_rooms_sel = st.multiselect(
                                            "Rooms to swap",
                                            [r["room"] for r in swappable],
                                            key=f"swap_rooms_{_gk}",
                                            label_visibility="collapsed",
                                            placeholder="Select rooms to move..."
                                        )
                                        if swap_rooms_sel and st.button(f"Move {len(swap_rooms_sel)} room(s) {swap_target}", key=f"do_swap_{_gk}", type="primary"):
                                            # Find the target group
                                            tgt_group = next((gg for gg in fg if gg.get("housekeeper","") == swap_target and gg.get("inspector","") == insp_name), None)
                                            if tgt_group:
                                                for rm_code in swap_rooms_sel:
                                                    src_room = next((r for r in g["rooms"] if r["room"] == rm_code), None)
                                                    if src_room:
                                                        # Move room from source to target group in session data
                                                        g["rooms"].remove(src_room)
                                                        tgt_group["rooms"].append(src_room)
                                                        # Update status tracking
                                                        old_hk = rs.get(rm_code,{}).get("housekeeper","")
                                                        _save_status(rm_code, {
                                                            "housekeeper": swap_target,
                                                            "group_label": tgt_group.get("label",""),
                                                            "swapped_from": old_hk,
                                                        })
                                                st.success(f"Moved {len(swap_rooms_sel)} room(s) to {swap_target}")
                                                st.rerun()

                        st.markdown("---")

                        # ── Room status rows ─────────────────────────────────
                        # Dedupe rooms within the group (a room can appear twice
                        # if upstream grouping duplicated it) so widget keys and
                        # the displayed list stay unique.
                        _seen_rm = set(); _dedup_rooms = []
                        for r in rooms_in_g:
                            _rc = r.get("room")
                            if _rc in _seen_rm: continue
                            _seen_rm.add(_rc); _dedup_rooms.append(r)
                        for _ridx, r in enumerate(_dedup_rooms):
                            rm = r["room"]
                            _rk = f"{_gk}_{_ridx}_{rm}" # fully unique row key
                            r_state = rs.get(rm, {"status":"pending"})
                            cur_status = r_state.get("status","pending")
                            sm = STATUS_META.get(cur_status, STATUS_META["pending"])

                            # Timestamps (shared formatter; keep the ts[:5] fallback)
                            def _fmt_ts(ts):
                                return _fmt_mtn(ts) or (ts[:5] if ts else "")

                            ts_start = _fmt_ts(r_state.get("started_at",""))
                            ts_clean = _fmt_ts(r_state.get("cleaned_at",""))
                            ts_insp = _fmt_ts(r_state.get("inspected_at",""))
                            ts_ac = _fmt_ts(r_state.get("marked_clean_at",""))

                            pet_icon = " " if r.get("pet") else ""
                            late_html = (f'<span style="font-family:\'DM Mono\',monospace;font-size:.6rem;'
                                         f'color:#f59e0b;background:rgba(245,158,11,.12);border-radius:4px;'
                                         f'padding:1px 5px;margin-left:4px"> {r.get("late_checkout","")}</span>'
                                         if r.get("late_checkout") else "")
                            guest_disp = r.get("guest","")
                            if len(guest_disp) > 22: guest_disp = guest_disp[:21]+"…"

                            # ── Info line: room + guest + animated status pill ──
                            _is_active = (cur_status == "cleaning_started")
                            _dot = (f'<span style="display:inline-block;width:6px;height:6px;'
                                    f'border-radius:50%;background:{sm["color"]};margin-right:5px;'
                                    f'vertical-align:middle;'
                                    + ('animation:pulseDot 1.1s ease-in-out infinite;' if _is_active else '')
                                    + '"></span>')
                            _ring = 'animation:statusPop .35s cubic-bezier(.34,1.56,.64,1) both' \
                                    + (',ringPulse 2s ease-out infinite' if _is_active else '')
                            st.markdown(
                                f'<div style="display:flex;align-items:center;justify-content:space-between;'
                                f'gap:8px;flex-wrap:wrap;padding:7px 10px;background:rgba(255,255,255,.02);'
                                f'border:1px solid rgba(99,102,241,.1);border-radius:8px;margin-bottom:4px">'
                                f' <div style="display:flex;align-items:center;gap:8px;min-width:0;flex:1">'
                                f' <span style="font-family:\'DM Mono\',monospace;font-size:.82rem;'
                                f'font-weight:600;color:#6366f1;white-space:nowrap">{rm}</span>'
                                f' <span style="font-size:.76rem;color:#94a3b8;overflow:hidden;'
                                f'text-overflow:ellipsis;white-space:nowrap">{guest_disp}{pet_icon}</span>'
                                f' {late_html}'
                                f' </div>'
                                f' <div style="background:{sm["bg"]};border:1px solid {sm["border"]};'
                                f'border-radius:6px;padding:3px 9px;font-size:.68rem;font-weight:600;'
                                f'color:{sm["color"]};white-space:nowrap;flex-shrink:0;{_ring}">'
                                f'{_dot}{sm["icon"]} {sm["label"]}</div>'
                                f'</div>', unsafe_allow_html=True)

                            # ── Action button row — 4 equal columns, stays horizontal ──
                            bc1, bc2, bc3, bc4 = st.columns(4)
                            with bc1:
                                if cur_status == "pending":
                                    if st.button("Clean", key=f"ac_{_rk}", use_container_width=True):
                                        _save_status(rm, {"status":"already_clean","marked_clean_at":_NOW()})
                                        st.session_state["_live_toast"] = f" {rm} marked Already Clean"
                                        st.rerun()
                                elif cur_status == "already_clean":
                                    if st.button("Undo", key=f"undo_ac_{_rk}", use_container_width=True):
                                        _save_status(rm, {"status":"pending","marked_clean_at":None})
                                        st.session_state["_live_toast"] = f" {rm} back to Pending"
                                        st.rerun()
                            with bc2:
                                if cur_status == "pending":
                                    if st.button("Start", key=f"start_{_rk}", use_container_width=True):
                                        _save_status(rm, {"status":"cleaning_started","started_at":_NOW()})
                                        st.session_state["_live_toast"] = f" {rm} — cleaning started"
                                        st.rerun()
                                elif cur_status == "cleaning_started":
                                    if st.button("Done", key=f"done_{_rk}", use_container_width=True):
                                        _save_status(rm, {"status":"cleaning_done","cleaned_at":_NOW()})
                                        st.session_state["_live_toast"] = f" {rm} — cleaning done, awaiting inspection"
                                        st.rerun()
                            with bc3:
                                if cur_status == "cleaning_done":
                                    if st.button("Inspect", key=f"insp_{_rk}", use_container_width=True):
                                        _save_status(rm, {"status":"inspected","inspected_at":_NOW()})
                                        st.session_state["_live_toast"] = f" {rm} inspected "
                                        st.rerun()
                                elif cur_status == "inspected":
                                    st.markdown(f'<div style="font-family:\'DM Mono\',monospace;font-size:.64rem;color:#a78bfa;padding:8px 0;text-align:center"> {ts_insp}</div>', unsafe_allow_html=True)
                            with bc4:
                                if cur_status != "pending":
                                    if st.button("Reset", key=f"reset_{_rk}", use_container_width=True, help="Reset to Pending"):
                                        _save_status(rm, {
                                            "status":"pending",
                                            "started_at":None,"cleaned_at":None,
                                            "inspected_at":None,"marked_clean_at":None
                                        })
                                        st.rerun()

                            # Timestamp trail under each room
                            ts_parts = []
                            if ts_ac: ts_parts.append(f" {ts_ac}")
                            if ts_start: ts_parts.append(f" {ts_start}")
                            if ts_clean: ts_parts.append(f" {ts_clean}")
                            if ts_insp: ts_parts.append(f" {ts_insp}")
                            if ts_parts:
                                st.markdown(
                                    f'<div style="font-family:\'DM Mono\',monospace;font-size:.6rem;'
                                    f'color:#334155;padding:0 0 4px 8px;letter-spacing:.04em">'
                                    f'{" · ".join(ts_parts)}</div>',
                                    unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════════
    st.markdown("---")
    # Service ordering for the download: Full Clean first, then IH, Daily Service,
    # Dust n Vac; uncertain rooms after those; stayover/verify rooms dead last.
    _SVC_ORDER = {SVC_FC:0, SVC_IH:1, SVC_DS:2, SVC_DV:3}
    export_rows=[]
    for g in fg:
        is_verify = g.get("verify_group", False)
        svc_rank = _SVC_ORDER.get(g.get("service_type",""), 4)
        for r in g["rooms"]:
            _guest = r.get("guest","")
            # "Unallocated" (and blank) rooms may already be clean — flag them so
            # they drop to the bottom of the download for manual review/assignment.
            _is_unalloc = str(_guest).strip().lower() in ("unallocated","---","")
            _svc_type = g.get("service_type","")
            # For Unallocated FULL CLEAN / IH rooms we don't pre-assign anyone —
            # leave HSKP and RQS blank so they can be assigned manually later.
            # Dust n Vac keeps its RQS 2, and Daily Service keeps the app's
            # assignment, as usual.
            _blank_assign = _is_unalloc and _svc_type in (SVC_FC, SVC_IH)
            _hskp = "" if (is_verify or _blank_assign) else g.get("housekeeper","")
            _rqs  = "" if (is_verify or _blank_assign) else g.get("inspector","")
            # Fold the late-checkout into Notes so it shows in the download. Format
            # as "Late Out: <time>" and keep any existing room note alongside it.
            _note = (r.get("notes","") or "").strip()
            _lc = (r.get("late_checkout","") or "").strip()
            if _lc:
                _lc_txt = _lc if _lc.lower().startswith("late out") else f"Late Out: {_lc}"
                _note = f"{_lc_txt}" + (f" · {_note}" if _note else "")
            export_rows.append({
                "Room":r.get("room",""),"Service":r.get("service",""),
                "Time (min)":r.get("time",""),"Pet":r.get("pet",""),
                "Current Guest or Status":_guest,
                "HSKP":_hskp,
                "RQS":_rqs,
                # Status is intentionally left BLANK in the downloaded file.
                "Notes":_note,"Status":"",
                "Carpet":"","Stripping":"","Arriving Guest":r.get("arriving",""),
                # kept only for internal sort ordering below (dropped before export)
                "_Group":("VERIFY — assign manually" if is_verify else g["label"]),
                "_Svc":svc_rank,
                "_RQS":(_rqs or "").lower(),
                "_HSKP":(_hskp or "").lower(),
                "_Unalloc":"Yes" if (_is_unalloc and not is_verify) else "No",
                "_Verify":"Yes" if is_verify else "No",
                "_Uncertain":"Yes" if r.get("uncertain") else "No",
            })
    export_df = pd.DataFrame(export_rows)
    # Order, top to bottom:
    #   1) confirmed rooms with a real guest — by service (FC→IH→DS→DV), RQS, HK
    #   2) "Unallocated" rooms (may already be clean) — dropped to the bottom for
    #      manual review, still grouped by service/RQS/HK
    #   3) uncertain rooms
    #   4) stayover / verify rooms (assign manually) — dead last
    if not export_df.empty and "_Verify" in export_df.columns:
        _sk = ["_Svc","_RQS","_HSKP","_Group"]
        base = export_df[(export_df["_Verify"]=="No") & (export_df["_Uncertain"]=="No")]
        normal      = base[base["_Unalloc"]=="No"].sort_values(_sk)
        unallocated = base[base["_Unalloc"]=="Yes"].sort_values(_sk)
        unconfirmed = export_df[(export_df["_Verify"]=="No") & (export_df["_Uncertain"]=="Yes")].sort_values(_sk)
        verify_rows = export_df[export_df["_Verify"]=="Yes"].sort_values("_Group")
        export_df   = pd.concat([normal,unallocated,unconfirmed,verify_rows],ignore_index=True)
    # Drop the internal sort-only helper columns so the file has exactly the
    # requested columns, in order.
    _EXPORT_COLS = ["Room","Service","Time (min)","Pet","Current Guest or Status",
                    "HSKP","RQS","Notes","Status","Carpet","Stripping","Arriving Guest"]
    if not export_df.empty:
        export_df = export_df[[c for c in _EXPORT_COLS if c in export_df.columns]]

    # ── Build a formatted Excel workbook: the 12-column schedule up top, a 7-row
    # gap, then a pivot summary (RQS HSKP rooms, with subtotals and a grand
    # total) matching the reference layout. ──────────────────────────────────
    def _build_excel(main_df, rows_for_pivot, free_staff=None):
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from collections import OrderedDict

        def _to_int(v):
            try: return int(float(v))
            except: return 0

        wb = Workbook(); ws = wb.active; ws.title = "Schedule"
        FONT = "Arial"
        reg = Font(name=FONT, size=10)
        bold = Font(name=FONT, size=10, bold=True)
        hdr_f = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        hdr_fl = PatternFill("solid", fgColor="2563A8")
        sub_fl = PatternFill("solid", fgColor="EEF0F3")
        rqs_fl = PatternFill("solid", fgColor="D9E1EC")
        thin = Side(style="thin", color="D0D5DD")
        border = Border(bottom=thin)

        # ── Main table ───────────────────────────────────────────────────────
        cols = list(main_df.columns)
        for ci, h in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=h); c.font = hdr_f; c.fill = hdr_fl
            c.alignment = Alignment(horizontal="left", vertical="center")
        r_i = 2
        for _, row in main_df.iterrows():
            for ci, h in enumerate(cols, 1):
                val = row[h]
                if h == "Time (min)":
                    val = _to_int(val) if str(val) != "" else ""
                ws.cell(row=r_i, column=ci, value=val).font = reg
            r_i += 1
        main_end = r_i - 1

        # Column widths for the main table
        widths = {"Room":10,"Service":16,"Time (min)":11,"Pet":6,
                  "Current Guest or Status":30,"HSKP":16,"RQS":14,"Notes":26,
                  "Status":12,"Carpet":9,"Stripping":10,"Arriving Guest":18}
        for ci, h in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 14)

        # ── Staff balance summary (no pivot — done manually after edits) ──────
        # A single table showing who is LIGHT (spare capacity) or HEAVY (near the
        # cap) so you can rebalance by hand: housekeepers by minutes, RQS by rooms.
        summary_start = main_end + 7 + 1   # leave exactly 7 blank rows
        pr = summary_start
        lw_hdr = Font(name=FONT, size=11, bold=True, color="16202E")
        # thresholds (mirrors the values used to build free_staff below)
        RQS_LOW_ROOMS_XL = 8

        if free_staff:
            ws.cell(row=pr, column=1, value="Staff with Spare Capacity").font = lw_hdr
            pr += 1
            ws.cell(row=pr, column=1, value=(
                f"Housekeepers under {LOW_MIN} min   \u00b7   "
                f"RQS under {RQS_LOW_ROOMS_XL} rooms   \u00b7   plus anyone free"
            )).font = reg
            pr += 1
            FS = ["Staff","Role","Status","Load","Rooms","Room List","Service"]
            for ci, h in enumerate(FS, 1):
                c = ws.cell(row=pr, column=ci, value=h); c.font = hdr_f; c.fill = hdr_fl
            pr += 1
            free_fill = PatternFill("solid", fgColor="EAF7EE")   # green = free
            low_fill2 = PatternFill("solid", fgColor="FFF4E5")   # amber = light
            for row in free_staff:
                status = row.get("status","")
                ws.cell(row=pr, column=1, value=row.get("name","")).font = reg
                ws.cell(row=pr, column=2, value=row.get("role","")).font = reg
                ws.cell(row=pr, column=3, value=status).font = reg
                ws.cell(row=pr, column=4, value=row.get("load","")).font = reg
                ws.cell(row=pr, column=5, value=row.get("rooms","")).font = reg
                ws.cell(row=pr, column=6, value=row.get("room_names","")).font = reg
                ws.cell(row=pr, column=7, value=row.get("service","")).font = reg
                fill = free_fill if status == "Free" else low_fill2
                for ci in range(1, 8): ws.cell(row=pr, column=ci).fill = fill
                pr += 1
        else:
            ws.cell(row=pr, column=1,
                    value="All staff at or above their thresholds.").font = reg

        ws.freeze_panes = "A2"
        buf = BytesIO(); wb.save(buf); return buf.getvalue()

    # Rows for the pivot, in the same confirmeduncertainverify order as the table
    _pivot_rows = export_df.to_dict("records") if not export_df.empty else []

    # ── Build the staff-rebalance list for the Excel (HKs + RQS) ───────────────
    # LIGHT = spare capacity (HK under LOW_MIN minutes, RQS under 8 rooms)
    # FREE  = present but unassigned
    _SVC_SHORT_XL = {SVC_FC:"FC", SVC_IH:"IH", SVC_DS:"DS", SVC_DV:"DV"}
    RQS_LOW_ROOMS_XL = 8
    _free_staff = []
    _hk_load = {}   # name -> {"time":int,"rooms":int,"svcs":set,"names":[...]}
    _rqs_load = {}  # name -> {"rooms":int,"svcs":set,"names":[...]}
    def _rnames(g):
        return [str(r.get("room","")) for r in g["rooms"] if r.get("room")]
    for g in fg:
        hk = g.get("housekeeper","")
        svc = g.get("service_type","")
        if hk and hk != "Manager":
            d = _hk_load.setdefault(hk, {"time":0,"rooms":0,"svcs":set(),"names":[]})
            d["time"] += g["time"]; d["rooms"] += len(g["rooms"]); d["svcs"].add(svc)
            d["names"].extend(_rnames(g))
        insp = g.get("inspector","")
        if insp:
            d = _rqs_load.setdefault(insp, {"rooms":0,"svcs":set(),"names":[]})
            d["rooms"] += len(g["rooms"]); d["svcs"].add(svc)
            d["names"].extend(_rnames(g))
    def _svcs_txt(svcs):
        return " ".join(_SVC_SHORT_XL[s] for s in [SVC_FC,SVC_IH,SVC_DS,SVC_DV] if s in svcs) or "—"
    def _names_txt(names):
        return ", ".join(sorted(names)) if names else ""
    # Housekeepers — light (under LOW_MIN minutes), lightest first
    for hk, d in sorted(_hk_load.items(), key=lambda kv:kv[1]["time"]):
        if d["time"] and d["time"] < LOW_MIN:
            _free_staff.append({"name":hk,"role":"HK","status":"Light","load":f'{d["time"]}m',
                                "rooms":d["rooms"],"room_names":_names_txt(d["names"]),
                                "service":_svcs_txt(d["svcs"]),"free":False})
    for hk in sorted(n for n in present_hk if n not in used_hk_set):
        _free_staff.append({"name":hk,"role":"HK","status":"Free","load":"Free",
                            "rooms":0,"room_names":"","service":"—","free":True})
    # RQS — light (fewer than 8 rooms), fewest first
    for insp, d in sorted(_rqs_load.items(), key=lambda kv:kv[1]["rooms"]):
        if d["rooms"] < RQS_LOW_ROOMS_XL:
            _free_staff.append({"name":insp,"role":"RQS","status":"Light","load":f'{d["rooms"]} rm',
                                "rooms":d["rooms"],"room_names":_names_txt(d["names"]),
                                "service":_svcs_txt(d["svcs"]),"free":False})
    for insp in sorted(n for n in present_insp if n not in _rqs_load):
        _free_staff.append({"name":insp,"role":"RQS","status":"Free","load":"Free",
                            "rooms":0,"room_names":"","service":"—","free":True})

    try:
        xlsx_bytes = _build_excel(export_df, _pivot_rows, free_staff=_free_staff)
        st.download_button("Download Excel", data=xlsx_bytes,
                           file_name="cleaning_schedule.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as _xl_ex:
        # Fall back to CSV if Excel generation ever fails, so the download button
        # is never dead.
        csv = export_df.to_csv(index=False).encode("utf-8-sig")
        st.download_button("Download CSV", data=csv,
                           file_name="cleaning_schedule.csv", mime="text/csv")
        st.caption(f"(Excel export unavailable: {_xl_ex})")
